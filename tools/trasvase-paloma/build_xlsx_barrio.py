import csv, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
p2b=json.load(open('Bases de datos/output_ciudades/bogota/bog-puesto-to-barrio.json'))
geo=json.load(open('/tmp/bogbar.json'))
bmeta={f['properties']['codigo']:(f['properties']['nombre'],f['properties'].get('loc_codigo','99'),f['properties'].get('loc_nombre','')) for f in geo['features']}
CMAP=[('Abelardo De La Espriella','ab'),('Iván Cepeda','ce'),('Paloma Valencia','pa'),('Sergio Fajardo','sf'),
('Santiago Botero','bo'),('Mauricio Lizcano','li'),('Miguel Uribe','mu'),('Sondra Macollins','ma'),
('Roy Barreras','ro'),('Gilberto Murillo','gm'),('Carlos Caicedo','ca'),('Gustavo Matamoros','mt')]
agg={}
with open('Bases de datos/nuevos archivos 1v 2026/PRECONTEO_1V_2026_MESA_nombres_corregidos.csv',encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        if r['cod_departamento']!='16': continue
        try: key=f"{int(r['zona']):02d}-{int(r['puesto']):02d}"
        except: continue
        bc=p2b.get(key)
        if not bc: continue
        a=agg.setdefault(bc,{k:0 for _,k in CMAP}|{'bl':0,'nu':0,'nm':0,'urna':0})
        for col,k in CMAP: a[k]+=int(r[col]or 0)
        a['bl']+=int(r['votos_blanco']or 0); a['nu']+=int(r['votos_nulos']or 0); a['nm']+=int(r['votos_no_marcados']or 0); a['urna']+=int(r['total_votos_urna']or 0)
for bc,a in agg.items(): a['cl']=a['urna']-(sum(a[k] for _,k in CMAP)+a['bl']+a['nu']+a['nm'])
rows=[(bmeta.get(bc,('?','99',''))[1],bmeta.get(bc,('?','99',''))[2],bc,bmeta.get(bc,('?',))[0],a) for bc,a in agg.items() if a['urna']>0]
rows.sort(key=lambda x:(str(x[0] or '99'), -x[4]['urna']))

LEFT=[('Localidad','loc'),('Código','cod'),('Barrio','bar'),('Votantes','urna')]
ITEMS=[('Abelardo','ab'),('Cepeda','ce'),('Paloma','pa'),('Fajardo','sf'),('Claudia López*','cl'),('Botero','bo'),
('Lizcano','li'),('Miguel Uribe','mu'),('Macollins','ma'),('Roy Barreras','ro'),('Murillo','gm'),('Caicedo','ca'),
('Matamoros','mt'),('Voto blanco','bl'),('Votos nulos','nu'),('No marcados','nm')]
COLS=[(t,'L') for t,_ in LEFT]
for t,k in ITEMS: COLS+=[(t,'V'),(t+' %','P')]
COLS.append(('Ganador','G'))
idx={c[0]:i+1 for i,c in enumerate(COLS)}
URNA_L=get_column_letter(idx['Votantes']); AB_L=get_column_letter(idx['Abelardo']); CE_L=get_column_letter(idx['Cepeda'])
itemkey={t:k for t,k in ITEMS}
wb=Workbook(); ws=wb.active; ws.title="Bogotá 1V por barrio"
ws.append([c[0] for c in COLS])
for c in range(1,len(COLS)+1):
    cell=ws.cell(1,c); cell.font=Font(name='Arial',bold=True,color="FFFFFF",size=8.5)
    cell.fill=PatternFill("solid",fgColor="22324d"); cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
R=2
for loc,locn,cod,bar,a in rows:
    for i,(t,typ) in enumerate(COLS,1):
        L=get_column_letter(i)
        if typ=='L':
            if t=='Localidad': ws.cell(R,i,locn or loc)
            elif t=='Código': ws.cell(R,i,cod)
            elif t=='Barrio': ws.cell(R,i,bar)
            elif t=='Votantes': ws.cell(R,i,a['urna'])
        elif typ=='V': ws.cell(R,i,a[itemkey[t[:-0] if False else t]])
        elif typ=='P': vb=get_column_letter(i-1); ws.cell(R,i,f"=IF({URNA_L}{R}=0,0,{vb}{R}/{URNA_L}{R})")
        elif typ=='G': ws.cell(R,i,f'=IF({AB_L}{R}>{CE_L}{R},"Abelardo","Cepeda")')
    R+=1
last=R-1
# total
for i,(t,typ) in enumerate(COLS,1):
    L=get_column_letter(i)
    if t=='Localidad': ws.cell(R,i,"BOGOTÁ (barrios mapeados)")
    elif typ=='V' or t=='Votantes': ws.cell(R,i,f"=SUM({L}2:{L}{last})")
    elif typ=='P': vb=get_column_letter(i-1); ws.cell(R,i,f"=IF({URNA_L}{R}=0,0,{vb}{R}/{URNA_L}{R})")
    elif typ=='G': ws.cell(R,i,f'=IF({AB_L}{R}>{CE_L}{R},"Abelardo","Cepeda")')
total_row=R
thin=Side(style='thin',color="EEEEEE")
for r in range(2,R+1):
    for i,(t,typ) in enumerate(COLS,1):
        cell=ws.cell(r,i); cell.font=Font(name='Arial',size=8.5,bold=(r==total_row)); cell.border=Border(bottom=thin)
        if r==total_row: cell.fill=PatternFill("solid",fgColor="EDEAF2")
        if typ=='V' or t=='Votantes': cell.number_format='#,##0'; cell.alignment=Alignment(horizontal='right')
        if typ=='P': cell.number_format='0.0%'; cell.alignment=Alignment(horizontal='right')
        if typ=='G': cell.alignment=Alignment(horizontal='center')
for i,(t,typ) in enumerate(COLS,1):
    w=20 if t=='Barrio' else 16 if t=='Localidad' else 8 if t=='Código' else 10 if t=='Votantes' else 11 if t=='Ganador' else 7 if typ=='P' else 9
    ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="D2"; ws.row_dimensions[1].height=40; ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}{last}"
ws2=wb.create_sheet("Fuentes y método")
for n in [["Bogotá — 1ª vuelta 2026 por barrio catastral",""],["",""],
["Fuente:","Preconteo Registraduría 1ª vuelta por mesa + mapeo puesto→barrio (bog-puesto-to-barrio) + cartografía catastral IDECA (1.001 barrios)."],
["Cobertura:","557 barrios con puesto de votación; cubren el 95,6% de los votos de Bogotá. Los barrios sin puesto (sus votantes votan en un puesto vecino) no aparecen."],
["% de cada columna:","Votos / total de votos en urna del barrio."],
["* Claudia López:","Recuperada del residual del preconteo (su columna llegó en 0; la urna sí la incluye). Ver detalle en el Excel por localidad."],
["Límite:","Un puesto sirve a varios barrios vecinos pero se asigna al barrio donde está físicamente. Preconteo preliminar."],
["Mapa:","ricardoruiz.co/bogota-1v-barrios.html"]]:
    ws2.append(n)
ws2.cell(1,1).font=Font(name='Arial',bold=True,size=12)
for r in range(3,8):
    ws2.cell(r,1).font=Font(name='Arial',bold=True,size=10); ws2.cell(r,1).alignment=Alignment(vertical='top')
    ws2.cell(r,2).font=Font(name='Arial',size=10); ws2.cell(r,2).alignment=Alignment(wrap_text=True,vertical='top')
ws2.column_dimensions['A'].width=20; ws2.column_dimensions['B'].width=100
out="Bases de datos/output_trasvase/Bogota-1V-2026-por-barrio.xlsx"; wb.save(out)
print("guardado:",out,"| barrios:",last-1,"| columnas:",len(COLS))
