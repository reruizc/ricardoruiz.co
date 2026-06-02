import csv, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
def ni(x):
    x=str(x).strip(); return int(x) if x.isdigit() else -1
def pk(d,m,z,p): return f"{d}-{m}-{ni(z)}-{ni(p)}"
cam=json.load(open('/tmp/cam16_puestos.json')); loc={}
for p in cam: loc[pk(p['dep_cod'],p['mun_cod'],p['zon_cod'],p['pue_cod_raw'])]=p['com_cod']
LOCS=[f'{i:03d}' for i in range(1,21)]
def bucket(zona,puesto):
    lc=loc.get(pk('16','001',zona,puesto),'000')
    if lc in LOCS: return lc
    z=ni(zona)
    return 'C90' if z==90 else 'C98' if z==98 else 'C99'
censo=json.load(open('Bases de datos/censos-puesto-2026.json'))['porPuesto']; cens={}
for k,v in censo.items():
    if not k.startswith('16-001-'): continue
    d,m,z,pp=k.split('-'); b=bucket(z,pp); cens[b]=cens.get(b,0)+v
CMAP=[('Abelardo De La Espriella','ab'),('Iván Cepeda','ce'),('Paloma Valencia','pa'),('Sergio Fajardo','sf'),
('Santiago Botero','bo'),('Mauricio Lizcano','li'),('Miguel Uribe','mu'),('Sondra Macollins','ma'),
('Roy Barreras','ro'),('Gilberto Murillo','gm'),('Carlos Caicedo','ca'),('Gustavo Matamoros','mt')]
agg={}
with open('Bases de datos/nuevos archivos 1v 2026/PRECONTEO_1V_2026_MESA_nombres_corregidos.csv',encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        if r['cod_departamento']!='16': continue
        b=bucket(r['zona'],r['puesto'])
        a=agg.setdefault(b,{k:0 for _,k in CMAP}|{'bl':0,'nu':0,'nm':0,'urna':0})
        for col,k in CMAP: a[k]+=int(r[col]or 0)
        a['bl']+=int(r['votos_blanco']or 0); a['nu']+=int(r['votos_nulos']or 0)
        a['nm']+=int(r['votos_no_marcados']or 0); a['urna']+=int(r['total_votos_urna']or 0)
for b,a in agg.items():
    a['cl']=a['urna']-(sum(a[k] for _,k in CMAP)+a['bl']+a['nu']+a['nm'])
NAMES={'001':'Usaquén','002':'Chapinero','003':'Santa Fe','004':'San Cristóbal','005':'Usme','006':'Tunjuelito',
'007':'Bosa','008':'Kennedy','009':'Fontibón','010':'Engativá','011':'Suba','012':'Barrios Unidos','013':'Teusaquillo',
'014':'Los Mártires','015':'Antonio Nariño','016':'Puente Aranda','017':'La Candelaria','018':'Rafael Uribe Uribe',
'019':'Ciudad Bolívar','020':'Sumapaz','C90':'Corferias / Puesto Censo (z.90)','C98':'Cárceles (z.98)','C99':'Otros especiales'}
order=LOCS+['C90','C98']+(['C99'] if 'C99' in agg else [])

# columnas
LEFT=[('#','num'),('Localidad','text'),('Censo electoral','censo'),('Votantes','urna'),('Participación','pctpart')]
ITEMS=[('Abelardo','ab'),('Cepeda','ce'),('Paloma','pa'),('Fajardo','sf'),('Claudia López*','cl'),('Botero','bo'),
('Lizcano','li'),('Miguel Uribe','mu'),('Macollins','ma'),('Roy Barreras','ro'),('Murillo','gm'),('Caicedo','ca'),
('Matamoros','mt'),('Voto blanco','bl'),('Votos nulos','nu'),('No marcados','nm')]
COLS=[]
for t,k in LEFT: COLS.append((t,k,'L'))
for t,k in ITEMS: COLS.append((t,k,'V')); COLS.append((t+' %','','P'))
COLS.append(('Ganador','','G'))
idx={c[0]:i+1 for i,c in enumerate(COLS)}
CENSO_L=get_column_letter(idx['Censo electoral']); URNA_L=get_column_letter(idx['Votantes'])
AB_L=get_column_letter(idx['Abelardo']); CE_L=get_column_letter(idx['Cepeda'])

wb=Workbook(); ws=wb.active; ws.title="Bogotá 1V por localidad"
ws.append([c[0] for c in COLS])
hdr=PatternFill("solid",fgColor="22324d")
for c in range(1,len(COLS)+1):
    cell=ws.cell(1,c); cell.font=Font(name='Arial',bold=True,color="FFFFFF",size=8.5)
    cell.fill=hdr; cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
def writerow(R,b=None,total=False):
    for i,(t,k,typ) in enumerate(COLS,1):
        L=get_column_letter(i)
        if typ=='L':
            if t=='#': ws.cell(R,i, (int(b) if b in LOCS else '—') if not total else '')
            elif t=='Localidad': ws.cell(R,i,'BOGOTÁ (TOTAL)' if total else NAMES.get(b,b))
            elif typ=='L' and t=='Censo electoral':
                ws.cell(R,i, f"=SUM({L}2:{L}{last})" if total else cens.get(b,0))
            elif t=='Votantes':
                ws.cell(R,i, f"=SUM({L}2:{L}{last})" if total else agg[b]['urna'])
            elif t=='Participación':
                ws.cell(R,i, f"=IF({CENSO_L}{R}=0,0,{URNA_L}{R}/{CENSO_L}{R})")
        elif typ=='V':
            ws.cell(R,i, f"=SUM({L}2:{L}{last})" if total else agg[b][k])
        elif typ=='P':
            vb=get_column_letter(i-1); ws.cell(R,i, f"=IF({URNA_L}{R}=0,0,{vb}{R}/{URNA_L}{R})")
        elif typ=='G':
            ws.cell(R,i, f'=IF({AB_L}{R}>{CE_L}{R},"Abelardo","Cepeda")')
row=2
for b in order:
    if b not in agg: continue
    writerow(row,b); row+=1
last=row-1
writerow(row,total=True); total_row=row
# formato
thin=Side(style='thin',color="ECECEC")
for r in range(2,row+1):
    for i,(t,k,typ) in enumerate(COLS,1):
        cell=ws.cell(r,i); cell.font=Font(name='Arial',size=8.5,bold=(r==total_row)); cell.border=Border(bottom=thin)
        if r==total_row: cell.fill=PatternFill("solid",fgColor="EDEAF2")
        if typ=='V' or (typ=='L' and t in('Censo electoral','Votantes')): cell.number_format='#,##0'; cell.alignment=Alignment(horizontal='right')
        if typ=='P' or (typ=='L' and t=='Participación'): cell.number_format='0.0%'; cell.alignment=Alignment(horizontal='right')
        if typ=='G': cell.alignment=Alignment(horizontal='center')
for i,(t,k,typ) in enumerate(COLS,1):
    w=22 if t=='Localidad' else 4 if t=='#' else 12 if t in('Censo electoral','Votantes') else 11 if t=='Ganador' else 7 if typ=='P' else 10
    ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="C2"; ws.row_dimensions[1].height=40

ws2=wb.create_sheet("Fuentes y método")
notes=[
["Primera vuelta presidencial 2026 — Bogotá por localidad (votación bruta y %)",""],["",""],
["Fuente votos:","Preconteo Registraduría 1ª vuelta (31-may-2026), por mesa — snapshot final de la noche (PRELIMINAR)."],
["Fuente censo:","Censo electoral por puesto 2026 (Divipole / COMUNAS_DATA)."],
["Puesto→localidad:","Escrutinio Cámara 2026 Bogotá (com_cod/com_nom por puesto)."],
["Unidad:","Mesa de votación, agregada a localidad. Bogotá = depto electoral 16, municipio 001."],
["Participación:","Votantes (= total de votos en urna) / censo electoral de la localidad."],
["% de cada columna:","Votos de la columna / total de votos en urna (la urna incluye blanco, nulos y no marcados)."],
["* Claudia López:","En el preconteo por mesa su columna llegó en CERO (error de mapeo). PERO el total de urna SÍ la incluye: nacional, urna − suma de columnas = 225.287 EXACTO = su votación oficial, sin una sola mesa negativa. Por eso participación y % de los demás NO están sesgados. Aquí se RECUPERA como ese residual por localidad (Bogotá ≈ 102.433; cifra por municipio más completa 106.420; diferencia ~96% = cobertura preliminar)."],
["Consistencia:","Por construcción, la suma de los 13 candidatos + blanco + nulos + no marcados = Votantes en cada fila."],
["Corferias / Puesto Censo (z.90):","Zona electoral 90. Censo muy alto y baja participación: son registros que mayoritariamente votan en otro lado."],
["Cárceles (z.98):","Zona electoral 98 (votación en establecimientos penitenciarios)."],
["Cobertura:","Preconteo preliminar; pueden faltar mesas. Los % son representativos; los conteos absolutos son un piso."],
["Análisis completo:","ricardoruiz.co/trasvase-paloma.html · ricardoruiz.co"],
]
for n in notes: ws2.append(n)
ws2.cell(1,1).font=Font(name='Arial',bold=True,size=12)
for r in range(3,len(notes)+1):
    ws2.cell(r,1).font=Font(name='Arial',bold=True,size=10); ws2.cell(r,1).alignment=Alignment(vertical='top')
    ws2.cell(r,2).font=Font(name='Arial',size=10); ws2.cell(r,2).alignment=Alignment(wrap_text=True,vertical='top')
ws2.column_dimensions['A'].width=26; ws2.column_dimensions['B'].width=98
out="Bases de datos/output_trasvase/Bogota-1V-2026-por-localidad.xlsx"; wb.save(out); print("guardado:",out,"| columnas:",len(COLS),"| filas datos:",last-1)
