#!/usr/bin/env python3
"""
Unifica los 32 CSV del escrutinio 1V 2026 (MMV por mesa, uno por depto)
en un solo CSV agregado POR PUESTO + un Excel amigable.

Entradas:  Bases de datos/ESCRUTINIO-1V/MMV_*.csv
           (DEP;DEPNOMBRE;MUN;MUNNOMBRE;ZONA;PUESTO;PUESNOMBRE;MESA;
            COMUCODIGO;COMUNOMBRE;CORCODIGO;CORNOMBRE;CIR;PAR;PARNOMBRE;
            CAN;CANCEDULA;CANNOMBRE;VOTOS)
Salidas:   Bases de datos/ESCRUTINIO-1V/ESCRUTINIO_1V_2026_PUESTO.csv
           Bases de datos/ESCRUTINIO-1V/Resultados_Escrutinio_1V_2026_por_puesto.xlsx

Notas:
- Las filas vienen sparse (solo candidatos con votos en la mesa);
  el agregado rellena con 0.
- Cobertura: 33 deptos del escrutinio oficial (118.350 mesas, 100% según
  el consolidado del sitio de escrutinios). Exterior/Consulados (88) NO
  fue publicado por la RNEC en escrutiniospresidente2026 (la comisión
  1033 está vacía y el MMV NACIONAL figura publicado=0), así que se
  integra desde el PRECONTEO por mesa (con Claudia recuperada). La
  columna `fuente` distingue escrutinio vs preconteo por fila.
- Murillo y Caicedo (renunciaron antes de 1V) tienen votos en el
  preconteo del exterior pero NO existen en el escrutinio: se excluyen
  y se reporta el monto excluido.
- Nombres de consulados (país + puesto): lookup desde los MMV declarados
  de Congreso (DEPTOS_DECLARADOS/MMV_XXX_88_*.csv, mismos códigos).
"""
import csv, glob, os, sys
from collections import defaultdict

BASE = '/Users/ricardoruiz/ricardoruiz.co/Bases de datos/ESCRUTINIO-1V'
PRECONTEO = '/Users/ricardoruiz/ricardoruiz.co/Bases de datos/nuevos archivos 1v 2026/PRECONTEO_1V_2026_MESA_con_Claudia.csv'
DECLARADOS_88 = '/Users/ricardoruiz/ricardoruiz.co/Bases de datos/DEPTOS_DECLARADOS'
OUT_CSV  = os.path.join(BASE, 'ESCRUTINIO_1V_2026_PUESTO.csv')
OUT_XLSX = os.path.join(BASE, 'Resultados_Escrutinio_1V_2026_por_puesto.xlsx')

# Candidatos en orden de votación nacional (escrutinio) + especiales
CANDS = [
    ('004', 'Abelardo De La Espriella'),
    ('001', 'Iván Cepeda Castro'),
    ('011', 'Paloma Valencia Laserna'),
    ('012', 'Sergio Fajardo Valderrama'),
    ('002', 'Claudia López'),
    ('003', 'Raúl Santiago Botero Jaramillo'),
    ('005', 'Óscar Mauricio Lizcano Arango'),
    ('006', 'Miguel Uribe Londoño'),
    ('007', 'Sondra Macollins Garvin Pinto'),
    ('008', 'Roy Leonardo Barreras Montealegre'),
    ('010', 'Gustavo Matamoros Camacho'),
]
SPECIALS = [('996', 'Votos en blanco'), ('997', 'Votos nulos'), ('998', 'Votos no marcados')]
ALL_CODES = [c for c, _ in CANDS] + [c for c, _ in SPECIALS]

# Columna del preconteo (header con nombres corregidos) → código CAN.
# Carlos Caicedo y Gilberto Murillo renunciaron: None = excluir.
PRECONTEO_MAP = {
    'Iván Cepeda': '001', 'Claudia López': '002', 'Santiago Botero': '003',
    'Abelardo De La Espriella': '004', 'Mauricio Lizcano': '005',
    'Miguel Uribe': '006', 'Sondra Macollins': '007', 'Roy Barreras': '008',
    'Carlos Caicedo': None, 'Gustavo Matamoros': '010',
    'Paloma Valencia': '011', 'Sergio Fajardo': '012', 'Gilberto Murillo': None,
    'votos_blanco': '996', 'votos_nulos': '997', 'votos_no_marcados': '998',
}

def load_nombres_exterior():
    """(mun, zona, puesto) → (país, nombre puesto) desde MMV declarados de Congreso."""
    look = {}
    for f in glob.glob(os.path.join(DECLARADOS_88, 'MMV_XXX_88_*.csv')):
        with open(f, encoding='utf-8-sig', newline='') as fh:
            rd = csv.reader(fh, delimiter=';')
            next(rd)
            for row in rd:
                if len(row) < 8:
                    continue
                key = (row[2], row[4], row[5])
                if key not in look:
                    look[key] = (row[3], row[6])
    return look

def agg_exterior_preconteo():
    """Agrega dep 88 del preconteo por puesto. Devuelve (rows_dict, excluidos)."""
    votos = defaultdict(lambda: defaultdict(int))
    mesas = defaultdict(set)
    excluidos = 0
    with open(PRECONTEO, encoding='utf-8-sig', newline='') as fh:
        rd = csv.DictReader(fh)
        cand_cols = [c for c in rd.fieldnames if c in PRECONTEO_MAP]
        for row in rd:
            if row['cod_departamento'] != '88':
                continue
            key = (row['cod_municipio'], row['zona'], row['puesto'])
            mesas[key].add(row['num_mesa'])
            for col in cand_cols:
                v = int(row[col] or 0)
                code = PRECONTEO_MAP[col]
                if code is None:
                    excluidos += v
                else:
                    votos[key][code] += v
    return votos, mesas, excluidos

def title_es(s):
    """Title-case suave para nombres en MAYÚSCULAS, respetando conectores."""
    minor = {'DE', 'DEL', 'LA', 'LAS', 'LOS', 'Y', 'EL', 'D.C.'}
    out = []
    for i, w in enumerate(s.split()):
        if w in minor and i > 0:
            out.append(w.lower() if w != 'D.C.' else w)
        else:
            out.append(w.capitalize())
    return ' '.join(out)

def main():
    files = sorted(glob.glob(os.path.join(BASE, 'MMV_*.csv')))
    print(f'{len(files)} archivos MMV')
    # key puesto = (dep, mun, zona, puesto)
    votos = defaultdict(lambda: defaultdict(int))   # key -> can -> votos
    mesas = defaultdict(set)                         # key -> set(mesa)
    meta  = {}                                       # key -> (depnom, munnom, puesnom, comucod, comunom)
    deps_seen = {}

    for f in files:
        with open(f, encoding='utf-8-sig', newline='') as fh:
            rd = csv.reader(fh, delimiter=';')
            header = next(rd)
            for row in rd:
                if len(row) < 19:
                    continue
                dep, depnom, mun, munnom, zona, pue, puenom, mesa = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
                comucod, comunom = row[8], row[9]
                can, vot = row[15], row[18]
                key = (dep, mun, zona, pue)
                try:
                    v = int(vot)
                except ValueError:
                    continue
                votos[key][can] += v
                mesas[key].add(mesa)
                if key not in meta:
                    meta[key] = (depnom, munnom, puenom, comucod, comunom)
                deps_seen[dep] = depnom

    print(f'{len(votos)} puestos · {sum(len(m) for m in mesas.values())} mesas · {len(deps_seen)} deptos')

    # sanity: candidatos no mapeados
    unmapped = set()
    for key, cv in votos.items():
        for c in cv:
            if c not in ALL_CODES:
                unmapped.add(c)
    if unmapped:
        print('⚠️ códigos de candidato no mapeados:', unmapped)
        sys.exit(1)

    header = (['cod_departamento', 'departamento', 'cod_municipio', 'municipio',
               'zona', 'cod_puesto', 'puesto', 'cod_comuna', 'comuna', 'mesas', 'fuente']
              + [n for _, n in CANDS]
              + [n for _, n in SPECIALS]
              + ['total_votos'])

    rows = []
    for key in sorted(votos.keys()):
        dep, mun, zona, pue = key
        depnom, munnom, puenom, comucod, comunom = meta[key]
        cv = votos[key]
        vals = [cv.get(c, 0) for c, _ in CANDS] + [cv.get(c, 0) for c, _ in SPECIALS]
        total = sum(vals)
        comunom_clean = '' if comunom.strip().upper() in ('NACIONAL', 'NULL', '') else title_es(comunom.strip())
        rows.append([dep, title_es(depnom), mun, title_es(munnom), zona, pue,
                     puenom.strip(), comucod, comunom_clean, len(mesas[key]), 'escrutinio']
                    + vals + [total])

    # ── Exterior (dep 88) desde el preconteo ──
    ext_votos, ext_mesas, ext_excluidos = agg_exterior_preconteo()
    nombres_ext = load_nombres_exterior()
    sin_nombre = 0
    pais_by_mun = {}
    for (mun, _z, _p), (munnom, _pn) in nombres_ext.items():
        pais_by_mun.setdefault(mun, munnom)
    for key in sorted(ext_votos.keys()):
        mun, zona, pue = key
        munnom, puenom = nombres_ext.get(key, ('', ''))
        if not munnom:
            # puesto nuevo sin homólogo en Congreso: al menos el país
            munnom = pais_by_mun.get(mun, '')
            sin_nombre += 1
        cv = ext_votos[key]
        vals = [cv.get(c, 0) for c, _ in CANDS] + [cv.get(c, 0) for c, _ in SPECIALS]
        total = sum(vals)
        rows.append(['88', 'Exterior (Consulados)', mun, title_es(munnom), zona, pue,
                     puenom.strip(), '', '', len(ext_mesas[key]), 'preconteo']
                    + vals + [total])
    print(f'Exterior: {len(ext_votos)} puestos · {sum(len(m) for m in ext_mesas.values())} mesas '
          f'· {sin_nombre} sin nombre · excluidos Murillo/Caicedo: {ext_excluidos} votos')

    # ── CSV (BOM utf-8, códigos quoted para preservar ceros) ──
    with open(OUT_CSV, 'w', encoding='utf-8-sig', newline='') as fh:
        w = csv.writer(fh, quoting=csv.QUOTE_NONNUMERIC)
        w.writerow(header)
        for r in rows:
            w.writerow(r)
    print(f'CSV → {OUT_CSV} ({len(rows)} filas)')

    # totales de control
    tot_by_cand = [sum(r[11 + i] for r in rows) for i in range(len(CANDS) + len(SPECIALS))]
    for (c, n), t in zip(CANDS + SPECIALS, tot_by_cand):
        print(f'  {n:42s} {t:>12,}')
    print(f'  {"TOTAL":42s} {sum(tot_by_cand):>12,}')

    # ── XLSX ──
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook(write_only=False)

    # Hoja 1 · Instrucciones
    ws = wb.active
    ws.title = 'Instrucciones'
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 110
    ox = Font(bold=True, size=14, color='8A1E16')
    bold = Font(bold=True, size=11)
    body = Font(size=11)
    L = [
        ('Escrutinio · Elección Presidencial 2026 · Primera vuelta (31 de mayo de 2026)', ox),
        ('Resultados oficiales del escrutinio, agregados por puesto de votación.', body),
        ('', None),
        ('Fuente', bold),
        ('Archivos MMV de escrutinio de la Registraduría Nacional del Estado Civil (un archivo por departamento,', body),
        ('nivel mesa), unificados y agregados por puesto por ricardoruiz.co.', body),
        ('', None),
        ('Qué contiene', bold),
        ('Una fila por puesto de votación. Columnas: códigos y nombres de departamento, municipio, zona,', body),
        ('puesto y comuna/localidad (cuando aplica), número de mesas escrutadas del puesto, votos por cada', body),
        ('uno de los 11 candidatos, votos en blanco, nulos, no marcados y total.', body),
        ('', None),
        ('Cobertura de esta versión', bold),
        ('Los 33 departamentos del territorio nacional con el ESCRUTINIO oficial (118.350 mesas, 100% del', body),
        ('consolidado publicado por la RNEC). La circunscripción Exterior/Consulados (cód. 88) NO fue', body),
        ('publicada en el sitio de escrutinios: sus filas vienen del PRECONTEO por mesa de la noche electoral', body),
        ('(dato real por mesa, incluida Claudia López reconstruida del total de urna). La columna "fuente"', body),
        ('indica de dónde sale cada fila: escrutinio | preconteo.', body),
        ('', None),
        ('Notas de lectura', bold),
        ('· Los códigos territoriales son los de la Registraduría (no DANE) y conservan ceros a la izquierda.', body),
        ('· Zona 90 = puesto censo (ej. Corferias) · zona 98 = cárceles. Son agregados especiales sin', body),
        ('  geografía de barrio; inclúyelos o descártalos según tu análisis.', body),
        ('· En el escrutinio solo aparecen los 11 candidatos en contienda. En las filas del exterior', body),
        ('  (preconteo) se excluyeron los votos de Murillo y Caicedo (renunciaron antes de la elección;', body),
        ('  el escrutinio no los contabiliza). En el exterior la zona/puesto puede ser alfanumérica y los', body),
        ('  nombres de consulado provienen del divipol de Congreso 2026 (mismos códigos).', body),
        ('· El % de cada candidato se calcula normalmente sobre votos válidos (candidatos + blanco).', body),
        ('', None),
        ('ricardoruiz.co · análisis electoral · datos: Registraduría Nacional', Font(size=10, italic=True, color='666666')),
    ]
    for i, (txt, ft) in enumerate(L, start=1):
        c = ws.cell(row=i, column=1, value=txt)
        if ft:
            c.font = ft
        c.alignment = Alignment(wrap_text=False, vertical='top')

    # Hoja 2 · Datos
    ws2 = wb.create_sheet('Datos por puesto')
    ws2.append(header)
    for r in rows:
        ws2.append(r)

    n_rows = len(rows) + 1
    n_cols = len(header)
    ref = f'A1:{get_column_letter(n_cols)}{n_rows}'
    tbl = Table(displayName='EscrutinioPuesto', ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name='TableStyleMedium4', showRowStripes=True)
    ws2.add_table(tbl)
    ws2.freeze_panes = 'G2'

    widths = [8, 18, 8, 22, 6, 8, 34, 8, 22, 7, 11] + [14] * (len(CANDS) + len(SPECIALS)) + [12]
    for i, wd in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = wd
    for col in range(12, n_cols + 1):
        for row in range(2, n_rows + 1):
            ws2.cell(row=row, column=col).number_format = '#,##0'

    wb.save(OUT_XLSX)
    print(f'XLSX → {OUT_XLSX}')

if __name__ == '__main__':
    main()
