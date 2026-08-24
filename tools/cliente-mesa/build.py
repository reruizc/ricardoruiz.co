#!/usr/bin/env python3
# Genera, a partir del preconteo por mesa con nombres corregidos:
#   1) un CSV idéntico pero con Claudia López RECUPERADA por mesa
#      (sus votos venían escondidos en total_votos_urna; residual = total - partes,
#       suma nacional exacta 225.287, sin negativos).
#   2) un Excel "bonito" para cliente: tabla con autofiltro, SIN códigos, con los
#      NOMBRES de departamento, municipio, zona/comuna y puesto (cruce PUESTOS_GEOREF).
import csv, re
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE='Bases de datos'
MESA=f'{BASE}/nuevos archivos 1v 2026/PRECONTEO_1V_2026_MESA_nombres_corregidos.csv'
GEOREF=f'{BASE}/PUESTOS_GEOREF.csv'
OUT_CSV=f'{BASE}/nuevos archivos 1v 2026/PRECONTEO_1V_2026_MESA_con_Claudia.csv'
OUT_XLSX=f'{BASE}/nuevos archivos 1v 2026/Resultados_1V_2026_por_mesa.xlsx'

def titlecase(s): return re.sub(r'\s+',' ',str(s or '').strip()).title()

# ── 1) Mapas de nombres desde PUESTOS_GEOREF (clave de 9 dígitos dep+mun+zona+puesto) ──
depname={}; munname={}; puename={}; comname={}
with open(GEOREF,newline='',encoding='utf-8-sig') as f:
    for row in csv.DictReader(f,delimiter=';'):
        cc=(row.get('CÓDIGO COMPLETO') or '').strip()
        if len(cc)!=9 or not cc.isdigit(): continue
        dep,mun=cc[:2],cc[2:5]
        depname.setdefault(dep,titlecase(row['DEPARTAMENTO']))
        munname.setdefault(dep+mun,titlecase(row['MUNICIPIO']))
        puename[cc]=titlecase(row.get('NOMBRE PUESTO'))
        com=re.sub(r'^\d+','',(row.get('NOMBRE COMUNA') or '').strip()).strip()
        comname[cc]=titlecase(com) if com.upper() not in ('NULL','NONE','SN','') else ''   # "NULL" → vacío
print(f'nombres: {len(depname)} deptos · {len(munname)} muns · {len(puename)} puestos')

# ── 2) Recorrer mesas: recuperar Claudia + reunir filas ──
CANDS=['Iván Cepeda','Santiago Botero','Abelardo De La Espriella','Mauricio Lizcano','Miguel Uribe',
       'Sondra Macollins','Roy Barreras','Carlos Caicedo','Gustavo Matamoros','Paloma Valencia',
       'Sergio Fajardo','Gilberto Murillo','Claudia López']
# orden de despliegue en el Excel (por votación nacional desc)
ORDER=['Iván Cepeda','Abelardo De La Espriella','Paloma Valencia','Sergio Fajardo','Claudia López',
       'Santiago Botero','Mauricio Lizcano','Miguel Uribe','Sondra Macollins','Roy Barreras',
       'Gilberto Murillo','Carlos Caicedo','Gustavo Matamoros']

f=open(MESA,newline='',encoding='utf-8-sig'); r=csv.reader(f); H=next(r)
idx={c:i for i,c in enumerate(H)}
ci=[idx[c] for c in CANDS]; iblan=idx['votos_blanco']; inul=idx['votos_nulos']; inma=idx['votos_no_marcados']
itot=idx['total_votos_urna']; iclau=idx['Claudia López']
quote={idx['cod_departamento'],idx['cod_municipio'],idx['zona'],idx['puesto'],idx['num_mesa'],idx['fecha_actualizacion']}

g=open(OUT_CSV,'w',newline='',encoding='utf-8-sig')
g.write(','.join('"'+c+'"' for c in H)+'\n')   # mismo header
xrows=[]; n=0; claudia_total=0; cov_pue=0
for row in r:
    dep=row[idx['cod_departamento']].zfill(2); mun=row[idx['cod_municipio']].zfill(3)
    zon=row[idx['zona']].zfill(2); pue=row[idx['puesto']].zfill(2)
    cc=dep+mun+zon+pue
    partes=sum(int(row[i] or 0) for i in ci)+int(row[iblan] or 0)+int(row[inul] or 0)+int(row[inma] or 0)
    claudia=max(0,int(row[itot] or 0)-partes)   # = total - partes (Claudia llega en 0 en su columna)
    row[iclau]=str(claudia); claudia_total+=claudia
    # CSV idéntico (códigos quoted, votos bare, fecha quoted)
    g.write(','.join(('"'+v+'"') if i in quote else v for i,v in enumerate(row))+'\n')
    # fila Excel con NOMBRES
    dnm=depname.get(dep) or ('Exterior' if dep=='88' else dep)
    mnm=munname.get(dep+mun) or (titlecase(mun) if dep=='88' else dep+mun)
    znm=comname.get(cc) or (f'Zona {int(zon)}' if zon.isdigit() else f'Zona {zon}')
    pnm=puename.get(cc)
    if pnm: cov_pue+=1
    else: pnm=(f'Puesto {int(pue)}' if pue.isdigit() else f'Puesto {pue}')
    mesa=row[idx['num_mesa']]; mesa=int(mesa) if mesa.isdigit() else mesa
    xrows.append([dnm,mnm,znm,pnm,mesa]+[int(row[idx[c]] or 0) for c in ORDER]
                 +[int(row[iblan] or 0),int(row[inul] or 0),int(row[inma] or 0),int(row[itot] or 0)])
    n+=1
    if n%20000==0: print(f'  {n:,} mesas…')
g.close(); f.close()
print(f'✓ CSV con Claudia: {OUT_CSV} · {n:,} mesas · Claudia recuperada {claudia_total:,} · puestos con nombre {cov_pue/n*100:.1f}%')

# ── 3) Excel bonito ──
SHORT={'Iván Cepeda':'Cepeda','Abelardo De La Espriella':'Abelardo','Paloma Valencia':'Paloma',
       'Sergio Fajardo':'Fajardo','Claudia López':'Claudia López','Santiago Botero':'Botero',
       'Mauricio Lizcano':'Lizcano','Miguel Uribe':'Miguel Uribe','Sondra Macollins':'Macollins',
       'Roy Barreras':'Roy Barreras','Gilberto Murillo':'Murillo','Carlos Caicedo':'Caicedo',
       'Gustavo Matamoros':'Matamoros'}
HEAD=['Departamento','Municipio','Zona / Comuna','Puesto de votación','Mesa']+[SHORT[c] for c in ORDER]+['Votos en blanco','Votos nulos','No marcados','Total votos']
OX='8A1E16'
wb=Workbook()

# --- Hoja Instrucciones ---
ins=wb.active; ins.title='Instrucciones'
def put(row,col,val,**kw):
    c=ins.cell(row=row,column=col,value=val)
    if kw.get('b') or kw.get('size') or kw.get('color'): c.font=Font(bold=kw.get('b',False),size=kw.get('size',11),color=kw.get('color','1A1510'))
    if kw.get('wrap'): c.alignment=Alignment(wrap_text=True,vertical='top')
    return c
put(1,1,'Resultados · Primera vuelta presidencial 2026',b=True,size=18,color=OX)
put(2,1,'Escrutinio (preconteo) por mesa de votación · 31 de mayo de 2026',size=11,color='6B6354')
rowi=4
bloques=[
 ('¿Qué contiene este archivo?','La hoja "Datos por mesa" trae el resultado de la 1ª vuelta presidencial en CADA mesa de votación del país (121.863 mesas), con los nombres de departamento, municipio, zona/comuna y puesto — sin códigos. Es una tabla con filtros.'),
 ('¿Cómo usarla?','Ve a la hoja "Datos por mesa". Cada columna del encabezado tiene una flecha de filtro: haz clic para filtrar por departamento, municipio, puesto, etc. Puedes ordenar por cualquier columna (p. ej. mayor votación de un candidato).'),
 ('Columnas','• Departamento / Municipio / Zona-Comuna / Puesto de votación: ubicación (con nombre).\n• Mesa: número de mesa dentro del puesto.\n• Una columna por candidato, con los votos en esa mesa.\n• Votos en blanco · nulos · no marcados · Total votos de la mesa.'),
 ('Nota sobre Claudia López','En el preconteo original, los votos de Claudia López llegaban sin asignar a su columna (en 0), aunque sí estaban contados en el total de cada mesa. Aquí se RECUPERARON mesa por mesa (total de la mesa menos el resto de votos). La suma nacional coincide exactamente con su total oficial: 225.287 votos.'),
 ('Fuente','Registraduría Nacional del Estado Civil — preconteo de 1ª vuelta (corte 31-may-2026, 23:30). Nombres de puestos y comunas: georreferenciación oficial de puestos de votación.'),
 ('Advertencia','Son cifras de PRECONTEO, sujetas a variación frente al escrutinio definitivo. Documento de trabajo.'),
]
for tit,txt in bloques:
    put(rowi,1,tit,b=True,size=12,color=OX); rowi+=1
    c=put(rowi,1,txt,size=10.5,wrap=True); ins.row_dimensions[rowi].height=14*(1+txt.count(chr(10))+len(txt)//95); rowi+=2
ins.column_dimensions['A'].width=110
for rr in range(1,rowi):
    if ins.cell(rr,1).alignment is None or not ins.cell(rr,1).alignment.wrap_text:
        pass

# --- Hoja Datos por mesa (tabla) ---
ws=wb.create_sheet('Datos por mesa')
ws.append(HEAD)
for xr in xrows: ws.append(xr)
last=get_column_letter(len(HEAD)); lastrow=ws.max_row
tab=Table(displayName='ResultadosMesa',ref=f'A1:{last}{lastrow}')
tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True,showColumnStripes=False,showFirstColumn=False,showLastColumn=False)
ws.add_table(tab)
ws.freeze_panes='F2'   # congela encabezado + las 5 columnas de ubicación
# anchos + formato de miles en columnas numéricas
widths=[16,24,26,34,7]+[11]*len(ORDER)+[10,9,9,11]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
ncol0=6  # primera columna de votos
for col in range(ncol0,len(HEAD)+1):
    for rr in range(2,lastrow+1):
        ws.cell(rr,col).number_format='#,##0'
ws.sheet_view.showGridLines=False
wb.save(OUT_XLSX)
print(f'✓ Excel: {OUT_XLSX} · {lastrow-1:,} filas · {len(HEAD)} columnas')
