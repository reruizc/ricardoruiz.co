#!/usr/bin/env python3
"""
RENADIA · lee las respuestas recogidas y las consolida para el equipo.

Baja los JSON que el HTML fue guardando en S3 (renadia-collect) y arma:
  - un Excel RENADIA_respuestas.xlsx con una hoja por reto (si hay openpyxl)
  - CSVs por reto + uno combinado (siempre; se abren en Excel con tildes OK)

Uso:
    python3 tools/renadia-collect/leer_respuestas.py
    python3 tools/renadia-collect/leer_respuestas.py --no-sync   # usa lo ya bajado

Requiere el AWS CLI configurado (el mismo que ya usas). openpyxl es opcional.
Salidas en: Bases de datos/DNP/respuestas-export/
"""
import os
import sys
import json
import glob
import csv
import subprocess

S3_PREFIX = "s3://elecciones-2026/renadia-collect/respuestas/"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))
OUT_DIR = os.path.join(REPO, "Bases de datos", "DNP", "respuestas-export")
RAW_DIR = os.path.join(OUT_DIR, "_raw")


def sync_s3():
    os.makedirs(RAW_DIR, exist_ok=True)
    print("Descargando respuestas desde S3…")
    try:
        subprocess.run(["aws", "s3", "sync", S3_PREFIX, RAW_DIR, "--quiet"], check=True)
    except FileNotFoundError:
        sys.exit("ERROR: no se encontró el comando 'aws'. Instala/configura el AWS CLI.")
    except subprocess.CalledProcessError as e:
        sys.exit("ERROR al sincronizar con S3: %s" % e)


def load_records():
    files = glob.glob(os.path.join(RAW_DIR, "**", "*.json"), recursive=True)
    recs = []
    for f in files:
        try:
            with open(f, encoding="utf-8") as fh:
                d = json.load(fh)
            if isinstance(d, dict):
                recs.append(d)
        except Exception:
            pass  # ignora archivos rotos
    # ordena por fecha de recepción
    recs.sort(key=lambda r: r.get("_recibido") or r.get("ts") or "")
    return recs


def fecha(r):
    return r.get("_recibido") or r.get("ts") or ""


def ans_at(r, i):
    a = r.get("respuestas") or []
    return a[i].get("respuesta", "") if i < len(a) else ""


# ---- filas por reto -------------------------------------------------------

CARTA_COLS = ["Fecha", "Usuario", "Grupo", "OVR", "Posición",
              "Datos", "IA", "Gobernanza", "Cocreación",
              "P1", "P2", "P3", "P4", "P5", "Sesión", "Origen", "IP"]

def row_carta(r):
    res = r.get("resultado") or {}
    st = res.get("stats") or {}
    return [fecha(r), r.get("nombre", ""), r.get("grupo", ""),
            res.get("ovr", ""), res.get("posicion", ""),
            st.get("DAT", ""), st.get("IA", ""), st.get("GOB", ""), st.get("COC", ""),
            ans_at(r, 0), ans_at(r, 1), ans_at(r, 2), ans_at(r, 3), ans_at(r, 4),
            r.get("sid", ""), r.get("origen", ""), r.get("_ip", "")]


MAD_COLS = ["Fecha", "Usuario", "Sector",
            "Desafío más urgente", "Datos que faltan",
            "Datos que ya tenemos", "Iniciativas de IA",
            "Sesión", "Origen", "IP"]

def row_madurez(r):
    return [fecha(r), r.get("nombre", ""), r.get("grupo", ""),
            ans_at(r, 0), ans_at(r, 1), ans_at(r, 2), ans_at(r, 3),
            r.get("sid", ""), r.get("origen", ""), r.get("_ip", "")]


PK_COLS = ["Fecha", "Usuario", "Atajadas", "Total",
           "P1", "P2", "P3", "P4", "P5", "Sesión", "Origen", "IP"]

def _pk(r, i):
    a = r.get("respuestas") or []
    if i >= len(a):
        return ""
    x = a[i]
    mark = " ✓" if x.get("correcto") else " ✗"
    return (x.get("respuesta", "") + mark).strip()

def row_penaltis(r):
    res = r.get("resultado") or {}
    return [fecha(r), r.get("nombre", ""), res.get("atajadas", ""), res.get("total", ""),
            _pk(r, 0), _pk(r, 1), _pk(r, 2), _pk(r, 3), _pk(r, 4),
            r.get("sid", ""), r.get("origen", ""), r.get("_ip", "")]


def flat_ans(r):
    parts = []
    for a in (r.get("respuestas") or []):
        s = str(a.get("pregunta", "")) + " → " + str(a.get("respuesta", ""))
        if "correcto" in a:
            s += " [ok]" if a.get("correcto") else " [x]"
        parts.append(s)
    return " | ".join(parts)

def flat_res(r):
    res = r.get("resultado") or {}
    return " · ".join("%s: %s" % (k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
                      for k, v in res.items())

TODAS_COLS = ["Fecha", "Reto", "Usuario", "Grupo/Sector", "Resultado", "Respuestas",
              "Sesión", "Origen", "IP"]

def row_todas(r):
    j = {"carta": "Reto 1 · Carta", "madurez": "Reto 2 · Diagnóstico",
         "penaltis": "Reto 3 · Penaltis"}.get(r.get("juego"), r.get("juego", ""))
    return [fecha(r), j, r.get("nombre", ""), r.get("grupo", ""),
            flat_res(r), flat_ans(r), r.get("sid", ""), r.get("origen", ""), r.get("_ip", "")]


def write_csv(name, cols, rows):
    path = os.path.join(OUT_DIR, name)
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:  # BOM: Excel muestra tildes
        w = csv.writer(fh)
        w.writerow(cols)
        w.writerows(rows)
    return path


def write_xlsx(sheets):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        print("  (openpyxl no está instalado → me salto el Excel; usa los CSV.)")
        return None
    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="12182A")
    head_font = Font(bold=True, color="FFFFFF")
    for title, cols, rows in sheets:
        ws = wb.create_sheet(title[:31])
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"
        # ancho de columnas (aprox por contenido)
        for c in range(1, len(cols) + 1):
            L = get_column_letter(c)
            maxlen = len(str(cols[c - 1]))
            for r in rows[:200]:
                if c - 1 < len(r):
                    maxlen = max(maxlen, len(str(r[c - 1])))
            ws.column_dimensions[L].width = min(60, max(10, maxlen + 2))
        if ws.max_row > 1:
            ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), ws.max_row)
    path = os.path.join(OUT_DIR, "RENADIA_respuestas.xlsx")
    wb.save(path)
    return path


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    if "--no-sync" not in sys.argv:
        sync_s3()

    recs = load_records()
    carta = [r for r in recs if r.get("juego") == "carta"]
    mad = [r for r in recs if r.get("juego") == "madurez"]
    pk = [r for r in recs if r.get("juego") == "penaltis"]

    r_carta = [row_carta(r) for r in carta]
    r_mad = [row_madurez(r) for r in mad]
    r_pk = [row_penaltis(r) for r in pk]
    r_todas = [row_todas(r) for r in recs]

    write_csv("reto1_carta.csv", CARTA_COLS, r_carta)
    write_csv("reto2_diagnostico.csv", MAD_COLS, r_mad)
    write_csv("reto3_penaltis.csv", PK_COLS, r_pk)
    write_csv("todas_las_respuestas.csv", TODAS_COLS, r_todas)

    xlsx = write_xlsx([
        ("Reto 1 · Carta", CARTA_COLS, r_carta),
        ("Reto 2 · Diagnóstico", MAD_COLS, r_mad),
        ("Reto 3 · Penaltis", PK_COLS, r_pk),
        ("Todas", TODAS_COLS, r_todas),
    ])

    print("\nRespuestas leídas: %d  (Reto 1: %d · Reto 2: %d · Reto 3: %d)"
          % (len(recs), len(carta), len(mad), len(pk)))
    print("Participantes únicos (por sesión): %d"
          % len({r.get("sid") for r in recs if r.get("sid")}))
    print("\nArchivos en: %s" % OUT_DIR)
    for f in ["reto1_carta.csv", "reto2_diagnostico.csv", "reto3_penaltis.csv", "todas_las_respuestas.csv"]:
        print("  - " + f)
    if xlsx:
        print("  - RENADIA_respuestas.xlsx  (una hoja por reto)")


if __name__ == "__main__":
    main()
