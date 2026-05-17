#!/usr/bin/env python3
"""
build-arquetipos-medellin/build.py

Procesa los 3 Excel del paquete de Nury para el módulo 05 (Arquetipos
territoriales) de proyecto-dc y emite JSONs listos para subir a S3.

Uso:
    python3 tools/build-arquetipos-medellin/build.py \\
        "Insumos /" \\
        "Bases de datos/proyecto-dc/arquetipos"

Inputs (rutas absolutas relativas al primer argumento):
    ANALISIS DE EVOLUCIÓN ARQUETIPOS.xlsx
    EVOLUCIÓN ARQUETIPOS MEDELLÍN.xlsx
    VOTOS 2027 ARQUETIPO BARRIO.xlsx

Outputs (en el segundo argumento):
    arquetipos.json
    por-barrio.json
    por-comuna.json
    proyeccion-2027-resumen.json
    correlaciones-top.json
    transiciones.json
    metodologia.json
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

import openpyxl


S3_ESCUDOS = (
    "https://elecciones-2026.s3.us-east-1.amazonaws.com/ricardoruiz.co/"
    "bases+de+datos/Proyecto+DC/arquetipos/escudos"
)

FAMILY_MAP = {
    "Protección y orden cotidiano": "proteccion",
    "Estabilidad y continuidad": "continuidad",
    "Supervivencia económica y servicios básicos": "supervivencia",
    "Desconfianza y castigo": "castigo",
    "Pertenencia y dignidad territorial": "pertenencia",
    "Protección con resultados y orden competente": "proteccion",
    "Continuidad pragmática y gestión barrial": "continuidad",
    "Supervivencia económica y servicios cotidianos": "supervivencia",
    "Castigo a la restauración y demanda de alternancia": "castigo",
    "Pertenencia comunitaria y autonomía territorial": "pertenencia",
    "Sin dato": None,
}

SCORE_KEY_MAP = {
    "Score Protección y orden cotidiano": "proteccion",
    "Score Estabilidad y continuidad": "continuidad",
    "Score Supervivencia económica y servicios básicos": "supervivencia",
    "Score Desconfianza y castigo": "castigo",
    "Score Pertenencia y dignidad territorial": "pertenencia",
}

ARQUETIPOS_CANON = {
    "proteccion": {
        "color": "#2563eb",
        "color_dark": "#1d4ed8",
        "label_corto": "Protección",
        "orden": 1,
        "base": {
            "id": "proteccion",
            "nombre": "Protección y orden cotidiano",
            "emocion": "Miedo gestionado, búsqueda de autoridad y alivio",
            "deseo": "Que el Estado controle el barrio, reduzca la amenaza y ordene lo cotidiano.",
            "miedo": "Desorden, extorsión, criminalidad, abandono o pérdida de control territorial.",
            "sesgo": "Sesgo de seguridad y preferencia por figuras con reputación de control.",
            "escudo": f"{S3_ESCUDOS}/proteccion-orden.jpg",
            "vigencia": "2015-2023",
        },
        "evol": {
            "id": "proteccion_evol",
            "nombre": "Protección con resultados y orden competente",
            "emocion": "Demanda de orden con eficacia probada",
            "deseo": "Mantener la seguridad pero exigir resultados visibles y gestión competente.",
            "miedo": "Retroceso del orden, criminalidad reactiva, fallas en la ejecución.",
            "sesgo": "Sesgo de continuidad bajo desempeño percibido.",
            "escudo": f"{S3_ESCUDOS}/proteccion-resultados.jpg",
            "vigencia": "2027",
        },
    },
    "continuidad": {
        "color": "#16a34a",
        "color_dark": "#15803d",
        "label_corto": "Continuidad",
        "orden": 2,
        "base": {
            "id": "continuidad",
            "nombre": "Estabilidad y continuidad",
            "emocion": "Confianza conservadora, continuidad y reducción de incertidumbre",
            "deseo": "Mantener beneficios, obras, relaciones institucionales o estilos de gobierno conocidos.",
            "miedo": "Cambio abrupto, pérdida de programas, ruptura de redes o improvisación.",
            "sesgo": "Sesgo de statu quo y aversión a la pérdida.",
            "escudo": f"{S3_ESCUDOS}/estabilidad-continuidad.jpg",
            "vigencia": "2015-2023",
        },
        "evol": {
            "id": "continuidad_evol",
            "nombre": "Continuidad pragmática y gestión barrial",
            "emocion": "Pragmatismo de redes locales conocidas",
            "deseo": "Que la red barrial siga resolviendo lo cotidiano sin disrupción.",
            "miedo": "Quiebre de la red local, llegada de actores externos.",
            "sesgo": "Sesgo de proximidad y confianza en gestores conocidos.",
            "escudo": f"{S3_ESCUDOS}/continuidad-gestion.jpg",
            "vigencia": "2027",
        },
    },
    "supervivencia": {
        "color": "#b45309",
        "color_dark": "#92400e",
        "label_corto": "Supervivencia",
        "orden": 3,
        "base": {
            "id": "supervivencia",
            "nombre": "Supervivencia económica y servicios básicos",
            "emocion": "Ansiedad material y pragmatismo de necesidad",
            "deseo": "Resolver empleo, servicios, ayudas, trámites, transporte y economía cotidiana.",
            "miedo": "Quedar sin acceso a recursos, ayudas, ingresos o intermediación de gestión.",
            "sesgo": "Sesgo de utilidad inmediata y reciprocidad territorial.",
            "escudo": f"{S3_ESCUDOS}/supervivencia-basicos.jpg",
            "vigencia": "2015-2023",
        },
        "evol": {
            "id": "supervivencia_evol",
            "nombre": "Supervivencia económica y servicios cotidianos",
            "emocion": "Apuro material sostenido, exigencia de gestión efectiva",
            "deseo": "Acceso a ingresos, ayudas y trámites con menos fricción.",
            "miedo": "Pérdida de ayudas y deterioro de servicios públicos.",
            "sesgo": "Sesgo de utilidad inmediata combinado con desconfianza institucional.",
            "escudo": f"{S3_ESCUDOS}/supervivencia-cotidianos.jpg",
            "vigencia": "2027",
        },
    },
    "castigo": {
        "color": "#dc2626",
        "color_dark": "#b91c1c",
        "label_corto": "Castigo",
        "orden": 4,
        "base": {
            "id": "castigo",
            "nombre": "Desconfianza y castigo",
            "emocion": "Rabia contenida, sospecha y autoprotección",
            "deseo": "Castigar élites, maquinarias o gestiones percibidas como abusivas o desconectadas.",
            "miedo": "Ser usado por políticos, quedar marcado o entregar demasiado poder a una red.",
            "sesgo": "Sesgo de negatividad, castigo retrospectivo y voto defensivo.",
            "escudo": f"{S3_ESCUDOS}/desconfianza-castigo.jpg",
            "vigencia": "2015-2023",
        },
        "evol": {
            "id": "castigo_evol",
            "nombre": "Castigo a la restauración y demanda de alternancia",
            "emocion": "Indignación contra la restauración hegemónica",
            "deseo": "Romper la continuidad del orden percibido como excluyente y abrir alternancia.",
            "miedo": "Consolidación de una élite excluyente o regresión autoritaria.",
            "sesgo": "Sesgo de castigo retrospectivo, sensibilidad a la fatiga del oficialismo.",
            "escudo": f"{S3_ESCUDOS}/castigo-restauracion.jpg",
            "vigencia": "2027",
        },
    },
    "pertenencia": {
        "color": "#a21caf",
        "color_dark": "#86198f",
        "label_corto": "Pertenencia",
        "orden": 5,
        "base": {
            "id": "pertenencia",
            "nombre": "Pertenencia y dignidad territorial",
            "emocion": "Orgullo barrial, reconocimiento y dignidad",
            "deseo": "Que el barrio sea visible, respetado y representado por sus liderazgos propios.",
            "miedo": "Ser tratado como periferia, botín electoral o territorio sin voz.",
            "sesgo": "Sesgo identitario, proximidad comunitaria y representación local.",
            "escudo": f"{S3_ESCUDOS}/pertenencia-dignidad.jpg",
            "vigencia": "2015-2023",
        },
        "evol": {
            "id": "pertenencia_evol",
            "nombre": "Pertenencia comunitaria y autonomía territorial",
            "emocion": "Reivindicación de autonomía local y voz propia",
            "deseo": "Reconocimiento institucional y autonomía decisoria sobre el territorio.",
            "miedo": "Tratamiento como periferia subordinada o pérdida de la voz barrial.",
            "sesgo": "Sesgo identitario reforzado y desconfianza al centralismo.",
            "escudo": f"{S3_ESCUDOS}/pertenencia-comunitaria.jpg",
            "vigencia": "2027",
        },
    },
}

YEARS = (2015, 2019, 2023)


def r4(v):
    if v is None:
        return None
    if isinstance(v, float):
        return round(v, 4)
    return v


def family(name):
    if name is None:
        return None
    return FAMILY_MAP.get(str(name).strip())


def family_or_unknown(name):
    fam = family(name)
    return fam  # None means "Sin dato"


def load_sheet(wb, name, header_row=0):
    """Return list of dicts using header_row as header."""
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[header_row]
    out = []
    for r in rows[header_row + 1:]:
        if all(v is None for v in r):
            continue
        out.append({header[i]: r[i] for i in range(len(header)) if header[i] is not None})
    return out


def build_arquetipos_json():
    return {
        "familias": list(ARQUETIPOS_CANON.keys()),
        "arquetipos": ARQUETIPOS_CANON,
    }


def build_por_barrio(wb_a, wb_e, wb_v):
    """Master per-barrio JSON, keyed by Código DAP."""
    matrix = load_sheet(wb_a, "03_matriz_evolucion_barrio")
    base_larga = load_sheet(wb_a, "04_base_larga_barrio_anio")
    pivot_scores = load_sheet(wb_a, "11_scores_pivot_barrio")
    p2027 = load_sheet(wb_e, "2027")
    votos_wide = load_sheet(wb_v, "02_votos_barrio_wide")

    base_by_year = {y: {} for y in YEARS}
    for r in base_larga:
        y = r.get("Año")
        if y in base_by_year:
            base_by_year[y][r["Código DAP"]] = r

    pivot_by_dap = {r["Código DAP"]: r for r in pivot_scores}
    p2027_by_dap = {r["Código DAP"]: r for r in p2027}
    votos_by_dap = {r["Código DAP"]: r for r in votos_wide}

    out = {}
    for row in matrix:
        dap = row["Código DAP"]
        comuna = row["Comuna"]
        barrio = row["Barrio"]

        arquetipo = {str(y): family_or_unknown(row.get(f"{y} · Arquetipo final")) for y in YEARS}

        ganadores = {}
        for y in YEARS:
            ganadores[str(y)] = {
                "alcaldia": row.get(f"{y} · Ganador Alcaldía"),
                "pct_alc": r4(row.get(f"{y} · % ganador Alcaldía")),
                "concejo": row.get(f"{y} · Ganador Concejo"),
                "pct_con": r4(row.get(f"{y} · % ganador Concejo")),
                "jal": row.get(f"{y} · Ganador JAL"),
                "pct_jal": r4(row.get(f"{y} · % ganador JAL")),
            }

        correlaciones = {}
        for y in YEARS:
            correlaciones[str(y)] = {
                "alc_con": r4(row.get(f"{y} · Corr Alcaldía-Concejo")),
                "alc_jal": r4(row.get(f"{y} · Corr Alcaldía-JAL")),
                "jal_con": r4(row.get(f"{y} · Corr JAL-Concejo")),
                "score_tri": r4(row.get(f"{y} · Score triádico ganador")),
                "nivel": row.get(f"{y} · Nivel relación emocional"),
            }

        contexto = {}
        for y in YEARS:
            contexto[str(y)] = {
                "riesgo": row.get(f"{y} · Riesgo analítico"),
                "zona_criminal": row.get(f"{y} · Zona criminal-territorial"),
                "agenda": row.get(f"{y} · Tema agenda más sensible"),
            }

        # Scores normalizados — 2015 está en escala 0-100 en la base larga;
        # 2019/2023 vienen ya 0-1 en pivot. Normalizamos todo a 0-1.
        scores = {str(y): {} for y in YEARS}
        for y in YEARS:
            br = base_by_year[y].get(dap, {})
            for col, fam_id in SCORE_KEY_MAP.items():
                v = br.get(col)
                if v is None:
                    continue
                # 2015 viene 0-100, normalizar
                if y == 2015 and isinstance(v, (int, float)) and v > 1.5:
                    v = v / 100.0
                scores[str(y)][fam_id] = round(float(v), 4)

        # Trayectoria
        trayectoria_raw = row.get("Trayectoria arquetípica 2015→2019→2023")
        trayectoria_ids = None
        if trayectoria_raw:
            parts = [p.strip() for p in trayectoria_raw.split("→")]
            trayectoria_ids = [family_or_unknown(p) for p in parts]

        # Votos por año (incluyendo proyección 2027)
        v = votos_by_dap.get(dap, {})
        votos = {
            "2015": v.get("2015 · Total válidos Alcaldía"),
            "2019": v.get("2019 · Total válidos Alcaldía"),
            "2023": v.get("2023 · Total válidos Alcaldía"),
            "2027": v.get("Votos válidos Alcaldía 2027 proyectados"),
        }

        # Proyección 2027 detallada
        p27 = p2027_by_dap.get(dap, {})
        probs_2027 = {}
        votos_2027_por_arq = {}
        prob_cols = {
            "proteccion": (
                "Prob 2027 · Protección con resultados y orden competente",
                "Votos 2027 · Protección con resultados y orden competente",
            ),
            "castigo": (
                "Prob 2027 · Castigo a la restauración y demanda de alternancia",
                "Votos 2027 · Castigo a la restauración y demanda de alternancia",
            ),
            "continuidad": (
                "Prob 2027 · Continuidad pragmática y gestión barrial",
                "Votos 2027 · Continuidad pragmática y gestión barrial",
            ),
            "supervivencia": (
                "Prob 2027 · Supervivencia económica y servicios cotidianos",
                "Votos 2027 · Supervivencia económica y servicios cotidianos",
            ),
            "pertenencia": (
                "Prob 2027 · Pertenencia comunitaria y autonomía territorial",
                "Votos 2027 · Pertenencia comunitaria y autonomía territorial",
            ),
        }
        for fam_id, (pcol, vcol) in prob_cols.items():
            pv = p27.get(pcol)
            vv = v.get(vcol)
            if pv is not None:
                probs_2027[fam_id] = round(float(pv), 4)
            if vv is not None:
                votos_2027_por_arq[fam_id] = int(vv)

        proyeccion_2027 = {
            "arquetipo_proy": family_or_unknown(p27.get("Arquetipo proyectado 2027")),
            "prob_proy": r4(p27.get("Prob arquetipo proyectado 2027")),
            "arquetipo_alt": family_or_unknown(p27.get("Arquetipo alterno 2027")),
            "prob_alt": r4(p27.get("Prob arquetipo alterno 2027")),
            "riesgo_cambio": r4(p27.get("Riesgo de cambio 2023→2027")),
            "nivel_riesgo": p27.get("Nivel riesgo cambio"),
            "comportamiento": p27.get("Comportamiento probable 2027"),
            "probs": probs_2027,
            "votos_por_arquetipo": votos_2027_por_arq,
            "factor_crecimiento": r4(v.get("Factor crecimiento 2023→2027")),
            "escenario_participacion": v.get("Escenario participación"),
        }

        out[dap] = {
            "dap": dap,
            "comuna": comuna,
            "barrio": barrio,
            "arquetipo": arquetipo,
            "scores": scores,
            "ganadores": ganadores,
            "correlaciones": correlaciones,
            "contexto": contexto,
            "trayectoria": trayectoria_ids,
            "trayectoria_raw": trayectoria_raw,
            "tipo_evolucion": row.get("Tipo de evolución emocional"),
            "transicion_15_19": row.get("Transición 2015→2019"),
            "transicion_19_23": row.get("Transición 2019→2023"),
            "cambio_15_23": row.get("Cambio 2015→2023"),
            "votos": votos,
            "proyeccion_2027": proyeccion_2027,
        }
    return out


def build_por_comuna(wb_a, wb_v):
    """Comuna-level aggregates for 2015/19/23 (from ANALISIS) and 2027 (from VOTOS)."""
    resumen = load_sheet(wb_a, "12_resumen_comuna")
    resumen_2027 = load_sheet(wb_v, "04_resumen_comuna_arq")

    out = {}
    for r in resumen:
        comuna = r["Comuna"]
        year = str(r["Año"])
        fam = family_or_unknown(r["Arquetipo final"])
        if fam is None:
            continue
        out.setdefault(comuna, {}).setdefault(year, {"distribucion": {}, "votos_por_arq": {}, "score_tri_por_arq": {}, "barrios_total": 0, "votos_total": 0})
        slot = out[comuna][year]
        slot["distribucion"][fam] = slot["distribucion"].get(fam, 0) + (r["Barrios"] or 0)
        slot["votos_por_arq"][fam] = slot["votos_por_arq"].get(fam, 0) + (r["Votantes_alcaldia"] or 0)
        slot["score_tri_por_arq"][fam] = r4(r["Score_triadico_promedio"])
        slot["barrios_total"] += (r["Barrios"] or 0)
        slot["votos_total"] += (r["Votantes_alcaldia"] or 0)

    # Dominante por año
    for com, years in out.items():
        for yr, slot in years.items():
            if slot["distribucion"]:
                slot["dominante"] = max(slot["distribucion"].items(), key=lambda kv: kv[1])[0]

    # 2027 desde VOTOS
    for r in resumen_2027:
        comuna = r["Comuna"]
        fam = family_or_unknown(r["Arquetipo 2027"])
        if fam is None:
            continue
        out.setdefault(comuna, {}).setdefault("2027", {"distribucion": {}, "votos_por_arq": {}, "barrios_total": 0, "votos_total": 0})
        slot = out[comuna]["2027"]
        slot["distribucion"][fam] = (r["Barrios_dominantes"] or 0)
        slot["votos_por_arq"][fam] = (r["Votos_2027_proyectados"] or 0)
        slot["votos_total"] += (r["Votos_2027_proyectados"] or 0)
        slot["barrios_total"] += (r["Barrios_dominantes"] or 0)

    for com, years in out.items():
        slot = years.get("2027")
        if slot and slot["distribucion"]:
            slot["dominante"] = max(slot["distribucion"].items(), key=lambda kv: kv[1])[0]

    return out


def build_proyeccion_2027_resumen(wb_v, wb_a):
    """National aggregates for the 5 evolved 2027 archetypes + base 2015/19/23 totals."""
    r01 = load_sheet(wb_v, "01_resumen_arquetipos")
    arqs_2027 = []
    total_votos = 0
    for r in r01:
        fam = family_or_unknown(r["Arquetipo 2027"])
        if fam is None:
            continue
        votos = int(r.get("Votos_2027_proyectados") or 0)
        total_votos += votos
        arqs_2027.append({
            "id": f"{fam}_evol",
            "familia": fam,
            "nombre": r["Arquetipo 2027"],
            "votos": votos,
            "pct_ciudad": r4(r["Pct votos ciudad 2027 proyectado"]),
            "pct_barrial_prom": r4(r["Pct_promedio_barrial"]),
            "barrios_dominantes": int(r.get("Barrios_dominantes") or 0),
        })

    # Resumen base 2015/19/23 desde 02_resumen_anual_arquetipos
    base_anual = load_sheet(wb_a, "02_resumen_anual_arquetipos")
    por_anio = {str(y): [] for y in YEARS}
    for r in base_anual:
        fam = family_or_unknown(r["Arquetipo final"])
        if fam is None:
            continue
        y = str(r["Año"])
        if y not in por_anio:
            continue
        por_anio[y].append({
            "id": fam,
            "familia": fam,
            "nombre": r["Arquetipo final"],
            "barrios": int(r.get("Barrios") or 0),
            "votantes": int(r.get("Votantes_alcaldia") or 0),
            "pct_barrios": r4(r.get("% barrios del año")),
        })

    return {
        "ciudad_2027": {
            "votos_proyectados_total": total_votos,
            "arquetipos": arqs_2027,
        },
        "base_por_anio": por_anio,
    }


def build_correlaciones_top(wb_a):
    top = load_sheet(wb_a, "14_top_correlaciones_anio")
    out = {"2015": [], "2019": [], "2023": []}
    for r in top:
        y = str(r["Año"])
        if y not in out:
            continue
        out[y].append({
            "tipo": r["Tipo cruce"],
            "alcaldia": r["Candidato Alcaldía"],
            "concejo": r["Candidato Concejo"],
            "jal": r["Candidato JAL"],
            "corr_jal_alc": r4(r["Corr JAL-Alcaldía"]),
            "corr_jal_con": r4(r["Corr JAL-Concejo"]),
            "corr_alc_con": r4(r["Corr Alcaldía-Concejo"]),
            "score_directo": r4(r["Score directo JAL→Alc+Con"]),
            "cohesion": r4(r["Cohesión triádica absoluta"]),
            "nivel": r["Nivel relación"],
            "patron": r["Patrón de signos"],
        })
    # Cap a top 50 por año
    for y in out:
        out[y] = sorted(out[y], key=lambda x: -(x["cohesion"] or 0))[:50]
    return out


def build_transiciones(wb_a):
    out = {"matrices": {}, "trayectorias": []}

    for label, sheet in (
        ("2015_2019", "06_transicion_2015_2019"),
        ("2019_2023", "07_transicion_2019_2023"),
        ("2015_2023", "08_transicion_2015_2023"),
    ):
        ws = wb_a[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        # First col is "from" arquetipo, rest are "to" archetypes
        cells = []
        for r in rows[1:]:
            if r[0] is None:
                continue
            src_fam = family_or_unknown(r[0])
            for j in range(1, len(header)):
                tgt_fam = family_or_unknown(header[j])
                count = r[j]
                if count and count > 0:
                    cells.append({
                        "from": src_fam,
                        "to": tgt_fam,
                        "from_raw": r[0],
                        "to_raw": header[j],
                        "count": int(count),
                    })
        out["matrices"][label] = cells

    tray = load_sheet(wb_a, "09_trayectorias")
    for r in tray:
        raw = r["Trayectoria arquetípica 2015→2019→2023"]
        parts = [family_or_unknown(p.strip()) for p in raw.split("→")] if raw else None
        out["trayectorias"].append({
            "trayectoria": parts,
            "trayectoria_raw": raw,
            "tipo": r["Tipo de evolución emocional"],
            "barrios": int(r["Barrios"] or 0),
        })
    out["trayectorias"].sort(key=lambda x: -x["barrios"])
    return out


def build_metodologia(wb_a):
    meto = load_sheet(wb_a, "00_metodologia")
    dicc = load_sheet(wb_a, "15_diccionario")
    arq_def = load_sheet(wb_a, "01_5_arquetipos_emocionales")
    return {
        "secciones": [
            {"campo": r["Campo"], "descripcion": r["Descripción"]} for r in meto if r.get("Campo")
        ],
        "diccionario": [
            {"campo": r["Campo"], "definicion": r["Definición"]} for r in dicc if r.get("Campo")
        ],
        "arquetipos_definicion": [
            {
                "nombre": r["Arquetipo"],
                "emocion": r["Emoción dominante"],
                "deseo": r["Deseo político"],
                "miedo": r["Miedo dominante"],
                "sesgo": r["Sesgo emocional"],
            }
            for r in arq_def if r.get("Arquetipo")
        ],
    }


def write_json(path: Path, data: Any):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  → {path}  ({path.stat().st_size:,} bytes)")


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)

    src = Path(sys.argv[1])
    out = Path(sys.argv[2])

    f_analisis = src / "ANALISIS DE EVOLUCIÓN ARQUETIPOS.xlsx"
    f_evolucion = src / "EVOLUCIÓN ARQUETIPOS MEDELLÍN.xlsx"
    f_votos = src / "VOTOS 2027 ARQUETIPO BARRIO.xlsx"
    for f in (f_analisis, f_evolucion, f_votos):
        if not f.exists():
            sys.exit(f"ERROR: no existe {f}")

    print("Cargando Excel…")
    wb_a = openpyxl.load_workbook(f_analisis, data_only=True)
    wb_e = openpyxl.load_workbook(f_evolucion, data_only=True)
    wb_v = openpyxl.load_workbook(f_votos, data_only=True)

    print("Construyendo arquetipos.json…")
    write_json(out / "arquetipos.json", build_arquetipos_json())

    print("Construyendo por-barrio.json…")
    pb = build_por_barrio(wb_a, wb_e, wb_v)
    write_json(out / "por-barrio.json", pb)
    print(f"    {len(pb)} barrios procesados")

    print("Construyendo por-comuna.json…")
    pc = build_por_comuna(wb_a, wb_v)
    write_json(out / "por-comuna.json", pc)
    print(f"    {len(pc)} comunas procesadas")

    print("Construyendo proyeccion-2027-resumen.json…")
    write_json(out / "proyeccion-2027-resumen.json", build_proyeccion_2027_resumen(wb_v, wb_a))

    print("Construyendo correlaciones-top.json…")
    write_json(out / "correlaciones-top.json", build_correlaciones_top(wb_a))

    print("Construyendo transiciones.json…")
    write_json(out / "transiciones.json", build_transiciones(wb_a))

    print("Construyendo metodologia.json…")
    write_json(out / "metodologia.json", build_metodologia(wb_a))

    print("\nListo.")


if __name__ == "__main__":
    main()
