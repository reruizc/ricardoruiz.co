#!/usr/bin/env python3
"""
json_to_xlsx.py · Exporta el banco de preguntas del test-presidencial-2026 a Excel
para que Ricardo pueda revisar y corregir los enunciados a mano.

Lee desde:
  Bases de datos/test-presidencial/{arquetipos,candidatos,registros,mini_test,preguntas}.json

Escribe a:
  Bases de datos/test-presidencial/banco-preguntas-v1.xlsx

Hojas:
  - arquetipos          (5 filas)
  - candidatos          (6 filas)
  - registros           (3 filas)
  - mini_test           (las 4 preguntas + opciones con scoring por candidato)
  - preguntas_base      (los 7 enunciados de pregunta × 3 registros = 21 filas)
  - opciones_<registro> (3 hojas, una por registro: ~210 filas cada una)
                        Estructura: pregunta_id · opcion_id · arquetipo · arq_secundario
                        · candidato_id · candidato_nombre · enunciado · needs_review

Uso:
  python3 tools/build-banco-preguntas/json_to_xlsx.py
"""

import json
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "Bases de datos" / "test-presidencial"
OUT = DATA / "banco-preguntas-v1.xlsx"
VARIANTES_FILE = DATA / "variantes-tematicas.json"

REGISTROS = ["analitico", "popular", "digital"]

# Estilos
HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def load(name):
    with open(DATA / f"{name}.json", encoding="utf-8") as f:
        return json.load(f)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def wrap_rows(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER


# ---------- ARQUETIPOS ----------
def sheet_arquetipos(wb, data):
    ws = wb.create_sheet("arquetipos")
    headers = [
        "id", "nombre_corto", "nombre_2015", "nombre_2027",
        "emocion", "miedo", "deseo", "sesgo", "color", "marco",
    ]
    ws.append(headers)
    for a in data["arquetipos"]:
        ws.append([a.get(h, "") for h in headers])
        # color en la celda "color"
        cell = ws.cell(row=ws.max_row, column=headers.index("color") + 1)
        try:
            cell.fill = PatternFill("solid", fgColor=a["color"].lstrip("#"))
            cell.font = Font(color="FFFFFF", bold=True)
        except Exception:
            pass
    style_header(ws, len(headers))
    autosize(ws, [14, 16, 30, 36, 38, 38, 38, 38, 12, 38])
    wrap_rows(ws)
    ws.freeze_panes = "A2"


# ---------- CANDIDATOS ----------
def sheet_candidatos(wb, data):
    ws = wb.create_sheet("candidatos")
    headers = ["id", "nombre", "partido", "eje", "lente", "tono_propio", "color"]
    ws.append(headers)
    for c in data["candidatos"]:
        ws.append([c.get(h, "") for h in headers])
        cell = ws.cell(row=ws.max_row, column=headers.index("color") + 1)
        try:
            cell.fill = PatternFill("solid", fgColor=c["color"].lstrip("#"))
            cell.font = Font(color="FFFFFF", bold=True)
        except Exception:
            pass
    style_header(ws, len(headers))
    autosize(ws, [8, 22, 36, 18, 60, 60, 12])
    wrap_rows(ws)
    ws.freeze_panes = "A2"


# ---------- REGISTROS ----------
def sheet_registros(wb, data):
    ws = wb.create_sheet("registros")
    headers = ["id", "icono", "titulo", "lead", "tono_redaccion"]
    ws.append(headers)
    for r in data["registros"]:
        ws.append([r.get(h, "") for h in headers])
    style_header(ws, len(headers))
    autosize(ws, [12, 8, 22, 50, 60])
    wrap_rows(ws)
    ws.freeze_panes = "A2"


# ---------- MINI TEST ----------
def sheet_mini_test(wb, mt_data, cand_ids):
    ws = wb.create_sheet("mini_test")
    headers = (
        ["pregunta_id", "tipo", "tema", "enunciado_analitico", "enunciado_popular",
         "enunciado_digital", "opcion_id_o_valor", "opcion_etiqueta"]
        + [f"pts_{cid}" for cid in cand_ids]
    )
    ws.append(headers)
    for q in mt_data["preguntas"]:
        for i, op in enumerate(q["opciones"]):
            opid = str(op.get("id", op.get("valor", "")))
            ws.append([
                q["id"] if i == 0 else "",
                q["tipo"] if i == 0 else "",
                q["tema"] if i == 0 else "",
                q["enunciado"]["analitico"] if i == 0 else "",
                q["enunciado"]["popular"] if i == 0 else "",
                q["enunciado"]["digital"] if i == 0 else "",
                opid,
                op["etiqueta"],
            ] + [op["puntos"].get(cid, 0) for cid in cand_ids])
    style_header(ws, len(headers))
    autosize(ws, [11, 12, 28, 50, 50, 50, 12, 32] + [8] * len(cand_ids))
    wrap_rows(ws)
    ws.freeze_panes = "B2"


# ---------- PREGUNTAS BASE ----------
def sheet_preguntas_base(wb, preg_data):
    ws = wb.create_sheet("preguntas_base")
    headers = ["pregunta_id", "tema", "registro", "enunciado_pregunta", "needs_review"]
    ws.append(headers)
    for p in preg_data["preguntas"]:
        for reg in REGISTROS:
            texto = p["enunciado"].get(reg)
            ws.append([
                p["id"], p["tema"], reg,
                texto if texto else "",
                "FALSE" if texto else "TRUE",
            ])
    style_header(ws, len(headers))
    autosize(ws, [12, 28, 12, 70, 14])
    wrap_rows(ws)
    ws.freeze_panes = "A2"


# ---------- OPCIONES POR REGISTRO ----------
def sheet_opciones_registro(wb, preg_data, cands, registro):
    ws = wb.create_sheet(f"opciones_{registro}")
    headers = [
        "pregunta_id", "tema", "opcion_id", "arquetipo", "arq_secundario",
        "candidato_id", "candidato_nombre", "enunciado", "needs_review",
    ]
    ws.append(headers)
    cand_by_id = {c["id"]: c["nombre"] for c in cands}
    for p in preg_data["preguntas"]:
        for op in p["opciones"]:
            for c in cands:
                cid = c["id"]
                txt = (op["enunciados"].get(cid) or {}).get(registro)
                ws.append([
                    p["id"], p["tema"], op["id"],
                    op["arquetipo"], op.get("arquetipo_secundario") or "",
                    cid, cand_by_id[cid],
                    txt if txt else "",
                    "FALSE" if txt else "TRUE",
                ])
                # tinte de fondo según arquetipo (sutil)
                last_row = ws.max_row
                arq_color_map = {
                    "proteccion": "e8f1f8",
                    "estabilidad": "ebf4ee",
                    "supervivencia": "fbeee2",
                    "castigo": "f6e0e0",
                    "pertenencia": "efe5f3",
                }
                fill = arq_color_map.get(op["arquetipo"])
                if fill:
                    for c_idx in range(1, len(headers) + 1):
                        ws.cell(row=last_row, column=c_idx).fill = PatternFill("solid", fgColor=fill)
    style_header(ws, len(headers))
    autosize(ws, [11, 26, 9, 14, 14, 9, 22, 88, 14])
    wrap_rows(ws)
    ws.freeze_panes = "C2"


# ---------- PRIORIDAD TEMÁTICA ----------
def sheet_prioridad(wb, preg_data):
    prio = preg_data.get("prioridad_tematica")
    if not prio:
        return
    ws = wb.create_sheet("prioridad_tematica")
    headers = ["id", "registro", "tipo", "campo", "enunciado", "needs_review"]
    ws.append(headers)
    # Pregunta enunciado por registro
    for reg in REGISTROS:
        texto = (prio.get("enunciado") or {}).get(reg)
        ws.append([prio["id"], reg, prio.get("tipo", ""), "pregunta", texto or "", "FALSE" if texto else "TRUE"])
    # Opciones
    for op in prio.get("opciones", []):
        for reg in REGISTROS:
            txt = (op.get("etiqueta") or {}).get(reg)
            ws.append([op["id"], reg, "opcion", "etiqueta", txt or "", "FALSE" if txt else "TRUE"])
    style_header(ws, len(headers))
    autosize(ws, [16, 12, 18, 14, 75, 14])
    wrap_rows(ws)
    ws.freeze_panes = "A2"


def sheet_variantes_registro(wb, variantes_data, cands, registro):
    """Una hoja por registro con todas las opciones de las 5 variantes
    temáticas (P1-salud, P2-costo_vida, etc.). Misma estructura que las
    hojas opciones_*."""
    ws = wb.create_sheet(f"variantes_{registro}")
    headers = [
        "pregunta_id", "tema_variante", "opcion_id", "arquetipo", "arq_secundario",
        "candidato_id", "candidato_nombre", "enunciado", "needs_review",
    ]
    ws.append(headers)
    cand_by_id = {c["id"]: c["nombre"] for c in cands}
    arq_color_map = {
        "proteccion": "e8f1f8", "estabilidad": "ebf4ee",
        "supervivencia": "fbeee2", "castigo": "f6e0e0", "pertenencia": "efe5f3",
    }
    for v in variantes_data.get("variantes", []):
        for op in v.get("opciones", []):
            for c in cands:
                cid = c["id"]
                txt = (op["enunciados"].get(cid) or {}).get(registro)
                ws.append([
                    v["pregunta_id"], v.get("tema_nombre") or v["tema_id"], op["id"],
                    op["arquetipo"], op.get("arquetipo_secundario") or "",
                    cid, cand_by_id[cid],
                    txt if txt else "",
                    "FALSE" if txt else "TRUE",
                ])
                last_row = ws.max_row
                fill = arq_color_map.get(op["arquetipo"])
                if fill:
                    for c_idx in range(1, len(headers) + 1):
                        ws.cell(row=last_row, column=c_idx).fill = PatternFill("solid", fgColor=fill)
    style_header(ws, len(headers))
    autosize(ws, [12, 22, 16, 14, 14, 9, 22, 88, 14])
    wrap_rows(ws)
    ws.freeze_panes = "C2"


def load_variantes():
    if VARIANTES_FILE.exists():
        with open(VARIANTES_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {"variantes": []}


def main():
    arq = load("arquetipos")
    cand = load("candidatos")
    reg = load("registros")
    mt = load("mini_test")
    preg = load("preguntas")
    variantes = load_variantes()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quita la hoja default

    sheet_arquetipos(wb, arq)
    sheet_candidatos(wb, cand)
    sheet_registros(wb, reg)
    sheet_mini_test(wb, mt, [c["id"] for c in cand["candidatos"]])
    sheet_preguntas_base(wb, preg)
    for r in REGISTROS:
        sheet_opciones_registro(wb, preg, cand["candidatos"], r)
    sheet_prioridad(wb, preg)
    for r in REGISTROS:
        sheet_variantes_registro(wb, variantes, cand["candidatos"], r)

    wb.save(OUT)
    print(f"OK · {OUT}")
    print(f"   hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
