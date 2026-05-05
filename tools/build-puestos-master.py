#!/usr/bin/env python3
"""
tools/build-puestos-master.py

Cruza las dos hojas de Divipole_Congreso_CON DATOS.xlsx en un único CSV
maestro (puestos-master.csv) que reemplaza a PUESTOS_GEOREF.csv y a
COMUNAS_DATA.csv para los pipelines aguas abajo:

  Hoja Divipole_2026         → fuente PRIMARIA de censo (mujeres, hombres,
                               total, mesas) y de Longitud. Cubre 41,3 M
                               de censo (= padrón nacional oficial).
  Hoja reporte_HVP_27012026  → fuente PRIMARIA de LATITUD (la otra hoja
                               solo trae longitud por bug del Excel original).
                               También cubre LONGITUD como respaldo.

Cruce por `dd-mm-zz-pp` (= "Cod unico" sin formato). Cada registro lleva:

  dd, mm, zz, pp, codigo, departamento, municipio, nombre_puesto, direccion,
  comuna, mujeres, hombres, total, mesas, lat, lng

Cobertura esperada:
  - 13.491 puestos en territorio con lat + lng + censo (vista mapa completa).
  - ~255 puestos consulares (dep 88) con censo pero sin lat (no van al mapa
    pero entran en agregaciones nacionales). Quedan en el CSV con lat/lng vacíos.

Uso:
  python3 tools/build-puestos-master.py <input.xlsx> <out.csv>

Ejemplo:
  python3 tools/build-puestos-master.py \
    "/Users/ricardoruiz/ricardoruiz.co/Bases de datos/Divipole_Congreso_CON DATOS.xlsx" \
    "/Users/ricardoruiz/ricardoruiz.co/Bases de datos/puestos-master.csv"
"""

import csv
import sys
from pathlib import Path

import openpyxl


def to_int(v):
    try:
        return int(v) if v is not None else 0
    except (TypeError, ValueError):
        return 0


def to_float(v):
    try:
        if v in (None, ""):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def pad(v, n):
    if v is None:
        return "0" * n
    s = str(v).strip()
    return s.zfill(n) if s.isdigit() else s.zfill(n)


def read_divipole(ws):
    """Hoja Divipole_2026 → dict[key] = {censo, mesas, ..., lng}."""
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    cols = {h: i for i, h in enumerate(header)}
    out = {}
    for row in rows:
        dd = pad(row[cols["dd"]], 2)
        mm = pad(row[cols["mm"]], 3)
        zz = pad(row[cols["zz"]], 2)
        pp = pad(row[cols["pp"]], 2)
        key = f"{dd}-{mm}-{zz}-{pp}"
        out[key] = {
            "dd": dd, "mm": mm, "zz": zz, "pp": pp,
            "codigo": row[cols["Cod unico"]] or "",
            "departamento": row[cols["departamento"]] or "",
            "municipio": row[cols["municipio"]] or "",
            "nombre_puesto": row[cols["puesto"]] or "",
            "direccion": row[cols["dirección"]] or "",
            "comuna": row[cols["comuna"]] or "",
            "mujeres": to_int(row[cols["mujeres"]]),
            "hombres": to_int(row[cols["hombres"]]),
            "total": to_int(row[cols["total"]]),
            "mesas": to_int(row[cols["mesas"]]),
            # Longitud está en col Longitud (16) y Longitud2 (17). Ambas son LONGITUD.
            "lng_divipole": to_float(row[cols["Longitud"]]),
        }
    return out


def read_reporte_hvp(ws):
    """Hoja reporte_HVP_27012026 → dict[key] = {lat, lng}."""
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    cols = {h: i for i, h in enumerate(header)}
    out = {}
    cod_idx = cols["CÓDIGO COMPLETO"]
    lat_idx = cols["LATITUD"]
    lng_idx = cols["LONGITUD"]
    for row in rows:
        codcomp = row[cod_idx]
        if not codcomp:
            continue
        s = str(codcomp).zfill(9)
        if len(s) != 9:
            continue
        key = f"{s[0:2]}-{s[2:5]}-{s[5:7]}-{s[7:9]}"
        out[key] = {
            "lat": to_float(row[lat_idx]),
            "lng": to_float(row[lng_idx]),
        }
    return out


def main():
    if len(sys.argv) != 3:
        print("Uso: python3 tools/build-puestos-master.py <input.xlsx> <out.csv>", file=sys.stderr)
        sys.exit(1)

    inp = Path(sys.argv[1])
    out = Path(sys.argv[2])
    if not inp.exists():
        print(f"No existe: {inp}", file=sys.stderr)
        sys.exit(1)
    out.parent.mkdir(parents=True, exist_ok=True)

    print(f"\n[build-puestos-master] leyendo {inp.name}")
    wb = openpyxl.load_workbook(inp, read_only=True, data_only=True)
    print(f"  hojas detectadas: {wb.sheetnames}")

    div = read_divipole(wb["Divipole_2026"])
    print(f"  Divipole_2026         : {len(div):,} puestos · censo total = {sum(r['total'] for r in div.values()):,}")
    hvp = read_reporte_hvp(wb["reporte_HVP_27012026"])
    print(f"  reporte_HVP_27012026  : {len(hvp):,} puestos con lat/lng")

    # Cruce: tomamos la base de Divipole (más completa en censo).
    rows_out = []
    sin_lat = sin_lng = 0
    for key, r in div.items():
        coords = hvp.get(key, {})
        lat = coords.get("lat")
        lng = coords.get("lng") if coords.get("lng") is not None else r["lng_divipole"]
        if lat is None: sin_lat += 1
        if lng is None: sin_lng += 1
        rows_out.append({
            "dd": r["dd"], "mm": r["mm"], "zz": r["zz"], "pp": r["pp"],
            "codigo": r["codigo"],
            "departamento": r["departamento"],
            "municipio": r["municipio"],
            "nombre_puesto": r["nombre_puesto"],
            "direccion": r["direccion"],
            "comuna": r["comuna"],
            "mujeres": r["mujeres"],
            "hombres": r["hombres"],
            "total": r["total"],
            "mesas": r["mesas"],
            "lat": "" if lat is None else f"{lat:.10g}",
            "lng": "" if lng is None else f"{lng:.10g}",
        })

    # Reporte de cobertura.
    total_censo = sum(r["total"] for r in rows_out)
    en_terr = [r for r in rows_out if r["dd"] != "88"]
    consul = [r for r in rows_out if r["dd"] == "88"]
    total_terr = sum(r["total"] for r in en_terr)
    total_consul = sum(r["total"] for r in consul)
    con_lat = sum(1 for r in rows_out if r["lat"])
    con_lat_terr = sum(1 for r in en_terr if r["lat"])

    # Escribir CSV.
    fieldnames = ["dd","mm","zz","pp","codigo","departamento","municipio","nombre_puesto","direccion","comuna","mujeres","hombres","total","mesas","lat","lng"]
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        w.writeheader()
        for r in rows_out:
            w.writerow(r)

    sz = out.stat().st_size
    print(f"\n  ✓ {out.name}   {len(rows_out):,} filas · {sz/1024:.0f} KB")
    print(f"\n  Cobertura:")
    print(f"    Total puestos     : {len(rows_out):,}  · censo total = {total_censo:,}")
    print(f"    En territorio (≠88): {len(en_terr):,}  · censo = {total_terr:,}  · con lat = {con_lat_terr:,}")
    print(f"    Consulares  (dep 88): {len(consul):,}  · censo = {total_consul:,}  · con lat = {sum(1 for r in consul if r['lat']):,}")
    print(f"    Sin LAT           : {sin_lat:,} (sin estos no van al mapa, sí entran en agregados)")
    print(f"    Sin LNG           : {sin_lng:,}")
    print()


if __name__ == "__main__":
    main()
