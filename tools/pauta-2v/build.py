# -*- coding: utf-8 -*-
"""
Plan de Pauta Digital · Segunda Vuelta 2026 (campaña Cepeda)
============================================================
Re-corta el pipeline del informe Pacto (output_pacto_1v_2026) + el análisis
de edad (output_edad_1v) + PUESTOS_GEOREF en un FORMATO DE COMPRA DE MEDIOS:

  - Lista priorizada de municipios para MOVILIZACIÓN broad (geo+demo) ponderada
    por el NETO de izquierda (solo donde el bloque Petro-2V gana), porque prender
    abstención donde gana la derecha le suma a Abelardo.
  - Lista priorizada de municipios para PERSUASIÓN de centro (Fajardo+Claudia).
  - Universo de RECUPERACIÓN (techo Petro-2V) para custom audiences.
  - Hiperlocal: comunas/zonas de las ciudades grandes con centroide lat/lon
    para targeting por radio en Meta.
  - Edad (Cepeda gana jóvenes) y género (brecha) como palancas de segmento/mensaje.

Salidas (en Bases de datos/output_pauta_2v/):
  - Plan_Pauta_Digital_2V.xlsx   (entregable Excel, 6 hojas)
  - pauta-data.json              (se inyecta INLINE en el HTML privado)
"""
import csv, json, os, unicodedata
from collections import defaultdict

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
PACTO = os.path.join(ROOT, "Bases de datos", "output_pacto_1v_2026")
EDAD  = os.path.join(ROOT, "Bases de datos", "output_edad_1v")
GEOREF = os.path.join(ROOT, "Bases de datos", "PUESTOS_GEOREF.csv")
OUT   = os.path.join(ROOT, "Bases de datos", "output_pauta_2v")
os.makedirs(OUT, exist_ok=True)

def norm(s):
    if not s: return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode("ascii")
    return s.upper().strip()

# ── 1. Modelo nacional + territorial del informe Pacto ──────────────────────
model = json.load(open(os.path.join(PACTO, "twov_model.json")))
terr  = json.load(open(os.path.join(PACTO, "twov_territorial.json")))
estr  = json.load(open(os.path.join(PACTO, "twov_estrategias.json")))

muni_terr = { m["cod"]: m for m in terr["muni"] }   # cod = dep2+mun3 (ej "16001")
puestos   = terr["puesto"]                            # pcode 9 dígitos + barrio/comuna

# ── 2. Georef: censo (M/H), centroide y lat/lon por puesto ──────────────────
geo_pue = {}                       # 9-dígitos -> (lat, lon, cM, cH)
muni_geo = defaultdict(lambda: {"cM":0,"cH":0,"sumla":0.0,"sumlo":0.0,"w":0.0,"npue":0})
with open(GEOREF, encoding="utf-8-sig") as f:
    r = csv.DictReader(f, delimiter=";")
    for row in r:
        cc = (row.get("CÓDIGO COMPLETO") or "").strip()
        if len(cc) != 9: continue
        zona = cc[5:7]
        try: la = float(row.get("LATITUD") or "nan"); lo = float(row.get("LONGITUD") or "nan")
        except: la = lo = float("nan")
        try: cM = int(float(row.get("MUJERES") or 0)); cH = int(float(row.get("HOMBRES") or 0))
        except: cM = cH = 0
        good_geo = (-5 < la < 14) and (-82 < lo < -66)
        if good_geo:
            geo_pue[cc] = (la, lo, cM, cH)
        if zona in ("90","98"):    # puesto censo / cárceles: ruido geográfico
            continue
        mcod = cc[:5]
        g = muni_geo[mcod]
        g["cM"] += cM; g["cH"] += cH; g["npue"] += 1
        if good_geo and (cM+cH) > 0:
            w = cM + cH
            g["sumla"] += la*w; g["sumlo"] += lo*w; g["w"] += w

# ── 3. Votantes 2026 por municipio (master) para abstención ─────────────────
master = json.load(open(os.path.join(PACTO, "master_2026_puesto.json")))
votantes_muni = defaultdict(int)
for p in master:
    dep = str(p.get("dep","")).zfill(2); mun = str(p.get("mun","")).zfill(3)
    zona = str(p.get("zona","")).zfill(2)
    if dep == "88" or zona in ("90","98"): continue   # exterior / ruido
    votantes_muni[dep+mun] += int(p.get("total_votos_urna",0) or 0)

# ── 4. Edad: share de Cepeda entre 18-35 por departamento (ei-deptos.csv) ────
cep_young_dep = {}     # dep2 -> share Cepeda 18-35
with open(os.path.join(EDAD, "ei-deptos.csv"), encoding="utf-8") as f:
    for row in csv.DictReader(f):
        if row["grupo"] == "18-35":
            cep_young_dep[row["dep"].zfill(2)] = float(row["cepeda"])

# Composición etaria del electorado por municipio (w26-puesto.csv, bandas b0..b9)
young_cnt = defaultdict(float); tot_cnt = defaultdict(float)
with open(os.path.join(EDAD, "w26-puesto.csv"), encoding="utf-8") as f:
    for row in csv.DictReader(f):
        pc = row["pcode"].split("-")          # dep-mun-zona-puesto
        if len(pc) < 2: continue
        mcod = pc[0].zfill(2) + pc[1].zfill(3)
        try:
            bands = [float(row[f"b{i}"]) for i in range(10)]
        except: continue
        young_cnt[mcod] += sum(bands[:4])      # 18-20,21-25,26-30,31-35
        tot_cnt[mcod]   += sum(bands)

# Edad nacional + por ciudad (ei-ciudades.csv) para el HTML
edad_ciudades = defaultdict(dict)
with open(os.path.join(EDAD, "ei-ciudades.csv"), encoding="utf-8") as f:
    for row in csv.DictReader(f):
        edad_ciudades[row["unit"]][row["grupo"]] = {
            "cepeda": round(float(row["cepeda"]),3),
            "abelardo": round(float(row["abelardo"]),3),
        }

# ── 5. Tabla maestra por municipio ──────────────────────────────────────────
munis = []
for cod, m in muni_terr.items():
    base   = m.get("base",0) or 0
    techo  = m.get("techo",0) or 0
    cepnow = m.get("cep_now",0) or 0
    recup  = max(0, m.get("recuperar",0) or 0)      # techo - cepeda (universo)
    centro = max(0, m.get("centro",0) or 0)          # 0.55·Faj + 0.65·Cla
    g = muni_geo.get(cod, {})
    censo  = (g.get("cM",0) + g.get("cH",0)) if g else 0
    cM = g.get("cM",0) if g else 0; cH = g.get("cH",0) if g else 0
    lat = round(g["sumla"]/g["w"],5) if g and g.get("w") else None
    lon = round(g["sumlo"]/g["w"],5) if g and g.get("w") else None
    votantes = votantes_muni.get(cod, 0)
    share_izq2v = (techo/base) if base else 0          # techo Petro-2V / total
    abst_count  = max(0, censo - votantes) if censo else 0
    # NETO de movilización: solo positivo donde la izquierda gana la 2V
    abst_neto   = int(abst_count * (2*share_izq2v - 1)) if share_izq2v > 0.5 else 0
    young_share = (young_cnt.get(cod,0)/tot_cnt[cod]) if tot_cnt.get(cod) else None
    munis.append({
        "cod": cod, "dep": m.get("dep",""), "muni": m.get("muni",""),
        "cep_now": cepnow, "techo": techo, "recuperar": recup, "centro": centro,
        "base": base, "votantes": votantes, "censo": censo,
        "share_izq2v": round(share_izq2v,3), "abst_count": abst_count,
        "abst_neto": abst_neto,
        "men_share": round(cH/censo,3) if censo else None,
        "young_share": round(young_share,3) if young_share is not None else None,
        "cep_young": round(cep_young_dep.get(cod[:2],0),3) if cod[:2] in cep_young_dep else None,
        "lat": lat, "lon": lon,
    })

# Pesos de presupuesto
sum_neto   = sum(x["abst_neto"]  for x in munis) or 1
sum_centro = sum(x["centro"]     for x in munis) or 1
sum_recup  = sum(x["recuperar"]  for x in munis) or 1
for x in munis:
    x["mob_w"] = round(x["abst_neto"]/sum_neto, 5)      # movilización broad (neto)
    x["per_w"] = round(x["centro"]/sum_centro, 5)        # persuasión centro
    x["rec_w"] = round(x["recuperar"]/sum_recup, 5)      # universo recuperación

n_mob = sum(1 for x in munis if x["abst_neto"] > 0)
print(f"municipios: {len(munis)} | con neto>0 (movilización broad): {n_mob} | "
      f"Σneto={sum_neto:,} | Σcentro={sum_centro:,} | Σrecuperar={sum_recup:,}")

# ── 6. Hiperlocal: comunas/zonas de ciudades grandes (join puesto↔georef) ────
CITY_NAMES = {  # cod muni -> nombre ciudad
    "16001":"Bogotá","01001":"Medellín","31001":"Cali","03001":"Barranquilla",
    "05001":"Cartagena","09001":"Manizales","66001":"Pereira","68001":"Bucaramanga",
    "54001":"Cúcuta","41001":"Neiva","23001":"Montería","52001":"Pasto",
    "50001":"Villavicencio","47001":"Santa Marta","76001":"Cali",
}
comuna_agg = defaultdict(lambda: {"recuperar":0,"censo":0,"sumla":0.0,"sumlo":0.0,"w":0.0,"npue":0})
for p in puestos:
    cod = str(p.get("dep","")).zfill(2) + str(p.get("mun","")).zfill(3)
    if cod not in CITY_NAMES: continue
    comuna = (p.get("comuna") or "").strip() or "(sin comuna)"
    pc = str(p.get("pcode","")).zfill(9)
    key = (cod, comuna)
    a = comuna_agg[key]
    a["recuperar"] += max(0, p.get("recuperar",0) or 0)
    a["npue"] += 1
    gp = geo_pue.get(pc)
    if gp:
        la, lo, cM, cH = gp
        w = cM + cH
        a["censo"] += w
        if w > 0:
            a["sumla"] += la*w; a["sumlo"] += lo*w; a["w"] += w

comunas = defaultdict(list)
for (cod, comuna), a in comuna_agg.items():
    if a["recuperar"] <= 0: continue
    comunas[cod].append({
        "comuna": comuna, "recuperar": int(a["recuperar"]), "censo": int(a["censo"]),
        "npue": a["npue"],
        "lat": round(a["sumla"]/a["w"],5) if a["w"] else None,
        "lon": round(a["sumlo"]/a["w"],5) if a["w"] else None,
    })
for cod in comunas:
    comunas[cod].sort(key=lambda r:-r["recuperar"])

# ── 7. JSON para el HTML (inline) ───────────────────────────────────────────
meta = {
    "fecha_2v": "2026-06-21",
    "total_1v": model["total_1v"],
    "cepeda_1v": model["votos"]["cepeda"], "abelardo_1v": model["votos"]["abelardo"],
    "abe_floor": model["abe_floor"], "cep_ceiling": model["cep_ceiling"],
    "gap": model["gap"], "center_contrib": model["center_contrib"],
    "need_over_1v": model["need_over_1v"],
    "recuperar_total": estr["recuperar_total"], "abst_target": estr["abst_target"],
    "abst_n_muni": estr["abst_n_muni"], "centro_total": estr["centro_total"],
    "sum_neto": sum_neto, "n_mob": n_mob,
    "supuestos": model["supuestos"],
}
# munis compactos para el HTML (solo lo que pinta el mapa/tabla)
munis_html = [{
    "cod":x["cod"],"dep":x["dep"],"muni":x["muni"],
    "rec":x["recuperar"],"cen":x["centro"],"net":x["abst_neto"],"base":x["base"],"techo":x["techo"],
    "censo":x["censo"],"part": round(x["votantes"]/x["censo"],3) if x["censo"] else None,
    "izq":x["share_izq2v"],"young":x["young_share"],"cy":x["cep_young"],
    "men":x["men_share"],"lat":x["lat"],"lon":x["lon"],
    "mw":x["mob_w"],"pw":x["per_w"],
} for x in munis]

# nombres reales (se inyectan en runtime tras el gate; el HTML público queda neutro)
NAMES = {"cand":"Cepeda","rival":"Abelardo","izq22":"Petro","faj":"Fajardo","cla":"Claudia"}
# índice de la capa por barrio (lo genera build_barrios.py)
_bidx_path = os.path.join(OUT, "barrios", "_index.json")
barrio_index = json.load(open(_bidx_path)) if os.path.exists(_bidx_path) else []

data_html = {
    "meta": meta,
    "names": NAMES,
    "edad": {k:dict(v) for k,v in edad_ciudades.items()},
    "munis": munis_html,
    "comunas": {k:v[:25] for k,v in comunas.items()},  # top 25 comunas por ciudad
    "city_names": CITY_NAMES,
    "barrio_index": barrio_index,
}
with open(os.path.join(OUT, "pauta-data.json"), "w", encoding="utf-8") as f:
    json.dump(data_html, f, ensure_ascii=False, separators=(",",":"))
print("→ pauta-data.json", round(os.path.getsize(os.path.join(OUT,"pauta-data.json"))/1024,1), "KB")

# ── 8. Excel entregable ─────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OX = "8A1E16"; CREAM = "F4F0E7"; INK = "14110A"
hdr_fill = PatternFill("solid", fgColor=OX)
hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
zebra = PatternFill("solid", fgColor=CREAM)
thin = Side(style="thin", color="DDD6C8")
border = Border(left=thin,right=thin,top=thin,bottom=thin)

def style_sheet(ws, ncols, freeze="A2"):
    for c in range(1, ncols+1):
        cell = ws.cell(row=1, column=c); cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"
    for r in range(2, ws.max_row+1):
        for c in range(1, ncols+1):
            cell = ws.cell(row=r, column=c); cell.border = border
            if r % 2 == 0: cell.fill = zebra
            v = cell.value
            if isinstance(v,(int,float)) and not isinstance(v,bool) and abs(v) >= 1000:
                cell.number_format = "#,##0"

wb = openpyxl.Workbook()

# Hoja 0 · Léeme
ws = wb.active; ws.title = "Léeme"
ws.column_dimensions["A"].width = 110
LEE = [
    ("Plan de Pauta Digital · Segunda Vuelta 2026", True),
    ("Documento de trabajo · uso interno de campaña · borrador automático.", False),
    ("", False),
    ("CÓMO LEER ESTE EXCEL", True),
    ("Cada hoja es una lista de targeting lista para cargar en el administrador de anuncios de Meta / Google Ads.", False),
    ("Las cifras salen del preconteo 1V por mesa + el modelo de 2ª vuelta (techo Petro-2V) + el análisis de edad por inferencia ecológica.", False),
    ("", False),
    ("LAS DOS CAMPAÑAS (no se suman igual)", True),
    (f"· MOVILIZACIÓN (GOTV): universo de la izquierda demovilizada. Objetivo nacional ≈ {estr['abst_target']:,} votos en {estr['abst_n_muni']} municipios.", False),
    ("  REGLA DE ORO: la pauta broad por geografía se concentra SOLO donde la izquierda GANA la 2ª vuelta (columna 'Neto izq.').", False),
    ("  Prender abstención donde gana Abelardo le suma a Abelardo. Para llegar a simpatizantes en zona adversa se usan listas propias (custom audiences), no geo abierto.", False),
    (f"· PERSUASIÓN DE CENTRO: voto transferible de Fajardo+Claudia ≈ {estr['centro_total']:,}. Mensaje moderado / anti-extremos. Sobre todo urbano.", False),
    ("", False),
    ("Centro (~0,7M) SE SUMA a Movilización (~1,9M) = ~2,65M (lo que Cepeda debe sumar sobre su 1V).", False),
    ("Recuperación (techo, ~2,05M) y Movilización son el MISMO universo visto distinto: no se suman entre sí.", False),
    ("", False),
    ("COLUMNAS", True),
    ("Recuperar = techo Petro-2V 2022 − Cepeda 1V 2026 (universo de la izquierda en ese territorio).", False),
    ("Neto izq. = abstención × (2·share_izq2v − 1), solo si la izquierda gana la 2V. Es el voto NETO que rinde movilizar broad ahí.", False),
    ("Centro disp. = 0,55·Fajardo + 0,65·Claudia (supuestos de transferencia del modelo).", False),
    ("% pres. = peso del municipio en el presupuesto de esa campaña (proporción del universo).", False),
    ("Lat/Lon = centroide del municipio (ponderado por censo) para targeting por PIN + RADIO en Meta.", False),
    ("Cepeda jóvenes = share de Cepeda entre 18-35 en el depto (inferencia ecológica). Alto = pauta a jóvenes muy eficiente.", False),
    ("% jóvenes = composición 18-35 del electorado proyectado del municipio.", False),
    ("", False),
    ("ANTES DE PAUTAR — 3 ALERTAS", True),
    ("1. Meta exige AUTORIZACIÓN de anuncios políticos (verificación + disclaimer 'pagado por'). Tarda días. Verificar HOY si la cuenta ya está autorizada.", False),
    ("2. El gasto digital cuenta para los topes de campaña (CNE) y se reporta a Cuentas Claras. Cruzar con el contador.", False),
    ("3. TikTok NO acepta pauta política pagada (solo orgánico + creadores). Veda/propaganda: confirmar fechas con la jurídica.", False),
]
for i,(txt,bold) in enumerate(LEE, start=1):
    c = ws.cell(row=i, column=1, value=txt)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if bold: c.font = Font(bold=True, color=OX, size=12 if i==1 else 11)

def add_sheet(title, headers, rows, widths):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for r in rows: ws.append(r)
    style_sheet(ws, len(headers))
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

dep_name = {x["cod"]: x["dep"] for x in munis}

# Hoja 1 · Movilización (broad, neto>0), ordenada por neto desc
mob = sorted([x for x in munis if x["abst_neto"]>0], key=lambda x:-x["abst_neto"])
rows = [[x["muni"], x["dep"], x["abst_neto"], x["recuperar"], x["censo"],
         round(x["votantes"]/x["censo"],3) if x["censo"] else None,
         round(x["share_izq2v"]*100,1), round(x["mob_w"]*100,2),
         x["cep_young"], x["young_share"], x["lat"], x["lon"]] for x in mob]
add_sheet("Movilización (municipios)",
    ["Municipio","Departamento","Neto izq.","Recuperar (techo)","Censo","Participación 1V",
     "Share izq. 2V (%)","% presupuesto","Cepeda jóvenes","% jóvenes","Lat","Lon"],
    rows, [26,20,12,14,12,12,12,12,11,10,11,11])

# Hoja 2 · Persuasión centro
per = sorted([x for x in munis if x["centro"]>0], key=lambda x:-x["centro"])
rows = [[x["muni"], x["dep"], x["centro"], x["censo"], round(x["per_w"]*100,2),
         round(x["share_izq2v"]*100,1), x["lat"], x["lon"]] for x in per]
add_sheet("Persuasión centro (municipios)",
    ["Municipio","Departamento","Centro disp.","Censo","% presupuesto",
     "Share izq. 2V (%)","Lat","Lon"],
    rows, [26,20,13,12,12,14,11,11])

# Hoja 3 · Universo recuperación (todos, para custom audiences)
rec = sorted(munis, key=lambda x:-x["recuperar"])
rows = [[x["muni"], x["dep"], x["recuperar"], x["cep_now"], x["techo"], x["censo"],
         round(x["share_izq2v"]*100,1)] for x in rec if x["recuperar"]>0]
add_sheet("Universo recuperación",
    ["Municipio","Departamento","Recuperar (techo)","Cepeda 1V","Techo Petro-2V","Censo","Share izq. 2V (%)"],
    rows, [26,20,16,12,15,12,14])

# Hoja 4 · Hiperlocal ciudades (comunas con centroide para radio)
rows = []
for cod, lst in comunas.items():
    city = CITY_NAMES.get(cod, cod)
    for r in lst:
        rows.append([city, r["comuna"], r["recuperar"], r["censo"], r["npue"], r["lat"], r["lon"]])
rows.sort(key=lambda r:(r[0], -r[2]))
add_sheet("Hiperlocal (comunas)",
    ["Ciudad","Comuna / Localidad","Recuperar","Censo","# puestos","Lat","Lon"],
    rows, [16,30,12,12,11,11,11])

# Hoja 5 · Presupuesto y plataformas (plantilla con CPM)
ws = wb.create_sheet("Presupuesto y plataformas")
ws.column_dimensions["A"].width = 34
for c in "BCDE": ws.column_dimensions[c].width = 18
plan = [
    ["Plataforma","Rol","Tipo de targeting","Split sugerido","Nota"],
    ["Meta (FB/IG)","Movilización + persuasión","Geo (municipio + pin/radio) · edad · género","45%","Backbone. Requiere autorización política."],
    ["YouTube / Google","Alcance + persuasión","Depto/ciudad · edad · afinidad","20%","Video corto (bumper 6s) + skippable."],
    ["TikTok (orgánico)","Movilización jóvenes","Creadores + orgánico (NO pauta)","—","Base joven de Cepeda. Sin pauta pagada."],
    ["WhatsApp / territorial","GOTV puerta a puerta digital","Listas propias por municipio","15%","Difusión + clic-to-WhatsApp desde Meta."],
    ["X (Twitter)","Conversación / élite","Orgánico + amplificación","—","Premium no es herramienta de campaña."],
    ["Reserva / testing","A/B + contingencia","—","20%","Reasignar a lo que mejor convierta."],
    [],
    ["CALCULADORA DE ALCANCE (editable)","","","",""],
    ["Presupuesto total (COP)","10000000","","",""],
    ["% a movilización","70%","","",""],
    ["% a persuasión","30%","","",""],
    ["CPM Meta (COP / 1000 impres.)","12000","Benchmark $2-4 USD; sube en 2V por competencia",""],
    ["Frecuencia objetivo","3.5","veces que ve el anuncio cada persona",""],
    ["→ Impresiones estimadas","=B11/B14*1000","",""],
    ["→ Personas alcanzadas (aprox)","=B16/B15","",""],
]
for r in plan: ws.append(r)
for c in range(1,6):
    cell = ws.cell(row=1,column=c); cell.fill=hdr_fill; cell.font=hdr_font
    cell.alignment=Alignment(horizontal="center",wrap_text=True)
ws.cell(row=9,column=1).font = Font(bold=True, color=OX, size=11)
ws.cell(row=10,column=2).number_format = "#,##0"
ws.cell(row=16,column=2).number_format = "#,##0"
ws.cell(row=17,column=2).number_format = "#,##0"

path = os.path.join(OUT, "Plan_Pauta_Digital_2V.xlsx")
wb.save(path)
print("→ Plan_Pauta_Digital_2V.xlsx", round(os.path.getsize(path)/1024,1), "KB")
print("\nTOP 8 movilización (neto):")
for x in mob[:8]:
    print(f"  {x['muni'][:24]:24} {x['dep'][:14]:14} neto={x['abst_neto']:>8,} izq2v={x['share_izq2v']*100:4.1f}%  cy={x['cep_young']}")
print("\nTOP 8 persuasión centro:")
for x in per[:8]:
    print(f"  {x['muni'][:24]:24} {x['dep'][:14]:14} centro={x['centro']:>7,}")
