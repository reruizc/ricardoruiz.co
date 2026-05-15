#!/usr/bin/env python3
"""
tools/build-bog-barrio.py

Genera señales electorales + censos por BARRIO CATASTRAL de Bogotá.
Mismo patrón que el flujo UPL: point-in-polygon de puestos contra el
GeoJSON oficial BOG-BARRIOS-CATASTRALES (1.001 features).

Inputs:
  PUESTOS_GEOREF.csv             (puestos Bogotá con lat/lon)
  BOG-BARRIOS-CATASTRALES.geojson (1001 barrios con codigo, nombre, loc_*)
  COMUNAS_DATA.csv                (censo 2026 por puesto)
  Divipol 23.09.2021.xlsx         (censo 2022)
  out_dir                          (donde escribir los JSONs)

Outputs (en out_dir):
  bog-barrio-<sig>.json × 10    Señales históricas + senado/cámara 2026.
  censo-barrio-{2018,2022,2026}.json
  bog-puesto-to-barrio.json     Mapeo crudo {puestoKey: barrioCod}.

NOTAS:
- 1001 barrios vs 1084 puestos: muchos barrios sin puestos asignados.
  El frontend hereda del vecino más cercano (kNN por centroide).
- Las 4 señales nacionales por puesto (pres-2010/14/18/22 + 4 consultas)
  cargan ~12 MB. Aquí se descargan y filtran solo Bogotá → puesto → barrio.
"""

import csv, json, os, sys, re, urllib.request
from collections import defaultdict

S3 = 'https://elecciones-2026.s3.us-east-1.amazonaws.com/ricardoruiz.co/congreso-2026/output'

SIGNALS = [
    ('pres-2010',                'nac-pres',     f'{S3}/historicos/pres-2010-v1/por-puesto.json'),
    ('pres-2014',                'nac-pres',     f'{S3}/historicos/pres-2014-v1/por-puesto.json'),
    ('pres-2018',                'nac-pres',     f'{S3}/historicos/pres-2018-v1/por-puesto.json'),
    ('pres-2022',                'nac-pres',     f'{S3}/historicos/pres-2022-v1/por-puesto.json'),
    ('consulta-2025-pacto',      'nac-consulta', f'{S3}/historicos/consulta-2025-pacto/por-puesto.json'),
    ('consulta-2026-gran',       'nac-consulta', f'{S3}/historicos/consulta-2026-gran/por-puesto.json'),
    ('consulta-2026-frente',     'nac-consulta', f'{S3}/historicos/consulta-2026-frente/por-puesto.json'),
    ('consulta-2026-soluciones', 'nac-consulta', f'{S3}/historicos/consulta-2026-soluciones/por-puesto.json'),
    ('senado-2026',              'cong-2026',    f'{S3}/senado/departamentos/16/puestos.json'),
    ('camara-2026',              'cong-2026',    f'{S3}/camara/departamentos/16/puestos.json'),
]

BOG_DEP_ELEC, BOG_MUN_ELEC = '16', '001'
BOG_DEP_DANE = '11'
PFX_NAC = f'{BOG_DEP_DANE}-{BOG_MUN_ELEC}-'    # archivos por-puesto.json usan DANE
CACHE_DIR = '/tmp/bog-barrio-cache'

# Censos
BOG_2018_DEP_CENSO = 5_703_232
BOG_2022_DEP_CENSO = 5_630_748


def pad2(s): return str(s).zfill(2)
def pad3(s): return str(s).zfill(3)


def fetch_cached(url):
    os.makedirs(CACHE_DIR, exist_ok=True)
    fname = os.path.join(CACHE_DIR, re.sub(r'[^a-zA-Z0-9]+', '_', url) + '.json')
    if os.path.exists(fname) and os.path.getsize(fname) > 100:
        return json.load(open(fname))
    print(f'[fetch] {url}', file=sys.stderr)
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=120) as r:
        data = r.read()
    with open(fname, 'wb') as f: f.write(data)
    return json.loads(data)


# ── point-in-polygon ────────────────────────────────────────────────
def point_in_ring(x, y, ring):
    inside = False; n = len(ring); j = n - 1
    for i in range(n):
        xi, yi = ring[i][0], ring[i][1]
        xj, yj = ring[j][0], ring[j][1]
        if ((yi > y) != (yj > y)) and \
           (x < (xj - xi) * (y - yi) / ((yj - yi) or 1e-12) + xi):
            inside = not inside
        j = i
    return inside


def point_in_feature(x, y, geom):
    polys = [geom['coordinates']] if geom['type'] == 'Polygon' else geom['coordinates']
    for poly in polys:
        if not poly or not point_in_ring(x, y, poly[0]): continue
        in_hole = any(point_in_ring(x, y, ring) for ring in poly[1:])
        if not in_hole: return True
    return False


def feature_bbox(geom):
    minx, miny, maxx, maxy = 1e9, 1e9, -1e9, -1e9
    polys = [geom['coordinates']] if geom['type'] == 'Polygon' else geom['coordinates']
    for poly in polys:
        for ring in poly:
            for x, y in ring:
                if x < minx: minx = x
                if x > maxx: maxx = x
                if y < miny: miny = y
                if y > maxy: maxy = y
    return minx, miny, maxx, maxy


# ── Carga puestos Bogotá ────────────────────────────────────────────
def norm_barrio(s):
    """Normaliza un nombre de barrio para matching. Idéntica lógica del
    frontend (normBarrioFront): strip tildes + UPPER + colapsar no-alfanum."""
    if not s: return ''
    import unicodedata
    s = unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode('ascii').upper()
    return re.sub(r'[^A-Z0-9 ]+', ' ', s).strip()


def load_puestos_bogota(csv_path):
    # Zonas especiales (puestos censo / cárceles) — NO se asignan a barrios:
    #   90 = PUESTO CENSO (FERIA EXPOSICIÓN / CORFERIAS).
    #        Recoge votos de ciudadanos sin asignación específica; al hacer
    #        PIP cae artificialmente en el barrio donde está físicamente
    #        (Centro Nariño / La Esmeralda) e infla esas señales.
    #   98 = Cárceles (Picota, Modelo, Buen Pastor, Cárcel Distrital).
    SPECIAL_ZONAS = {'90', '98'}
    out = {}
    n_skip = 0
    with open(csv_path, encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        header = next(reader); header[0] = header[0].replace('﻿', '')
        col = lambda n: next(i for i, h in enumerate(header) if h.strip().upper() == n.upper())
        iDep = col('DEPARTAMENTO'); iZon = col('ZONA'); iPue = col('PUESTO')
        iLat = col('LATITUD'); iLon = col('LONGITUD'); iNom = col('NOMBRE PUESTO')
        iBar = col('BARRIO')
        for row in reader:
            if not row or len(row) <= iLon: continue
            if 'BOGOTA' not in (row[iDep] or '').upper(): continue
            zz = pad2(row[iZon]); pp = pad2(row[iPue])
            if zz in SPECIAL_ZONAS:
                n_skip += 1
                continue
            try:
                lat = float((row[iLat] or '').replace(',', '.'))
                lon = float((row[iLon] or '').replace(',', '.'))
            except ValueError:
                continue
            out[f'{zz}-{pp}'] = {
                'lat': lat, 'lon': lon,
                'nombre': row[iNom] or '',
                'barrio': (row[iBar] or '').strip(),
                'barrio_norm': norm_barrio(row[iBar] or ''),
            }
    if n_skip:
        print(f'    {n_skip} puestos especiales (zona 90 censo, 98 cárceles) excluidos', file=sys.stderr)
    return out


def build_puesto_to_barrio(puestos, barrio_geo):
    """Asigna cada puesto a un barrio catastral.

    Estrategia (en orden de preferencia):
    1) Match por NOMBRE: si el campo BARRIO del CSV de Registraduría
       coincide (normalizado) con el `nombre` del GeoJSON catastral.
       Es la fuente más confiable porque ambos vienen de listas
       oficiales y los nombres se corresponden bien.
    2) Match con sufijo: a veces el CSV trae "QUINTA PAREDES B" y el
       catastral sólo "QUINTA PAREDES". Probamos quitar el último token
       si es una letra/dígito suelto.
    3) Fallback PIP: punto-en-polígono del lat/lon contra el catastral.
       Es el viejo método; sólo se usa cuando los pasos 1-2 fallan."""
    # Index por nombre normalizado
    by_name = {}
    barrios_geo = []
    for f in barrio_geo['features']:
        cod = f['properties'].get('codigo')
        if not cod: continue
        nm = norm_barrio(f['properties'].get('nombre', ''))
        if nm:
            by_name.setdefault(nm, []).append(cod)
        barrios_geo.append((cod, f['properties'], feature_bbox(f['geometry']), f['geometry']))

    out = {}; unmatched = []
    n_by_name = 0; n_by_suffix = 0; n_by_pip = 0
    for zzpp, p in puestos.items():
        assigned = None
        # 1) Match por nombre exacto
        bn = p.get('barrio_norm', '')
        if bn and bn in by_name and len(by_name[bn]) == 1:
            assigned = by_name[bn][0]; n_by_name += 1
        # 2) Sufijo de una letra/dígito ("QUINTA PAREDES B" → "QUINTA PAREDES")
        if not assigned and bn:
            toks = bn.split()
            if len(toks) >= 2 and len(toks[-1]) <= 2:
                bn2 = ' '.join(toks[:-1])
                if bn2 in by_name and len(by_name[bn2]) == 1:
                    assigned = by_name[bn2][0]; n_by_suffix += 1
        # 3) Fallback PIP por lat/lon
        if not assigned:
            x, y = p['lon'], p['lat']
            for cod, props, bb, geom in barrios_geo:
                if x < bb[0] or x > bb[2] or y < bb[1] or y > bb[3]: continue
                if point_in_feature(x, y, geom):
                    assigned = cod; n_by_pip += 1; break
        if assigned: out[zzpp] = assigned
        else: unmatched.append(zzpp)
    print(f'    match: {n_by_name} por nombre + {n_by_suffix} por sufijo + {n_by_pip} por PIP', file=sys.stderr)
    return out, unmatched


def iter_puestos_signal(json_data, fmt):
    if fmt == 'nac-pres':
        cands = json_data.get('candidatos', {})
        for k, v in (json_data.get('puestos') or {}).items():
            if not k.startswith(PFX_NAC): continue
            parts = k.split('-')
            zzpp = f'{pad2(parts[2])}-{pad2(parts[3])}'
            votos = {}
            for cod, n in (v.get('v') or {}).items():
                nombre = cands.get(cod, {}).get('nombre')
                if not nombre: continue
                votos[nombre] = votos.get(nombre, 0) + (n or 0)
            yield zzpp, v.get('vv', 0) or 0, votos
    elif fmt == 'nac-consulta':
        for k, v in (json_data.get('puestos') or {}).items():
            if not k.startswith(PFX_NAC): continue
            parts = k.split('-')
            zzpp = f'{pad2(parts[2])}-{pad2(parts[3])}'
            votos = {}
            for nombre, n in (v.get('v') or {}).items():
                votos[nombre] = votos.get(nombre, 0) + (n or 0)
            yield zzpp, v.get('vv', 0) or 0, votos
    elif fmt == 'cong-2026':
        for p in (json_data or []):
            if p.get('dep_cod') != BOG_DEP_ELEC or p.get('mun_cod') != BOG_MUN_ELEC: continue
            zzpp = f"{pad2(p.get('zon_cod'))}-{pad2(p.get('pue_cod_raw'))}"
            votos = {}
            for k, n in (p.get('partidos') or {}).items():
                if re.match(r'^99[6-9]$', str(k)): continue
                votos[k] = votos.get(k, 0) + (n or 0)
            yield zzpp, p.get('votval', 0) or 0, votos


# ── Censos ──────────────────────────────────────────────────────────
def load_censo_2026_bogota(csv_path):
    out = {}
    with open(csv_path, encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        header = next(reader); header[0] = header[0].replace('﻿', '')
        col = lambda n: next(i for i, h in enumerate(header) if h.strip().lower() == n.lower())
        iDd = col('dd'); iMm = col('mm'); iZz = col('zz'); iPp = col('pp')
        iTot = col('total'); iMuj = col('mujeres'); iHom = col('hombres')
        for row in reader:
            if not row or len(row) <= iTot: continue
            if pad2(row[iDd]) != BOG_DEP_ELEC or pad3(row[iMm]) != BOG_MUN_ELEC: continue
            zz = pad2(row[iZz]); pp = pad2(row[iPp])
            try: tot = int(row[iTot] or 0)
            except: tot = 0
            if tot <= 0: continue
            try: muj = int(row[iMuj] or 0)
            except: muj = 0
            try: hom = int(row[iHom] or 0)
            except: hom = 0
            out[f'{zz}-{pp}'] = {'total': tot, 'mujeres': muj, 'hombres': hom}
    return out


def load_censo_2022_bogota(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = None; header_idx = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if row and str(row[0]).strip().lower() == 'dd':
            for j, h in enumerate(row):
                if h is not None: header_idx[str(h).strip().lower()] = j
            header_row = i; break
    iDd = header_idx['dd']; iMm = header_idx['mm']
    iZz = header_idx['zz']; iPp = header_idx['pp']; iTot = header_idx['total']
    iMuj = header_idx.get('mujeres'); iHom = header_idx.get('hombres')
    out = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or row[iDd] is None: continue
        if pad2(row[iDd]) != BOG_DEP_ELEC or pad3(row[iMm]) != BOG_MUN_ELEC: continue
        zz = pad2(row[iZz]); pp = pad2(row[iPp])
        try: tot = int(row[iTot] or 0)
        except: tot = 0
        if tot <= 0: continue
        try: muj = int(row[iMuj] or 0) if iMuj is not None else 0
        except: muj = 0
        try: hom = int(row[iHom] or 0) if iHom is not None else 0
        except: hom = 0
        out[f'{zz}-{pp}'] = {'total': tot, 'mujeres': muj, 'hombres': hom}
    return out


def main():
    if len(sys.argv) < 6:
        print('Uso: build-bog-barrio.py <PUESTOS_GEOREF.csv> <BOG-BARRIOS.geojson> <COMUNAS_DATA.csv> <Divipol2021.xlsx> <out-dir>', file=sys.stderr)
        sys.exit(1)
    csv_path, geo_path, censo26_path, divipol22_path, out_dir = sys.argv[1:6]
    os.makedirs(out_dir, exist_ok=True)

    print('[1] cargando puestos Bogotá…')
    puestos = load_puestos_bogota(csv_path)
    print(f'    {len(puestos)} puestos con coords')

    print('[2] cargando GeoJSON barrios…')
    barrio_geo = json.load(open(geo_path))
    barrio_info = {}
    for f in barrio_geo['features']:
        cod = f['properties'].get('codigo')
        if not cod: continue
        barrio_info[cod] = {
            'cod': cod, 'nombre': f['properties'].get('nombre', ''),
            'loc_cod': f['properties'].get('loc_codigo', ''),
            'loc_nombre': f['properties'].get('loc_nombre', ''),
        }
    print(f'    {len(barrio_info)} barrios')

    print('[3] point-in-polygon puesto → barrio…')
    pue_to_barrio, unmatched = build_puesto_to_barrio(puestos, barrio_geo)
    print(f'    matched: {len(pue_to_barrio)}/{len(puestos)} · unmatched: {len(unmatched)}')

    # Mapeo crudo (útil para frontend)
    with open(os.path.join(out_dir, 'bog-puesto-to-barrio.json'), 'w') as f:
        json.dump(pue_to_barrio, f, separators=(',', ':'))

    # 4) Procesa cada señal
    print('[4] agregando 10 señales por barrio…')
    for sig, fmt, url in SIGNALS:
        try: data = fetch_cached(url)
        except Exception as e: print(f'    [err] {sig}: {e}', file=sys.stderr); continue
        by_b = defaultdict(lambda: {'vv': 0, 'votos': defaultdict(int), 'n_puestos': 0})
        n_in_bog = 0; n_matched = 0
        for zzpp, vv, votos in iter_puestos_signal(data, fmt):
            n_in_bog += 1
            bcod = pue_to_barrio.get(zzpp)
            if not bcod: continue
            n_matched += 1
            b = by_b[bcod]
            b['vv'] += vv; b['n_puestos'] += 1
            for k, n in votos.items(): b['votos'][k] += n
        por_barrio = {}
        for cod, b in by_b.items():
            vv = b['vv']
            cands = [{'nombre': k, 'votos': n, 'pct': round(100 * n / vv, 3) if vv else 0}
                     for k, n in b['votos'].items()]
            cands.sort(key=lambda x: -x['votos'])
            info = barrio_info.get(cod, {})
            por_barrio[cod] = {
                'cod': cod, 'nombre': info.get('nombre', ''),
                'loc_cod': info.get('loc_cod', ''), 'loc_nombre': info.get('loc_nombre', ''),
                'votos_validos': vv, 'n_puestos': b['n_puestos'],
                'candidatos': cands,
            }
        out_path = os.path.join(out_dir, f'bog-barrio-{sig}.json')
        with open(out_path, 'w') as f:
            json.dump({'sig': sig, 'n_barrios': len(por_barrio), 'por_barrio': por_barrio,
                       'puestos_total': n_in_bog, 'puestos_asignados': n_matched}, f,
                      separators=(',', ':'))
        print(f'    ✓ {sig:30s} {len(por_barrio):4d} barrios · {n_matched}/{n_in_bog}')

    # 5) Censos por barrio
    print('[5] censo 2026 (COMUNAS_DATA)…')
    c26 = load_censo_2026_bogota(censo26_path)
    print(f'    {len(c26)} puestos con censo')
    def agg_censo(censo_by_pue, year):
        by_b = defaultdict(lambda: {'total': 0, 'mujeres': 0, 'hombres': 0, 'n_puestos': 0})
        for zzpp, c in censo_by_pue.items():
            bcod = pue_to_barrio.get(zzpp)
            if not bcod: continue
            b = by_b[bcod]; b['total'] += c['total']; b['mujeres'] += c['mujeres']
            b['hombres'] += c['hombres']; b['n_puestos'] += 1
        por = {}; total = 0
        for cod, b in by_b.items():
            info = barrio_info.get(cod, {})
            por[cod] = {'cod': cod, 'nombre': info.get('nombre', ''),
                        'loc_cod': info.get('loc_cod', ''), 'loc_nombre': info.get('loc_nombre', ''),
                        'censo': b['total'], 'mujeres': b['mujeres'], 'hombres': b['hombres'],
                        'n_puestos': b['n_puestos']}
            total += b['total']
        return por, total
    p26, t26 = agg_censo(c26, 2026)
    print(f'    {len(p26)} barrios · censo total {t26:,}')
    print('[6] censo 2022 (Divipol 2021)…')
    c22 = load_censo_2022_bogota(divipol22_path)
    p22, t22 = agg_censo(c22, 2022)
    print(f'    {len(p22)} barrios · censo total {t22:,}')
    print('[7] censo 2018 (escalado desde 2022)…')
    ratio = BOG_2018_DEP_CENSO / max(BOG_2022_DEP_CENSO, 1)
    p18 = {cod: {**e, 'censo': round(e['censo'] * ratio),
                       'mujeres': round(e['mujeres'] * ratio),
                       'hombres': round(e['hombres'] * ratio)}
           for cod, e in p22.items()}
    t18 = sum(e['censo'] for e in p18.values())
    print(f'    {len(p18)} barrios · censo total {t18:,} (ratio {ratio:.4f})')

    for year, por, total in [(2026, p26, t26), (2022, p22, t22), (2018, p18, t18)]:
        with open(os.path.join(out_dir, f'censo-barrio-{year}.json'), 'w') as f:
            json.dump({'year': year, 'ciudad_total': total, 'n_barrios': len(por), 'por_barrio': por},
                      f, separators=(',', ':'))
        print(f'✓ censo-barrio-{year}.json')


if __name__ == '__main__':
    main()
