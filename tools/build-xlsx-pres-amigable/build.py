#!/usr/bin/env python3
"""
XLSX 'amigables' del centro de descargas: con NOMBRES (sin códigos) y CON
ESTILO para gente que no es experta en datos. Encabezado de color, fuente
tamaño 13, filas alternadas (zebra), encabezado congelado, autofiltro,
anchos de columna y separador de miles en Votos.

Sobrescribe los nombres que SIRVE el frontend (descargas.html · _resolveTarget):
  mesa    → presidencial-1v/<año>/GCS_<año>PRES1V.xlsx
  puesto  → presidencial-1v/<año>/GCS_<año>PRES1V_PUESTO.xlsx

Los CSV crudos (con códigos) NO se tocan: quedan para usuarios avanzados
(Python/R/Stata). Sólo los XLSX se vuelven amigables.

Reusa la resolución de nombres de tools/build-xlsx-con-nombres/build.py
(fuente única de verdad para el cruce código→nombre).

Uso:
  python3 tools/build-xlsx-pres-amigable/build.py                  # 4 años, mesa, sin subir
  python3 tools/build-xlsx-pres-amigable/build.py --puesto         # por puesto
  python3 tools/build-xlsx-pres-amigable/build.py --puesto --upload
  python3 tools/build-xlsx-pres-amigable/build.py --only=GCS_2022PRES1V.csv --puesto
Flags:
  --puesto     agrega por puesto (default: por mesa)
  --upload     sube a S3 sobrescribiendo el archivo servido
  --no-zebra   sin filas alternadas (útil para los mesa gigantes)
  --only=a,b   procesa sólo esos nombres de CSV
"""
import csv
import importlib.util
import subprocess
import sys
import time
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
# Reusar la resolución de nombres del módulo hermano (single source of truth).
_spec = importlib.util.spec_from_file_location(
    "con_nombres", str(HERE.parent / "build-xlsx-con-nombres" / "build.py"))
con = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(con)

SRC = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos")
GCS_DIR = SRC / "FINAL SUBIDA GCS"
OUT = SRC / "output_xlsx_amigable"
OUT.mkdir(parents=True, exist_ok=True)
S3_BASE = "s3://elecciones-2026/ricardoruiz.co/DESCARGAS/raw"

SHEET_ROWS = 1_000_000

DEFAULTS = [
    ("GCS_2010PRES1V.csv", "presidencial-1v", 2010),
    ("GCS_2014PRES1V.csv", "presidencial-1v", 2014),
    ("GCS_2018PRES1V.csv", "presidencial-1v", 2018),
    ("GCS_2022PRES1V.csv", "presidencial-1v", 2022),
    ("GCS_2010PRES2V.csv", "presidencial-2v", 2010),
    ("GCS_2014PRES2V.csv", "presidencial-2v", 2014),
    ("GCS_2018PRES2V.csv", "presidencial-2v", 2018),
    ("GCS_2022PRES2V.csv", "presidencial-2v", 2022),
]

HEADER_PUESTO = [
    "Fuente", "Fecha elección", "Corporación", "Circunscripción",
    "Departamento", "Municipio", "Comuna", "Puesto", "Barrio",
    "Partido", "Candidato", "Votos",
]
HEADER_MESA = [
    "Fuente", "Fecha elección", "Corporación", "Circunscripción",
    "Departamento", "Municipio", "Comuna", "Puesto", "Barrio", "Mesa",
    "Partido", "Candidato", "Votos",
]

# ── Estilo (instanciado una vez → openpyxl dedupe de estilos) ──
FONT_SIZE = 13
HEADER_FILL = PatternFill("solid", fgColor="8A1E16")    # oxblood (acento del sitio)
HEADER_FONT = Font(name="Calibri", size=FONT_SIZE, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_BORDER = Border(bottom=Side(style="medium", color="5a130d"))
DATA_FONT = Font(name="Calibri", size=FONT_SIZE)
ZEBRA_FILL = PatternFill("solid", fgColor="F4F0E7")     # crema (filas pares)
NUM_ALIGN = Alignment(horizontal="right")
NUM_FMT = "#,##0"

COL_WIDTH = {
    "Fuente": 11, "Fecha elección": 15, "Corporación": 17, "Circunscripción": 18,
    "Departamento": 22, "Municipio": 24, "Comuna": 22, "Puesto": 34, "Barrio": 26,
    "Mesa": 9, "Partido": 32, "Candidato": 32, "Votos": 12,
}


def write_styled(header, rows_iter, xlsx_out, zebra=True):
    """Escribe un XLSX con estilo en write_only (memoria acotada para mesa)."""
    wb = Workbook(write_only=True)
    votos_idx = len(header) - 1                # 'Votos' siempre va última
    last_col = get_column_letter(len(header))

    def new_sheet(name):
        ws = wb.create_sheet(name)
        for i, h in enumerate(header, 1):
            ws.column_dimensions[get_column_letter(i)].width = COL_WIDTH.get(h, 16)
        ws.freeze_panes = "A2"
        hcells = []
        for h in header:
            c = WriteOnlyCell(ws, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = HEADER_ALIGN
            c.border = HEADER_BORDER
            hcells.append(c)
        ws.append(hcells)
        return ws

    sheets = []
    ws = new_sheet("Datos")
    rows_in_sheet = 1            # cuenta el header
    data_row = 0                 # filas de datos (para el zebra)
    for vals in rows_iter:
        if rows_in_sheet >= SHEET_ROWS:
            sheets.append((ws, rows_in_sheet))
            ws = new_sheet(f"Datos {len(sheets) + 1}")
            rows_in_sheet = 1
        data_row += 1
        shade = zebra and (data_row % 2 == 0)
        cells = []
        for j, v in enumerate(vals):
            c = WriteOnlyCell(ws, value=v)
            c.font = DATA_FONT
            if shade:
                c.fill = ZEBRA_FILL
            if j == votos_idx:
                c.number_format = NUM_FMT
                c.alignment = NUM_ALIGN
            cells.append(c)
        ws.append(cells)
        rows_in_sheet += 1
    sheets.append((ws, rows_in_sheet))

    for ws, nrows in sheets:
        ws.auto_filter.ref = f"A1:{last_col}{nrows}"

    wb.save(xlsx_out)
    return {
        "filas": sum(n - 1 for _, n in sheets),
        "sheets": len(sheets),
        "mb": round(xlsx_out.stat().st_size / (1024 * 1024), 2),
    }


def iter_puesto(gcs_in, lookup_data):
    """Agrega por puesto y emite filas resueltas a nombre (sin códigos)."""
    lookup_full, depto_by_cod, mun_by_cod = lookup_data
    agg = defaultdict(int)
    with gcs_in.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";")
        headers = next(reader)
        col = {h: i for i, h in enumerate(headers)}
        for needed in ("COD_DDE", "COD_MME", "COD_ZZ", "COD_PP", "NUM_VOT",
                       "DES_COR", "DES_CIR", "DES_PAR", "DES_CAN", "FUENTE", "FEC_ELEC"):
            if needed not in col:
                raise SystemExit(f"Falta columna en {gcs_in.name}: {needed}")
        for row in reader:
            try:
                vot = int((row[col["NUM_VOT"]] or "0").strip() or "0")
            except (ValueError, IndexError):
                vot = 0
            try:
                key = (
                    row[col["FUENTE"]], row[col["FEC_ELEC"]],
                    row[col["DES_COR"]], row[col["DES_CIR"]],
                    row[col["COD_DDE"]].strip().zfill(2),
                    row[col["COD_MME"]].strip().zfill(3),
                    row[col["COD_ZZ"]].strip().zfill(2),
                    row[col["COD_PP"]].strip().zfill(2),
                    row[col["DES_PAR"]], row[col["DES_CAN"]],
                )
            except IndexError:
                continue
            agg[key] += vot
    for key, votos in agg.items():
        (fuente, fecha, corp, circ, dde, mme, zz, pp, partido, candidato) = key
        depto, municipio, puesto, comuna, barrio, _hit = con.resolve_names(
            dde, mme, zz, pp, lookup_full, depto_by_cod, mun_by_cod)
        yield [fuente, fecha, corp, circ, depto, municipio, comuna, puesto,
               barrio, partido, candidato, votos]


def iter_mesa(gcs_in, lookup_data):
    """Streamea por mesa y emite filas resueltas a nombre (sin códigos)."""
    lookup_full, depto_by_cod, mun_by_cod = lookup_data
    with gcs_in.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";")
        headers = next(reader)
        col = {h: i for i, h in enumerate(headers)}
        for needed in ("COD_DDE", "COD_MME", "COD_ZZ", "COD_PP", "NUM_VOT",
                       "DES_COR", "DES_CIR", "DES_PAR", "DES_CAN", "DES_MS",
                       "FUENTE", "FEC_ELEC"):
            if needed not in col:
                raise SystemExit(f"Falta columna en {gcs_in.name}: {needed}")
        for row in reader:
            try:
                dde = row[col["COD_DDE"]].strip().zfill(2)
                mme = row[col["COD_MME"]].strip().zfill(3)
                zz = row[col["COD_ZZ"]].strip().zfill(2)
                pp = row[col["COD_PP"]].strip().zfill(2)
            except IndexError:
                continue
            depto, municipio, puesto, comuna, barrio, _hit = con.resolve_names(
                dde, mme, zz, pp, lookup_full, depto_by_cod, mun_by_cod)
            try:
                vraw = row[col["NUM_VOT"]].strip()
                votos = int(vraw) if vraw else 0
            except (ValueError, IndexError):
                votos = 0
            yield [row[col["FUENTE"]], row[col["FEC_ELEC"]],
                   row[col["DES_COR"]], row[col["DES_CIR"]],
                   depto, municipio, comuna, puesto, barrio,
                   row[col["DES_MS"]], row[col["DES_PAR"]],
                   row[col["DES_CAN"]], votos]


def s3_upload(local_path, cat, año):
    key = f"{S3_BASE}/{cat}/{año}/{local_path.name}"
    cmd = ["aws", "s3", "cp", str(local_path), key,
           "--content-type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
           "--cache-control", "public, max-age=86400", "--no-progress"]
    subprocess.run(cmd, check=True, capture_output=True, text=True)
    return key.replace("s3://elecciones-2026/",
                       "https://elecciones-2026.s3.us-east-1.amazonaws.com/")


def main():
    args = sys.argv[1:]
    por_puesto = "--puesto" in args
    upload = "--upload" in args
    zebra = "--no-zebra" not in args
    only_set = set()
    for a in args:
        if a.startswith("--only="):
            only_set = {x.strip() for x in a.split("=", 1)[1].split(",") if x.strip()}

    archivos = [(n, c, y) for (n, c, y) in DEFAULTS if not only_set or n in only_set]
    if not archivos:
        raise SystemExit("Sin archivos para procesar (revisa --only=).")

    header = HEADER_PUESTO if por_puesto else HEADER_MESA
    suffix = "_PUESTO" if por_puesto else ""
    mode = "POR PUESTO" if por_puesto else "POR MESA"
    print(f"Modo: {mode} · zebra={'on' if zebra else 'off'} · upload={'on' if upload else 'off'}", flush=True)

    lookup_data = con.load_geo_lookup()

    for idx, (csv_name, cat, año) in enumerate(archivos, 1):
        csv_in = GCS_DIR / csv_name
        if not csv_in.exists():
            print(f"[{idx}/{len(archivos)}] SKIP (no existe): {csv_name}", flush=True)
            continue
        out_dir = OUT / cat / str(año)
        out_dir.mkdir(parents=True, exist_ok=True)
        out_name = csv_name.replace(".csv", f"{suffix}.xlsx")
        xlsx_out = out_dir / out_name
        print(f"[{idx}/{len(archivos)}] {csv_name} → {out_name}", flush=True)
        t0 = time.time()
        rows_iter = iter_puesto(csv_in, lookup_data) if por_puesto else iter_mesa(csv_in, lookup_data)
        stats = write_styled(header, rows_iter, xlsx_out, zebra=zebra)
        stats["sec"] = round(time.time() - t0, 1)
        print(f"   OK · {stats['filas']:,} filas · {stats['sheets']} hoja(s) · {stats['mb']} MB · {stats['sec']}s", flush=True)
        if upload:
            url = s3_upload(xlsx_out, cat, año)
            print(f"   ↑ {url}", flush=True)


if __name__ == "__main__":
    main()
