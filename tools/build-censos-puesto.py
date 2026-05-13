#!/usr/bin/env python3
"""
tools/build-censos-puesto.py

Genera el censo electoral por PUESTO para 2026, 2022 y 2018, optimizado
para consumo del cliente del módulo Oportunidad (Nivel C — bias propio
por puesto).

Inputs:
  - COMUNAS_DATA.csv (Divipole 2026 vigente)
  - Divipol 23.09.2021.xlsx (Divipole previo a pres-2022)
  - wiki-pres-2018-deptos.json (censo 2018 reconstruido por depto)

Outputs (en out_dir):
  censos-puesto-2026.json   { year, nacional, porPuesto: { 'dd-mm-zz-pp': int } }
  censos-puesto-2022.json   idem
  censos-puesto-2018.json   idem (escalado por ratio depto 2018/2022)

Uso:
  python3 tools/build-censos-puesto.py \
    "Bases de datos/COMUNAS_DATA.csv" \
    "Bases de datos/Divipol 23.09.2021.xlsx" \
    "tools/wiki-pres-2018-deptos.json" \
    "Bases de datos"
"""

import csv, json, os, sys
from collections import defaultdict
import openpyxl

# Mapeo nombre Wikipedia → código electoral (misma tabla que build-censo-2018-wiki.py).
DEPTO_CODE = {
    'Amazonas':'60','Antioquia':'01','Arauca':'40','Atlántico':'03',
    'Bogotá':'16','Bolívar':'05','Boyacá':'07','Caldas':'09','Caquetá':'44',
    'Casanare':'46','Cauca':'11','Cesar':'12','Chocó':'17','Consulados':'88',
    'Córdoba':'13','Cundinamarca':'15','Guainía':'50','Guaviare':'54',
    'Huila':'19','La Guajira':'48','Magdalena':'21','Meta':'52','Nariño':'23',
    'Norte de Santander':'25','Putumayo':'64','Quindío':'26','Risaralda':'24',
    'San Andrés y Providencia':'56','Santander':'27','Sucre':'28','Tolima':'29',
    'Valle del Cauca':'31','Vaupés':'68','Vichada':'72',
}

def pad2(s): return str(s).zfill(2)
def pad3(s): return str(s).zfill(3)


def load_csv_2026(path):
    out = {}
    nac = 0
    with open(path, encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        header = next(reader); header[0] = header[0].replace('﻿', '')
        col = lambda n: next(i for i, h in enumerate(header) if h.strip().lower() == n.lower())
        iDd = col('dd'); iMm = col('mm'); iZz = col('zz'); iPp = col('pp'); iTot = col('total')
        for row in reader:
            if not row or len(row) <= iTot: continue
            try: tot = int(row[iTot] or 0)
            except ValueError: tot = 0
            if tot <= 0: continue
            k = f'{pad2(row[iDd])}-{pad3(row[iMm])}-{pad2(row[iZz])}-{pad2(row[iPp])}'
            out[k] = tot
            nac += tot
    return out, nac


def load_xlsx_2022(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_idx = {}; header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if row and str(row[0]).strip().lower() == 'dd':
            for j, h in enumerate(row):
                if h is not None: header_idx[str(h).strip().lower()] = j
            header_row = i; break
    iDd = header_idx['dd']; iMm = header_idx['mm']
    iZz = header_idx['zz']; iPp = header_idx['pp']; iTot = header_idx['total']
    out = {}; nac = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or row[iDd] is None: continue
        try: tot = int(row[iTot] or 0)
        except (TypeError, ValueError): tot = 0
        if tot <= 0: continue
        k = f'{pad2(row[iDd])}-{pad3(row[iMm])}-{pad2(row[iZz])}-{pad2(row[iPp])}'
        out[k] = tot
        nac += tot
    return out, nac


def main():
    if len(sys.argv) < 5:
        print('Uso: build-censos-puesto.py <COMUNAS_DATA.csv> <Divipol2021.xlsx> <wiki-2018-deptos.json> <out-dir>', file=sys.stderr)
        sys.exit(1)
    csv26, xlsx22, wiki18_path, out_dir = sys.argv[1:5]
    os.makedirs(out_dir, exist_ok=True)

    print('[1] censo 2026 (COMUNAS_DATA)…')
    p26, nac26 = load_csv_2026(csv26)
    print(f'    {len(p26):,} puestos · nacional {nac26:,}')

    print('[2] censo 2022 (Divipol 23-09-2021)…')
    p22, nac22 = load_xlsx_2022(xlsx22)
    print(f'    {len(p22):,} puestos · nacional {nac22:,}')

    print('[3] censo 2018 (escalado por ratio de depto desde Wikipedia)…')
    wiki18 = json.load(open(wiki18_path))
    by_dep_18 = {DEPTO_CODE[d['depto']]: d['censo'] for d in wiki18 if d['depto'] in DEPTO_CODE}
    # Censo 2022 por depto (para sacar ratio)
    by_dep_22 = defaultdict(int)
    for k, v in p22.items():
        by_dep_22[k.split('-')[0]] += v
    ratio_by_dep = {d: by_dep_18[d] / max(by_dep_22.get(d, 0), 1) for d in by_dep_18}
    p18 = {}
    nac18 = 0
    huérfanos = 0
    for k, v22 in p22.items():
        dep = k.split('-')[0]
        r = ratio_by_dep.get(dep)
        if r is None:
            huérfanos += 1; continue
        v18 = round(v22 * r)
        p18[k] = v18
        nac18 += v18
    print(f'    {len(p18):,} puestos · nacional {nac18:,} · huérfanos {huérfanos}')

    for year, data, nac in [(2026, p26, nac26), (2022, p22, nac22), (2018, p18, nac18)]:
        path = os.path.join(out_dir, f'censos-puesto-{year}.json')
        json.dump({'year': year, 'nacional': nac, 'porPuesto': data},
                  open(path, 'w'), separators=(',', ':'))
        print(f'✓ {path} ({os.path.getsize(path)/1024:.0f} KB)')


if __name__ == '__main__':
    main()
