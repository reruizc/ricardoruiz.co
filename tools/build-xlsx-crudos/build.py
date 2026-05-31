#!/usr/bin/env python3
"""
Convierte los CSVs crudos GCS de la Registraduría a XLSX para el centro
de descargas de ricardoruiz.co.

Estrategia:
- Archivos chicos (todas las filas caben en una hoja Excel): un .xlsx con
  una sola hoja `Datos` y header.
- Archivos que exceden el límite Excel (1.048.576 filas/hoja, dejamos
  margen y partimos cada 1.000.000): varias hojas `Datos 1`, `Datos 2`,
  cada una con header replicado para que sea autocontenida.

Usa openpyxl en write_only mode para no cargar todo a memoria.
"""
import csv
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

SRC = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos/FINAL SUBIDA GCS")
OUT = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos/output_crudos_xlsx")
OUT.mkdir(parents=True, exist_ok=True)

# Archivos chicos a convertir (los grandes Congreso/Territoriales son ZIP por depto en otra sesión).
# Mapeo nombre csv → (categoría, año, sub-path en S3).
CHICOS = [
    # nombre_csv, categoria_s3, año, label
    ("GCS_2010PRES1V.csv",   "presidencial-1v",     2010),
    ("GCS_2014PRES1V.csv",   "presidencial-1v",     2014),
    ("GCS_2018PRES1V.csv",   "presidencial-1v",     2018),
    ("GCS_2022PRES1V.csv",   "presidencial-1v",     2022),
    ("GCS_2010PRES2V.csv",   "presidencial-2v",     2010),
    ("GCS_2014PRES2V.csv",   "presidencial-2v",     2014),
    ("GCS_2018PRES2V.csv",   "presidencial-2v",     2018),
    ("GCS_2022PRES2V.csv",   "presidencial-2v",     2022),
    ("GCS_2016PLEB.csv",     "plebiscito",          2016),
    ("GCS_2019JAL.csv",      "jal",                 2019),
    ("GCS_2023JAL.csv",      "jal",                 2023),
    ("GCS_2021CLMJ.csv",     "consejos-juventud",   2021),
    ("GCS_2025CLMJ.csv",     "consejos-juventud",   2025),
    ("GCS_2022CONSU.csv",    "consultas",           2022),
    ("GCS_2025CONSU.csv",    "consultas",           2025),
    ("GCS_2025CONSU_CAM.csv","consultas",           2025),
    ("GCS_2025CONSU_SEN.csv","consultas",           2025),
    ("GCS_2025ATIP_MAG.csv", "atipicas",            2025),
]

SHEET_ROWS = 1_000_000  # margen sobre 1.048.576 del límite Excel
HEADER_FONT = Font(bold=True)


def convert_one(csv_path: Path, out_path: Path) -> dict:
    """Convierte un CSV a XLSX. Devuelve estadísticas (filas, hojas, MB)."""
    wb = Workbook(write_only=True)
    ws = None
    sheet_idx = 0
    rows_in_sheet = 0
    total_rows = 0
    header = None

    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        for i, row in enumerate(reader):
            if i == 0:
                header = row
                continue
            if ws is None or rows_in_sheet >= SHEET_ROWS:
                sheet_idx += 1
                ws = wb.create_sheet(f"Datos {sheet_idx}" if sheet_idx > 1 else "Datos")
                ws.append(header)
                rows_in_sheet = 1
            ws.append(row)
            rows_in_sheet += 1
            total_rows += 1

    # Si el CSV estaba vacío (sólo header), creamos hoja vacía con header igual.
    if ws is None:
        ws = wb.create_sheet("Datos")
        ws.append(header or [])

    wb.save(out_path)
    return {
        "filas": total_rows,
        "hojas": sheet_idx if sheet_idx > 0 else 1,
        "mb": round(out_path.stat().st_size / (1024 * 1024), 2),
    }


def main():
    print(f"Output dir: {OUT}")
    total = len(CHICOS)
    results = []
    for idx, (csv_name, cat, año) in enumerate(CHICOS, 1):
        csv_path = SRC / csv_name
        if not csv_path.exists():
            print(f"[{idx}/{total}] SKIP (no existe): {csv_name}")
            continue
        out_dir = OUT / cat / str(año)
        out_dir.mkdir(parents=True, exist_ok=True)
        xlsx_name = csv_name.replace(".csv", ".xlsx")
        out_path = out_dir / xlsx_name
        print(f"[{idx}/{total}] {csv_name} → {cat}/{año}/{xlsx_name} ...", flush=True)
        try:
            stats = convert_one(csv_path, out_path)
            results.append((csv_name, stats))
            print(f"   OK · {stats['filas']:,} filas · {stats['hojas']} hoja(s) · {stats['mb']} MB", flush=True)
        except Exception as e:
            print(f"   FAIL: {e}", flush=True)
            results.append((csv_name, {"error": str(e)}))

    print("\n=== RESUMEN ===")
    for name, stats in results:
        if "error" in stats:
            print(f"  {name}: ERROR {stats['error']}")
        else:
            print(f"  {name}: {stats['filas']:,} filas · {stats['hojas']} hojas · {stats['mb']} MB")


if __name__ == "__main__":
    main()
