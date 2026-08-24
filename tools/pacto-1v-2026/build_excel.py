#!/usr/bin/env python3
import json,csv
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
OUT='Bases de datos/output_pacto_1v_2026'
B=json.load(open(f'{OUT}/blocks_all.json')); DIF=json.load(open(f'{OUT}/dif_2022.json'))
OX='8A1E16'; thin=Side(style='thin',color='DDDDDD'); bd=Border(thin,thin,thin,thin)
# Claudia por depto (municipio CSV)
clau={}
with open('Bases de datos/nuevos archivos 1v 2026/Base nombres corregidos primera vuelta 2026.csv',newline='',encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        if 'CLAUDIA' in (row['Candidato'] or '').upper():
            d=(row['cod_departamento'] or '').zfill(2); clau[d]=clau.get(d,0)+int(row['Suma de Votos'] or 0)
wb=Workbook()
def hdr(ws,headers,row=1):
    for i,h in enumerate(headers,1):
        c=ws.cell(row=row,column=i,value=h); c.font=Font(bold=True,color='FFFFFF',size=10.5)
        c.fill=PatternFill('solid',fgColor=OX); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bd
def fit(ws):
    for col in ws.columns:
        w=max((len(str(c.value)) for c in col if c.value is not None),default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(w+2,11),34)
def pct(v): return (v/100) if isinstance(v,(int,float)) else None
deptos=sorted(B['depto'].items(),key=lambda x:-(x[1]['base_votos']))

# Resumen
ws=wb.active; ws.title='Resumen'
ws.append(['Análisis Nacional Electoral · Primera vuelta 2026 — Soporte'])
ws['A1'].font=Font(bold=True,size=14,color=OX); ws.append([])
ws.append(['Bloque','Hallazgo clave','Dato'])
hdr(ws,['Bloque','Hallazgo clave','Dato'],row=3)
for r in [['1 · Petro 2V→1V 2022','Creció +10 pp entre vueltas; más en Caribe/Bogotá/Pacífico (centro + anti-Rodolfo)','40,4% → 50,4%'],
          ['2 · Cepeda vs Petro 1V','Igualó a Petro nacional; cayó en las grandes ciudades, creció en la frontera','Bogotá −4,3 pp'],
          ['2 · Techo de recuperación','Petro 2V-22 marca el techo de la izquierda por ciudad','Bgtá 58,6 · BAQ 63,4 · Cali 63,9 · MDE 33,7'],
          ['3 · Centro (Fajardo+Claudia)','Voto paisa/andino/urbano: Antioquia, Bogotá, Eje Cafetero','Fajardo 4,3% + Claudia 0,95%'],
          ['3 · Oviedo (consulta)','Voto de centro anti-Abelardo; 42% en Bogotá; 64% ya votó Abelardo','Disputable ~180k en Bogotá'],
          ['4 · Derecha (Abe+Paloma)','No creció: se consolidó. Avanzó donde cayó la izquierda (Bogotá +6,5)','51,1% vs 51,9% (Fico+Rodolfo 22)']]:
    ws.append(r)
fit(ws); ws.column_dimensions['B'].width=62; ws.column_dimensions['C'].width=34

# B1
ws=wb.create_sheet('B1 · Petro 2V vs 1V')
hdr(ws,['Departamento','Petro 1V-22 %','Petro 2V-22 %','Crecimiento 2V−1V (pp)'])
for nm,d in sorted(B['depto'].items(),key=lambda x:-(x[1]['dif_petro2v1v'] or -99)):
    ws.append([nm,pct(d['petro1v']),pct(d['petro2v']),d['dif_petro2v1v']])
for r in ws.iter_rows(min_row=2):
    r[1].number_format=r[2].number_format='0.0%'
    for c in r: c.border=bd
fit(ws)

# B2
ws=wb.create_sheet('B2 · Cepeda vs Petro')
hdr(ws,['Departamento','Cepeda 26 %','Petro 1V-22 %','Diferencial (pp)','Techo Petro 2V-22 %','Espacio recuperación (pp)','Abstención 26 %'])
for nm,d in sorted(B['depto'].items(),key=lambda x:(x[1]['dif_cep'] if x[1]['dif_cep'] is not None else 99)):
    esp=round((d['petro2v'] or 0)-(d['cep26'] or 0),1) if d['petro2v'] and d['cep26'] else None
    ws.append([nm,pct(d['cep26']),pct(d['petro1v']),d['dif_cep'],pct(d['petro2v']),esp,pct(d['abst26'])])
for r in ws.iter_rows(min_row=2):
    for i in (1,2,4,6): r[i].number_format='0.0%'
    for c in r: c.border=bd
fit(ws)

# B2 Bogotá localidad
ws=wb.create_sheet('B2 · Bogotá localidad')
hdr(ws,['Localidad','Cepeda 26 %','Petro 1V-22 %','Diferencial (pp)','Techo Petro 2V-22 %'])
for L,c,p1,dif,p2 in sorted(DIF['bogota_loc'],key=lambda r:r[3]):
    if L=='Ciudad Bolivar': continue
    ws.append([L,pct(c),pct(p1),round(dif,1),pct(p2)])
for r in ws.iter_rows(min_row=2):
    for i in (1,2,4): r[i].number_format='0.0%'
    for c in r: c.border=bd
fit(ws)

# B3 Centro
ws=wb.create_sheet('B3 · Centro')
hdr(ws,['Departamento','Fajardo 26 %','Fajardo votos','Claudia votos','Centro (Faj+Cla) votos'])
DEPCOD={v['cod']:k for k,v in B['depto'].items()}
fajv={}
import json as _j
# fajardo votos por depto desde blocks (no guardado) -> recompute desde % * base
for nm,d in deptos:
    cod=d['cod']; fv=round((d['centro26'] or 0)/100*d['base_votos']) if d['centro26'] else 0
    cv=clau.get(cod,0)
    ws.append([nm,pct(d['centro26']),fv,cv,fv+cv])
for r in ws.iter_rows(min_row=2):
    r[1].number_format='0.0%'
    for c in r: c.border=bd
fit(ws); ws.append([]); ws.append(['Nota: Claudia a nivel municipio (no disponible a mesa en el preconteo). Sumar también Oviedo (consulta) para el universo de centro.'])

# B4 Derecha
ws=wb.create_sheet('B4 · Derecha')
hdr(ws,['Departamento','Derecha 26 (Abe+Pal) %','Fico+Rodolfo 1V-22 %','Avance (pp)'])
for nm,d in sorted(B['depto'].items(),key=lambda x:-(x[1]['dif_der'] if x[1]['dif_der'] is not None else -99)):
    ws.append([nm,pct(d['der26']),pct(d['der22']),d['dif_der']])
for r in ws.iter_rows(min_row=2):
    r[1].number_format=r[2].number_format='0.0%'
    for c in r: c.border=bd
fit(ws)

# Ciudades por comuna
ws=wb.create_sheet('Ciudades · comuna')
hdr(ws,['Ciudad','Comuna/Localidad','Cepeda 26 %','Petro 1V-22 %','Dif (pp)','Derecha 26 %','Abstención 26 %'])
for city,d in B['city_comuna'].items():
    for com,v in sorted(d.items(),key=lambda x:-(x[1]['base_votos'])):
        ws.append([city,com,pct(v['cep26']),pct(v['petro1v']),v['dif_cep'],pct(v['der26']),pct(v['abst26'])])
for r in ws.iter_rows(min_row=2):
    for i in (2,3,5,6): r[i].number_format='0.0%'
    for c in r: c.border=bd
fit(ws)

wb.save(f'{OUT}/Soporte_Analisis_Pacto_1V_2026.xlsx')
print('✓ Soporte_Analisis_Pacto_1V_2026.xlsx')
for s in wb.sheetnames: print('  ·',s)
