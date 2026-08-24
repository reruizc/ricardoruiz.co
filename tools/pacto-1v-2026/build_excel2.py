#!/usr/bin/env python3
import json,csv,unicodedata
_norm=lambda s: ''.join(c for c in unicodedata.normalize('NFD',str(s or '').upper().strip()) if unicodedata.category(c)!='Mn')
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
OUT='Bases de datos/output_pacto_1v_2026'
BA=json.load(open(f'{OUT}/blocks_all.json')); BF=json.load(open(f'{OUT}/blocks_full.json')); DIF=json.load(open(f'{OUT}/dif_2022.json'))
EST=json.load(open(f'{OUT}/estrato_bogota.json'))
TER=json.load(open(f'{OUT}/twov_territorial.json'))
OX='8A1E16'; thin=Side(style='thin',color='E2DDD0'); bd=Border(thin,thin,thin,thin)
ZEBRA=PatternFill('solid',fgColor='F4F0E7')   # crema suave para filas alternas
clau={}
with open('Bases de datos/nuevos archivos 1v 2026/Base nombres corregidos primera vuelta 2026.csv',newline='',encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        if 'CLAUDIA' in (row['Candidato'] or '').upper(): d=(row['cod_departamento'] or '').zfill(2); clau[d]=clau.get(d,0)+int(row['Suma de Votos'] or 0)
wb=Workbook()
def sheet(name): return wb.create_sheet(name)
def hdr(ws,headers,row=1):
    for i,h in enumerate(headers,1):
        c=ws.cell(row=row,column=i,value=h); c.font=Font(bold=True,color='FFFFFF',size=10); c.fill=PatternFill('solid',fgColor=OX); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bd
    ws.row_dimensions[row].height=30
def fin(ws,pcols=(),filt=True,hdrrow=1,zebra=True):
    for ri,row in enumerate(ws.iter_rows(min_row=hdrrow+1),start=hdrrow+1):
        zeb = zebra and (ri-hdrrow)%2==0
        for c in row:
            i=c.column-1; c.border=bd
            if i in pcols and isinstance(c.value,(int,float)): c.number_format='0.0%'
            elif isinstance(c.value,(int,float)) and abs(c.value)>=1000: c.number_format='#,##0'   # miles automáticos
            if zeb: c.fill=ZEBRA
    for col in ws.columns:
        w=max((len(str(c.value)) for c in col if c.value is not None),default=8); ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(w+2,11),38)
    if filt and ws.max_row>hdrrow+1: ws.freeze_panes=f'A{hdrrow+1}'; ws.auto_filter.ref=ws.dimensions
def cf_div(ws,colletter,first=2,lo=-15,hi=15):   # escala roja→blanca→verde para diferencias (pp)
    ws.conditional_formatting.add(f'{colletter}{first}:{colletter}{ws.max_row}', ColorScaleRule(
        start_type='num',start_value=lo,start_color='D06B57', mid_type='num',mid_value=0,mid_color='FBF6EC',
        end_type='num',end_value=hi,end_color='4F9E68'))
def cf_bar(ws,colletter,first=2,color='8A1E16'):  # barra de datos para conteos
    ws.conditional_formatting.add(f'{colletter}{first}:{colletter}{ws.max_row}', DataBarRule(start_type='min',end_type='max',color=color,showValue=True))
P=lambda v:(v/100) if isinstance(v,(int,float)) else None
DEPN=lambda code: BA['depto'].get(code,{})

# Resumen
ws=wb.active; ws.title='Resumen'
ws.append(['Análisis Nacional Electoral · Primera vuelta 2026 — Soporte']); ws['A1'].font=Font(bold=True,size=14,color=OX); ws.append([])
ws.append(['Capítulo','Hallazgo','Dato']); hdr(ws,['Capítulo','Hallazgo','Dato'],3)
for r in [['1 · Petro 2V→1V','Creció +10pp; Caribe/Bogotá/Pacífico (centro+anti-Rodolfo)','40,4→50,4%'],
          ['2 · Cepeda vs Petro','Igualó a Petro; cayó en ciudades, creció en frontera','Bogotá −4,3pp'],
          ['2 · Techo recuperación','Petro 2V-22 = techo izquierda','Bgtá 58,6 · BAQ 63,4 · Cali 63,9 · MDE 33,7'],
          ['2 · Crece+abstención','187 municipios crece izquierda y cae participación (Santander/Boyacá/Chocó)','ver hoja'],
          ['3 · Centro+Oviedo','Solo ~8% de Oviedo votó Paloma; 67% Cepeda, 20% Fajardo (mesa a mesa)','Oviedo→Paloma 8%'],
          ['4 · Derecha','No creció, se consolidó; avanzó donde cayó izquierda (Bogotá +6,5)','51,1 vs 51,9%'],
          ['5 · Bogotá estrato','El voto se ordena por estrato; la izquierda cayó en la clase media (E3-E4)','Empate en estrato 3'],
          ['8 · Camino 2V','Piso de la derecha consolidada ~12,33M; Cepeda con el centro ~10,41M','Faltan ~1,9M'],
          ['8 · Dónde están','~2,05M por recuperar hasta el techo de Petro-2V, mapeados por municipio y puesto','ver hojas 8·Camino 2V']]: ws.append(r)
fin(ws,filt=False,zebra=False); ws.column_dimensions['B'].width=64; ws.column_dimensions['C'].width=34

# 1 Petro depto + municipio
ws=sheet('1·Petro 2V-1V (depto)'); hdr(ws,['Departamento','Petro 1V %','Petro 2V %','Creció (pp)'])
for k,v in sorted(BA['depto'].items(),key=lambda x:-(x[1]['dif_petro2v1v'] or -99)): ws.append([k,P(v['petro1v']),P(v['petro2v']),v['dif_petro2v1v']])
fin(ws,(1,2))
ws=sheet('1·Petro 2V-1V (municipio)'); hdr(ws,['Departamento','Municipio','Petro 1V %','Petro 2V %','Creció (pp)'])
for code,v in sorted(BF['muni'].items(),key=lambda x:-(x[1]['dif_petro2v1v'] or -99)):
    if v['dif_petro2v1v'] is None: continue
    ws.append([v['dep'],v['muni'],P(v['petro1v']),P(v['petro2v']),v['dif_petro2v1v']])
fin(ws,(2,3))

# 2 Cepeda depto + municipio + Bogotá loc + abstención
ws=sheet('2·Cepeda vs Petro (depto)'); hdr(ws,['Departamento','Cepeda 26 %','Petro 1V-22 %','Dif (pp)','Techo Petro 2V %','Espacio (pp)','Particip. 26 %'])
for k,v in sorted(BA['depto'].items(),key=lambda x:(x[1]['dif_cep'] if x[1]['dif_cep'] is not None else 99)):
    esp=round((v['petro2v'] or 0)-(v['cep26'] or 0),1) if v['petro2v'] and v['cep26'] else None
    ws.append([k,P(v['cep26']),P(v['petro1v']),v['dif_cep'],P(v['petro2v']),esp,P(v['abst26'])])
fin(ws,(1,2,4,6))
ws=sheet('2·Cepeda vs Petro (municipio)'); hdr(ws,['Departamento','Municipio','Cepeda 26 %','Petro 1V-22 %','Dif (pp)','Techo Petro 2V %'])
for code,v in sorted(BF['muni'].items(),key=lambda x:(x[1]['dif_cep'] if x[1]['dif_cep'] is not None else 99)):
    if v['dif_cep'] is None: continue
    ws.append([v['dep'],v['muni'],P(v['cep26']),P(v['petro1v']),v['dif_cep'],P(v['petro2v'])])
fin(ws,(2,3,5))
ws=sheet('2·Bogotá localidad'); hdr(ws,['Localidad','Cepeda 26 %','Petro 1V-22 %','Dif (pp)','Techo Petro 2V %'])
for L,c,p1,dif,p2 in sorted(DIF['bogota_loc'],key=lambda r:r[3]):
    if L=='Ciudad Bolivar': continue
    ws.append([L,P(c),P(p1),round(dif,1),P(p2)])
fin(ws,(1,2,4))
ws=sheet('2·Crece+abstención (muni)'); hdr(ws,['Departamento','Municipio','Cepeda crece (pp)','Particip. 2022 %','Particip. 2026 %','Caída particip. (pp)'])
for r in BF['muni_abst']: ws.append([r['dep'],r['muni'],r['dif_cep'],P(r['part22']),P(r['part26']),r['caida_part']])
fin(ws,(3,4))

# 3 Centro depto
ws=sheet('3·Centro (depto)'); hdr(ws,['Departamento','Fajardo 26 %','Fajardo votos','Claudia votos','Centro votos'])
for code,v in sorted(BA['depto'].items(),key=lambda x:-(x[1]['centro26'] or 0)):
    fv=round((v['centro26'] or 0)/100*v['base_votos']) if v['centro26'] else 0; cv=clau.get(v['cod'],0)
    ws.append([code,P(v['centro26']),fv,cv,fv+cv])
fin(ws,(1,))

# 4 Derecha depto + municipio
ws=sheet('4·Derecha (depto)'); hdr(ws,['Departamento','Derecha 26 %','Fico+Rodolfo 1V-22 %','Avance (pp)'])
for k,v in sorted(BA['depto'].items(),key=lambda x:-(x[1]['dif_der'] or -99)): ws.append([k,P(v['der26']),P(v['der22']),v['dif_der']])
fin(ws,(1,2))
ws=sheet('4·Derecha (municipio)'); hdr(ws,['Departamento','Municipio','Derecha 26 %','Fico+Rodolfo 22 %','Avance (pp)'])
for code,v in sorted(BF['muni'].items(),key=lambda x:-(x[1]['dif_der'] or -99)):
    if v['dif_der'] is None: continue
    ws.append([v['dep'],v['muni'],P(v['der26']),P(v['der22']),v['dif_der']])
fin(ws,(2,3))

# Ciudades comuna + barrio — voto por recuperar (hasta el techo Petro 2V) + centro disponible
FAJ_T=0.55  # transferencia del centro (Fajardo) a Cepeda · mismo supuesto del Cap 8
def _noise(s):  # agregados especiales que NO son territorio (zona 90/98, exterior, sin asignar)
    u=str(s or '').upper().strip()
    return any(k in u for k in ('PUESTO CENSO','CONSULADO','CARCEL','CÁRCEL','EXTERIOR')) or u in ('OTROS','CORR','CIUDAD','SN','NO APLICA','')
ws=sheet('Ciudades·comuna'); hdr(ws,['Ciudad','Comuna/Localidad','Cepeda 1V','Techo (Petro 2V)','Votos por recuperar','Centro disponible','Votos 26 (base)','Cepeda %','Centro %'])
rows_c=[]
for city,dd in BA['city_comuna'].items():
    agg={}  # fusiona duplicados por tilde (Ciudad Bolívar/Ciudad Bolivar, etc.)
    for com,v in dd.items():
        if _noise(com): continue
        base=v['base_votos']; k=_norm(com)
        a=agg.setdefault(k,{'n':com,'base':0,'cep':0,'techo':0,'cdisp':0,'craw':0})
        a['base']+=base; a['cep']+=round(v['cep26']/100*base); a['techo']+=round((v['petro2v'] or 0)/100*base)
        a['cdisp']+=round((v['centro26'] or 0)/100*base*FAJ_T); a['craw']+=(v['centro26'] or 0)/100*base
        if any(ord(c)>127 for c in com) and not any(ord(c)>127 for c in a['n']): a['n']=com   # prefiere el nombre con tilde
    for a in agg.values():
        rec=max(0,a['techo']-a['cep']); b=a['base'] or 1
        rows_c.append((rec,[city,a['n'],a['cep'],a['techo'],rec,a['cdisp'],a['base'],a['cep']/b,a['craw']/b]))
for _,row in sorted(rows_c,key=lambda x:-x[0]): ws.append(row)
fin(ws,(7,8))
# fusiona variantes tilde/mayúscula del MISMO barrio (espejo de la fusión por comuna de arriba).
# Llave = (ciudad, norm(comuna), norm(barrio)): no mezcla homónimos de comunas distintas
# (San Francisco en Suba vs Ciudad Bolívar son barrios diferentes por la lógica dd-mm-zz).
def _bestname(variants):  # grafía más limpia: con tilde > mixta (no TODO-MAYÚS) > más capitalizada
    def sc(s): return (any(ord(c)>127 for c in s),0 if s.isupper() else 1,sum(w[:1].isupper() for w in s.split()),sum(c.isupper() for c in s))
    b=max(variants,key=sc); return b.title() if (b.isupper() or b.islower()) else b
ws=sheet('Ciudades·barrio'); hdr(ws,['Ciudad','Comuna/Localidad','Barrio','Cepeda 1V','Techo (Petro 2V)','Votos por recuperar','Centro disponible','Votos 26 (base)','Cepeda %','Centro %'])
agg_b={}
for city,rows in BF['city_barrio'].items():
    for r in rows:
        if _noise(r['comuna']) or _noise(r['barrio']): continue
        base=r['base']; k=(city,_norm(r['comuna']),_norm(r['barrio']))
        a=agg_b.setdefault(k,{'city':city,'comv':set(),'barv':set(),'base':0,'cep':0,'techo':0,'cdisp':0,'craw':0})
        a['comv'].add(r['comuna']); a['barv'].add(r['barrio'])
        a['base']+=base; a['cep']+=round(r['cep26']/100*base); a['techo']+=round((r.get('petro2v') or 0)/100*base)
        a['cdisp']+=round((r.get('centro26') or 0)/100*base*FAJ_T); a['craw']+=(r.get('centro26') or 0)/100*base
rows_b=[]
for a in agg_b.values():
    rec=max(0,a['techo']-a['cep']); b=a['base'] or 1
    rows_b.append((rec,[a['city'],_bestname(a['comv']),_bestname(a['barv']),a['cep'],a['techo'],rec,a['cdisp'],a['base'],a['cep']/b,a['craw']/b]))
for _,row in sorted(rows_b,key=lambda x:-x[0]): ws.append(row)
fin(ws,(8,9))

# Bogotá por estrato socioeconómico
ws=sheet('Bogotá·estrato'); hdr(ws,['Estrato','Puestos','Censo electoral','% ciudad','Cepeda 26 %','Abelardo 26 %','Paloma 26 %','Fajardo 26 %','Petro 1V-22 %','Petro 2V-22 %','Dif Cep−Petro1V (pp)','Techo Petro 2V (pp)','Particip. 26 %'])
_tot=sum(EST['estratos'][str(e)]['censo'] for e in range(1,7)) or 1
for e in range(1,7):
    v=EST['estratos'][str(e)]
    ws.append([v['nombre'],v['puestos'],v['censo'],v['censo']/_tot,P(v['cepeda']),P(v['abelardo']),P(v['paloma']),P(v['fajardo']),P(v['pet1v22']),P(v['pet2v22']),v['dif_cep_pet'],v['techo_2v'],P(v['participacion'])])
fin(ws,(3,4,5,6,7,8,9,12),filt=False)
ws.append([])
m=EST['meta']; note=ws.cell(row=ws.max_row+1,column=1,value=f"Cruce punto-en-polígono de {m['puestos_asignados']} puestos georreferenciados de Bogotá contra la capa oficial de estratificación por manzana (SDP/IDECA, 44.260 manzanas, oct-2025). Distancia mediana puesto–manzana {m['dist_mediana_m']} m. El diferencial 2026 vs 2022 usa los {m['overlap_2022']} puestos apareados entre años. Estrato = manzana del puesto, proxy del entorno socioeconómico.")
note.font=Font(italic=True,size=8,color='5A5448'); note.alignment=Alignment(wrap_text=True); ws.merge_cells(start_row=note.row,start_column=1,end_row=note.row,end_column=13)

# 8 · Camino 2ª vuelta (municipio + puesto) — votos por recuperar hasta el techo
ws=sheet('8·Camino 2V (municipio)')
hdr(ws,['Departamento','Municipio','Cepeda 1V','Techo (Petro 2V)','Votos por recuperar','Centro disponible','Censo (válidos+blanco)','Abstención %'])
for r in TER['muni']:
    ws.append([r['dep'],r['muni'],r['cep_now'],r['techo'],r['recuperar'],r['centro'],r['base'],r['abst']/100])
fin(ws,(7,))
ws=sheet('8·Camino 2V (puesto)')
hdr(ws,['Clave puesto','Departamento','Municipio','Barrio','Comuna/Localidad','Cepeda 1V','Techo (Petro 2V)','Votos por recuperar','Censo puesto'])
for r in TER['puesto']:
    ws.append([r['pcode'],r.get('dep_nombre') or r['dep'],r.get('mun_nombre') or r['mun'],r['barrio'],r['comuna'],r['cep_now'],r['techo'],r['recuperar'],r['base']])
fin(ws)

# formato condicional automático: escala roja→verde en columnas (pp), barra en 'recuperar'/'centro disponible'
def cf_auto(ws,hdrrow=1):
    if ws.max_row<=hdrrow+1: return
    for cell in ws[hdrrow]:
        h=str(cell.value or '').lower(); L=get_column_letter(cell.column)
        if '(pp)' in h: cf_div(ws,L,first=hdrrow+1)
        elif 'recuperar' in h or 'centro disponible' in h: cf_bar(ws,L,first=hdrrow+1)
for ws in wb.worksheets:
    cf_auto(ws, 3 if ws.title=='Resumen' else 1)

wb.save(f'{OUT}/Soporte_Analisis_Pacto_1V_2026.xlsx')
print('✓ Excel:',len(wb.sheetnames),'hojas'); [print('  ·',s) for s in wb.sheetnames]
