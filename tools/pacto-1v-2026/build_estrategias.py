#!/usr/bin/env python3
# Tres estrategias de campaña para la 2ª vuelta (Cepeda), cada una en su propio Excel + JSON para el Word.
#   1) CENTRO       — votos de Fajardo+Claudia transferibles (0,55F + 0,65C). Se SUMA aparte (~700K).
#   2) RECUPERACIÓN — techo Petro-2V 2022 − Cepeda 1V-2026 (el universo de la izquierda demovilizada, ~2,0M).
#   3) ABSTENCIÓN   — neto = abstención × (2·share_Petro2V − 1); solo positivo donde la izquierda gana la 2V (~1,9M).
# Recuperación y Abstención miran el MISMO universo (no se suman entre sí). Centro sí se suma.
import json,csv,unicodedata,collections
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule,DataBarRule
OUT='Bases de datos/output_pacto_1v_2026'; B='Bases de datos'

# ── supuestos de trasvase (idénticos a build_2v.py) ──
FAJ_CEP=0.55; CLA_CEP=0.65
ABST_TARGET=1_915_513   # gap del modelo (twov_model.json)

norm=lambda s:''.join(c for c in unicodedata.normalize('NFD',str(s or '').upper().strip()) if unicodedata.category(c)!='Mn')
import re
def cln(raw):  # comuna/localidad limpia
    s=re.sub(r'^\d+','',(raw or '')).strip(); s=re.sub(r'(?i)\b(comuna|localidad)\b','',s).strip()
    return s.title() or 'Sin dato'
def _bestname(vs):  # grafía más limpia: con tilde > mixta (no TODO-MAYÚS) > más capitalizada
    def sc(s): return (any(ord(c)>127 for c in s),0 if s.isupper() else 1,sum(w[:1].isupper() for w in s.split()),sum(c.isupper() for c in s))
    b=max(vs,key=sc); return b.title() if (b.isupper() or b.islower()) else b
def _noise(s):
    u=str(s or '').upper().strip()
    return any(k in u for k in ('PUESTO CENSO','CONSULADO','CARCEL','CÁRCEL','EXTERIOR','NULL')) or u in ('OTROS','CORR','CIUDAD','SN','NO APLICA','SIN DATO','')
def _clab(c): return '(Sin comuna)' if _noise(c) else c   # comuna desconocida (NULL/OTROS) → etiqueta (mixta, sobrevive al title-case de _bestname); NO se descarta un barrio válido por eso (caso Soledad: barrios OK, comuna NULL)

# ── ciudades (código electoral, resuelto por nombre) ──
CITIES=[('Bogotá','16001'),('Medellín','01001'),('Cali','31001'),('Barranquilla','03001'),
        ('Cartagena','05001'),('Cúcuta','25001'),('Bucaramanga','27001'),('Soledad','03052'),
        ('Soacha','15247'),('Ibagué','29001'),('Pereira','24001'),('Manizales','09001'),
        ('Pasto','23001'),('Villavicencio','52001'),('Santa Marta','21001'),('Montería','13001')]
CITY_OF={code:nm for nm,code in CITIES}
CODE_OF={nm:code for nm,code in CITIES}

# ── datos ──
BF=json.load(open(f'{OUT}/blocks_full.json'))['muni']
TER=json.load(open(f'{OUT}/twov_territorial.json'))
DEP={'16':'Bogotá','01':'Antioquia','03':'Atlántico','05':'Bolívar','07':'Boyacá','09':'Caldas','11':'Cauca','12':'Cesar','13':'Córdoba','15':'Cundinamarca','17':'Chocó','19':'Huila','21':'Magdalena','23':'Nariño','24':'Risaralda','25':'Norte de Santander','26':'Quindío','27':'Santander','28':'Sucre','29':'Tolima','31':'Valle del Cauca','40':'Arauca','44':'Caquetá','46':'Casanare','48':'La Guajira','50':'Guainía','52':'Meta','54':'Guaviare','56':'San Andrés','60':'Amazonas','64':'Putumayo','68':'Vaupés','72':'Vichada','88':'Exterior'}
MUNI={}  # nombre de municipio (electoral) desde el georef — cubre los municipios chicos que no están en blocks_full
with open(f'{B}/PUESTOS_GEOREF.csv',newline='',encoding='utf-8',errors='ignore') as f:
    for row in csv.DictReader(f,delimiter=';'):
        cc=(row.get('CÓDIGO COMPLETO') or '').strip()
        if len(cc)>=5: MUNI.setdefault(cc[:5],(row.get('MUNICIPIO') or '').strip().title())
def muni_name(code):
    if code in BF: return BF[code]['dep'],BF[code]['muni']
    return DEP.get(code[:2],code[:2]),MUNI.get(code,code)
M26=json.load(open(f'{OUT}/master_2026_puesto.json'))
M22=json.load(open(f'{OUT}/master_2022_puesto.json'))
CAND=['cepeda','abelardo','paloma','fajardo','botero','lizcano','miguel_uribe','macollins','roy','murillo','caicedo','matamoros','claudia']
def pc(p):
    try: return f"{int(p['dep']):02d}{int(p['mun']):03d}{int(p['zona']):02d}{int(p['puesto']):02d}"
    except: return None
# share Petro-2V por puesto (techo)
share22={}
for p in (M22 if isinstance(M22,list) else M22.values()):
    k=pc(p)
    if k and p.get('base2v'): share22[k]=(p['petro2v']/p['base2v']) if p['base2v'] else None

# Fajardo + Claudia oficiales por municipio (CSV nombres corregidos)
faj_mun=collections.defaultdict(int); cla_mun=collections.defaultdict(int)
with open(f'{B}/nuevos archivos 1v 2026/Base nombres corregidos primera vuelta 2026.csv',newline='',encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        cand=(r['Candidato'] or '').upper(); code=str(r['cod_departamento'] or '').zfill(2)+str(r['cod_municipio'] or '').zfill(3)
        v=int(r['Suma de Votos'] or 0)
        if 'FAJARDO' in cand: faj_mun[code]+=v
        elif 'CLAUDIA' in cand: cla_mun[code]+=v
FAJ_TOT=sum(faj_mun.values()); CLA_TOT=sum(cla_mun.values())

# ── tabla de puestos enriquecida (para el detalle de ciudades de las 3 campañas) ──
# claves agregables por (ciudad, comuna) y (ciudad, comuna, barrio)
def z(x,n): return str(x).zfill(n)
PUE=[]
for p in M26:
    dep=z(p['dep'],2); mun=z(p['mun'],3); zona=z(p['zona'],2)
    if zona in ('90','98') or dep=='88': continue
    code=dep+mun
    if code not in CITY_OF: continue
    k=pc(p); sh=share22.get(k)
    base=sum(int(p.get(c,0) or 0) for c in CAND)+int(p.get('votos_blanco',0) or 0)
    if base<1: continue
    PUE.append(dict(city=CITY_OF[code],comuna=cln(p.get('comuna','')),barrio=(p.get('barrio') or '').strip(),
        cep=int(p.get('cepeda',0) or 0),faj=int(p.get('fajardo',0) or 0),base=base,
        pot=int(p.get('pot',0) or 0),urna=int(p.get('total_votos_urna',0) or 0),
        techo=round((sh or 0)*base) if sh is not None else None,has_sh=sh is not None))

# censo (pot) y votantes (urna) por municipio para abstención (todo el país)
pot_mun=collections.defaultdict(int); urna_mun=collections.defaultdict(int)
for p in M26:
    if z(p['zona'],2) in ('90','98') or z(p['dep'],2)=='88': continue
    code=z(p['dep'],2)+z(p['mun'],3)
    pot_mun[code]+=int(p.get('pot',0) or 0); urna_mun[code]+=int(p.get('total_votos_urna',0) or 0)

# ═══════════════════ estilo Excel ═══════════════════
OX='8A1E16'; thin=Side(style='thin',color='E2DDD0'); bd=Border(thin,thin,thin,thin); ZEBRA=PatternFill('solid',fgColor='F4F0E7')
def hdr(ws,headers,row=1):
    for i,h in enumerate(headers,1):
        c=ws.cell(row=row,column=i,value=h); c.font=Font(bold=True,color='FFFFFF',size=10); c.fill=PatternFill('solid',fgColor=OX)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bd
    ws.row_dimensions[row].height=30
def fin(ws,pcols=(),filt=True,hdrrow=1,zebra=True):
    for ri,row in enumerate(ws.iter_rows(min_row=hdrrow+1),start=hdrrow+1):
        zeb=zebra and (ri-hdrrow)%2==0
        for c in row:
            i=c.column-1; c.border=bd
            if i in pcols and isinstance(c.value,(int,float)): c.number_format='0.0%'
            elif isinstance(c.value,(int,float)) and abs(c.value)>=1000: c.number_format='#,##0'
            if zeb: c.fill=ZEBRA
    for col in ws.columns:
        w=max((len(str(c.value)) for c in col if c.value is not None),default=8); ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(w+2,11),40)
    if filt and ws.max_row>hdrrow+1: ws.freeze_panes=f'A{hdrrow+1}'; ws.auto_filter.ref=ws.dimensions
def cf_bar(ws,L,first=2,color=OX): ws.conditional_formatting.add(f'{L}{first}:{L}{ws.max_row}',DataBarRule(start_type='min',end_type='max',color=color,showValue=True))
def cf_auto(ws,hdrrow=1):
    if ws.max_row<=hdrrow+1: return
    for cell in ws[hdrrow]:
        h=str(cell.value or '').lower(); L=get_column_letter(cell.column)
        if any(w in h for w in ('transferible','recuperar','neto cepeda','abstención')) and 'acum' not in h and '%' not in h: cf_bar(ws,L,first=hdrrow+1)
def leeme(wb,titulo,lineas):
    ws=wb.active; ws.title='Léeme'
    ws.append([titulo]); ws['A1'].font=Font(bold=True,size=14,color=OX); ws.append([])
    for ln in lineas:
        ws.append([ln]); c=ws.cell(row=ws.max_row,column=1)
        c.alignment=Alignment(wrap_text=True,vertical='top'); c.font=Font(size=10.5 if not ln.startswith('· ') else 10,bold=ln.endswith(':'))
    ws.column_dimensions['A'].width=110
    return ws
def acumcols(rows,validx):  # añade acumulado + % acum (sobre total) a una lista de filas (val en posición validx, 0-based)
    tot=sum(r[validx] for r in rows) or 1; acc=0; out=[]
    for r in rows:
        acc+=r[validx]; out.append(list(r)+[round(acc),acc/tot])
    return out,tot

# ═══════════════════ 1 · CENTRO ═══════════════════
def build_centro():
    wb=Workbook()
    leeme(wb,'Estrategia 1 · CENTRO — dónde están los ~700K votos a persuadir',[
        'QUÉ ES:',
        'El voto de centro (Sergio Fajardo + Claudia López) que se puede transferir a Cepeda en 2ª vuelta. Es una campaña de PERSUASIÓN a gente que ya votó, pero por el centro.',
        'CÓMO SE CALCULA:',
        f'· Centro transferible = 0,55 × Fajardo + 0,65 × Claudia (mismos supuestos del modelo de 2ª vuelta).',
        f'· Total nacional: Fajardo {FAJ_TOT:,.0f} + Claudia {CLA_TOT:,.0f}  →  {round(0.55*FAJ_TOT+0.65*CLA_TOT):,.0f} votos transferibles (~700K).',
        'CÓMO LEER LAS HOJAS:',
        '· "Por municipio (nacional)": ordenado de mayor a menor. La columna "Acumulado" suma de arriba hacia abajo; "% acum" dice qué fracción del total nacional llevas. Ej: con el top de ciudades ya tienes la mitad de los 700K.',
        '· Hojas de ciudades (Bogotá localidad/barrio, grandes ciudades comuna/barrio): el detalle fino para territorializar la persuasión.',
        'SALVEDAD IMPORTANTE:',
        '· Claudia López solo existe a nivel de MUNICIPIO en el preconteo (su columna llegó en 0 por mesa; se recuperó como residual agregado al municipio). En "Por municipio" va su dato REAL. En las hojas por localidad/comuna/barrio se ESTIMA: se reparte el total municipal de Claudia en proporción a cómo se distribuye Fajardo (ambos son voto de centro urbano), en la columna "Claudia (est.)". El "Centro transferible" submunicipal ya la incluye: 0,55 × Fajardo + 0,65 × Claudia (est.).',
        '· El voto del exterior (consulados) va en una sola fila agregada "Exterior (consulados)": son votos reales de centro, pero no se trabajan en territorio como un municipio.',
        'CÓMO ENCAJA CON LAS OTRAS DOS CAMPAÑAS:',
        '· Esta campaña SÍ se SUMA a la de Abstención (son votantes distintos: aquí persuades a quien votó centro; allá movilizas a quien no votó). 700K (centro) + 1,9M (movilización) = 2,65M, lo que Cepeda necesita sobre su 1ª vuelta para empatar el piso de la derecha.'])
    # por municipio
    ws=wb.create_sheet('Por municipio (nacional)'); hdr(ws,['#','Departamento','Municipio','Fajardo','Claudia','Centro transferible','Acumulado','% acum'])
    rows=[]; ext_f=ext_c=0
    for code in set(faj_mun)|set(cla_mun):   # CSV oficial como columna vertebral (incluye municipios chicos)
        f=faj_mun.get(code,0); c=cla_mun.get(code,0)
        if code[:2]=='88': ext_f+=f; ext_c+=c; continue   # exterior: una sola fila agregada (no territorializable)
        tr=round(0.55*f+0.65*c)
        if tr<=0: continue
        dep,mun=muni_name(code); rows.append([dep,mun,f,c,tr])
    if ext_f or ext_c: rows.append(['Exterior','Exterior (consulados)',ext_f,ext_c,round(0.55*ext_f+0.65*ext_c)])
    rows.sort(key=lambda r:-r[4])
    rows,tot=acumcols(rows,4)
    for i,r in enumerate(rows,1): ws.append([i]+r)
    fin(ws,(7,)); cf_auto(ws)
    # detalle ciudades: Bogotá localidad, ciudades comuna, ciudades barrio (Fajardo×0,55)
    _centro_localidad(wb); _centro_comuna(wb); _centro_barrio(wb)
    wb.save(f'{OUT}/Estrategia_1_Centro.xlsx'); return tot

# Claudia (est.) submunicipal = total municipal de Claudia repartido en proporción a Fajardo
def _cla_est(code,faj,cityfaj): return round(cla_mun.get(code,0)*faj/cityfaj) if cityfaj else 0

def _centro_localidad(wb):
    ws=wb.create_sheet('Bogotá · localidad'); hdr(ws,['#','Localidad','Fajardo','Claudia (est.)','Centro transferible','Acumulado','% acum'])
    agg=collections.defaultdict(int)
    for p in PUE:
        if p['city']=='Bogotá' and not _noise(p['comuna']): agg[p['comuna']]+=p['faj']
    cf=sum(agg.values())
    rows=[]
    for com,f in agg.items():
        cla=_cla_est(CODE_OF['Bogotá'],f,cf); rows.append([com,f,cla,round(0.55*f+0.65*cla)])
    rows.sort(key=lambda r:-r[3]); rows,_=acumcols(rows,3)
    for i,r in enumerate(rows,1): ws.append([i]+r)
    fin(ws,(6,)); cf_auto(ws)

def _centro_comuna(wb):
    ws=wb.create_sheet('Grandes ciudades · comuna'); hdr(ws,['Ciudad','Comuna/Localidad','Fajardo','Claudia (est.)','Centro transferible'])
    agg=collections.defaultdict(int)
    for p in PUE:
        agg[(p['city'],_clab(p['comuna']))]+=p['faj']
    cityfaj=collections.defaultdict(int)
    for (c,com),f in agg.items(): cityfaj[c]+=f
    rows=[]
    for (c,com),f in agg.items():
        cla=_cla_est(CODE_OF[c],f,cityfaj[c]); rows.append([c,com,f,cla,round(0.55*f+0.65*cla)])
    rows.sort(key=lambda r:(r[0],-r[4]))
    for r in rows: ws.append(r)
    fin(ws); cf_auto(ws)

def _centro_barrio(wb):
    ws=wb.create_sheet('Grandes ciudades · barrio'); hdr(ws,['Ciudad','Comuna/Localidad','Barrio','Fajardo','Claudia (est.)','Centro transferible'])
    agg={}
    for p in PUE:
        if _noise(p['barrio']): continue
        com=_clab(p['comuna']); k=(p['city'],norm(com),norm(p['barrio']))
        a=agg.setdefault(k,{'city':p['city'],'comv':set(),'barv':set(),'faj':0}); a['comv'].add(com); a['barv'].add(p['barrio']); a['faj']+=p['faj']
    cityfaj=collections.defaultdict(int)
    for a in agg.values(): cityfaj[a['city']]+=a['faj']
    rows=[]
    for a in agg.values():
        cla=_cla_est(CODE_OF[a['city']],a['faj'],cityfaj[a['city']]); rows.append([a['city'],_bestname(a['comv']),_bestname(a['barv']),a['faj'],cla,round(0.55*a['faj']+0.65*cla)])
    rows.sort(key=lambda r:(r[0],-r[5]))
    for r in rows: ws.append(r)
    fin(ws); cf_auto(ws)

# ═══════════════════ 2 · RECUPERACIÓN (techo Petro-2V) ═══════════════════
def build_recuperacion():
    wb=Workbook()
    rec_tot=sum(r['recuperar'] for r in TER['muni'])
    leeme(wb,'Estrategia 2 · RECUPERACIÓN — el techo de la izquierda (~2,0M)',[
        'QUÉ ES:',
        'El universo de votos que la izquierda YA TUVO en su mejor noche reciente (Petro, 2ª vuelta 2022) y que Cepeda todavía no recupera. Es la campaña para la gente que se fue por desilusión o por los acuerdos de la 2ª vuelta pasada.',
        'CÓMO SE CALCULA:',
        '· Techo = % de Petro en 2ª vuelta 2022 aplicado al electorado de 2026.',
        '· Votos por recuperar = máx(0, Techo − Cepeda 1ª vuelta 2026).',
        f'· Total nacional (suma municipio por municipio): {rec_tot:,.0f} votos.',
        'CÓMO LEER LAS HOJAS:',
        '· "Por municipio" y "Por puesto": ordenados por votos a recuperar. Es el mapa de PRIORIDAD: dónde la izquierda fue más grande que Cepeda hoy.',
        '· Hojas de ciudades por comuna y barrio: el mismo cálculo, territorializado.',
        'CÓMO ENCAJA CON LAS OTRAS DOS CAMPAÑAS:',
        '· OJO — esta campaña NO se suma con la de Abstención: ambas describen el MISMO universo (la coalición de Petro-2022 que hoy no está con Cepeda), desde dos ángulos. Recuperación lo DIMENSIONA (cuánto y dónde); Abstención dice CÓMO recuperarlo (subiendo participación donde ganamos la 2V).',
        '· La parte de este universo que se recupera movilizando abstención son ~1,9M (ver Estrategia 3).'])
    ws=wb.create_sheet('Por municipio'); hdr(ws,['#','Departamento','Municipio','Cepeda 1V','Techo (Petro 2V)','Votos por recuperar','Acumulado','% acum'])
    rows=[[r['dep'],r['muni'],r['cep_now'],r['techo'],r['recuperar']] for r in TER['muni'] if r['recuperar']>0]
    rows.sort(key=lambda r:-r[4]); rows,_=acumcols(rows,4)
    for i,r in enumerate(rows,1): ws.append([i]+r)
    fin(ws,(7,)); cf_auto(ws)
    ws=wb.create_sheet('Por puesto'); hdr(ws,['Clave puesto','Departamento','Municipio','Comuna/Localidad','Barrio','Cepeda 1V','Techo (Petro 2V)','Votos por recuperar'])
    for r in TER['puesto']:
        if r['recuperar']<=0: continue
        ws.append([r['pcode'],r.get('dep_nombre') or r['dep'],r.get('mun_nombre') or r['mun'],r['comuna'],r['barrio'],r['cep_now'],r['techo'],r['recuperar']])
    fin(ws); cf_auto(ws)
    _rec_comuna(wb); _rec_barrio(wb)
    wb.save(f'{OUT}/Estrategia_2_Recuperacion.xlsx'); return rec_tot

def _rec_comuna(wb):
    ws=wb.create_sheet('Grandes ciudades · comuna'); hdr(ws,['Ciudad','Comuna/Localidad','Cepeda 1V','Techo (Petro 2V)','Votos por recuperar'])
    agg=collections.defaultdict(lambda:[0,0])
    for p in PUE:
        if not p['has_sh']: continue
        a=agg[(p['city'],_clab(p['comuna']))]; a[0]+=p['cep']; a[1]+=p['techo']
    rows=sorted(([c,com,cep,tec,max(0,tec-cep)] for (c,com),(cep,tec) in agg.items()),key=lambda r:(r[0],-r[4]))
    for r in rows: ws.append(r)
    fin(ws); cf_auto(ws)

def _rec_barrio(wb):
    ws=wb.create_sheet('Grandes ciudades · barrio'); hdr(ws,['Ciudad','Comuna/Localidad','Barrio','Cepeda 1V','Techo (Petro 2V)','Votos por recuperar'])
    agg={}
    for p in PUE:
        if _noise(p['barrio']) or not p['has_sh']: continue
        com=_clab(p['comuna']); k=(p['city'],norm(com),norm(p['barrio']))
        a=agg.setdefault(k,{'city':p['city'],'comv':set(),'barv':set(),'cep':0,'tec':0}); a['comv'].add(com); a['barv'].add(p['barrio']); a['cep']+=p['cep']; a['tec']+=p['techo']
    rows=sorted(([a['city'],_bestname(a['comv']),_bestname(a['barv']),a['cep'],a['tec'],max(0,a['tec']-a['cep'])] for a in agg.values()),key=lambda r:(r[0],-r[5]))
    for r in rows: ws.append(r)
    fin(ws); cf_auto(ws)

# ═══════════════════ 3 · ABSTENCIÓN (neto sobre techo 2V) ═══════════════════
def abst_muni_rows():
    rows=[]
    for code,v in BF.items():
        if v['dep']=='Exterior': continue
        censo=pot_mun.get(code,0); vot=urna_mun.get(code,0); ab=censo-vot
        if censo<=0 or ab<=0: continue
        sh=(v.get('petro2v') or 0)/100; cep=(v.get('cep26') or 0)/100
        neto=max(0,round(ab*(2*sh-1)))
        rows.append(dict(code=code,dep=v['dep'],muni=v['muni'],censo=censo,vot=vot,ab=ab,abpct=ab/censo,sh=sh,cep=cep,neto=neto))
    rows.sort(key=lambda r:-r['neto'])
    return rows

def build_abstencion():
    wb=Workbook()
    rows=abst_muni_rows(); tot=sum(r['neto'] for r in rows)
    # ¿cuántos municipios para juntar el objetivo?
    acc=0; n_target=0
    for r in rows:
        acc+=r['neto']; n_target+=1
        if acc>=ABST_TARGET: break
    leeme(wb,'Estrategia 3 · ABSTENCIÓN — movilizar donde somos fuertes (~1,9M)',[
        'QUÉ ES:',
        'El voto que se gana SACANDO A VOTAR a la gente que hoy se abstiene, pero SOLO en las zonas donde la izquierda gana de verdad. Es una campaña de MOVILIZACIÓN (GOTV), no de persuasión.',
        'CÓMO SE CALCULA (y por qué es "neto"):',
        '· En 2ª vuelta, cada nuevo votante alimenta a los DOS bloques. Si en un territorio la izquierda gana 60-40, por cada 10 abstencionistas que saco a votar gano 6 y pierdo 4: neto +2.',
        '· Neto Cepeda = abstención × (2 × share − 1), donde share = % de Petro en 2ª vuelta 2022 (el resultado cabeza a cabeza).',
        '· Esto es POSITIVO solo donde share > 50% (donde la izquierda ya ganó la 2V de 2022). En zonas que perdemos, movilizar le ayuda a Abelardo: por eso no aparecen.',
        f'· Total nacional neto disponible: {tot:,.0f}. El objetivo del modelo (la brecha a cerrar) es {ABST_TARGET:,.0f}.',
        'RESPUESTA A TU PREGUNTA ("colocar municipios hasta llegar a 1,9M"):',
        f'· Los {ABST_TARGET:,.0f} netos se juntan con los PRIMEROS {n_target} municipios de la lista. La columna "Suma el objetivo 1,9M" marca con ✓ exactamente esos municipios prioritarios.',
        'CÓMO LEER LAS HOJAS:',
        '· Ordenadas por "Neto Cepeda". "Censo" = potencial electoral; "Votantes" = quienes votaron; "Abstención" = la diferencia. "Share Petro 2V" mide qué tan fuerte es la izquierda ahí.',
        '· El detalle baja a 4 niveles: por municipio (nacional) · por localidad (Bogotá) · por comuna (16 ciudades) · POR BARRIO (16 ciudades). Los barrios se fusionan por (ciudad, comuna, barrio) para no duplicar por diferencias de tilde/mayúscula.',
        '· OJO con sumar entre niveles: NO cuadran exacto entre sí, por dos razones. (1) El neto se recorta a cero donde la izquierda no gana, así que al bajar a barrio afloran bolsones fuertes dentro de comunas mixtas y el total por barrio suele salir más alto (mobilización quirúrgica solo en los barrios fuertes). (2) El detalle por comuna/barrio solo incluye puestos cruzados con la 2ª vuelta 2022, así que en municipios con cruce parcial puede salir más bajo. Usa cada hoja para UBICAR dónde concentrar; el objetivo nacional de 1,9M se mide SIEMPRE en la hoja por municipio (no se re-suma desde barrio).',
        '· Municipios cuya comuna llega como "NULL" en el preconteo (p. ej. Soledad) sí entran a barrio: sus barrios son válidos y se agrupan bajo "(sin comuna)".',
        'CÓMO ENCAJA CON LAS OTRAS DOS CAMPAÑAS:',
        '· SÍ se suma con Centro (700K + 1,9M = 2,65M). NO se suma con Recuperación (son el mismo universo visto distinto).'])
    ws=wb.create_sheet('Por municipio'); hdr(ws,['#','Departamento','Municipio','Censo','Votantes','Abstención','Abst. %','Share Petro 2V','Share Cepeda 1V','Neto Cepeda','Acumulado','% acum','Suma el objetivo 1,9M'])
    acc=0
    for i,r in enumerate(rows,1):
        acc+=r['neto']
        ws.append([i,r['dep'],r['muni'],r['censo'],r['vot'],r['ab'],r['abpct'],r['sh'],r['cep'],r['neto'],round(acc),acc/tot,'✓' if acc<=ABST_TARGET or (acc-r['neto'])<ABST_TARGET else ''])
    fin(ws,(6,7,8,11)); cf_auto(ws)
    # cierre exacto en n_target
    _abst_localidad(wb); _abst_comuna(wb); _abst_barrio(wb)
    wb.save(f'{OUT}/Estrategia_3_Abstencion.xlsx'); return tot,n_target

def _abst_city_agg(only_bogota=False):
    agg=collections.defaultdict(lambda:[0,0,0,0])  # pot, urna, base, techo
    for p in PUE:
        if only_bogota and p['city']!='Bogotá': continue
        if not p['has_sh']: continue
        a=agg[(p['city'],_clab(p['comuna']))]; a[0]+=p['pot']; a[1]+=p['urna']; a[2]+=p['base']; a[3]+=p['techo']
    out=[]
    for (c,com),(pot,urna,base,techo) in agg.items():
        if pot<=0 or base<=0: continue
        ab=pot-urna; sh=techo/base
        if ab<=0: continue
        out.append([c,com,pot,urna,ab,ab/pot,sh,max(0,round(ab*(2*sh-1)))])
    return out

def _abst_localidad(wb):
    ws=wb.create_sheet('Bogotá · localidad'); hdr(ws,['Localidad','Censo','Votantes','Abstención','Abst. %','Share Petro 2V','Neto Cepeda'])
    rows=sorted(([r[1],r[2],r[3],r[4],r[5],r[6],r[7]] for r in _abst_city_agg(only_bogota=True)),key=lambda r:-r[6])
    for r in rows: ws.append(r)
    fin(ws,(4,5)); cf_auto(ws)

def _abst_comuna(wb):
    ws=wb.create_sheet('Grandes ciudades · comuna'); hdr(ws,['Ciudad','Comuna/Localidad','Censo','Votantes','Abstención','Abst. %','Share Petro 2V','Neto Cepeda'])
    rows=sorted(_abst_city_agg(),key=lambda r:(r[0],-r[7]))
    for r in rows: ws.append(r)
    fin(ws,(5,6)); cf_auto(ws)

def _abst_barrio(wb):  # abstención por barrio · dedup por (ciudad, norm(comuna), norm(barrio)) — mismo fix de tildes
    ws=wb.create_sheet('Grandes ciudades · barrio'); hdr(ws,['Ciudad','Comuna/Localidad','Barrio','Censo','Votantes','Abstención','Abst. %','Share Petro 2V','Neto Cepeda'])
    agg={}
    for p in PUE:
        if _noise(p['barrio']) or not p['has_sh']: continue
        com=_clab(p['comuna']); k=(p['city'],norm(com),norm(p['barrio']))
        a=agg.setdefault(k,{'city':p['city'],'comv':set(),'barv':set(),'pot':0,'urna':0,'base':0,'techo':0})
        a['comv'].add(com); a['barv'].add(p['barrio'])
        a['pot']+=p['pot']; a['urna']+=p['urna']; a['base']+=p['base']; a['techo']+=p['techo']
    rows=[]
    for a in agg.values():
        if a['pot']<=0 or a['base']<=0: continue
        ab=a['pot']-a['urna']; sh=a['techo']/a['base']
        if ab<=0: continue
        rows.append([a['city'],_bestname(a['comv']),_bestname(a['barv']),a['pot'],a['urna'],ab,ab/a['pot'],sh,max(0,round(ab*(2*sh-1)))])
    rows.sort(key=lambda r:(r[0],-r[8]))
    for r in rows: ws.append(r)
    fin(ws,(6,7)); cf_auto(ws)

# ═══════════════════ run ═══════════════════
centro_tot=build_centro()
rec_tot=build_recuperacion()
abst_tot,abst_n=build_abstencion()

M2V=json.load(open(f'{OUT}/twov_model.json'))
# top listas para el Word
centro_top=[]
for code,v in BF.items():
    if v['dep']=='Exterior': continue
    f=faj_mun.get(code,0); c=cla_mun.get(code,0); tr=round(0.55*f+0.65*c)
    if tr>0: centro_top.append({'muni':v['muni'],'dep':v['dep'],'faj':f,'cla':c,'tr':tr})
centro_top.sort(key=lambda r:-r['tr']); centro_top=centro_top[:12]
arows=abst_muni_rows()
abst_top=[{'muni':r['muni'],'dep':r['dep'],'censo':r['censo'],'ab':r['ab'],'sh':round(r['sh']*100,1),'neto':r['neto']} for r in arows[:12]]
rec_top=[{'muni':r['muni'],'dep':r['dep'],'cep_now':r['cep_now'],'techo':r['techo'],'recuperar':r['recuperar']} for r in TER['muni'][:10]]
# totales municipales por ciudad (las 9 con mapa por barrio) — el ANCLA que citan los párrafos del Word
CITY9=[('Bogotá','16001'),('Medellín','01001'),('Cali','31001'),('Barranquilla','03001'),('Cartagena','05001'),
       ('Manizales','09001'),('Pereira','24001'),('Bucaramanga','27001'),('Cúcuta','25001'),
       # zonas fuertes de abstención añadidas (mapa por barrio Voronoi)
       ('Pasto','23001'),('Buenaventura','31019'),('Santa Marta','21001'),('Palmira','31079'),('Tumaco','23139'),('Sincelejo','28001'),('Soledad','03052'),('Quibdó','17001')]
_rec_by={r['cod']:r['recuperar'] for r in TER['muni']}
_abst_by={r['code']:r for r in arows}
city9={}
for nm,code in CITY9:
    f=faj_mun.get(code,0); c=cla_mun.get(code,0)
    ar=_abst_by.get(code,{})
    city9[code]={'name':nm,'centro':round(0.55*f+0.65*c),'recuperar':_rec_by.get(code,0),
                 'abst_neto':ar.get('neto',0),'abst_share':round(ar.get('sh',0)*100,1) if ar else 0}
EST={'centro_total':round(0.55*FAJ_TOT+0.65*CLA_TOT),'fajardo_total':FAJ_TOT,'claudia_total':CLA_TOT,
     'city9':city9,
     'recuperar_total':rec_tot,'abst_target':ABST_TARGET,'abst_total_neto':round(abst_tot),'abst_n_muni':abst_n,
     'gap':M2V['gap'],'center_contrib':M2V['center_contrib'],'need_over_1v':M2V['need_over_1v'],
     'abe_floor':M2V['abe_floor'],'cep_ceiling':M2V['cep_ceiling'],
     'centro_top':centro_top,'abst_top':abst_top,'rec_top':rec_top}
json.dump(EST,open(f'{OUT}/twov_estrategias.json','w'),ensure_ascii=False,indent=1)

print('✓ Estrategia_1_Centro.xlsx       · centro transferible nacional =',f"{centro_tot:,.0f}")
print('✓ Estrategia_2_Recuperacion.xlsx · recuperar nacional (techo)   =',f"{rec_tot:,.0f}")
print('✓ Estrategia_3_Abstencion.xlsx   · neto nacional =',f"{abst_tot:,.0f}",f"· objetivo {ABST_TARGET:,} con top {abst_n} municipios")
print('✓ twov_estrategias.json')
