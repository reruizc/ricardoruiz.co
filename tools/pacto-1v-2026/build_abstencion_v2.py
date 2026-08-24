#!/usr/bin/env python3
# Estrategia 3 · Abstención — VERSIÓN 2.0 NACIONAL.
# A diferencia del Excel operativo (solo zonas fuertes, neto recortado a 0), esta versión trae
# TODOS los municipios y barrios del país, con el NETO FIRMADO (negativo donde la izquierda pierde
# la 2V = movilizar le ayuda al rival) y el SHARE Petro-2V con escala de colores (rojo→verde).
import json,collections,unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
OUT='Bases de datos/output_pacto_1v_2026'; B='Bases de datos'
norm=lambda s:''.join(c for c in unicodedata.normalize('NFD',str(s or '').upper().strip()) if unicodedata.category(c)!='Mn')
import re
def cln(raw):
    s=re.sub(r'^\d+','',(raw or '')).strip(); s=re.sub(r'(?i)\b(comuna|localidad)\b','',s).strip()
    return s.title() or ''
def _bestname(vs):
    def sc(s): return (any(ord(c)>127 for c in s),0 if s.isupper() else 1,sum(w[:1].isupper() for w in s.split()),sum(c.isupper() for c in s))
    b=max(vs,key=sc); return b.title() if (b.isupper() or b.islower()) else b
def _noise(s):
    u=norm(s); return any(k in u for k in ('PUESTO CENSO','CONSULADO','CARCEL','EXTERIOR','NULL')) or u in ('OTROS','CORR','CIUDAD','SN','NO APLICA','SIN DATO','')
def z(x,n): return str(x).zfill(n)
DEP={'16':'Bogotá','01':'Antioquia','03':'Atlántico','05':'Bolívar','07':'Boyacá','09':'Caldas','11':'Cauca','12':'Cesar','13':'Córdoba','15':'Cundinamarca','17':'Chocó','19':'Huila','21':'Magdalena','23':'Nariño','24':'Risaralda','25':'Norte de Santander','26':'Quindío','27':'Santander','28':'Sucre','29':'Tolima','31':'Valle del Cauca','40':'Arauca','44':'Caquetá','46':'Casanare','48':'La Guajira','50':'Guainía','52':'Meta','54':'Guaviare','56':'San Andrés','60':'Amazonas','64':'Putumayo','68':'Vaupés','72':'Vichada','88':'Exterior'}

BF=json.load(open(f'{OUT}/blocks_full.json'))['muni']
M26=json.load(open(f'{OUT}/master_2026_puesto.json')); M22=json.load(open(f'{OUT}/master_2022_puesto.json'))
def pc(p):
    try: return f"{int(p['dep']):02d}{int(p['mun']):03d}{int(p['zona']):02d}{int(p['puesto']):02d}"
    except: return None
sh22={pc(p):(p['petro2v']/p['base2v']) for p in M22 if pc(p) and p.get('base2v')}
CAND=['cepeda','abelardo','paloma','fajardo','botero','lizcano','miguel_uribe','macollins','roy','murillo','caicedo','matamoros','claudia']
# nombre de municipio desde georef (cubre municipios chicos)
MUNI={}
import csv
with open(f'{B}/PUESTOS_GEOREF.csv',newline='',encoding='utf-8',errors='ignore') as f:
    for row in csv.DictReader(f,delimiter=';'):
        cc=(row.get('CÓDIGO COMPLETO') or '').strip()
        if len(cc)>=5: MUNI.setdefault(cc[:5],(row.get('MUNICIPIO') or '').strip().title())
def muni_name(code): return BF[code]['muni'] if code in BF else MUNI.get(code,code)

def categoria(sh):
    if sh>=0.60: return 'Bastión izquierda (>60%)'
    if sh>=0.50: return 'Gana la 2V (50-60%)'
    if sh>=0.40: return 'Pierde la 2V (40-50%)'
    return 'Bastión derecha (<40%)'

# ── agregación municipal: censo/votantes desde master, share/cep desde blocks_full ──
pot_m=collections.defaultdict(int); urna_m=collections.defaultdict(int)
for p in M26:
    if z(p['zona'],2) in ('90','98') or z(p['dep'],2)=='88': continue
    c=z(p['dep'],2)+z(p['mun'],3); pot_m[c]+=int(p.get('pot',0) or 0); urna_m[c]+=int(p.get('total_votos_urna',0) or 0)
muni_rows=[]
for code,v in BF.items():
    if v['dep']=='Exterior' or v.get('petro2v') is None: continue
    censo=pot_m.get(code,0); vot=urna_m.get(code,0); ab=censo-vot
    if censo<=0: continue
    sh=v['petro2v']/100; cep=(v.get('cep26') or 0)/100
    neto=round(ab*(2*sh-1))   # FIRMADO
    muni_rows.append([v['dep'],v['muni'],censo,vot,ab,(ab/censo if censo else 0),sh,cep,neto,categoria(sh)])
muni_rows.sort(key=lambda r:-r[8])

# ── agregación barrial nacional: (dep,mun,norm(comuna),norm(barrio)) ──
agg={}
for p in M26:
    if z(p['zona'],2) in ('90','98') or z(p['dep'],2)=='88': continue
    bar=(p.get('barrio') or '').strip()
    if not bar or _noise(bar): continue
    sh=sh22.get(pc(p))
    if sh is None: continue
    dep=z(p['dep'],2); mun=z(p['mun'],3); com=cln(p.get('comuna',''))
    if _noise(com): com=''
    base=sum(int(p.get(c,0) or 0) for c in CAND)+int(p.get('votos_blanco',0) or 0)
    k=(dep,mun,norm(com),norm(bar))
    a=agg.setdefault(k,{'dep':dep,'mun':mun,'comv':set(),'barv':set(),'pot':0,'urna':0,'base':0,'techo':0})
    if com: a['comv'].add(com)
    a['barv'].add(bar); a['pot']+=int(p.get('pot',0) or 0); a['urna']+=int(p.get('total_votos_urna',0) or 0)
    a['base']+=base; a['techo']+=round(sh*base)
bar_rows=[]
for a in agg.values():
    if a['base']<=0 or a['pot']<=0: continue
    sh=a['techo']/a['base']; ab=a['pot']-a['urna']
    neto=round(ab*(2*sh-1))   # FIRMADO
    bar_rows.append([DEP.get(a['dep'],a['dep']),muni_name(a['dep']+a['mun']),
        (_bestname(a['comv']) if a['comv'] else ''),_bestname(a['barv']),
        a['pot'],a['urna'],ab,(ab/a['pot'] if a['pot'] else 0),sh,neto,categoria(sh)])
bar_rows.sort(key=lambda r:-r[9])

# ═══════════ Excel ═══════════
OX='8A1E16'; thin=Side(style='thin',color='E2DDD0'); bd=Border(thin,thin,thin,thin); ZEBRA=PatternFill('solid',fgColor='F4F0E7')
wb=Workbook()
def hdr(ws,headers):
    for i,h in enumerate(headers,1):
        c=ws.cell(row=1,column=i,value=h); c.font=Font(bold=True,color='FFFFFF',size=10); c.fill=PatternFill('solid',fgColor=OX)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bd
    ws.row_dimensions[1].height=30
def fin(ws,pcols=()):
    for ri,row in enumerate(ws.iter_rows(min_row=2),start=2):
        zeb=(ri%2==0)
        for c in row:
            i=c.column-1; c.border=bd
            if i in pcols and isinstance(c.value,(int,float)): c.number_format='0.0%'
            elif isinstance(c.value,(int,float)) and abs(c.value)>=1000: c.number_format='#,##0'
            if zeb: c.fill=ZEBRA
    for col in ws.columns:
        w=max((len(str(c.value)) for c in col if c.value is not None),default=8); ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(w+2,11),40)
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
def cf_share(ws,col):  # rojo (pierde) → crema (50%) → verde (gana)
    ws.conditional_formatting.add(f'{col}2:{col}{ws.max_row}',ColorScaleRule(
        start_type='num',start_value=0.35,start_color='D06B57',mid_type='num',mid_value=0.5,mid_color='FBF6EC',
        end_type='num',end_value=0.65,end_color='4F9E68'))
def cf_neto(ws,col):   # rojo (negativo) → crema (0) → verde (positivo)
    ws.conditional_formatting.add(f'{col}2:{col}{ws.max_row}',ColorScaleRule(
        start_type='num',start_value=-4000,start_color='D06B57',mid_type='num',mid_value=0,mid_color='FBF6EC',
        end_type='num',end_value=4000,end_color='4F9E68'))

# Léeme
ws=wb.active; ws.title='Léeme'
ws.append(['Estrategia 3 · Abstención — VERSIÓN 2.0 NACIONAL']); ws['A1'].font=Font(bold=True,size=14,color=OX); ws.append([])
for ln in [
 'QUÉ ES (y en qué se diferencia del Excel operativo):',
 '· El Excel operativo "Estrategia_3_Abstencion.xlsx" trae SOLO las zonas fuertes (share>50%) y recorta el neto a 0 donde la izquierda pierde la 2V, porque ahí movilizar no conviene.',
 '· Esta v2.0 trae TODOS los municipios (1.090) y barrios (5.083) del país, SIN recortar: el "Neto Cepeda (firmado)" puede ser NEGATIVO.',
 'QUÉ SIGNIFICA CADA COLUMNA CLAVE:',
 '· "Share Petro 2V": el % que sacó Petro (la izquierda) en la 2ª vuelta de 2022 en ese territorio. Es el resultado cabeza a cabeza izquierda vs. derecha. Por encima de 50% la izquierda gana la 2V ahí; por debajo, la pierde. Es la base de todo el cálculo.',
 '· "Neto Cepeda (firmado)": cuántos votos netos GANA (+) o PIERDE (−) Cepeda si sube la participación en ese territorio. Se calcula como abstención × (2 × Share Petro 2V − 1). Combina cuánta abstención hay con de qué lado está el territorio.',
 '· "Categoría 2V": etiqueta de fuerza según el share — Bastión izquierda (>60%) · Gana la 2V (50-60%) · Pierde la 2V (40-50%) · Bastión derecha (<40%). Sirve para filtrar rápido por tipo de territorio.',
 'CÓMO LEER EL NETO FIRMADO:',
 '· Neto = abstención × (2 × share − 1), con el share de Petro en la 2ª vuelta de 2022.',
 '· POSITIVO (verde): la izquierda gana la 2V ahí → sacar abstención SUMA a Cepeda. Es zona de movilización.',
 '· NEGATIVO (rojo): la izquierda pierde la 2V ahí → sacar abstención le SUMA a Abelardo. NO movilizar (o sería contraproducente).',
 '· Cerca de 0: territorio parejo (share ~50%), la movilización casi no mueve el saldo.',
 'COLORES:',
 '· La columna "Share Petro 2V" va con escala rojo→crema→verde (rojo = bastión rival <35%, verde = bastión propio >65%).',
 '· La columna "Neto Cepeda (firmado)" va con escala rojo (negativo) → crema (0) → verde (positivo).',
 '· La columna "Categoría 2V" clasifica: Bastión izquierda (>60%) · Gana la 2V (50-60%) · Pierde la 2V (40-50%) · Bastión derecha (<40%).',
 'ALCANCE:',
 '· Municipios: todos los que tienen referencia de 2ª vuelta 2022 (1.090). Barrios: todos los puestos con barrio georreferenciado y referencia 2022 (5.083 barrios; el resto del país es rural/vereda, sin barrio).',
 '· Barrios fusionados por (departamento, municipio, comuna, barrio) para no duplicar por diferencias de tilde/mayúscula.',
 'OJO: esta versión es para DIAGNÓSTICO/transparencia (ver todo el país). Para PRIORIZAR dónde movilizar, usar el Excel operativo (solo lo positivo, ordenado por objetivo).']:
    ws.append([ln]); c=ws.cell(row=ws.max_row,column=1); c.alignment=Alignment(wrap_text=True,vertical='top'); c.font=Font(size=14,bold=ln.endswith(':'))
ws.column_dimensions['A'].width=120

# Por municipio (nacional)
ws=wb.create_sheet('Por municipio (nacional)')
hdr(ws,['Departamento','Municipio','Censo','Votantes','Abstención','Abst. %','Share Petro 2V','Share Cepeda 1V','Neto Cepeda (firmado)','Categoría 2V'])
for r in muni_rows: ws.append(r)
fin(ws,(5,6,7)); cf_share(ws,'G'); cf_neto(ws,'I')

# Por barrio (nacional)
ws=wb.create_sheet('Por barrio (nacional)')
hdr(ws,['Departamento','Municipio','Comuna/Localidad','Barrio','Censo','Votantes','Abstención','Abst. %','Share Petro 2V','Neto Cepeda (firmado)','Categoría 2V'])
for r in bar_rows: ws.append(r)
fin(ws,(7,8)); cf_share(ws,'I'); cf_neto(ws,'J')

fn=f'{OUT}/Estrategia_3_Abstencion_NACIONAL_v2.xlsx'
wb.save(fn)
print('✓',fn)
print(f'  municipios: {len(muni_rows)} (positivos {sum(1 for r in muni_rows if r[8]>0)} · negativos {sum(1 for r in muni_rows if r[8]<0)})')
print(f'  barrios: {len(bar_rows)} (positivos {sum(1 for r in bar_rows if r[9]>0)} · negativos {sum(1 for r in bar_rows if r[9]<0)})')
