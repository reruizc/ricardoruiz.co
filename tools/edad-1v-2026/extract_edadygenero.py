#!/usr/bin/env python3
"""Extrae de Edadygenero.xlsx (645k filas, mesa-level) las elecciones
presidenciales a CSVs de caché para no volver a abrir el xlsx de 135 MB.

Salida: Bases de datos/output_edad_1v/cache/{p1v-2018,p1v-2022,p2v-2022}.csv
Mismas 47 columnas del original, una fila por mesa.
"""
import csv
import os
import sys
import time

import openpyxl

BASE = os.path.join(os.path.dirname(__file__), "..", "..", "Bases de datos")
SRC = os.path.join(BASE, "Edadygenero.xlsx")
OUT = os.path.join(BASE, "output_edad_1v", "cache")

# (año, fragmento del tipo de elección) -> nombre de archivo
TARGETS = {
    (2018, "Presidencia 1V"): "p1v-2018.csv",
    (2022, "Presidencia 1V"): "p1v-2022.csv",
    (2022, "Presidencia 2V"): "p2v-2022.csv",
}


def main():
    t0 = time.time()
    os.makedirs(OUT, exist_ok=True)
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    writers, files = {}, {}
    header = None
    counts = {k: 0 for k in TARGETS}
    n = 0
    for row in ws.iter_rows(values_only=True):
        n += 1
        if header is None:
            header = list(row)
            i_tipo = header.index("Datos de tipo de elección")
            i_anio = header.index("Año")
            for key, fname in TARGETS.items():
                f = open(os.path.join(OUT, fname), "w", newline="", encoding="utf-8")
                w = csv.writer(f)
                w.writerow(header)
                writers[key], files[key] = w, f
            continue
        tipo = str(row[i_tipo] or "")
        if "Presidencia" not in tipo:
            continue
        anio_val = row[i_anio]
        anio = anio_val.year if hasattr(anio_val, "year") else int(str(anio_val)[:4])
        vuelta = "1V" if ("1" in tipo) else ("2V" if "2" in tipo else "?")
        key = (anio, f"Presidencia {vuelta}")
        if key in writers:
            writers[key].writerow(row)
            counts[key] += 1
        if n % 100000 == 0:
            print(f"  ... {n} filas leídas ({time.time()-t0:.0f}s)", flush=True)

    for f in files.values():
        f.close()
    wb.close()
    print(f"Total filas: {n} en {time.time()-t0:.0f}s")
    for key, c in counts.items():
        print(f"  {key}: {c} mesas -> {TARGETS[key]}")


if __name__ == "__main__":
    sys.exit(main())
