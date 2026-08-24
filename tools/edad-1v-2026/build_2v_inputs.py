#!/usr/bin/env python3
"""Insumos para el análisis etario de 2ª vuelta (Cepeda vs Abelardo 2026,
Petro vs Rodolfo 2022).

Produce en Bases de datos/output_edad_1v/:
  edad-2v2022-puesto.csv   composición etaria OBSERVADA de votantes 2V-2022
                           (de Edadygenero.xlsx, mesa->puesto, fix Bogotá)
  votos-2v2022-puesto.csv  Petro vs Rodolfo 2V por puesto (GCS escrutinio)
  votos-2v2026-puesto.csv  Cepeda vs Abelardo 2V por puesto (master 2V)
"""
import csv
import json
import os
import sys
import time
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from probe_viabilidad import BANDS, BOG_LOC, nrm, zf  # noqa: E402

BASE = os.path.join(os.path.dirname(__file__), "..", "..", "Bases de datos")
OUT = os.path.join(BASE, "output_edad_1v")


# ------------------------------------------------------- edad 2V-2022 -> puesto
def edad_2v2022():
    import openpyxl
    src = os.path.join(BASE, "Edadygenero.xlsx")
    t0 = time.time()
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ix = {h: i for i, h in enumerate(header)}
    i_tipo = ix["Datos de tipo de elección"]
    i_anio = ix["Año"]
    i_dep = ix["Cód. Depto"]
    i_mun = ix["Cód. Municipio"]
    i_loc = ix["Cód. Comuna / Localidad"]
    i_pue = ix["Cód. Puesto de Votación"]
    i_dnom = ix["Estadosnoborrar"]
    i_suf = ix["Cantidad de Sufragantes"]
    band_ix = [ix[b] for b in BANDS]

    acc = defaultdict(lambda: [0] * (len(BANDS) + 1))  # bandas + sufragantes
    dname = {}
    n = kept = 0
    for row in rows:
        n += 1
        tipo = str(row[i_tipo] or "")
        if "Presidencia" not in tipo or "2" not in tipo:
            continue
        av = row[i_anio]
        anio = av.year if hasattr(av, "year") else int(str(av)[:4])
        if anio != 2022:
            continue
        kept += 1
        z = str(row[i_loc]).strip()
        zz = zf(z, 2) if z.isdigit() else BOG_LOC.get(nrm(z), z)
        pcode = f"{zf(row[i_dep],2)}-{zf(row[i_mun],3)}-{zz}-{zf(row[i_pue],2)}"
        a = acc[pcode]
        for j, bi in enumerate(band_ix):
            a[j] += int(row[bi] or 0)
        a[-1] += int(row[i_suf] or 0)
        dname.setdefault(pcode, row[i_dnom])
        if n % 100000 == 0:
            print(f"  ...{n:,} filas ({time.time()-t0:.0f}s)", flush=True)
    wb.close()
    with open(os.path.join(OUT, "edad-2v2022-puesto.csv"), "w", newline="",
              encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["pcode", "depname"] + BANDS + ["Cantidad de Sufragantes"])
        for k in sorted(acc):
            a = acc[k]
            w.writerow([k, dname[k]] + a[:-1] + [a[-1]])
    suf = sum(acc[k][-1] for k in acc)
    print(f"edad-2v2022-puesto.csv: {len(acc):,} puestos · {kept:,} mesas · "
          f"{suf:,} sufragantes ({time.time()-t0:.0f}s)")


# ------------------------------------------------------- votos 2V-2022 -> puesto
def votos_2v2022():
    acc = defaultdict(lambda: defaultdict(int))
    src = os.path.join(BASE, "FINAL SUBIDA GCS", "GCS_2022PRES2V.csv")
    with open(src, encoding="utf-8-sig") as f:
        r = csv.DictReader(f, delimiter=";")
        for row in r:
            dep = zf(row["COD_DDE"], 2)
            key = (f"{dep}-{zf(row['COD_MME'],3)}-"
                   f"{zf(row['COD_ZZ'],2)}-{zf(row['COD_PP'],2)}")
            v = int(row["NUM_VOT"] or 0)
            cod = str(row["COD_CAN"]).strip()
            a = acc[key]
            if cod == "2":
                a["petro"] += v
            elif cod == "1":
                a["rodolfo"] += v
            elif cod == "996":
                a["blanco"] += v
            elif cod == "997":
                a["nulos"] += v
            else:
                a["no_marcados"] += v
    with open(os.path.join(OUT, "votos-2v2022-puesto.csv"), "w", newline="",
              encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["pcode", "petro", "rodolfo", "blanco", "nulos",
                    "no_marcados", "total_dosc", "total_votos"])
        tot = defaultdict(int)
        for k in sorted(acc):
            a = acc[k]
            dosc = a["petro"] + a["rodolfo"]
            tt = dosc + a["blanco"] + a["nulos"] + a["no_marcados"]
            w.writerow([k, a["petro"], a["rodolfo"], a["blanco"], a["nulos"],
                        a["no_marcados"], dosc, tt])
            for kk, vv in a.items():
                tot[kk] += vv
    print(f"votos-2v2022-puesto.csv: {len(acc):,} puestos · "
          f"Petro {tot['petro']:,} · Rodolfo {tot['rodolfo']:,}")


# ------------------------------------------------------- votos 2V-2026 -> puesto
def votos_2v2026():
    m = json.load(open(os.path.join(OUT, "..", "output_2v",
                                    "master_unificado_puesto.json")))
    rows = []
    tc = ta = 0
    for r in m:
        pc = r["pcode"]  # 9 dígitos dep+mun+zona+puesto
        dep, mun, zz, pp = pc[:2], pc[2:5], pc[5:7], pc[7:9]
        key = f"{dep}-{mun}-{zz}-{pp}"
        cep, abe = int(r.get("cep2", 0)), int(r.get("abe2", 0))
        rows.append((key, cep, abe, cep + abe, int(r.get("urna2", 0))))
        tc += cep
        ta += abe
    with open(os.path.join(OUT, "votos-2v2026-puesto.csv"), "w", newline="",
              encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["pcode", "cepeda", "abelardo", "total_dosc", "total_votos"])
        for row in sorted(rows):
            w.writerow(row)
    print(f"votos-2v2026-puesto.csv: {len(rows):,} puestos · "
          f"Cepeda {tc:,} · Abelardo {ta:,}")


if __name__ == "__main__":
    votos_2v2026()
    votos_2v2022()
    edad_2v2022()  # el lento (lee el xlsx de 135 MB)
