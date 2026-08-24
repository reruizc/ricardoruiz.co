#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Mesas del preconteo 2V donde Abelardo De La Espriella obtuvo el 100% de la
votacion valida (Cepeda = 0 y Abelardo > 0), con:
  - datos del preconteo de 2V por mesa,
  - % que saco Abelardo en 1V en esa misma mesa,
  - participacion estimada en 1V y en 2V.

Definicion de "100%": de la votacion valida del balotaje (Cepeda + Abelardo),
el 100% fue para Abelardo, es decir Cepeda = 0 votos y Abelardo > 0 votos.

Fuentes:
  2V por mesa : Bases de datos/output_2v/detalle_nacional_presidencia_mesas.xlsx
                (preconteo; columnas Pre_*; codigos dep/mun/zona/puesto/mesa
                 extraidos del hipervinculo E14)
  1V por mesa : Bases de datos/nuevos archivos 1v 2026/PRECONTEO_1V_2026_MESA_con_Claudia.csv
  Censo       : Bases de datos/COMUNAS_DATA.csv  (censo y # de mesas por puesto)
  Nombres     : Bases de datos/test-presidencial/divipola.json  (fallback dep/mun)

Salida:
  Bases de datos/output_2v/Mesas_Abelardo_100pct_2V_vs_1V.xlsx
"""
import csv, re, json, unicodedata, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = os.path.join(os.path.dirname(__file__), '..', '..', 'Bases de datos')
BASE = os.path.normpath(BASE)
XLSX_2V = os.path.join(BASE, 'output_2v', 'detalle_nacional_presidencia_mesas.xlsx')
CSV_1V  = os.path.join(BASE, 'nuevos archivos 1v 2026', 'PRECONTEO_1V_2026_MESA_con_Claudia.csv')
COMUNAS = os.path.join(BASE, 'COMUNAS_DATA.csv')
DIVIPOLA= os.path.join(BASE, 'test-presidencial', 'divipola.json')
OUT     = os.path.join(BASE, 'output_2v', 'Mesas_Abelardo_100pct_2V_vs_1V.xlsx')

def norm(s):
    s = unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode().upper().strip()
    return re.sub(r'\s+', ' ', s)

# ---------- divipola: nombre -> codigo (fallback cuando falta el hipervinculo) ----------
dv = json.load(open(DIVIPOLA, encoding='utf-8'))
dep_by_name, mun_by = {}, {}
for d in dv['deptos']:
    dep_by_name[norm(d['nombre'])] = d['cod']
    for m in d['muns']:
        mun_by[(d['cod'], norm(m['nombre']))] = m['cod']

# ---------- censo por puesto ----------
# clave (dep,mun,zona,puesto) -> (censo_total, n_mesas, dep_nombre, mun_nombre)
censo = {}
with open(COMUNAS, encoding='utf-8-sig') as f:
    for row in csv.DictReader(f, delimiter=';'):
        try:
            k = (int(row['dd']), int(row['mm']), int(row['zz']), int(row['pp']))
        except (ValueError, KeyError):
            continue
        censo[k] = (int(row['total']), int(row['mesas']), norm(row['departamento']), norm(row['municipio']))

# ---------- 1V por mesa ----------
# clave (dep,mun,zona,puesto,mesa) -> (abe_1v, validos_1v, urna_1v)
CAND = ['Iván Cepeda','Santiago Botero','Abelardo De La Espriella','Mauricio Lizcano',
        'Miguel Uribe','Sondra Macollins','Roy Barreras','Carlos Caicedo','Gustavo Matamoros',
        'Paloma Valencia','Sergio Fajardo','Gilberto Murillo','Claudia López']
oneV = {}
with open(CSV_1V, encoding='utf-8-sig') as f:
    r = csv.reader(f); hdr = next(r); idx = {h: i for i, h in enumerate(hdr)}
    ia, ic, it = idx['Abelardo De La Espriella'], idx['Iván Cepeda'], idx['total_votos_urna']
    cand_idx = [idx[c] for c in CAND]
    for row in r:
        try:
            key = (int(row[0]), int(row[1]), int(row[2]), int(row[3]), int(row[4]))
        except ValueError:
            continue
        val = sum(int(row[j]) for j in cand_idx)
        oneV[key] = (int(row[ic]), int(row[ia]), val, int(row[it]))   # cep1, abe1, validos1, urna1

# ---------- 2V por mesa: detectar mesas Abelardo-100 ----------
rx = re.compile(r'/pdf/(\d+)/(\d+)/(\d+)/(\d+)/(\d+)/')
wb = openpyxl.load_workbook(XLSX_2V, read_only=True)
ws = wb['Presidencia_Mesas']
it = ws.iter_rows(values_only=True); next(it)

recs = []
warn_censo_mismatch = 0
for row in it:
    cep = row[10] or 0
    abe = row[11] or 0
    if not (cep == 0 and abe > 0):          # 100% valido de Abelardo
        continue
    dep_n, mun_n, zona_x, pto_x, pto_name, mesa_x = row[1], row[2], row[3], row[4], row[5], row[6]
    blanco, nomarc, nulos = row[7] or 0, row[8] or 0, row[9] or 0
    m = rx.search(str(row[24] or ''))
    if m:
        dep, mun, zona, pto, mesa = (int(g) for g in m.groups())
        src = 'E14'
    else:                                    # fallback por nombre + columnas del xlsx
        dc = dep_by_name.get(norm(dep_n)); mc = mun_by.get((dc, norm(mun_n))) if dc else None
        if not (dc and mc):
            print('  [FAIL] sin codigo:', dep_n, mun_n, zona_x, pto_x, mesa_x); continue
        dep, mun, zona, pto, mesa = int(dc), int(mc), int(zona_x), int(pto_x), int(mesa_x)
        src = 'nombre'

    key = (dep, mun, zona, pto, mesa)
    one = oneV.get(key)                       # (abe1v, val1v, urna1v) o None
    cz = censo.get((dep, mun, zona, pto))     # (censo, mesas, depn, munn) o None
    # guardia anti-colision: el nombre del censo debe coincidir con el del 2V
    if cz and (cz[2] != norm(dep_n) or cz[3] != norm(mun_n)):
        warn_censo_mismatch += 1
        cz = None

    tot2v = blanco + nomarc + nulos + cep + abe
    cep1v, abe1v, val1v, urna1v = (one if one else (None, None, None, None))

    censo_pto = cz[0] if cz else None
    n_mesas   = cz[1] if cz else None
    censo_mesa = (censo_pto / n_mesas) if (censo_pto and n_mesas) else None

    nota = ''
    if dep == 88: nota = 'Exterior / consulado'
    elif zona == 90: nota = 'Puesto censo (90)'
    elif zona == 98: nota = 'Carcel (98)'
    elif zona == 99: nota = 'Zona rural (99)'

    # Alerta de posible error de preconteo: si Cepeda tenía votación real en 1V
    # en esta misma mesa, un 0 absoluto en 2V es sospechoso (transcripción E14).
    alerta = (f'Revisar: Cepeda obtuvo {cep1v} votos en esta mesa en 1V'
              if (cep1v is not None and cep1v >= 10) else '')

    recs.append(dict(
        dep_n=dep_n, mun_n=mun_n, dep=dep, mun=mun, zona=zona, pto=pto,
        pto_name=pto_name, mesa=mesa,
        llave=f'{dep:02d}-{mun:03d}-{zona:02d}-{pto:02d}-{mesa:03d}',
        cep2=cep, abe2=abe, bl2=blanco, nu2=nulos, nm2=nomarc, tot2=tot2v,
        abe2_val=abe/(cep+abe), abe2_urna=(abe/tot2v if tot2v else None),
        cep1=cep1v, abe1=abe1v, val1=val1v, urna1=urna1v,
        abe1_val=(abe1v/val1v if val1v else None),
        abe1_urna=(abe1v/urna1v if urna1v else None),
        censo_pto=censo_pto, n_mesas=n_mesas, censo_mesa=censo_mesa,
        part1=(urna1v/censo_mesa if (censo_mesa and urna1v is not None) else None),
        part2=(tot2v/censo_mesa if censo_mesa else None),
        nota=nota, alerta=alerta, src=src,
    ))

recs.sort(key=lambda d: (-d['abe2'], d['dep_n'], d['mun_n']))
print(f'Mesas Abelardo-100: {len(recs)}  |  con 1V: {sum(1 for r in recs if r["abe1"] is not None)}'
      f'  |  con censo: {sum(1 for r in recs if r["censo_pto"])}  |  colisiones censo descartadas: {warn_censo_mismatch}')

# ---------- escribir xlsx ----------
wbo = openpyxl.Workbook()
ws = wbo.active; ws.title = 'Mesas Abelardo 100%'

COLS = [
    ('Departamento',        'dep_n',     22, None),
    ('Municipio',           'mun_n',     20, None),
    ('Cód. dep',            'dep',        7, '00'),
    ('Cód. mun',            'mun',        7, '000'),
    ('Zona',                'zona',       6, '00'),
    ('Cód. puesto',         'pto',        8, '00'),
    ('Puesto',              'pto_name',  26, None),
    ('Mesa',                'mesa',       6, '0'),
    ('Llave',               'llave',     16, None),
    ('2V Cepeda',           'cep2',       9, '#,##0'),
    ('2V Abelardo',         'abe2',      10, '#,##0'),
    ('2V Blanco',           'bl2',        8, '#,##0'),
    ('2V Nulos',            'nu2',        8, '#,##0'),
    ('2V No marc.',         'nm2',        9, '#,##0'),
    ('2V Votantes (urna)',  'tot2',      13, '#,##0'),
    ('2V Abelardo % válid.','abe2_val',  12, '0.0%'),
    ('2V Abelardo % urna',  'abe2_urna', 12, '0.0%'),
    ('1V Abelardo votos',   'abe1',      11, '#,##0'),
    ('1V Cepeda votos',     'cep1',      11, '#,##0'),
    ('1V Válidos',          'val1',       9, '#,##0'),
    ('1V Votantes (urna)',  'urna1',     13, '#,##0'),
    ('1V Abelardo % válid.','abe1_val',  12, '0.0%'),
    ('1V Abelardo % urna',  'abe1_urna', 12, '0.0%'),
    ('Censo puesto',        'censo_pto', 11, '#,##0'),
    ('Mesas en puesto',     'n_mesas',   10, '#,##0'),
    ('Censo mesa (est.)',   'censo_mesa',12, '#,##0'),
    ('Participación 1V (est.)','part1',  14, '0.0%'),
    ('Participación 2V (est.)','part2',  14, '0.0%'),
    ('Nota',                'nota',      20, None),
    ('Alerta',              'alerta',    34, None),
]

hdr_fill = PatternFill('solid', fgColor='8A1E16')   # oxblood del proyecto
hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
thin = Side(style='thin', color='D8D2C4')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, (title, _, _, _) in enumerate(COLS, 1):
    cell = ws.cell(1, c, title)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

for ri, rec in enumerate(recs, 2):
    for c, (_, field, _, fmt) in enumerate(COLS, 1):
        v = rec[field]
        cell = ws.cell(ri, c, v)
        if fmt: cell.number_format = fmt
        cell.border = border
        if ri % 2 == 0:
            cell.fill = PatternFill('solid', fgColor='F4F0E7')   # zebra crema

for c, (_, _, w, _) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = 'C2'
ws.row_dimensions[1].height = 30
last = f'{get_column_letter(len(COLS))}{len(recs)+1}'
tbl = Table(displayName='MesasAbelardo100', ref=f'A1:{last}')
tbl.tableStyleInfo = TableStyleInfo(name='TableStyleLight1', showRowStripes=False)
# (estilo manual ya aplicado; el Table aporta el autofiltro)
ws.add_table(tbl)

# ---------- hoja de metodologia ----------
wm = wbo.create_sheet('Metodología')
meth = [
    ('Mesas donde Abelardo obtuvo el 100% de la votación válida — Balotaje 2026', True),
    ('', False),
    ('Criterio de selección', True),
    ('Mesa del preconteo de 2ª vuelta en la que Iván Cepeda obtuvo 0 votos y Abelardo De La', False),
    ('Espriella obtuvo más de 0. Es decir, de la votación válida (Cepeda + Abelardo) el 100% fue', False),
    ('para Abelardo. Los votos en blanco, nulos y no marcados NO cuentan como votación a un', False),
    ('candidato; por eso "% válid." puede ser 100% aunque "% urna" sea menor.', False),
    ('', False),
    (f'Total de mesas que cumplen el criterio: {len(recs)}.', False),
    ('', False),
    ('Fuentes', True),
    ('• 2V por mesa: detalle_nacional_presidencia_mesas.xlsx (preconteo Registraduría, columnas Pre_*).', False),
    ('  Solo Cepeda y Abelardo vienen por mesa. Los códigos dep/mun/zona/puesto/mesa se extraen', False),
    ('  del hipervínculo al E14; cuando falta, se resuelven por nombre (divipola).', False),
    ('• 1V por mesa: PRECONTEO_1V_2026_MESA_con_Claudia.csv (preconteo Registraduría 1ª vuelta).', False),
    ('• Censo y nº de mesas por puesto: COMUNAS_DATA.csv (Divipole). Misma codificación que el', False),
    ('  preconteo (Nariño=23, Risaralda=24, Consulados=88), verificada antes del cruce.', False),
    ('', False),
    ('Definición de las columnas de %', True),
    ('• 2V/1V Abelardo % válid. = Abelardo / votos a candidatos (en 2V, Cepeda + Abelardo).', False),
    ('• 2V/1V Abelardo % urna  = Abelardo / total depositado en la urna (incluye blanco/nulo/no marcado).', False),
    ('', False),
    ('Participación (importante)', True),
    ('El censo electoral oficial está disponible por PUESTO, no por mesa. La participación por mesa', False),
    ('se ESTIMA dividiendo el censo del puesto entre su número de mesas:', False),
    ('   Censo mesa (est.) = Censo puesto / Mesas en puesto', False),
    ('   Participación = Votantes de la mesa / Censo mesa (est.)', False),
    ('En puestos de una sola mesa el valor es exacto; en puestos con varias mesas es aproximado', False),
    ('(la última mesa suele tener menos inscritos). Se dejan visibles el censo del puesto, el nº de', False),
    ('mesas y los votantes de cada mesa para que el cálculo sea auditable.', False),
    ('Mesas del exterior (consulados) y puestos sin censo en COMUNAS_DATA quedan con participación vacía.', False),
    ('', False),
    ('', False),
    ('Columna "Alerta" — posibles errores de preconteo', True),
    ('Una mesa con 100% de Abelardo debería ser una mesa diminuta donde casi nadie votó por Cepeda.', False),
    ('Si Cepeda obtuvo votación real en 1V en esa misma mesa (>= 10 votos) pero figura con 0 en 2V,', False),
    ('lo más probable es un error de transcripción del E14 en el preconteo (la cifra de Cepeda no se', False),
    ('cargó). Esas mesas quedan marcadas en "Alerta" y deberían corregirse con el escrutinio oficial.', False),
    ('Ejemplos: Armenia (Quindío) mesa 2 — 187 votantes en 2V con Cepeda en 0, pero 81 votos a Cepeda', False),
    ('en 1V; las otras 16 mesas del mismo puesto le dan a Cepeda entre 88 y 144 votos.', False),
    ('', False),
    ('Nota: preconteo, no escrutinio oficial. La columna "Nota" marca exterior y zonas especiales (90/98/99).', False),
]
for ri, (txt, bold) in enumerate(meth, 1):
    cell = wm.cell(ri, 1, txt)
    cell.font = Font(bold=bold, size=12 if (bold and ri == 1) else 10)
wm.column_dimensions['A'].width = 105

wbo.save(OUT)
print('OK ->', OUT)

# resumen en consola
ext = sum(1 for r in recs if r['dep'] == 88)
print(f'  de las cuales exterior/consulado: {ext}  |  territorio nacional: {len(recs)-ext}')
print('  Top 5 por votos de Abelardo en 2V:')
for r in recs[:5]:
    p1 = f'{r["abe1_val"]*100:.0f}%' if r['abe1_val'] is not None else 's/d'
    print(f'    {r["dep_n"][:18]:18} {r["mun_n"][:16]:16} mesa {r["mesa"]:>3}  abe2V={r["abe2"]:>3}  abe1V%={p1}')
