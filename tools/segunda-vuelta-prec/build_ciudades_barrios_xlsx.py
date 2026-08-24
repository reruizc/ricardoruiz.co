#!/usr/bin/env python3
"""
Resultados 2da vuelta presidencial 2026 (Cepeda vs Abelardo) por BARRIO en las
17 ciudades principales -> UN solo Excel, una hoja por ciudad.

FUENTES (mezcla, para no subcontar):
  - Votos 2V por mesa : output_2v/detalle_nacional_presidencia_mesas.xlsx  (autoritativo,
                        incluye corregimientos con codigo de puesto alfanumerico)
  - Barrio / comuna   : master_unificado_puesto.json  (resuelve barrio/comuna incl. BOGOTA,
                        que NO esta en PUESTOS_GEOREF.csv) por pcode 9 digitos
  - Fallback barrio   : PUESTOS_GEOREF.csv  (para los puestos alfanumericos que el master suelta)

Join puesto: pcode = cod5(5) + zona(2) + puesto(2)  ej '010010704' o '0100199A1'.
La 2V solo trae Cepeda y Abelardo; blanco/nulos/no-marcados van aparte por mesa.
"""
import json, os, csv, re, unicodedata
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(__file__), '..', '..')
BD   = os.path.join(ROOT, 'Bases de datos')
OUT  = os.path.join(BD, 'output_2v')
XLSX_2V = os.path.join(OUT, 'detalle_nacional_presidencia_mesas.xlsx')
MASTER  = os.path.join(OUT, 'master_unificado_puesto.json')
GEOREF  = os.path.join(BD, 'PUESTOS_GEOREF.csv')
OUTFILE = os.path.join(OUT, 'Resultados_2V_2026_por_barrio_ciudades.xlsx')

# cod5 electoral -> ciudad (mismo dict que build_master_2v); orden por votacion 2V
CITIES = [
    ('16001','Bogotá'),       ('01001','Medellín'),    ('31001','Cali'),
    ('03001','Barranquilla'), ('05001','Cartagena'),   ('25001','Cúcuta'),
    ('27001','Bucaramanga'),  ('29001','Ibagué'),      ('24001','Pereira'),
    ('52001','Villavicencio'),('15247','Soacha'),      ('21001','Santa Marta'),
    ('13001','Montería'),     ('23001','Pasto'),       ('09001','Manizales'),
    ('03052','Soledad'),      ('11001','Popayán'),
]
CITY = dict(CITIES)

# ---------------- normalizacion ----------------
def nrm(s):  # upper + sin tildes + colapsa espacios
    s = ''.join(c for c in unicodedata.normalize('NFD', str(s or '').upper().strip())
                if unicodedata.category(c) != 'Mn')
    return ' '.join(s.split())
def normm(s):  # nrm + quita ( ) .  (para mapear nombre depto/mun -> cod5)
    return ' '.join(nrm(s).replace('(', ' ').replace(')', ' ').replace('.', ' ').split())

SMALL = {'DE','DEL','LA','LAS','LOS','EL','Y','EN','A','PARA','POR','CON'}
ROMAN = {'I','II','III','IV','V','VI','VII','VIII','IX','X'}
def titlecase(s):
    s = ' '.join(str(s or '').split())
    out = []
    for i, w in enumerate(s.split(' ')):
        u = nrm(w)
        if i > 0 and u in SMALL: out.append(w.lower())
        elif u in ROMAN:         out.append(u)
        else:                    out.append(w.lower().capitalize())
    return ' '.join(out)

def clean_comuna(raw):
    """quita el prefijo de codigo ('04COMUNA 4 ARANJUEZ' -> code 4, 'Comuna 4 Aranjuez').
       Bogota no trae prefijo numerico -> code alto, nombre tal cual."""
    s = ' '.join(str(raw or '').split())
    if not s: return (999, 'Sin comuna asignada')
    m = re.match(r'^(\d{2,3})\s*(.+)$', s)
    code, name = (int(m.group(1)), m.group(2)) if m else (900, s)
    if nrm(name) in ('NULL', 'NAN', 'NONE', 'SN', ''): return (999, 'Sin comuna asignada')
    return (code, titlecase(name))

def pcode(cod5, zona, puesto):
    return f"{cod5}{str(zona).strip().zfill(2)}{str(puesto).strip().zfill(2)}"

# ---------------- modelo de voto por sexo (estimate_genero_2v.py) ----------------
BIG4N = {'Bogotá', 'Medellín', 'Cali', 'Barranquilla'}   # gap por comuna/localidad
MASTER_2022 = os.path.join(BD, 'output_pacto_1v_2026', 'master_2022_puesto.json')
GMODEL = json.load(open(os.path.join(OUT, 'genero-modelo-2v.json')))
GAPS = GMODEL['gaps']; WF = GMODEL['wfrac_puesto']
GAPS22 = GMODEL['gaps22']; WF22 = GMODEL['wfrac22_puesto']
FNAC = float(GMODEL['meta']['fmuj_nacional']); FNAC22 = float(GMODEL['meta']['fmuj_nacional_2022'])
GAP_NAC = GAPS['__national__']['gap']; GAP_NAC22 = GAPS22['__national__']['gap']
def gap_in(GG, ciudad, cdisp):
    if ciudad in BIG4N:
        k = f'{ciudad}|{nrm(cdisp)}'
        if k in GG: return GG[k]['gap']
    k2 = f'{ciudad}|*'
    return GG[k2]['gap'] if k2 in GG else GG['__national__']['gap']
def gap_for(ciudad, cdisp):   return gap_in(GAPS, ciudad, cdisp)
def gap_for22(ciudad, cdisp): return gap_in(GAPS22, ciudad, cdisp)
def decompose(cep, abe, fmuj, gap):
    """-> (fmuj, muj->Cep, muj->Abe, hom->Cep, hom->Abe). gap = brecha mujer-hombre en cuota Cepeda.
       Reconcilia: fmuj*muj_Cep + (1-fmuj)*hom_Cep = Cepeda real del barrio."""
    val = cep + abe
    if val <= 0: return (fmuj, 0, 0, 0, 0)
    C = cep / val
    mc = min(1.0, max(0.0, C + (1 - fmuj) * gap))
    hc = min(1.0, max(0.0, C - fmuj * gap))
    return (fmuj, mc, 1 - mc, hc, 1 - hc)

# ---------------- 1) master: pcode -> barrio/comuna/censo ----------------
master = json.load(open(MASTER))
M = {}
for p in master:
    M[p['pcode']] = {'barrio': (p.get('barrio') or '').strip(),
                     'comuna': (p.get('comuna') or '').strip(),
                     'censo':  int(p.get('pot') or 0),
                     'lat': p.get('lat'), 'lon': p.get('lon')}

# ---------------- 2) GEOREF: name->cod5  +  pcode->barrio/comuna/censo (fallback) ----------------
name2cod5 = {}
G = {}
with open(GEOREF, encoding='utf-8-sig') as f:
    for r in csv.DictReader(f, delimiter=';'):
        cc = (r['CÓDIGO COMPLETO'] or '').strip()
        if len(cc) < 9: continue
        name2cod5.setdefault((normm(r['DEPARTAMENTO']), normm(r['MUNICIPIO'])), cc[:5])
        try: censo = int(r['MUJERES'] or 0) + int(r['HOMBRES'] or 0)
        except: censo = 0
        G[cc] = {'barrio': (r['BARRIO'] or '').strip(),
                 'comuna': (r.get('NOMBRE COMUNA') or '').strip(), 'censo': censo,
                 'lat': r.get('LATITUD'), 'lon': r.get('LONGITUD')}

def cod5_of(dep, mun):
    nd, nm = normm(dep), normm(mun)
    if nd.startswith('BOGOTA') or nm.startswith('BOGOTA'): return '16001'
    return name2cod5.get((nd, nm))

# ---------------- 2b) reasignacion por lat/lon: barrios multi/junk y comunas faltantes ----------------
# Un puesto cuyo BARRIO concatena varios ("Dindalito, Ciudad de Cali, ...") o es basura ("N/A"),
# o cuya COMUNA llega "NULL", se reasigna por coordenadas al puesto LIMPIO mas cercano (regla del
# proyecto: coordenadas > nombre). Barrio malo -> barrio+comuna del vecino; solo comuna mala -> se
# conserva el barrio bueno y se toma la comuna del vecino con comuna valida.
def fnum(x):
    try: return float(x)
    except: return None
BAD_B = {'N/A','NA','SN','S/N','NULL','NAN','NONE','NO REGISTRA','SIN BARRIO','SIN BARRIO ASIGNADO','SIN DATO','-',''}
BAD_C = {'NULL','NAN','NONE','SN','SIN COMUNA','SIN COMUNA ASIGNADA','-',''}
def bad_barrio(b):
    bn = nrm(b)
    return (bn in BAD_B) or (',' in b) or bool(re.search(r'\s[-/]\s', b)) or bn.startswith('ENTRE ')
def bad_comuna(c):
    return nrm(c) in BAD_C

CITYP = defaultdict(list)   # cod5 -> [(code, lat, lon, comuna_raw, barrio_raw)]
for code in set(M) | set(G):
    if code[:5] not in CITY: continue
    g = M.get(code) or G.get(code)
    CITYP[code[:5]].append((code, fnum(g.get('lat')), fnum(g.get('lon')),
                            (g.get('comuna') or ''), (g.get('barrio') or '')))
def _nearest(la, lo, pool):
    best, bd = None, 1e18
    for xla, xlo, xc in pool:
        d = (xla-la)**2 + (xlo-lo)**2
        if d < bd: bd, best = d, xc
    return best
RESOLVE = {}; reasign = {'barrio': 0, 'comuna': 0}
for c5, plist in CITYP.items():
    rawmap = {code: (c, b) for code, la, lo, c, b in plist}
    cleanB = [(la,lo,code) for code,la,lo,c,b in plist if la is not None and not bad_barrio(b) and not bad_comuna(c)]
    cleanC = [(la,lo,code) for code,la,lo,c,b in plist if la is not None and not bad_comuna(c)]
    for code, la, lo, c, b in plist:
        bb, bc = bad_barrio(b), bad_comuna(c)
        if not bb and not bc:
            cc, cd = clean_comuna(c); RESOLVE[code] = (cc, cd, b.strip())
        elif bb:
            tgt = _nearest(la, lo, cleanB) if (la is not None and cleanB) else None
            if tgt:
                tc, tb = rawmap[tgt]; cc, cd = clean_comuna(tc)
                RESOLVE[code] = (cc, cd, tb.strip()); reasign['barrio'] += 1
            else:
                RESOLVE[code] = (999, 'Sin comuna asignada', 'Sin barrio asignado')
        else:
            tgt = _nearest(la, lo, cleanC) if (la is not None and cleanC) else None
            if tgt:
                tc, _ = rawmap[tgt]; cc, cd = clean_comuna(tc)
                RESOLVE[code] = (cc, cd, b.strip()); reasign['comuna'] += 1
            else:
                RESOLVE[code] = (900, 'Sin comuna asignada', b.strip())
print(f'reasignados por lat/lon: barrio {reasign["barrio"]} · comuna {reasign["comuna"]}')

# ---------------- 3) votos 2V por puesto (autoritativo, completo) ----------------
def ni(v):
    try: return int(v)
    except: return 0
V2 = defaultdict(lambda: {'cep':0,'abe':0,'bl':0,'nm':0,'nu':0,'mesas':0})
wb = load_workbook(XLSX_2V, read_only=True, data_only=True); ws = wb.worksheets[0]
nfilas = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    c5 = cod5_of(row[1], row[2])
    if c5 not in CITY: continue          # solo las 17 ciudades
    nfilas += 1
    code = pcode(c5, row[3], row[4])
    a = V2[code]
    a['bl'] += ni(row[7]); a['nm'] += ni(row[8]); a['nu'] += ni(row[9])
    a['cep'] += ni(row[10]); a['abe'] += ni(row[11]); a['mesas'] += 1
wb.close()
print(f'XLSX 2V: filas de las 17 ciudades = {nfilas:,}  | puestos con voto = {len(V2):,}')

# ---------------- 4) agrega por (ciudad, comuna, barrio) ----------------
def blank(): return {'cep':0,'abe':0,'bl':0,'nu':0,'nm':0,'mesas':0,'puestos':0,'censo':0,
                     'wmuj':0.0,'wtot':0.0,'fmuj':0,'mc':0,'ma':0,'hc':0,'ha':0,
                     'p22':None,'r22':None,'izq22':None,'mp22':None,'hp22':None,
                     'ccode':999,'cdisp':'','bdisp':''}
AGG = defaultdict(lambda: defaultdict(blank))   # ciudad -> key(comuna,barrio) -> agg
sin_geo = defaultdict(int)
CANON = {}   # (ciudad, nrm(comuna)) -> mejor display (prefiere la grafía con tildes)
def _dscore(s): return (1 if any(ord(c) > 127 for c in s) else 0, len(s))
for code, v in V2.items():
    c5 = code[:5]; ciudad = CITY[c5]
    g = M.get(code) or G.get(code)
    censo = g['censo'] if g else 0
    if code in RESOLVE:
        ccode, cdisp, barrio = RESOLVE[code]
    else:
        sin_geo[ciudad] += 1
        ccode, cdisp, barrio = 999, 'Sin comuna asignada', 'Sin barrio asignado'
    bdisp = titlecase(barrio)
    ck = nrm(cdisp)
    if (ciudad, ck) not in CANON or _dscore(cdisp) > _dscore(CANON[(ciudad, ck)]):
        CANON[(ciudad, ck)] = cdisp
    key = (ck, nrm(barrio))
    t = AGG[ciudad][key]
    t['cep'] += v['cep']; t['abe'] += v['abe']; t['bl'] += v['bl']; t['nu'] += v['nu']
    t['nm'] += v['nm']; t['mesas'] += v['mesas']; t['puestos'] += 1; t['censo'] += censo
    if ccode < t['ccode']: t['ccode'] = ccode
    t['cdisp'] = cdisp; t['bdisp'] = bdisp
    wfp = WF.get(code)
    if wfp: t['wmuj'] += wfp[0]; t['wtot'] += wfp[1]

# display de comuna canonico (misma localidad, una sola grafia)
for ciudad, rows in AGG.items():
    for a in rows.values():
        a['cdisp'] = CANON.get((ciudad, nrm(a['cdisp'])), a['cdisp'])

# ---------------- 5) comparativo: 2V 2022 (Petro vs Rodolfo) por barrio ----------------
def pc22(p):
    try: return f"{int(p['dep']):02d}{int(p['mun']):03d}{int(p['zona']):02d}{int(p['puesto']):02d}"
    except: return None
B22 = defaultdict(lambda: defaultdict(lambda: {'p22':0,'r22':0,'wmuj':0.0,'wtot':0.0}))
n22 = 0
for p in json.load(open(MASTER_2022)):
    code = pc22(p)
    if not code or code[:5] not in CITY or code not in RESOLVE: continue
    _, cdisp, barrio = RESOLVE[code]
    t = B22[CITY[code[:5]]][(nrm(cdisp), nrm(barrio))]
    t['p22'] += int(p.get('petro2v') or 0); t['r22'] += int(p.get('rodolfo2v') or 0); n22 += 1
    w = WF22.get(code)
    if w: t['wmuj'] += w[0]; t['wtot'] += w[1]
print(f'2022 2V: {n22:,} puestos mapeados a barrio en las 17 ciudades')

# ---------------- estilos (paleta del proyecto) ----------------
ROJO, AZUL = 'C0392B', '1F47CC'
HDR = PatternFill('solid', fgColor='8A1E16'); HDRF = Font(bold=True, color='FFFFFF', size=10)
ZEBRA = PatternFill('solid', fgColor='F4F0E7')
YELLOW = PatternFill('solid', fgColor='FFF1A6')   # "Sin dato disponible" (2022 faltante)
NODATA_F = Font(size=8, italic=True, color='7A6000')
CEPF = PatternFill('solid', fgColor='F7DAD5'); ABEF = PatternFill('solid', fgColor='D7DEF6')
WIN_C = PatternFill('solid', fgColor=ROJO); WIN_A = PatternFill('solid', fgColor=AZUL)
WINFT = Font(color='FFFFFF', bold=True)
thin = Side(style='thin', color='D9D2C5'); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment('center', vertical='center'); LEFT = Alignment('left', vertical='center')

COLS = [
    ('Barrio', 30, 'txt'), ('Comuna / Localidad', 26, 'txt'),
    ('Cepeda 2V', 11, 'int'), ('Abelardo 2V', 12, 'int'), ('Válidos 2V', 11, 'int'),
    ('Cepeda %', 9, 'pct'), ('Abelardo %', 10, 'pct'), ('Ganador', 11, 'txt'), ('Margen (pp)', 11, 'pp'),
    ('Blanco', 9, 'int'), ('Nulos', 8, 'int'), ('No marcad.', 10, 'int'), ('Total votos', 12, 'int'),
    ('Puestos', 8, 'int'), ('Mesas', 7, 'int'),
    ('% Mujeres', 9, 'pct'), ('Muj→Cepeda', 11, 'pct'), ('Muj→Abelardo', 12, 'pct'),
    ('Hom→Cepeda', 11, 'pct'), ('Hom→Abelardo', 12, 'pct'),
    ('Petro 2V 2022', 12, 'int'), ('Rodolfo 2V 2022', 14, 'int'), ('Izq 2022 %', 10, 'pct'),
    ('Swing izq 22→26 (pp)', 16, 'pp'), ('Muj→Petro 2022', 13, 'pct'), ('Hom→Petro 2022', 13, 'pct'),
]
GCOL = 8  # columna Ganador (1-based)
# genero 2026 (16-20): %Muj·MujCep·MujAbe·HomCep·HomAbe
# comparativo 2022 (21-26): Petro22·Rodolfo22·Izq2022%·Swing izq·Muj→Petro22·Hom→Petro22

def metricas(a):
    val = a['cep'] + a['abe']
    cp = a['cep']/val if val else 0; ap = a['abe']/val if val else 0
    gan = 'Cepeda' if a['cep'] >= a['abe'] else 'Abelardo'
    margen = abs(cp-ap)*100
    total = val + a['bl'] + a['nu'] + a['nm']
    swing = (cp - a['izq22']) * 100 if a['izq22'] is not None else None
    return [a['bdisp'], a['cdisp'], a['cep'], a['abe'], val, cp, ap, gan, margen,
            a['bl'], a['nu'], a['nm'], total, a['puestos'], a['mesas'],
            a['fmuj'], a['mc'], a['ma'], a['hc'], a['ha'],
            a['p22'], a['r22'], a['izq22'], swing, a['mp22'], a['hp22']]

def cellfmt(c, fmt):
    c.border = BORD; c.alignment = CEN
    if fmt == 'int': c.number_format = '#,##0'
    elif fmt == 'pct': c.number_format = '0.0%'
    elif fmt == 'pp': c.number_format = '+0.0;-0.0'

def pinta(ws, filas):
    for j, (t, w, _) in enumerate(COLS, 1):
        c = ws.cell(1, j, t); c.fill = HDR; c.font = HDRF; c.alignment = CEN; c.border = BORD
        ws.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for agg, total in filas:
        vals = metricas(agg)
        for j, (v, (t, w, fmt)) in enumerate(zip(vals, COLS), 1):
            if v is None and 21 <= j <= 26:        # 2022 sin dato -> amarillo
                c = ws.cell(r, j, 'Sin dato disponible')
                c.border = BORD; c.alignment = CEN; c.fill = YELLOW; c.font = NODATA_F
                continue
            c = ws.cell(r, j, v); cellfmt(c, fmt)
            if j <= 2: c.alignment = LEFT
            elif not total and r % 2 == 0: c.fill = ZEBRA
            if not total:
                if j in (3, 6, 17, 19, 21, 23, 25, 26): c.fill = CEPF
                if j in (4, 7, 18, 20, 22): c.fill = ABEF
        gc = ws.cell(r, GCOL); gc.fill = WIN_C if vals[7] == 'Cepeda' else WIN_A; gc.font = WINFT
        if total:
            for j in range(1, len(COLS)+1):
                cc = ws.cell(r, j); cc.font = Font(bold=True)
                if j != GCOL: cc.fill = PatternFill('solid', fgColor='ECE6D8')
        r += 1
    ws.freeze_panes = 'C2'; ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{r-1}"

# ---------------- hoja Leeme ----------------
def leeme(ws, totales):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3; ws.column_dimensions['B'].width = 116
    L = [('Resultados Segunda Vuelta Presidencial 2026 — por barrio · 17 ciudades', 't'),
         ('Cepeda vs Abelardo · 2da vuelta 21-jun-2026 · una hoja por ciudad', 's'), ('', ''),
         ('Qué hay aquí', 'h'),
         ('Una hoja por ciudad con el voto de 2da vuelta desagregado por barrio y comuna/localidad.', 'b'),
         ('Hoja "Resumen": totales de las 17 ciudades en una tabla. Orden de hojas: por votación 2V.', 'b'), ('', ''),
         ('Fuente y método', 'h'),
         ('Votos 2V: preconteo por mesa de la Registraduría (detalle_nacional_presidencia_mesas).', 'b'),
         ('Barrio y comuna de cada puesto: master por puesto del análisis 2V (resuelve Bogotá, que no', 'b'),
         ('está en PUESTOS_GEOREF). Para puestos de corregimiento con código alfanumérico se usa', 'b'),
         ('PUESTOS_GEOREF como respaldo. En 2da vuelta solo compitieron Cepeda y Abelardo.', 'b'),
         ('El barrio es el del puesto de votación (tal como lo nombra la Registraduría); se agregan', 'b'),
         ('todos los puestos del mismo barrio. Nombres unificados en mayúsc/minúsc y espacios. Los puestos', 'b'),
         ('cuyo barrio concatenaba varios ("Dindalito, Ciudad de Cali, ...") o llegaba como basura ("N/A"),', 'b'),
         ('o sin comuna, se reasignaron por COORDENADAS (lat/lon) al puesto bien identificado más cercano', 'b'),
         ('(regla del proyecto: coordenadas > nombre). Soledad no trae comunas en la fuente: queda "No Aplica".', 'b'), ('', ''),
         ('Columnas', 'h'),
         ('Válidos 2V = Cepeda + Abelardo. Cepeda %/Abelardo % = sobre válidos. Margen = diferencia en pp.', 'b'),
         ('Total votos = válidos + blanco + nulos + no marcados. Puestos = puestos de votación en el barrio.', 'b'),
         ('Mesas = mesas de votación sumadas. Ganador = quién obtuvo más votos en el barrio.', 'b'), ('', ''),
         ('Notas / límites', 'h'),
         ('No se incluye censo ni participación por barrio: la Registraduría reparte el censo por puesto de un', 'b'),
         ('modo que no coincide con dónde vota la gente (hay puestos con censo casi nulo y cientos de votos),', 'b'),
         ('así que una participación por barrio saldría sin sentido. Los votos sí son exactos.', 'b'),
         ('El barrio es el del puesto de votación; los puestos especiales (censo, cárceles) entran con su nombre.', 'b'),
         ('Para el desglose oficial de corregimientos de Medellín ver Resultados_2V_Medellin_por_barrio.xlsx.', 'b'), ('', ''),
         ('Voto por sexo — columnas % Mujeres / Muj→ / Hom→ (MODELO, no es conteo)', 'h'),
         ('El voto es secreto: no se observa por quién votó cada sexo. Se ESTIMA. Como las cédulas se asignan', 'b'),
         ('por rango y las mesas por cédula, dentro de un puesto hay mesas casi puras de mujeres y de hombres;', 'b'),
         ('comparar esas mesas del MISMO puesto (efectos fijos de puesto) da la brecha de sexo sin el sesgo', 'b'),
         ('ecológico. La brecha se estima del 1V 2022 (izquierda Petro vs derecha Fico+Rodolfo, el análogo', 'b'),
         ('binario del 2V), por comuna en Bogotá/Cali/Medellín/Barranquilla y por ciudad en el resto, con', 'b'),
         ('suavizado hacia región/nacional. Nacional: las mujeres votan ~6 pp menos a la izquierda (patrón', 'b'),
         ('latinoamericano, validado contra la proyección DANE 2026). % Mujeres = fracción del electorado del', 'b'),
         ('barrio según la proyección de población DANE 2026. Muj→Cepeda y Hom→Cepeda reconcilian con el', 'b'),
         ('resultado real: %muj·(Muj→) + %hom·(Hom→) = Cepeda real del barrio. Abelardo = 100 − Cepeda.', 'b'),
         ('LÍMITE: NO mide el comportamiento de género propio de cada barrio (con 1-2 puestos es inidentificable,', 'b'),
         ('ahí la inferencia ecológica se vuelve inestable). Impone la brecha de su comuna/ciudad sobre el', 'b'),
         ('resultado real y la mezcla de sexos del barrio; las diferencias entre barrios vienen de su resultado y', 'b'),
         ('su composición, no de dinámicas de género locales. Úsese como aproximación, no como dato observado.', 'b'), ('', ''),
         ('Comparativo con 2022 — Petro/Rodolfo 2V 2022, Izq 2022 %, Swing izq, Muj/Hom→Petro 2022', 'h'),
         ('La 2da vuelta fue izquierda vs derecha en ambos años: 2022 Petro vs Rodolfo, 2026 Cepeda vs Abelardo.', 'b'),
         ('Petro/Rodolfo 2V 2022 = votos reales de 2da vuelta 2022 por barrio (mismos puestos, mapeo idéntico).', 'b'),
         ('Izq 2022 % = Petro sobre válidos. Swing izq = Cepeda 2026 % − Petro 2022 % (+ = la izquierda creció;', 'b'),
         ('el de la derecha es el espejo, Abelardo − Rodolfo). Muj→Petro / Hom→Petro 2022 = el mismo modelo de', 'b'),
         ('sexo aplicado a 2022, pero con el gap REAL del 2V 2022 (Petro-Rodolfo) y la composición de sexo', 'b'),
         ('observada de 2022 — así cada año usa su propio dato. Dato de interés: la brecha de sexo se AMPLIÓ de', 'b'),
         ('~4 pp (2022) a ~6 pp (2026): Abelardo polariza más por género que Rodolfo. Barrios sin puesto en 2022', 'b'),
         ('quedan en blanco en estas columnas.', 'b'), ('', ''),
         ('Generado por tools/segunda-vuelta-prec/build_ciudades_barrios_xlsx.py (género: estimate_genero_2v.py)', 'n')]
    r = 1
    for txt, k in L:
        c = ws.cell(r, 2, txt)
        c.font = (Font(bold=True, size=15, color='8A1E16') if k == 't' else
                  Font(size=11, italic=True, color='555555') if k == 's' else
                  Font(bold=True, size=11, color='1F47CC') if k == 'h' else
                  Font(size=9, italic=True, color='888888') if k == 'n' else Font(size=10))
        r += 1

# ---------------- construir workbook ----------------
wb = Workbook()
leeme(wb.active, None); wb.active.title = 'Leeme'

# hoja Resumen (placeholder, se llena al final con totales por ciudad)
ws_res = wb.create_sheet('Resumen')

city_tot = []
GRAN = {'cep':0,'abe':0,'bl':0,'nu':0,'nm':0,'mesas':0,'puestos':0,'censo':0,'wmuj':0.0,'wtot':0.0,
        'p22':0,'r22':0,'wmuj22':0.0,'wtot22':0.0}
for c5, ciudad in CITIES:
    rows = AGG.get(ciudad, {})
    b22c = B22.get(ciudad, {})
    # adjunta el 2V 2022 a cada barrio por su misma llave + descompone por sexo con el gap REAL de 2022
    for key, a in rows.items():
        d = b22c.get(key)
        if d and (d['p22'] + d['r22']) > 0:
            f22 = d['wmuj']/d['wtot'] if d['wtot'] > 0 else FNAC22
            a['p22'], a['r22'] = d['p22'], d['r22']; a['izq22'] = d['p22']/(d['p22']+d['r22'])
            _, a['mp22'], _, a['hp22'], _ = decompose(d['p22'], d['r22'], f22, gap_for22(ciudad, a['cdisp']))
        else:
            a['p22'] = a['r22'] = a['izq22'] = a['mp22'] = a['hp22'] = None
    items = sorted(rows.values(), key=lambda a: (a['ccode'], a['cdisp'], -(a['cep']+a['abe'])))
    tot = blank(); tot['bdisp'] = f'TOTAL {ciudad.upper()}'; tot['cdisp'] = ''
    for a in items:
        for k in ('cep','abe','bl','nu','nm','mesas','puestos','censo','wmuj','wtot'): tot[k] += a[k]
    # genero 2026 por barrio: descompone el resultado real con el gap de su comuna/ciudad y su mezcla de sexos
    gp_city = GAPS.get(f'{ciudad}|*', {}).get('gap', GAP_NAC)
    for a in items:
        fb = a['wmuj']/a['wtot'] if a['wtot'] > 0 else FNAC
        a['fmuj'], a['mc'], a['ma'], a['hc'], a['ha'] = decompose(a['cep'], a['abe'], fb, gap_for(ciudad, a['cdisp']))
    fbt = tot['wmuj']/tot['wtot'] if tot['wtot'] > 0 else FNAC
    tot['fmuj'], tot['mc'], tot['ma'], tot['hc'], tot['ha'] = decompose(tot['cep'], tot['abe'], fbt, gp_city)
    # total ciudad 2022 (todos los puestos 2022 de la ciudad)
    tp = sum(d['p22'] for d in b22c.values()); tr = sum(d['r22'] for d in b22c.values())
    twm = sum(d['wmuj'] for d in b22c.values()); twt = sum(d['wtot'] for d in b22c.values())
    tot['p22'], tot['r22'], tot['wmuj22'], tot['wtot22'] = tp, tr, twm, twt
    if tp + tr > 0:
        f22 = twm/twt if twt > 0 else FNAC22
        tot['izq22'] = tp/(tp+tr)
        _, tot['mp22'], _, tot['hp22'], _ = decompose(tp, tr, f22, GAPS22.get(f'{ciudad}|*', {}).get('gap', GAP_NAC22))
    filas = [(a, False) for a in items] + [(tot, True)]
    ws = wb.create_sheet(ciudad[:31])
    pinta(ws, filas)
    val = tot['cep'] + tot['abe']
    city_tot.append((ciudad, len(items), tot, val))
    for k in GRAN: GRAN[k] += tot[k]
    print(f'  {ciudad:<14} barrios={len(items):<4} Cepeda={tot["cep"]:>9,} Abelardo={tot["abe"]:>9,} '
          f'({100*tot["abe"]/val:.1f}% Abe)  puestos={tot["puestos"]}')

# ---- llenar hoja Resumen ----
RCOLS = [('Ciudad', 16, 'txt'), ('Barrios', 9, 'int'), ('Cepeda 2V', 12, 'int'), ('Abelardo 2V', 12, 'int'),
         ('Válidos 2V', 12, 'int'), ('Cepeda %', 9, 'pct'), ('Abelardo %', 10, 'pct'), ('Ganador', 10, 'txt'),
         ('Margen (pp)', 11, 'pp'), ('Muj→Cepeda', 11, 'pct'), ('Hom→Cepeda', 11, 'pct'),
         ('Petro 2022 %', 11, 'pct'), ('Swing izq (pp)', 13, 'pp'), ('Muj→Petro 22', 12, 'pct'), ('Hom→Petro 22', 12, 'pct'),
         ('Puestos', 8, 'int'), ('Mesas', 7, 'int')]
for j, (t, w, _) in enumerate(RCOLS, 1):
    c = ws_res.cell(1, j, t); c.fill = HDR; c.font = HDRF; c.alignment = CEN; c.border = BORD
    ws_res.column_dimensions[get_column_letter(j)].width = w
def resrow(ws, r, name, nb, t, val, bold=False):
    cp = t['cep']/val if val else 0; ap = t['abe']/val if val else 0
    gan = 'Cepeda' if t['cep'] >= t['abe'] else 'Abelardo'
    izq22 = t.get('izq22'); swing = (cp - izq22) * 100 if izq22 is not None else None
    vals = [name, nb, t['cep'], t['abe'], val, cp, ap, gan, abs(cp-ap)*100,
            t.get('mc', 0), t.get('hc', 0), izq22, swing, t.get('mp22'), t.get('hp22'),
            t['puestos'], t['mesas']]
    for j, (v, (tt, w, fmt)) in enumerate(zip(vals, RCOLS), 1):
        c = ws.cell(r, j, v); cellfmt(c, fmt)
        if j == 1: c.alignment = LEFT
        elif j in (10, 11, 12, 14, 15) and not bold: c.fill = CEPF
    g = ws.cell(r, 8); g.fill = WIN_C if gan == 'Cepeda' else WIN_A; g.font = WINFT
    if bold:
        for j in range(1, len(RCOLS)+1):
            cc = ws.cell(r, j); cc.font = Font(bold=True)
            if j != 8: cc.fill = PatternFill('solid', fgColor='ECE6D8')
r = 2
for ciudad, nb, tot, val in city_tot:
    resrow(ws_res, r, ciudad, nb, tot, val); r += 1
_fb = GRAN['wmuj']/GRAN['wtot'] if GRAN['wtot'] > 0 else FNAC
GRAN['fmuj'], GRAN['mc'], GRAN['ma'], GRAN['hc'], GRAN['ha'] = decompose(GRAN['cep'], GRAN['abe'], _fb, GAP_NAC)
_fb22 = GRAN['wmuj22']/GRAN['wtot22'] if GRAN['wtot22'] > 0 else FNAC22
GRAN['izq22'] = GRAN['p22']/(GRAN['p22']+GRAN['r22']) if (GRAN['p22']+GRAN['r22']) > 0 else None
_, GRAN['mp22'], _, GRAN['hp22'], _ = decompose(GRAN['p22'], GRAN['r22'], _fb22, GAP_NAC22)
resrow(ws_res, r, 'TOTAL 17 CIUDADES', sum(x[1] for x in city_tot), GRAN, GRAN['cep']+GRAN['abe'], bold=True)
ws_res.freeze_panes = 'B2'; ws_res.auto_filter.ref = f"A1:{get_column_letter(len(RCOLS))}{r}"
ws_res.sheet_view.showGridLines = False

wb.save(OUTFILE)
gval = GRAN['cep'] + GRAN['abe']
print(f'\nOK -> {OUTFILE}')
print(f'  17 ciudades · {sum(x[1] for x in city_tot):,} filas-barrio · '
      f'Cepeda {GRAN["cep"]:,} / Abelardo {GRAN["abe"]:,} ({100*GRAN["abe"]/gval:.1f}% Abe)')
if sin_geo:
    print('  puestos sin barrio (especiales):', dict(sin_geo))
