#!/usr/bin/env python3
"""
Resultados 2da vuelta presidencial 2026 (Cepeda vs Abelardo) en MEDELLIN,
por comuna/corregimiento y por barrio -> dos Excel.

FUENTES AUTORITATIVAS (no el master, que omite los corregimientos con codigo
de puesto alfanumerico A1-A6/B1-B7 y subcuenta ~67k votos):
  - Votos 2V por mesa : output_2v/detalle_nacional_presidencia_mesas.xlsx
  - Votos 1V por mesa : nuevos archivos 1v 2026/PRECONTEO_1V_2026_MESA_con_Claudia.csv
  - Puesto->barrio/comuna/censo : PUESTOS_GEOREF.csv (clave codigo 9 digitos)

La 2V solo trae Cepeda y Abelardo; blanco/nulos/no-marcados vienen aparte por mesa.
Join puesto: pcode = dep(2)+mun(3)+zona(2)+puesto  ej '0100199A1'.
"""
import json, os, csv
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(__file__), '..', '..')
BD   = os.path.join(ROOT, 'Bases de datos')
OUT  = os.path.join(BD, 'output_2v')
XLSX_2V = os.path.join(OUT, 'detalle_nacional_presidencia_mesas.xlsx')
CSV_1V  = os.path.join(BD, 'nuevos archivos 1v 2026', 'PRECONTEO_1V_2026_MESA_con_Claudia.csv')
GEOREF  = os.path.join(BD, 'PUESTOS_GEOREF.csv')

# zona numerica -> comuna (1..16)
Z2C = {1:'01',2:'01',3:'02',4:'02',5:'03',6:'03',7:'04',8:'04',9:'05',10:'05',
       11:'06',12:'06',13:'07',14:'07',15:'08',16:'08',17:'09',18:'09',19:'10',20:'10',
       21:'11',22:'11',23:'12',24:'12',25:'13',26:'13',27:'14',28:'14',29:'15',
       30:'16',31:'16',32:'16'}
NOMBRE = {
 '01':'Comuna 1 - Popular','02':'Comuna 2 - Santa Cruz','03':'Comuna 3 - Manrique',
 '04':'Comuna 4 - Aranjuez','05':'Comuna 5 - Castilla','06':'Comuna 6 - Doce de Octubre',
 '07':'Comuna 7 - Robledo','08':'Comuna 8 - Villa Hermosa','09':'Comuna 9 - Buenos Aires',
 '10':'Comuna 10 - La Candelaria','11':'Comuna 11 - Laureles Estadio','12':'Comuna 12 - La America',
 '13':'Comuna 13 - San Javier','14':'Comuna 14 - El Poblado','15':'Comuna 15 - Guayabal',
 '16':'Comuna 16 - Belen',
 '17':'Corregimiento Altavista','18':'Corregimiento San Antonio de Prado',
 '19':'Corregimiento Palmitas','20':'Corregimiento San Cristobal','21':'Corregimiento Santa Elena',
 '90':'Especial - Puesto Censo','98':'Especial - Carceles / Reclusion'}
TIPO = {c:('Comuna' if c.isdigit() and 1<=int(c)<=16 else
           'Especial' if c in ('90','98') else 'Corregimiento') for c in NOMBRE}
ORDEN = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16',
         '17','18','19','20','21','90','98']

def pcode(dep, mun, zona, puesto):
    return f"{str(dep).zfill(2)}{str(mun).zfill(3)}{str(zona).zfill(2)}{str(puesto).zfill(2)}"

def corr_de_texto(t):
    t=(t or '').upper()
    if 'ALTAVISTA' in t: return '17'
    if 'SAN ANTONIO' in t or 'PRADO' in t: return '18'
    if 'PALMITAS' in t: return '19'
    if 'CRISTOBAL' in t: return '20'
    if 'SANTA ELENA' in t: return '21'
    return None

# -------- GEOREF registry: pcode -> barrio/comuna/censo --------
GREG={}
with open(GEOREF, encoding='utf-8-sig') as f:
    for r in csv.DictReader(f, delimiter=';'):
        if str(r.get('DEPARTAMENTO','')).strip()!='ANTIOQUIA': continue
        if 'MEDELL' not in str(r.get('MUNICIPIO','')).upper(): continue
        z=r['ZONA']; pu=r['PUESTO']
        key=pcode('01','001',z,pu)
        try: censo=int(r['MUJERES'] or 0)+int(r['HOMBRES'] or 0)
        except: censo=0
        zi=int(z) if z.isdigit() else 0
        if zi in Z2C: cc=Z2C[zi]
        elif zi==90: cc='90'
        elif zi==98: cc='98'
        elif zi==99: cc=corr_de_texto(r.get('NOMBRE COMUNA') or r.get('BARRIO'))
        else: cc=None
        GREG[key]={'barrio':(r['BARRIO'] or '').strip() or 'SIN BARRIO',
                   'comuna':cc, 'censo':censo, 'pnom':(r.get('NOMBRE PUESTO') or '').strip()}

# -------- 2V votes por puesto --------
V2=defaultdict(lambda: dict(cep=0,abe=0,bl=0,nu=0,nm=0,mesas=0,pnom=''))
wb=load_workbook(XLSX_2V, read_only=True); ws=wb.active
it=ws.iter_rows(values_only=True); next(it)
def num(x): return x if isinstance(x,(int,float)) else 0
for r in it:
    if 'MEDELL' not in str(r[2] or '').upper(): continue
    key=pcode('01','001', r[3], r[4])
    a=V2[key]; a['cep']+=num(r[10]); a['abe']+=num(r[11])
    a['bl']+=num(r[7]); a['nm']+=num(r[8]); a['nu']+=num(r[9]); a['mesas']+=1
    a['pnom']=str(r[5] or '')
wb.close()

# -------- 1V votes por puesto (cara a cara cepeda/abelardo) --------
V1=defaultdict(lambda: dict(cep=0,abe=0))
with open(CSV_1V, encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        if row.get('cod_departamento')!='01' or row.get('cod_municipio')!='001': continue
        key=pcode('01','001', row['zona'], row['puesto'])
        V1[key]['cep']+=int(row.get('Iván Cepeda') or 0)
        V1[key]['abe']+=int(row.get('Abelardo De La Espriella') or 0)

# -------- agregacion comuna / barrio --------
def blank(): return dict(cep2=0,abe2=0,bl=0,nu=0,nm=0,mesas=0,cep1=0,abe1=0,censo=0)
por_comuna=defaultdict(blank); por_barrio=defaultdict(blank)
unmatched=[]
allkeys=set(V2)|set(GREG)
for key in allkeys:
    v2=V2.get(key); g=GREG.get(key)
    zona=int(key[5:7]) if key[5:7].isdigit() else 0
    # comuna
    if g and g['comuna']: c=g['comuna']
    elif zona in Z2C: c=Z2C[zona]
    elif zona==90: c='90'
    elif zona==98: c='98'
    elif zona==99:
        pu=key[7:9]   # puestos serie A = San Antonio de Prado, serie B = San Cristobal (codigo georef)
        c=('18' if pu[:1]=='A' else '20' if pu[:1]=='B'
           else corr_de_texto((v2 or {}).get('pnom')) or '21')
    else: c='??'
    barrio=g['barrio'] if g else ((v2 or {}).get('pnom') or 'SIN BARRIO').strip()
    censo=g['censo'] if g else 0
    if v2 is None:           # puesto en georef sin votos (cerrado) -> aporta censo
        por_comuna[c]['censo']+=censo
        por_barrio[(c,barrio)]['censo']+=censo
        continue
    if g is None: unmatched.append((key, v2['pnom']))
    v1=V1.get(key, {'cep':0,'abe':0})
    for tgt in (por_comuna[c], por_barrio[(c,barrio)]):
        tgt['cep2']+=v2['cep']; tgt['abe2']+=v2['abe']
        tgt['bl']+=v2['bl']; tgt['nu']+=v2['nu']; tgt['nm']+=v2['nm']; tgt['mesas']+=v2['mesas']
        tgt['cep1']+=v1['cep']; tgt['abe1']+=v1['abe']; tgt['censo']+=censo

# ------------------------------ estilos -----------------------------
ROJO,AZUL='C0392B','1F47CC'
HDR=PatternFill('solid',fgColor='8A1E16'); HDRF=Font(bold=True,color='FFFFFF',size=10)
ZEBRA=PatternFill('solid',fgColor='F4F0E7')
CEPF=PatternFill('solid',fgColor='F7DAD5'); ABEF=PatternFill('solid',fgColor='D7DEF6')
WIN_C=PatternFill('solid',fgColor=ROJO); WIN_A=PatternFill('solid',fgColor=AZUL)
WINFT=Font(color='FFFFFF',bold=True)
thin=Side(style='thin',color='D9D2C5'); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment('center',vertical='center'); LEFT=Alignment('left',vertical='center')

COLS_BASE=[
 ('Cepeda 2V (votos)',12,'int'),('Abelardo 2V (votos)',13,'int'),('Validos 2V',11,'int'),
 ('Cepeda %',9,'pct'),('Abelardo %',9,'pct'),('Ganador',11,'txt'),('Margen (pp)',11,'pp'),
 ('Blanco 2V',10,'int'),('Nulos 2V',9,'int'),('No marcados 2V',12,'int'),('Total votos 2V',12,'int'),
 ('Censo electoral',13,'int'),('Participacion %',13,'pct'),('Mesas',7,'int'),
 ('Cepeda 1V % (cara a cara)',15,'pct'),('Abelardo 1V % (cara a cara)',15,'pct'),
 ('Swing Cepeda 1V→2V (pp)',15,'pp')]

def metricas(a):
    val=a['cep2']+a['abe2']
    cp=a['cep2']/val if val else 0; ap=a['abe2']/val if val else 0
    gan='Cepeda' if a['cep2']>=a['abe2'] else 'Abelardo'
    margen=abs(cp-ap)*100
    total=val+a['bl']+a['nu']+a['nm']
    part=total/a['censo'] if a['censo'] else 0
    v1=a['cep1']+a['abe1']; cep1=a['cep1']/v1 if v1 else 0; abe1=a['abe1']/v1 if v1 else 0
    swing=(cp-cep1)*100 if v1 else 0
    return [a['cep2'],a['abe2'],val,cp,ap,gan,margen,a['bl'],a['nu'],a['nm'],total,
            a['censo'],part,a['mesas'],cep1,abe1,swing]

def cellfmt(c,fmt):
    c.border=BORD; c.alignment=CEN
    if fmt=='int': c.number_format='#,##0'
    elif fmt=='pct': c.number_format='0.0%'
    elif fmt=='pp': c.number_format='+0.0;-0.0'

def pinta(ws, pre_cols, filas):
    cols=pre_cols+COLS_BASE; npre=len(pre_cols)
    for j,(t,w,*_) in enumerate(cols,1):
        c=ws.cell(1,j,t); c.fill=HDR; c.font=HDRF; c.alignment=CEN; c.border=BORD
        ws.column_dimensions[get_column_letter(j)].width=w
    r=2; gcol=npre+6
    for pre,agg,total in filas:
        m=metricas(agg); vals=list(pre)+m
        for j,(v,(t,w,fmt)) in enumerate(zip(vals,cols),1):
            c=ws.cell(r,j,v); cellfmt(c,fmt)
            if j<=npre: c.alignment=LEFT
            elif not total and r%2==0: c.fill=ZEBRA
            if not total:
                if j in (npre+1,npre+4): c.fill=CEPF
                if j in (npre+2,npre+5): c.fill=ABEF
        gc=ws.cell(r,gcol); gc.fill=WIN_C if m[5]=='Cepeda' else WIN_A; gc.font=WINFT
        if total:
            for j in range(1,len(cols)+1):
                cc=ws.cell(r,j); cc.font=Font(bold=True)
                if j!=gcol: cc.fill=PatternFill('solid',fgColor='ECE6D8')
        r+=1
    ws.freeze_panes='C2'; ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{r-1}"

def leeme(ws,nivel,nfilas,tot):
    ws.sheet_view.showGridLines=False
    ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=112
    ab=100*tot['abe2']/(tot['cep2']+tot['abe2'])
    L=[('Resultados Segunda Vuelta Presidencial 2026 — Medellin','t'),
       (f'Nivel: {nivel}  ·  {nfilas} filas  ·  Cepeda vs Abelardo  ·  2da vuelta 21-jun-2026','s'),('',''),
       ('Fuente y metodo','h'),
       ('Votos 2V: preconteo por mesa de la Registraduria (detalle_nacional_presidencia_mesas).','b'),
       ('Votos 1V: preconteo 1V por mesa (con Claudia recuperada). Puesto->barrio/comuna/censo:','b'),
       ('PUESTOS_GEOREF de la Registraduria. La comuna se asigna por la zona electoral del puesto;','b'),
       ('los 5 corregimientos (Altavista, San Antonio de Prado, Palmitas, San Cristobal, Santa','b'),
       ('Elena) salen de la zona 99. En 2V solo compitieron Ivan Cepeda y Abelardo De La Espriella.','b'),('',''),
       ('Columnas','h'),
       ('Validos 2V = Cepeda + Abelardo. Cepeda %/Abelardo % = sobre validos. Margen = diferencia pp.','b'),
       ('Total votos 2V = validos + blanco + nulos + no marcados. Participacion = total / censo.','b'),
       ('Cepeda/Abelardo 1V % (cara a cara) = su votacion de 1ra vuelta entre los dos solamente.','b'),
       ('Swing Cepeda = Cepeda 2V % menos Cepeda 1V % cara a cara (cuanto crecio/cayo entre vueltas).','b'),('',''),
       ('Totales y notas','h'),
       (f'Control de ciudad (preconteo real): Cepeda {tot["cep2"]:,} / Abelardo {tot["abe2"]:,} '
        f'= {ab:.1f}% Abelardo. Abelardo gano TODAS las comunas menos Popular (casi empate).','b'),
       ('Las filas "Especial - Puesto Censo" (zona 90) y "Especial - Carceles" (zona 98) recogen','b'),
       ('votantes sin barrio asignado; se listan aparte para que los totales cuadren con la ciudad.','b'),
       ('En El Poblado la participacion supera 100%: hay centros de votacion (p.ej. Los Naranjos) con','b'),
       ('censo 0 asignado pero votacion real; es un artefacto de como la Registraduria reparte el censo','b'),
       ('por puesto, no un error de conteo. Los votos son correctos.','b'),('',''),
       ('Generado por tools/segunda-vuelta-prec/build_medellin_excel.py','n')]
    r=1
    for txt,k in L:
        c=ws.cell(r,2,txt)
        c.font=(Font(bold=True,size=15,color='8A1E16') if k=='t' else
                Font(size=11,italic=True,color='555555') if k=='s' else
                Font(bold=True,size=11,color='1F47CC') if k=='h' else
                Font(size=9,italic=True,color='888888') if k=='n' else Font(size=10))
        r+=1

# ----------------------------- COMUNA -----------------------------
tot=blank()
for c in ORDEN:
    if c in por_comuna:
        for k in tot: tot[k]+=por_comuna[c][k]
wb=Workbook(); leeme(wb.active,'comuna / corregimiento',len([c for c in ORDEN if c in por_comuna]),tot)
wb.active.title='Leeme'
ws=wb.create_sheet('Por comuna')
filas=[([NOMBRE[c],TIPO[c]], por_comuna[c], False) for c in ORDEN if c in por_comuna]
filas.append((['TOTAL MEDELLIN',''], tot, True))
pinta(ws,[('Comuna / Corregimiento',32,'txt'),('Tipo',14,'txt')],filas)
out1=os.path.join(OUT,'Resultados_2V_Medellin_por_comuna.xlsx'); wb.save(out1)

# ----------------------------- BARRIO -----------------------------
wb=Workbook(); leeme(wb.active,'barrio',len(por_barrio),tot); wb.active.title='Leeme'
ws=wb.create_sheet('Por barrio')
def sk(item):
    (c,b),a=item
    return (ORDEN.index(c) if c in ORDEN else 99, -(a['cep2']+a['abe2']))
filas=[([NOMBRE.get(c,c), b], a, False) for (c,b),a in sorted(por_barrio.items(),key=sk)]
filas.append((['TOTAL MEDELLIN',''], tot, True))
pinta(ws,[('Comuna / Corregimiento',30,'txt'),('Barrio',28,'txt')],filas)
out2=os.path.join(OUT,'Resultados_2V_Medellin_por_barrio.xlsx'); wb.save(out2)

print('OK  ->', out1)
print('    ->', out2)
print(f'  comunas/corr/esp: {len([c for c in ORDEN if c in por_comuna])} | barrios: {len(por_barrio)}')
print(f'  CONTROL ciudad: Cepeda {tot["cep2"]:,} / Abelardo {tot["abe2"]:,} '
      f'({100*tot["abe2"]/(tot["cep2"]+tot["abe2"]):.2f}% Abelardo) | mesas {tot["mesas"]:,}')
if unmatched:
    print(f'  puestos con voto sin georef (barrio=nombre puesto): {len(unmatched)}')
    for k,n in unmatched[:8]: print('     ',k,n)
