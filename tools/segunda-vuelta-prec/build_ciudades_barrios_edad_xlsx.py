#!/usr/bin/env python3
"""
Composición ETARIA de los votantes por BARRIO en las 17 ciudades principales,
presidencial 2026 (1ª y 2ª vuelta) -> UN solo Excel, una hoja por ciudad.
Misma estructura que los Excel de resultados por barrio (filas = barrio).

QUÉ ES: cuántos votantes de cada franja de edad hay en cada barrio. NO es
"por quién votó cada edad" (eso solo se estima por ciudad/depto vía inferencia
ecológica; a nivel barrio, con 1-2 puestos, es inidentificable).

FUENTES:
  - Composición etaria 1V 2026 por puesto : output_edad_1v/w26-puesto.csv
        (proyección 2026 = perfil 2022 del puesto × envejecimiento DANE × raking
         IPF a los votantes reales del preconteo 1V por puesto). 10 bandas b0..b9.
  - Total de votantes 2V por puesto        : master_unificado_puesto.json (urna2)
  - Barrio / comuna                         : geo_barrios.py

MODELO 2V: la mezcla de edad por puesto se toma de la proyección 1V y se escala
al total de votantes reales de 2V del puesto (2V tuvo más participación). El %
por barrio puede diferir algo entre vueltas porque cambia el peso relativo de
los puestos del barrio. Supuesto declarado: no modelamos un corrimiento etario
propio 1V->2V (habría que rehacer el raking para 2V); los absolutos 2V sí están
anclados a la participación real de 2V.

Bandas Edadygenero (10) -> grupos (5):
  18-25 = b0(18-20)+b1(21-25) · 26-35 = b2+b3 · 36-45 = b4+b5
  46-60 = b6+b7+b8 · 61+ = b9(Mayor a 60)
"""
import os, csv, json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import geo_barrios as GEO

ROOT = os.path.join(os.path.dirname(__file__), '..', '..')
BD   = os.path.join(ROOT, 'Bases de datos')
OUT  = os.path.join(BD, 'output_2v')
W26  = os.path.join(BD, 'output_edad_1v', 'w26-puesto.csv')
OUTFILE = os.path.join(OUT, 'Composicion_Edad_2026_por_barrio_ciudades.xlsx')

GRUPOS = ['18-25', '26-35', '36-45', '46-60', '61+']
GMAP = {'18-25': ['b0', 'b1'], '26-35': ['b2', 'b3'], '36-45': ['b4', 'b5'],
        '46-60': ['b6', 'b7', 'b8'], '61+': ['b9']}

g = GEO.build()
CITY, RESOLVE = g.CITY, g.RESOLVE
CITIES = GEO.CITIES

# ---- total de votantes 2V por puesto (master) ----
URNA2 = {}
for p in json.load(open(GEO.MASTER)):
    URNA2[p['pcode']] = int(p.get('urna2') or 0)

# ---- 1) composición 1V por puesto -> grupos de edad + total ----
def blankp(): return {gr: 0.0 for gr in GRUPOS}
P1 = {}   # pcode -> {grupo: votantes 1V}
nfaltan2v = 0
with open(W26) as f:
    for r in csv.DictReader(f):
        pc = r['pcode'].replace('-', '')
        if pc[:5] not in CITY: continue
        d = blankp()
        for gr, bs in GMAP.items():
            d[gr] = sum(float(r[b]) for b in bs)
        P1[pc] = d
print(f'w26: puestos de las 17 ciudades = {len(P1):,}')

# ---- 2) agrega por (ciudad, comuna, barrio): 1V y 2V (2V = 1V × urna2/total1V) ----
def blank():
    return {'v1': {gr: 0.0 for gr in GRUPOS}, 'v2': {gr: 0.0 for gr in GRUPOS},
            't1': 0.0, 't2': 0.0, 'puestos': 0, 'ccode': 999, 'cdisp': '', 'bdisp': '',
            'sin2v': 0}
AGG = defaultdict(lambda: defaultdict(blank))
sin_geo = defaultdict(int)
for pc, comp in P1.items():
    ciudad = CITY[pc[:5]]
    if pc in RESOLVE:
        ccode, cdisp, barrio = RESOLVE[pc]
    else:
        sin_geo[ciudad] += 1
        ccode, cdisp, barrio = 999, 'Sin comuna asignada', 'Sin barrio asignado'
    g.register_canon(ciudad, cdisp)
    tot1 = sum(comp.values())
    u2 = URNA2.get(pc, 0)
    f2 = (u2 / tot1) if tot1 > 0 else 0.0   # escala 1V->2V por participación real
    key = (GEO.nrm(cdisp), GEO.nrm(barrio))
    t = AGG[ciudad][key]
    for gr in GRUPOS:
        t['v1'][gr] += comp[gr]
        t['v2'][gr] += comp[gr] * f2
    t['t1'] += tot1; t['t2'] += tot1 * f2
    t['puestos'] += 1
    if u2 <= 0: t['sin2v'] += 1
    if ccode < t['ccode']: t['ccode'] = ccode
    t['cdisp'] = cdisp; t['bdisp'] = GEO.titlecase(barrio)
for ciudad, rows in AGG.items():
    for a in rows.values():
        a['cdisp'] = g.canon(ciudad, a['cdisp'])

# ---------------- estilos ----------------
HDR = PatternFill('solid', fgColor='8A1E16'); HDRF = Font(bold=True, color='FFFFFF', size=10)
HDR1 = PatternFill('solid', fgColor='2E5A34'); HDR2 = PatternFill('solid', fgColor='3A4A7A')  # sub-bloques 1V/2V
ZEBRA = PatternFill('solid', fgColor='F4F0E7')
# gradiente por franja (joven->rojo cálido, mayor->azul frío) para las columnas %
GRCLR = {'18-25': 'F7DAD5', '26-35': 'F9E7D0', '36-45': 'F1EEDC', '46-60': 'DCE6F0', '61+': 'D2DBEF'}
thin = Side(style='thin', color='D9D2C5'); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment('center', vertical='center'); LEFT = Alignment('left', vertical='center')

# columnas: Barrio, Comuna, [1V abs x5], Tot 1V, [1V % x5], [2V abs x5], Tot 2V, [2V % x5], Puestos
def cols_def():
    c = [('Barrio', 30, 'txt'), ('Comuna / Localidad', 26, 'txt')]
    for gr in GRUPOS: c.append((f'1V {gr}', 9, 'int'))
    c.append(('Votantes 1V', 11, 'int'))
    for gr in GRUPOS: c.append((f'1V % {gr}', 9, 'pct'))
    for gr in GRUPOS: c.append((f'2V {gr}', 9, 'int'))
    c.append(('Votantes 2V', 11, 'int'))
    for gr in GRUPOS: c.append((f'2V % {gr}', 9, 'pct'))
    c.append(('Puestos', 8, 'int'))
    return c
COLS = cols_def()

def metricas(a):
    row = [a['bdisp'], a['cdisp']]
    for gr in GRUPOS: row.append(round(a['v1'][gr]))
    row.append(round(a['t1']))
    for gr in GRUPOS: row.append(a['v1'][gr] / a['t1'] if a['t1'] > 0 else 0)
    for gr in GRUPOS: row.append(round(a['v2'][gr]))
    row.append(round(a['t2']))
    for gr in GRUPOS: row.append(a['v2'][gr] / a['t2'] if a['t2'] > 0 else 0)
    row.append(a['puestos'])
    return row

def cellfmt(c, fmt):
    c.border = BORD; c.alignment = CEN
    if fmt == 'int': c.number_format = '#,##0'
    elif fmt == 'pct': c.number_format = '0.0%'

def pinta(ws, filas):
    for j, (t, w, _) in enumerate(COLS, 1):
        c = ws.cell(1, j, t); c.font = HDRF; c.alignment = CEN; c.border = BORD
        # color de header por bloque
        c.fill = HDR1 if t.startswith('1V') else HDR2 if t.startswith('2V') else HDR
        ws.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for agg, total in filas:
        vals = metricas(agg)
        for j, (v, (t, w, fmt)) in enumerate(zip(vals, COLS), 1):
            c = ws.cell(r, j, v); cellfmt(c, fmt)
            if j <= 2: c.alignment = LEFT
            elif not total and r % 2 == 0: c.fill = ZEBRA
            if not total and fmt == 'pct':
                gr = t.split('% ')[-1]
                if gr in GRCLR: c.fill = PatternFill('solid', fgColor=GRCLR[gr])
        if total:
            for j in range(1, len(COLS) + 1):
                cc = ws.cell(r, j); cc.font = Font(bold=True)
                cc.fill = PatternFill('solid', fgColor='ECE6D8')
        r += 1
    ws.freeze_panes = 'C2'; ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{r-1}"

# ---------------- hoja Leeme ----------------
def leeme(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3; ws.column_dimensions['B'].width = 116
    L = [('Composición etaria de los votantes por barrio — presidencial 2026 · 17 ciudades', 't'),
         ('1ª y 2ª vuelta · una hoja por ciudad · cuántos votantes de cada edad hay en cada barrio', 's'), ('', ''),
         ('Qué es (y qué NO es)', 'h'),
         ('Es la COMPOSICIÓN POR EDAD del electorado de cada barrio: qué porción de los votantes está en', 'b'),
         ('cada franja (18-25, 26-35, 36-45, 46-60, 61+). NO dice por quién votó cada edad: el voto es', 'b'),
         ('secreto y "quién votó a quién por edad" solo se estima por ciudad/depto con inferencia ecológica;', 'b'),
         ('a nivel barrio (1-2 puestos) es inidentificable. Aquí solo está la mezcla de edades que asistió.', 'b'), ('', ''),
         ('Fuente y método', 'h'),
         ('Composición etaria por puesto: proyección 2026 (perfil de edad 2022 del puesto × envejecimiento', 'b'),
         ('DANE por departamento × raking IPF a los votantes reales del preconteo 1V del puesto). Insumo:', 'b'),
         ('Edadygenero (RNEC) trae sufragantes por mesa × 10 bandas de edad para 2018/2022; 2026 aún no', 'b'),
         ('existe, por eso se proyecta. El nivel (cuántos votaron) es dato real del preconteo; solo la mezcla', 'b'),
         ('interna por edad es modelo. Barrio y comuna del puesto: mismo método que los otros Excel por barrio', 'b'),
         ('(master por puesto + reasignación por coordenadas cuando el nombre del barrio es basura o falta).', 'b'), ('', ''),
         ('1ª vuelta vs 2ª vuelta', 'h'),
         ('La 2V tuvo más participación (63,6% vs 58,0% nacional). Los ABSOLUTOS de 2V por franja escalan la', 'b'),
         ('mezcla de edad del puesto al total de votantes REALES de 2V del puesto, así que muestran cuántos', 'b'),
         ('votantes más de cada edad aparecieron. LÍMITE declarado: no modelamos un corrimiento etario propio', 'b'),
         ('de 1V->2V (eso exigiría rehacer el raking para 2V); se asume que dentro de cada puesto la mezcla de', 'b'),
         ('edad es la misma en ambas vueltas. El % por barrio puede diferir algo entre vueltas porque cambia el', 'b'),
         ('peso relativo de los puestos del barrio, no por un cambio de comportamiento por edad.', 'b'), ('', ''),
         ('Columnas', 'h'),
         ('1V 18-25 ... 61+ = votantes de esa franja en 1ª vuelta. Votantes 1V = total del barrio (base de los %).', 'b'),
         ('1V % 18-25 ... = porción de cada franja sobre el total del barrio. Igual para el bloque 2V. Puestos =', 'b'),
         ('puestos de votación del barrio con proyección de edad disponible.', 'b'), ('', ''),
         ('Notas / límites', 'h'),
         ('Cifras de edad redondeadas (la proyección da valores fraccionarios). Puestos especiales (censo,', 'b'),
         ('cárceles) y unos pocos sin proyección de edad no entran; por eso el total de votantes de este archivo', 'b'),
         ('puede quedar ligeramente por debajo del total de los Excel de resultados. Úsese como aproximación', 'b'),
         ('de la estructura de edad, no como conteo censal.', 'b'), ('', ''),
         ('Generado por tools/segunda-vuelta-prec/build_ciudades_barrios_edad_xlsx.py', 'n')]
    r = 1
    for txt, k in L:
        c = ws.cell(r, 2, txt)
        c.font = (Font(bold=True, size=15, color='8A1E16') if k == 't' else
                  Font(size=11, italic=True, color='555555') if k == 's' else
                  Font(bold=True, size=11, color='1F47CC') if k == 'h' else
                  Font(size=9, italic=True, color='888888') if k == 'n' else Font(size=10))
        r += 1

# ---------------- construir workbook ----------------
wb = Workbook()
leeme(wb.active); wb.active.title = 'Leeme'
ws_res = wb.create_sheet('Resumen')

city_tot = []
GRAN = blank(); GRAN['bdisp'] = 'TOTAL 17 CIUDADES'
for c5, ciudad in CITIES:
    rows = AGG.get(ciudad, {})
    items = sorted(rows.values(), key=lambda a: (a['ccode'], a['cdisp'], -a['t1']))
    tot = blank(); tot['bdisp'] = f'TOTAL {ciudad.upper()}'; tot['cdisp'] = ''
    for a in items:
        for gr in GRUPOS:
            tot['v1'][gr] += a['v1'][gr]; tot['v2'][gr] += a['v2'][gr]
        tot['t1'] += a['t1']; tot['t2'] += a['t2']; tot['puestos'] += a['puestos']
    filas = [(a, False) for a in items] + [(tot, True)]
    ws = wb.create_sheet(ciudad[:31])
    pinta(ws, filas)
    city_tot.append((ciudad, len(items), tot))
    for gr in GRUPOS:
        GRAN['v1'][gr] += tot['v1'][gr]; GRAN['v2'][gr] += tot['v2'][gr]
    GRAN['t1'] += tot['t1']; GRAN['t2'] += tot['t2']; GRAN['puestos'] += tot['puestos']
    p61 = tot['v1']['61+'] / tot['t1'] * 100 if tot['t1'] else 0
    pjov = tot['v1']['18-25'] / tot['t1'] * 100 if tot['t1'] else 0
    print(f'  {ciudad:<14} barrios={len(items):<4} votantes1V={round(tot["t1"]):>9,} '
          f'18-25={pjov:4.1f}%  61+={p61:4.1f}%  puestos={tot["puestos"]}')

# ---- hoja Resumen: composición % por ciudad (1V y 2V) ----
RCOLS = [('Ciudad', 16, 'txt'), ('Barrios', 8, 'int'), ('Votantes 1V', 12, 'int')] + \
        [(f'1V % {gr}', 9, 'pct') for gr in GRUPOS] + \
        [('Votantes 2V', 12, 'int')] + [(f'2V % {gr}', 9, 'pct') for gr in GRUPOS] + \
        [('Puestos', 8, 'int')]
for j, (t, w, _) in enumerate(RCOLS, 1):
    c = ws_res.cell(1, j, t); c.font = HDRF; c.alignment = CEN; c.border = BORD
    c.fill = HDR1 if t.startswith('1V') else HDR2 if t.startswith('2V') else HDR
    ws_res.column_dimensions[get_column_letter(j)].width = w
def resrow(r, name, nb, t, bold=False):
    vals = [name, nb, round(t['t1'])] + [t['v1'][gr] / t['t1'] if t['t1'] else 0 for gr in GRUPOS] + \
           [round(t['t2'])] + [t['v2'][gr] / t['t2'] if t['t2'] else 0 for gr in GRUPOS] + [t['puestos']]
    for j, (v, (tt, w, fmt)) in enumerate(zip(vals, RCOLS), 1):
        c = ws_res.cell(r, j, v); cellfmt(c, fmt)
        if j == 1: c.alignment = LEFT
        if not bold and fmt == 'pct':
            gr = tt.split('% ')[-1]
            if gr in GRCLR: c.fill = PatternFill('solid', fgColor=GRCLR[gr])
    if bold:
        for j in range(1, len(RCOLS) + 1):
            cc = ws_res.cell(r, j); cc.font = Font(bold=True)
            cc.fill = PatternFill('solid', fgColor='ECE6D8')
r = 2
for ciudad, nb, tot in city_tot:
    resrow(r, ciudad, nb, tot); r += 1
resrow(r, 'TOTAL 17 CIUDADES', sum(x[1] for x in city_tot), GRAN, bold=True)
ws_res.freeze_panes = 'B2'; ws_res.auto_filter.ref = f"A1:{get_column_letter(len(RCOLS))}{r}"
ws_res.sheet_view.showGridLines = False

wb.save(OUTFILE)
print(f'\nOK -> {OUTFILE}')
tj = GRAN['v1']['18-25'] / GRAN['t1'] * 100; tm = GRAN['v1']['61+'] / GRAN['t1'] * 100
print(f'  17 ciudades · {sum(x[1] for x in city_tot):,} filas-barrio · votantes 1V {round(GRAN["t1"]):,} '
      f'/ 2V {round(GRAN["t2"]):,} · 18-25 {tj:.1f}% · 61+ {tm:.1f}%')
if sin_geo: print('  puestos sin barrio:', dict(sin_geo))
