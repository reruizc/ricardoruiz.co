#!/usr/bin/env python3
"""
Resultados PRIMERA vuelta presidencial 2026 por BARRIO en las 17 ciudades
principales -> UN solo Excel, una hoja por ciudad. Espejo del Excel 2V
(build_ciudades_barrios_xlsx.py) pero con los 13 candidatos de 1V.

FUENTES:
  - Votos 1V por mesa : "nuevos archivos 1v 2026/PRECONTEO_1V_2026_MESA_con_Claudia.csv"
                        (autoritativo · Claudia López recuperada exacta por mesa)
  - Barrio / comuna   : geo_barrios.py (master por puesto + reasignación por lat/lon)

Join puesto: pcode = dep(2)+mun(3)+zona(2)+puesto(2) = 9 dígitos.
"""
import os, csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import geo_barrios as GEO

ROOT = os.path.join(os.path.dirname(__file__), '..', '..')
BD   = os.path.join(ROOT, 'Bases de datos')
OUT  = os.path.join(BD, 'output_2v')
CSV1V = os.path.join(BD, 'nuevos archivos 1v 2026', 'PRECONTEO_1V_2026_MESA_con_Claudia.csv')
OUTFILE = os.path.join(OUT, 'Resultados_1V_2026_por_barrio_ciudades.xlsx')

g = GEO.build()
CITY, RESOLVE = g.CITY, g.RESOLVE
CITIES = GEO.CITIES

# ---- candidatos: nombre en el CSV -> etiqueta corta (orden se define por votación nacional) ----
CAND_CSV = ['Iván Cepeda', 'Abelardo De La Espriella', 'Paloma Valencia', 'Sergio Fajardo',
            'Santiago Botero', 'Claudia López', 'Mauricio Lizcano', 'Roy Barreras',
            'Gilberto Murillo', 'Carlos Caicedo', 'Miguel Uribe', 'Gustavo Matamoros',
            'Sondra Macollins']
SHORT = {'Iván Cepeda': 'Cepeda', 'Abelardo De La Espriella': 'Abelardo', 'Paloma Valencia': 'Paloma',
         'Sergio Fajardo': 'Fajardo', 'Santiago Botero': 'Botero', 'Claudia López': 'Claudia',
         'Mauricio Lizcano': 'Lizcano', 'Roy Barreras': 'Roy', 'Gilberto Murillo': 'Murillo',
         'Carlos Caicedo': 'Caicedo', 'Miguel Uribe': 'M. Uribe', 'Gustavo Matamoros': 'Matamoros',
         'Sondra Macollins': 'Macollins'}

def ni(v):
    try: return int(v)
    except: return 0

# ---------------- 1) votos 1V por puesto (autoritativo) ----------------
def blankp():
    d = {c: 0 for c in CAND_CSV}; d.update({'bl': 0, 'nu': 0, 'nm': 0, 'total': 0, 'mesas': 0}); return d
V1 = defaultdict(blankp)
NAC = defaultdict(int)
nfilas = 0
with open(CSV1V, encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        cod5 = f"{str(r['cod_departamento']).strip().zfill(2)}{str(r['cod_municipio']).strip().zfill(3)}"
        if cod5 not in CITY: continue
        nfilas += 1
        code = GEO.pcode(cod5, r['zona'], r['puesto'])
        a = V1[code]
        for c in CAND_CSV:
            v = ni(r[c]); a[c] += v; NAC[c] += v
        a['bl'] += ni(r['votos_blanco']); a['nu'] += ni(r['votos_nulos'])
        a['nm'] += ni(r['votos_no_marcados']); a['total'] += ni(r['total_votos_urna']); a['mesas'] += 1
print(f'CSV 1V: filas (mesas) de las 17 ciudades = {nfilas:,}  | puestos con voto = {len(V1):,}')

# orden nacional de candidatos (desc) dentro de las 17 ciudades
CAND = sorted(CAND_CSV, key=lambda c: -NAC[c])
print('orden candidatos (17 ciudades):', ' > '.join(f'{SHORT[c]} {NAC[c]:,}' for c in CAND))

# ---------------- 2) agrega por (ciudad, comuna, barrio) ----------------
def blank():
    d = {c: 0 for c in CAND}
    d.update({'bl': 0, 'nu': 0, 'nm': 0, 'total': 0, 'mesas': 0, 'puestos': 0, 'censo': 0,
              'ccode': 999, 'cdisp': '', 'bdisp': ''})
    return d
AGG = defaultdict(lambda: defaultdict(blank))
sin_geo = defaultdict(int)
for code, v in V1.items():
    ciudad = CITY[code[:5]]
    gg = g.M.get(code) or g.G.get(code)
    censo = gg['censo'] if gg else 0
    if code in RESOLVE:
        ccode, cdisp, barrio = RESOLVE[code]
    else:
        sin_geo[ciudad] += 1
        ccode, cdisp, barrio = 999, 'Sin comuna asignada', 'Sin barrio asignado'
    g.register_canon(ciudad, cdisp)
    key = (GEO.nrm(cdisp), GEO.nrm(barrio))
    t = AGG[ciudad][key]
    for c in CAND: t[c] += v[c]
    t['bl'] += v['bl']; t['nu'] += v['nu']; t['nm'] += v['nm']; t['total'] += v['total']
    t['mesas'] += v['mesas']; t['puestos'] += 1; t['censo'] += censo
    if ccode < t['ccode']: t['ccode'] = ccode
    t['cdisp'] = cdisp; t['bdisp'] = GEO.titlecase(barrio)
for ciudad, rows in AGG.items():
    for a in rows.values():
        a['cdisp'] = g.canon(ciudad, a['cdisp'])

# ---------------- estilos (paleta del proyecto) ----------------
HDR = PatternFill('solid', fgColor='8A1E16'); HDRF = Font(bold=True, color='FFFFFF', size=10)
ZEBRA = PatternFill('solid', fgColor='F4F0E7')
# color del ganador por candidato (los que tienen identidad; resto gris)
CANDCLR = {'Cepeda': 'C0392B', 'Abelardo': '1F47CC', 'Paloma': '1866DF', 'Fajardo': 'EEAA22',
           'Claudia': 'B9A81E', 'Roy': '3D8B3D', 'Botero': '7A7A7A', 'Lizcano': '7A7A7A',
           'Murillo': '3D8B3D', 'Caicedo': '3D8B3D', 'M. Uribe': '1F47CC', 'Matamoros': '7A7A7A',
           'Macollins': '7A7A7A'}
CEPF = PatternFill('solid', fgColor='F7DAD5'); ABEF = PatternFill('solid', fgColor='D7DEF6')
WINFT = Font(color='FFFFFF', bold=True)
thin = Side(style='thin', color='D9D2C5'); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment('center', vertical='center'); LEFT = Alignment('left', vertical='center')

def cols_def():
    c = [('Barrio', 30, 'txt'), ('Comuna / Localidad', 26, 'txt')]
    for cd in CAND: c.append((SHORT[cd], 10, 'int'))
    c += [('Válidos', 11, 'int'), ('Ganador', 12, 'txt'), ('Ganador %', 10, 'pct'),
          ('2º lugar', 12, 'txt'), ('Margen (pp)', 11, 'pp'),
          ('Blanco', 9, 'int'), ('Nulos', 8, 'int'), ('No marcad.', 10, 'int'),
          ('Total votos', 12, 'int'), ('Puestos', 8, 'int'), ('Mesas', 7, 'int')]
    return c
COLS = cols_def()
NC = len(CAND)
GCOL = 2 + NC + 2   # columna 'Ganador' (1-based): Barrio,Comuna,[cands],Válidos,Ganador...

def metricas(a):
    votos = [(c, a[c]) for c in CAND]
    validos = sum(a[c] for c in CAND) + a['bl']
    ordv = sorted(votos, key=lambda x: -x[1])
    gan, gv = ordv[0]
    seg, sv = ordv[1] if len(ordv) > 1 else ('', 0)
    ganp = gv / validos if validos else 0
    margen = (gv - sv) / validos * 100 if validos else 0
    row = [a['bdisp'], a['cdisp']] + [a[c] for c in CAND] + [
        validos, SHORT[gan], ganp, SHORT[seg], margen,
        a['bl'], a['nu'], a['nm'], a['total'], a['puestos'], a['mesas']]
    return row, SHORT[gan]

def cellfmt(c, fmt):
    c.border = BORD; c.alignment = CEN
    if fmt == 'int': c.number_format = '#,##0'
    elif fmt == 'pct': c.number_format = '0.0%'
    elif fmt == 'pp': c.number_format = '+0.0;-0.0'

def pinta(ws, filas):
    for j, (t, w, _) in enumerate(COLS, 1):
        c = ws.cell(1, j, t); c.fill = HDR; c.font = HDRF; c.alignment = CEN; c.border = BORD
        ws.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for agg, total in filas:
        vals, ganshort = metricas(agg)
        for j, (v, (t, w, fmt)) in enumerate(zip(vals, COLS), 1):
            c = ws.cell(r, j, v); cellfmt(c, fmt)
            if j <= 2: c.alignment = LEFT
            elif not total and r % 2 == 0: c.fill = ZEBRA
        # tinte suave para las columnas de Cepeda / Abelardo
        if not total:
            for name, fill in (('Iván Cepeda', CEPF), ('Abelardo De La Espriella', ABEF)):
                if name in CAND:
                    ws.cell(r, 3 + CAND.index(name)).fill = fill
        gc = ws.cell(r, GCOL); clr = CANDCLR.get(ganshort, '7A7A7A')
        gc.fill = PatternFill('solid', fgColor=clr); gc.font = WINFT
        if total:
            for j in range(1, len(COLS) + 1):
                cc = ws.cell(r, j); cc.font = Font(bold=True)
                if j != GCOL: cc.fill = PatternFill('solid', fgColor='ECE6D8')
        r += 1
    ws.freeze_panes = 'C2'; ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{r-1}"

# ---------------- hoja Leeme ----------------
def leeme(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3; ws.column_dimensions['B'].width = 116
    L = [('Resultados Primera Vuelta Presidencial 2026 — por barrio · 17 ciudades', 't'),
         ('13 candidatos · 1ª vuelta 31-may-2026 · una hoja por ciudad', 's'), ('', ''),
         ('Qué hay aquí', 'h'),
         ('Una hoja por ciudad con el voto de 1ª vuelta desagregado por barrio y comuna/localidad.', 'b'),
         ('Hoja "Resumen": totales de las 17 ciudades. Orden de hojas y de candidatos: por votación.', 'b'), ('', ''),
         ('Fuente y método', 'h'),
         ('Votos 1V: preconteo por mesa de la Registraduría (PRECONTEO_1V_2026_MESA_con_Claudia).', 'b'),
         ('Claudia López llegó en 0 por mesa en el preconteo crudo; aquí va RECUPERADA EXACTA por mesa', 'b'),
         ('(su voto estaba dentro del total de la urna; el residual da su cifra oficial sin negativos).', 'b'),
         ('Barrio y comuna de cada puesto: master por puesto del análisis (resuelve Bogotá, que no está', 'b'),
         ('en PUESTOS_GEOREF). Los puestos cuyo barrio concatenaba varios o llegaba como basura ("N/A"),', 'b'),
         ('o sin comuna, se reasignaron por COORDENADAS (lat/lon) al puesto bien identificado más cercano', 'b'),
         ('(regla del proyecto: coordenadas > nombre). Soledad no trae comunas en la fuente: queda "No Aplica".', 'b'), ('', ''),
         ('Columnas', 'h'),
         ('Cada columna de candidato = votos en el barrio. Válidos = suma de candidatos + Blanco.', 'b'),
         ('Ganador = candidato más votado del barrio; Ganador % sobre válidos; 2º lugar = quién quedó detrás;', 'b'),
         ('Margen (pp) = diferencia 1º−2º sobre válidos. Total votos = suma de la urna (incluye nulos y no marcados).', 'b'),
         ('Puestos = puestos de votación del barrio; Mesas = mesas sumadas.', 'b'), ('', ''),
         ('Notas / límites', 'h'),
         ('No se incluye censo ni participación por barrio: la Registraduría reparte el censo por puesto de un', 'b'),
         ('modo que no coincide con dónde vota la gente, así que una participación por barrio saldría sin sentido.', 'b'),
         ('Los votos sí son exactos. El barrio es el del puesto de votación; los puestos especiales (censo,', 'b'),
         ('cárceles) entran con su nombre.', 'b'), ('', ''),
         ('Generado por tools/segunda-vuelta-prec/build_ciudades_barrios_1v_xlsx.py', 'n')]
    r = 1
    for txt, k in L:
        c = ws.cell(r, 2, txt)
        c.font = (Font(bold=True, size=15, color='8A1E16') if k == 't' else
                  Font(size=11, italic=True, color='555555') if k == 's' else
                  Font(bold=True, size=11, color='1F47CC') if k == 'h' else
                  Font(size=9, italic=True, color='888888') if k == 'n' else Font(size=10))
        r += 1

# ---------------- construir workbook ----------------
wb = Workbook()
leeme(wb.active); wb.active.title = 'Leeme'
ws_res = wb.create_sheet('Resumen')

city_tot = []
GRAN = blank(); GRAN['bdisp'] = 'TOTAL 17 CIUDADES'
for c5, ciudad in CITIES:
    rows = AGG.get(ciudad, {})
    items = sorted(rows.values(), key=lambda a: (a['ccode'], a['cdisp'], -sum(a[c] for c in CAND)))
    tot = blank(); tot['bdisp'] = f'TOTAL {ciudad.upper()}'; tot['cdisp'] = ''
    for a in items:
        for c in CAND: tot[c] += a[c]
        for k in ('bl', 'nu', 'nm', 'total', 'mesas', 'puestos', 'censo'): tot[k] += a[k]
    filas = [(a, False) for a in items] + [(tot, True)]
    ws = wb.create_sheet(ciudad[:31])
    pinta(ws, filas)
    city_tot.append((ciudad, len(items), tot))
    for c in CAND: GRAN[c] += tot[c]
    for k in ('bl', 'nu', 'nm', 'total', 'mesas', 'puestos', 'censo'): GRAN[k] += tot[k]
    val = sum(tot[c] for c in CAND) + tot['bl']
    win = max(CAND, key=lambda c: tot[c])
    print(f'  {ciudad:<14} barrios={len(items):<4} ganador={SHORT[win]:<9} '
          f'Cepeda={tot["Iván Cepeda"]:>9,} Abelardo={tot["Abelardo De La Espriella"]:>9,} puestos={tot["puestos"]}')

# ---- hoja Resumen ----
RCOLS = [('Ciudad', 16, 'txt'), ('Barrios', 9, 'int')] + [(SHORT[c], 10, 'int') for c in CAND] + \
        [('Válidos', 11, 'int'), ('Ganador', 12, 'txt'), ('Ganador %', 10, 'pct'),
         ('Margen (pp)', 11, 'pp'), ('Puestos', 8, 'int'), ('Mesas', 7, 'int')]
for j, (t, w, _) in enumerate(RCOLS, 1):
    c = ws_res.cell(1, j, t); c.fill = HDR; c.font = HDRF; c.alignment = CEN; c.border = BORD
    ws_res.column_dimensions[get_column_letter(j)].width = w
RGCOL = 2 + NC + 2
def resrow(r, name, nb, t, bold=False):
    validos = sum(t[c] for c in CAND) + t['bl']
    ordv = sorted(((c, t[c]) for c in CAND), key=lambda x: -x[1])
    gan, gv = ordv[0]; seg, sv = ordv[1]
    vals = [name, nb] + [t[c] for c in CAND] + [
        validos, SHORT[gan], gv / validos if validos else 0,
        (gv - sv) / validos * 100 if validos else 0, t['puestos'], t['mesas']]
    for j, (v, (tt, w, fmt)) in enumerate(zip(vals, RCOLS), 1):
        c = ws_res.cell(r, j, v); cellfmt(c, fmt)
        if j == 1: c.alignment = LEFT
    gc = ws_res.cell(r, RGCOL); gc.fill = PatternFill('solid', fgColor=CANDCLR.get(SHORT[gan], '7A7A7A')); gc.font = WINFT
    if bold:
        for j in range(1, len(RCOLS) + 1):
            cc = ws_res.cell(r, j); cc.font = Font(bold=True)
            if j != RGCOL: cc.fill = PatternFill('solid', fgColor='ECE6D8')
r = 2
for ciudad, nb, tot in city_tot:
    resrow(r, ciudad, nb, tot); r += 1
resrow(r, 'TOTAL 17 CIUDADES', sum(x[1] for x in city_tot), GRAN, bold=True)
ws_res.freeze_panes = 'B2'; ws_res.auto_filter.ref = f"A1:{get_column_letter(len(RCOLS))}{r}"
ws_res.sheet_view.showGridLines = False

wb.save(OUTFILE)
gval = sum(GRAN[c] for c in CAND)
print(f'\nOK -> {OUTFILE}')
print(f'  17 ciudades · {sum(x[1] for x in city_tot):,} filas-barrio · Cepeda {GRAN["Iván Cepeda"]:,} / '
      f'Abelardo {GRAN["Abelardo De La Espriella"]:,} (válidos-cand {gval:,})')
if sin_geo: print('  puestos sin barrio:', dict(sin_geo))
