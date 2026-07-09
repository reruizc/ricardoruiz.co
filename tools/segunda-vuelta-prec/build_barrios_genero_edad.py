#!/usr/bin/env python3
"""
Cuatro Excel por BARRIO (17 ciudades), presidencial 2026, misma estructura que
los otros archivos por barrio (Leeme + Resumen + una hoja por ciudad):

  1) Genero_1V_2026_por_barrio_ciudades.xlsx
       votos de los 13 candidatos + % de mujeres del barrio + % de mujeres de
       cada candidato principal (Cepeda/Abelardo/Paloma/Fajardo).
  2) Genero_2V_2026_por_barrio_ciudades.xlsx
       votos Cepeda/Abelardo + % de mujeres del barrio + % de mujeres de cada
       uno + cuánto varió respecto a la 1ª vuelta (voto y composición de mujeres).
  3) Edad_1V_2026_por_barrio_ciudades.xlsx
       votos de los 13 candidatos + composición etaria (18-30/31-60/61+) del
       barrio y de cada candidato principal.
  4) Edad_2V_2026_por_barrio_ciudades.xlsx
       votos Cepeda/Abelardo + composición etaria del barrio y de cada uno +
       cuánto varió el perfil de edad respecto a la 1ª vuelta.

MÉTODO (declarado, no es conteo — el voto es secreto):
  - % de mujeres / composición etaria DEL BARRIO = proyección demográfica 2026
    del electorado del puesto (raking a votantes reales), agregada por barrio.
  - % de mujeres / edad POR CANDIDATO = se impone el sesgo NACIONAL del candidato
    (de la inferencia ecológica ya calibrada) sobre el resultado real del barrio
    y su composición demográfica, y se reconcilia con el voto observado. NO mide
    el comportamiento propio del barrio (con 1-2 puestos es inidentificable);
    las diferencias entre barrios vienen de su resultado y su demografía.
    Aproximación pedida explícitamente; úsese como lectura, no como dato observado.

Fuentes: geo_barrios.py (barrio/comuna) · PRECONTEO 1V por mesa (con Claudia) ·
master (votos 2V) · genero-modelo-2v.json (fracción mujeres por puesto + gap 2V) ·
mujeres-ei.json (gap de género por candidato 1V) · ei-final.csv + ei-2v-final.csv
(perfil de edad por candidato).
"""
import os, csv, json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import geo_barrios as GEO

ROOT = os.path.join(os.path.dirname(__file__), '..', '..')
BD   = os.path.join(ROOT, 'Bases de datos')
OUT  = os.path.join(BD, 'output_2v')
CSV1V = os.path.join(BD, 'nuevos archivos 1v 2026', 'PRECONTEO_1V_2026_MESA_con_Claudia.csv')
W26   = os.path.join(BD, 'output_edad_1v', 'w26-puesto.csv')
GMODEL = os.path.join(OUT, 'genero-modelo-2v.json')
MUJEI  = os.path.join(BD, 'output_mujeres_1v', 'mujeres-ei.json')
EI1V   = os.path.join(BD, 'output_edad_1v', 'ei-final.csv')
EI2V   = os.path.join(BD, 'output_edad_1v', 'ei-2v-final.csv')

g = GEO.build()
CITY, RESOLVE, CITIES = g.CITY, g.RESOLVE, GEO.CITIES

# ------------------------------------------------------------------ candidatos
CAND_CSV = ['Iván Cepeda', 'Abelardo De La Espriella', 'Paloma Valencia', 'Sergio Fajardo',
            'Santiago Botero', 'Claudia López', 'Mauricio Lizcano', 'Roy Barreras',
            'Gilberto Murillo', 'Carlos Caicedo', 'Miguel Uribe', 'Gustavo Matamoros',
            'Sondra Macollins']
SHORT = {'Iván Cepeda': 'Cepeda', 'Abelardo De La Espriella': 'Abelardo', 'Paloma Valencia': 'Paloma',
         'Sergio Fajardo': 'Fajardo', 'Santiago Botero': 'Botero', 'Claudia López': 'Claudia',
         'Mauricio Lizcano': 'Lizcano', 'Roy Barreras': 'Roy', 'Gilberto Murillo': 'Murillo',
         'Carlos Caicedo': 'Caicedo', 'Miguel Uribe': 'M. Uribe', 'Gustavo Matamoros': 'Matamoros',
         'Sondra Macollins': 'Macollins'}
MAIN = ['Iván Cepeda', 'Abelardo De La Espriella', 'Paloma Valencia', 'Sergio Fajardo']

def ni(v):
    try: return int(v)
    except: return 0

# ------------------------------------------------------------------ modelos de sesgo
# (a) género por candidato 1V (multi): brecha muj-hom en la cuota del candidato.
#     Usamos el modelo ROBUSTO (mesa-EF · escenario_1v de mujeres-refinado), NO la EI
#     por puesto (mujeres-ei.sexo_nacional), cuyo signo está inflado/invertido por
#     falacia ecológica (brecha ~+18 para Cepeda vs ~+0.4 robusto). En 1V el método
#     robusto NO halla diferencias de género relevantes por candidato; el sesgo de
#     género vive en el DUELO de 2V (ver archivo 2V).
MREF = json.load(open(MUJREF := os.path.join(BD, 'output_mujeres_1v', 'mujeres-refinado.json')))['escenario_1v']
GAP_G_1V = {'Iván Cepeda': MREF['cepeda']['brecha'] / 100, 'Abelardo De La Espriella': MREF['abelardo']['brecha'] / 100,
            'Paloma Valencia': MREF['paloma']['brecha'] / 100, 'Sergio Fajardo': MREF['fajardo26']['brecha'] / 100}

# (b) género binario 2V: fracción mujeres por puesto + gap por ciudad/comuna
GMOD = json.load(open(GMODEL))
WF = GMOD['wfrac_puesto']; GAPS = GMOD['gaps']; GAP_NAC = GAPS['__national__']['gap']
BIG4 = set(GMOD['meta']['big4']); FNAC = float(GMOD['meta']['fmuj_nacional'])
def gap2v(ciudad, cdisp):
    if ciudad in BIG4:
        k = f'{ciudad}|{GEO.nrm(cdisp)}'
        if k in GAPS: return GAPS[k]['gap']
    k2 = f'{ciudad}|*'
    return GAPS[k2]['gap'] if k2 in GAPS else GAP_NAC

# (c) edad por candidato: P(vota c | franja) en 5 grupos EI
SHORTREV = {'Cepeda': 'Iván Cepeda', 'Abelardo': 'Abelardo De La Espriella',
            'Paloma': 'Paloma Valencia', 'Fajardo': 'Sergio Fajardo'}
BETA1 = defaultdict(dict)   # nombre-csv -> {grupo5: beta}
for r in csv.DictReader(open(EI1V)):
    if r['year'] != '2026': continue
    c = SHORTREV.get(r['cand'])
    if c: BETA1[c][r['grupo']] = float(r['beta'])
# 2V (Cepeda binario) + 1V cara a cara (para el Δ del archivo 2V)
B2 = defaultdict(dict); B1H = defaultdict(dict)
for r in csv.DictReader(open(EI2V)):
    if r['contest'] == '2V-2026': B2[r['grupo']] = float(r['izq_share'])
    elif r['contest'] == '1V-2026 (cara a cara)': B1H[r['grupo']] = float(r['izq_share'])

EI5 = ['18-25', '26-35', '36-45', '46-60', '61+']
# mapa banda-10 -> grupo EI-5 (para aplicar el beta) y -> banda-3 usuario
B10_EI = ['18-25', '18-25', '26-35', '26-35', '36-45', '36-45', '46-60', '46-60', '46-60', '61+']
B10_3  = ['18-30', '18-30', '18-30', '31-60', '31-60', '31-60', '31-60', '31-60', '31-60', '61+']
G3 = ['18-30', '31-60', '61+']

# ------------------------------------------------------------------ 1) votos 1V por puesto
def blank1(): return {c: 0 for c in CAND_CSV} | {'bl': 0, 'nu': 0, 'nm': 0, 'tot': 0, 'mesas': 0}
V1 = defaultdict(blank1)
with open(CSV1V, encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        cod5 = f"{str(r['cod_departamento']).strip().zfill(2)}{str(r['cod_municipio']).strip().zfill(3)}"
        if cod5 not in CITY: continue
        a = V1[GEO.pcode(cod5, r['zona'], r['puesto'])]
        for c in CAND_CSV: a[c] += ni(r[c])
        a['bl'] += ni(r['votos_blanco']); a['nu'] += ni(r['votos_nulos'])
        a['nm'] += ni(r['votos_no_marcados']); a['tot'] += ni(r['total_votos_urna']); a['mesas'] += 1

# votos 2V + edad(1V, 10 bandas) por puesto, desde master y w26
V2 = {}; MES2 = {}
for p in json.load(open(GEO.MASTER)):
    V2[p['pcode']] = (int(p.get('cep2') or 0), int(p.get('abe2') or 0), int(p.get('mesas2') or 0))
AGE10 = {}
with open(W26) as f:
    for r in csv.DictReader(f):
        pc = r['pcode'].replace('-', '')
        if pc[:5] in CITY: AGE10[pc] = [float(r[f'b{i}']) for i in range(10)]

# ------------------------------------------------------------------ agrega por barrio
def blank():
    d = {c: 0 for c in CAND_CSV}
    d.update({'bl': 0, 'nu': 0, 'nm': 0, 'tot': 0, 'mesas1': 0, 'cep2': 0, 'abe2': 0, 'mesas2': 0,
              'wmuj': 0.0, 'wtot': 0.0, 'age': [0.0] * 10, 'puestos': 0,
              'ccode': 999, 'cdisp': '', 'bdisp': ''})
    return d
AGG = defaultdict(lambda: defaultdict(blank))
codes = set(V1) | set(V2) | set(AGE10)
for code in codes:
    if code[:5] not in CITY: continue
    ciudad = CITY[code[:5]]
    if code in RESOLVE:
        ccode, cdisp, barrio = RESOLVE[code]
    else:
        ccode, cdisp, barrio = 999, 'Sin comuna asignada', 'Sin barrio asignado'
    g.register_canon(ciudad, cdisp)
    t = AGG[ciudad][(GEO.nrm(cdisp), GEO.nrm(barrio))]
    v1 = V1.get(code)
    if v1:
        for c in CAND_CSV: t[c] += v1[c]
        t['bl'] += v1['bl']; t['nu'] += v1['nu']; t['nm'] += v1['nm']; t['tot'] += v1['tot']; t['mesas1'] += v1['mesas']
    v2 = V2.get(code)
    if v2: t['cep2'] += v2[0]; t['abe2'] += v2[1]; t['mesas2'] += v2[2]
    wf = WF.get(code)
    if wf: t['wmuj'] += wf[0]; t['wtot'] += wf[1]
    ag = AGE10.get(code)
    if ag:
        for i in range(10): t['age'][i] += ag[i]
    t['puestos'] += 1
    if ccode < t['ccode']: t['ccode'] = ccode
    t['cdisp'] = cdisp; t['bdisp'] = GEO.titlecase(barrio)
for ciudad, rows in AGG.items():
    for a in rows.values(): a['cdisp'] = g.canon(ciudad, a['cdisp'])

# ------------------------------------------------------------------ cálculos demográficos
def femfrac(a):
    return a['wmuj'] / a['wtot'] if a['wtot'] > 0 else FNAC

def fem_share_of(v_c, base, f, gap):
    """% de mujeres entre los votantes del candidato. gap = brecha muj-hom en su cuota."""
    if base <= 0 or v_c <= 0: return f
    s = v_c / base
    m = min(1.0, max(0.0, s + (1 - f) * gap))   # cuota del cand entre mujeres
    return min(1.0, max(0.0, f * m / s))         # P(mujer | votó al cand)

def barrio_edad3(a):
    tot = sum(a['age']) or 1
    return {'18-30': (a['age'][0] + a['age'][1] + a['age'][2]) / tot,
            '31-60': sum(a['age'][3:9]) / tot,
            '61+':   a['age'][9] / tot}

def cand_edad3(a, beta5):
    """composición 18-30/31-60/61+ de los votantes del candidato, dado su beta P(vota|edad)."""
    w = [a['age'][i] * beta5.get(B10_EI[i], 0.0) for i in range(10)]
    tot = sum(w) or 1
    return {'18-30': (w[0] + w[1] + w[2]) / tot, '31-60': sum(w[3:9]) / tot, '61+': w[9] / tot}

def base1v(a):  # base de válidos 1V (13 cand + blanco) para las cuotas de género
    return sum(a[c] for c in CAND_CSV) + a['bl']

# ================================================================== estilos comunes
HDR = PatternFill('solid', fgColor='8A1E16'); HDRF = Font(bold=True, color='FFFFFF', size=10)
HDRW = PatternFill('solid', fgColor='9B3B6E')   # bloque mujeres (magenta)
HDRH = PatternFill('solid', fgColor='3A4A7A')   # bloque hombres (azul)
HDRE = PatternFill('solid', fgColor='2E5A34')   # bloque edad (verde)
HDRD = PatternFill('solid', fgColor='7A5A00')   # bloque variación (ámbar)
ZEBRA = PatternFill('solid', fgColor='F4F0E7')
MUJF = PatternFill('solid', fgColor='F7D9E6')   # tinte columnas % mujeres
HOMF = PatternFill('solid', fgColor='D7E3F7')   # tinte columnas % hombres
GRCLR = {'18-30': 'F7DAD5', '31-60': 'F1EEDC', '61+': 'D2DBEF'}
CANDCLR = {'Cepeda': 'C0392B', 'Abelardo': '1F47CC', 'Paloma': '1866DF', 'Fajardo': 'EEAA22',
           'Claudia': 'B9A81E', 'Roy': '3D8B3D', 'Botero': '7A7A7A', 'Lizcano': '7A7A7A',
           'Murillo': '3D8B3D', 'Caicedo': '3D8B3D', 'M. Uribe': '1F47CC', 'Matamoros': '7A7A7A',
           'Macollins': '7A7A7A'}
WINFT = Font(color='FFFFFF', bold=True)
thin = Side(style='thin', color='D9D2C5'); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment('center', vertical='center'); LEFT = Alignment('left', vertical='center')

def cellfmt(c, fmt):
    c.border = BORD; c.alignment = CEN
    if fmt == 'int': c.number_format = '#,##0'
    elif fmt == 'pct': c.number_format = '0.0%'
    elif fmt == 'pp':  c.number_format = '+0.0;-0.0'

def hdr_fill(title):
    if title.startswith('% Muj') or title == '% Mujeres': return HDRW
    if title.startswith('% Hom') or title == '% Hombres': return HDRH
    if title.startswith(('18-30', '31-60', '61+')) or ('·' in title and any(b in title for b in G3)): return HDRE
    if title.startswith('Δ'): return HDRD
    return HDR

def write_sheet(ws, COLS, filas, rowfn, gcol=None, ganfn=None):
    for j, (t, w, _) in enumerate(COLS, 1):
        c = ws.cell(1, j, t); c.font = HDRF; c.alignment = CEN; c.border = BORD
        c.fill = hdr_fill(t); ws.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for agg, total in filas:
        vals = rowfn(agg)
        for j, (v, (t, w, fmt)) in enumerate(zip(vals, COLS), 1):
            c = ws.cell(r, j, v); cellfmt(c, fmt)
            if j <= 2: c.alignment = LEFT
            elif not total and r % 2 == 0: c.fill = ZEBRA
            if not total:
                if t.startswith('% Muj') or t == '% Mujeres': c.fill = MUJF
                elif t.startswith('% Hom') or t == '% Hombres': c.fill = HOMF
                elif fmt == 'pct':
                    gr = t.split('·')[-1].strip() if '·' in t else t.split()[-1]
                    if gr in GRCLR: c.fill = PatternFill('solid', fgColor=GRCLR[gr])
        if gcol and ganfn:
            gc = ws.cell(r, gcol); gc.fill = PatternFill('solid', fgColor=CANDCLR.get(ganfn(agg), '7A7A7A')); gc.font = WINFT
        if total:
            for j in range(1, len(COLS) + 1):
                cc = ws.cell(r, j); cc.font = Font(bold=True)
                if not (gcol and j == gcol): cc.fill = PatternFill('solid', fgColor='ECE6D8')
        r += 1
    ws.freeze_panes = 'C2'; ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{r-1}"

def leeme_sheet(ws, lines):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3; ws.column_dimensions['B'].width = 116
    r = 1
    for txt, k in lines:
        c = ws.cell(r, 2, txt)
        c.font = (Font(bold=True, size=15, color='8A1E16') if k == 't' else
                  Font(size=11, italic=True, color='555555') if k == 's' else
                  Font(bold=True, size=11, color='1F47CC') if k == 'h' else
                  Font(size=9, italic=True, color='888888') if k == 'n' else Font(size=10))
        r += 1

def order_items(rows):
    return sorted(rows.values(), key=lambda a: (a['ccode'], a['cdisp'], -base1v(a)))

def add_totals(items, extra_sum):
    tot = blank(); tot['cdisp'] = ''
    for a in items:
        for c in CAND_CSV: tot[c] += a[c]
        for k in extra_sum: tot[k] += a[k]
        for i in range(10): tot['age'][i] += a['age'][i]
    return tot

MET = ['bl', 'nu', 'nm', 'tot', 'mesas1', 'cep2', 'abe2', 'mesas2', 'wmuj', 'wtot', 'puestos']

# ================================================================== FILE 1 · Género 1V
def build_genero_1v():
    COLS = [('Barrio', 30, 'txt'), ('Comuna / Localidad', 24, 'txt')]
    for c in CAND_CSV: COLS.append((SHORT[c], 10, 'int'))
    COLS += [('Válidos', 11, 'int'), ('% Mujeres', 10, 'pct'), ('% Hombres', 10, 'pct')]
    for c in MAIN: COLS += [(f'% Muj {SHORT[c]}', 12, 'pct'), (f'% Hom {SHORT[c]}', 12, 'pct')]
    COLS += [('Puestos', 8, 'int'), ('Mesas', 7, 'int')]

    def rowfn(a):
        f = femfrac(a); base = base1v(a)
        row = [a['bdisp'], a['cdisp']] + [a[c] for c in CAND_CSV] + [base, f, 1 - f]
        for c in MAIN:
            fs = fem_share_of(a[c], base, f, GAP_G_1V[c]); row += [fs, 1 - fs]
        row += [a['puestos'], a['mesas1']]
        return row

    wb = Workbook(); leeme_sheet(wb.active, LEEME_G1); wb.active.title = 'Leeme'
    ws_res = wb.create_sheet('Resumen')
    tots = []
    for c5, ciudad in CITIES:
        items = order_items(AGG.get(ciudad, {}))
        tot = add_totals(items, MET); tot['bdisp'] = f'TOTAL {ciudad.upper()}'
        write_sheet(wb.create_sheet(ciudad[:31]), COLS, [(a, False) for a in items] + [(tot, True)], rowfn)
        tots.append((ciudad, len(items), tot))
    # Resumen
    RC = [('Ciudad', 16, 'txt'), ('Barrios', 8, 'int')] + [(SHORT[c], 10, 'int') for c in CAND_CSV] + \
         [('Válidos', 11, 'int'), ('% Mujeres', 10, 'pct'), ('% Hombres', 10, 'pct')]
    for c in MAIN: RC += [(f'% Muj {SHORT[c]}', 12, 'pct'), (f'% Hom {SHORT[c]}', 12, 'pct')]
    RC += [('Puestos', 8, 'int')]
    gran = add_totals([t for _, _, t in tots], MET); gran['bdisp'] = 'TOTAL 17 CIUDADES'
    def rr(a, name, nb):
        f = femfrac(a); base = base1v(a); row = [name, nb] + [a[c] for c in CAND_CSV] + [base, f, 1 - f]
        for c in MAIN:
            fs = fem_share_of(a[c], base, f, GAP_G_1V[c]); row += [fs, 1 - fs]
        return row + [a['puestos']]
    _resumen(ws_res, RC, [(rr(t, ciu, nb), False) for ciu, nb, t in tots] + [(rr(gran, 'TOTAL 17 CIUDADES', sum(x[1] for x in tots)), True)])
    save(wb, 'Genero_1V_2026_por_barrio_ciudades.xlsx', tots)

# ================================================================== FILE 2 · Género 2V
def build_genero_2v():
    COLS = [('Barrio', 30, 'txt'), ('Comuna / Localidad', 24, 'txt'),
            ('Cepeda 2V', 11, 'int'), ('Abelardo 2V', 12, 'int'), ('Válidos 2V', 11, 'int'),
            ('Cepeda %', 9, 'pct'), ('Abelardo %', 10, 'pct'), ('Ganador', 11, 'txt'),
            ('% Mujeres', 10, 'pct'), ('% Hombres', 10, 'pct'),
            ('% Muj Cepeda', 12, 'pct'), ('% Hom Cepeda', 12, 'pct'),
            ('% Muj Abelardo', 13, 'pct'), ('% Hom Abelardo', 13, 'pct'),
            ('Δ Cepeda voto vs 1V', 16, 'pp'), ('Δ % Muj Cepeda vs 1V', 17, 'pp'),
            ('Puestos', 8, 'int'), ('Mesas', 7, 'int')]
    GCOL = 8

    def h2h(cep, abe): return cep / (cep + abe) if (cep + abe) else 0
    def rowfn(a):
        f = femfrac(a); gap = gap2v(a['_ciudad'], a['cdisp'])
        v = a['cep2'] + a['abe2']
        cp = a['cep2'] / v if v else 0
        s2 = h2h(a['cep2'], a['abe2']); s1 = h2h(a['Iván Cepeda'], a['Abelardo De La Espriella'])
        # composición de mujeres de Cepeda (2-way, mismo gap binario) en 2V y 1V
        def femcep(s):
            if s <= 0: return f
            m = min(1.0, max(0.0, s + (1 - f) * gap)); return min(1.0, max(0.0, f * m / s))
        def femabe(s):  # Abelardo = complemento; su cuota entre mujeres = 1-m_cep
            if s >= 1: return f
            m = min(1.0, max(0.0, s + (1 - f) * gap)); ma = 1 - m
            return min(1.0, max(0.0, f * ma / (1 - s))) if (1 - s) > 0 else f
        mc, ma = femcep(s2), femabe(s2)
        return [a['bdisp'], a['cdisp'], a['cep2'], a['abe2'], v, cp, (1 - cp) if v else 0,
                'Cepeda' if a['cep2'] >= a['abe2'] else 'Abelardo',
                f, 1 - f, mc, 1 - mc, ma, 1 - ma,
                (s2 - s1) * 100, (femcep(s2) - femcep(s1)) * 100,
                a['puestos'], a['mesas2']]

    wb = Workbook(); leeme_sheet(wb.active, LEEME_G2); wb.active.title = 'Leeme'
    ws_res = wb.create_sheet('Resumen'); tots = []
    for c5, ciudad in CITIES:
        rows = AGG.get(ciudad, {})
        for a in rows.values(): a['_ciudad'] = ciudad
        items = sorted(rows.values(), key=lambda a: (a['ccode'], a['cdisp'], -(a['cep2'] + a['abe2'])))
        tot = add_totals(items, MET); tot['bdisp'] = f'TOTAL {ciudad.upper()}'; tot['_ciudad'] = ciudad
        write_sheet(wb.create_sheet(ciudad[:31]), COLS, [(a, False) for a in items] + [(tot, True)], rowfn,
                    gcol=GCOL, ganfn=lambda a: 'Cepeda' if a['cep2'] >= a['abe2'] else 'Abelardo')
        tots.append((ciudad, len(items), tot))
    RC = [('Ciudad', 16, 'txt'), ('Barrios', 8, 'int'), ('Cepeda 2V', 12, 'int'), ('Abelardo 2V', 12, 'int'),
          ('Cepeda %', 9, 'pct'), ('% Mujeres', 10, 'pct'), ('% Hombres', 10, 'pct'),
          ('% Muj Cepeda', 12, 'pct'), ('% Hom Cepeda', 12, 'pct'), ('% Muj Abelardo', 13, 'pct'),
          ('% Hom Abelardo', 13, 'pct'), ('Δ % Muj Cep vs 1V', 16, 'pp'), ('Puestos', 8, 'int')]
    gran = add_totals([t for _, _, t in tots], MET); gran['bdisp'] = 'TOTAL 17 CIUDADES'; gran['_ciudad'] = 'Nacional'
    def rr(a, name, nb):
        f = femfrac(a); gap = GAP_NAC; v = a['cep2'] + a['abe2']
        s2 = a['cep2'] / v if v else 0; s1 = (a['Iván Cepeda'] / (a['Iván Cepeda'] + a['Abelardo De La Espriella'])) if (a['Iván Cepeda'] + a['Abelardo De La Espriella']) else 0
        def fc(s):
            if s <= 0: return f
            m = min(1.0, max(0.0, s + (1 - f) * gap)); return f * m / s
        def fa(s):
            if s >= 1: return f
            m = min(1.0, max(0.0, s + (1 - f) * gap)); return f * (1 - m) / (1 - s)
        return [name, nb, a['cep2'], a['abe2'], s2, f, 1 - f, fc(s2), 1 - fc(s2), fa(s2), 1 - fa(s2),
                (fc(s2) - fc(s1)) * 100, a['puestos']]
    _resumen(ws_res, RC, [(rr(t, ciu, nb), False) for ciu, nb, t in tots] +
             [(rr(gran, 'TOTAL 17 CIUDADES', sum(x[1] for x in tots)), True)],
             gcol=None)
    save(wb, 'Genero_2V_2026_por_barrio_ciudades.xlsx', tots, twoc=True)

# ================================================================== FILE 3 · Edad 1V
def build_edad_1v():
    COLS = [('Barrio', 30, 'txt'), ('Comuna / Localidad', 24, 'txt')]
    for c in CAND_CSV: COLS.append((SHORT[c], 10, 'int'))
    COLS.append(('Válidos', 11, 'int'))
    for gr in G3: COLS.append((f'Barrio · {gr}', 11, 'pct'))
    for c in MAIN:
        for gr in G3: COLS.append((f'{SHORT[c]} · {gr}', 11, 'pct'))
    COLS.append(('Puestos', 8, 'int'))

    def rowfn(a):
        base = base1v(a); bc = barrio_edad3(a)
        row = [a['bdisp'], a['cdisp']] + [a[c] for c in CAND_CSV] + [base]
        row += [bc[gr] for gr in G3]
        for c in MAIN:
            cc = cand_edad3(a, BETA1.get(c, {}))
            row += [cc[gr] for gr in G3]
        row.append(a['puestos'])
        return row

    wb = Workbook(); leeme_sheet(wb.active, LEEME_E1); wb.active.title = 'Leeme'
    ws_res = wb.create_sheet('Resumen'); tots = []
    for c5, ciudad in CITIES:
        items = order_items(AGG.get(ciudad, {}))
        tot = add_totals(items, MET); tot['bdisp'] = f'TOTAL {ciudad.upper()}'
        write_sheet(wb.create_sheet(ciudad[:31]), COLS, [(a, False) for a in items] + [(tot, True)], rowfn)
        tots.append((ciudad, len(items), tot))
    RC = [('Ciudad', 16, 'txt'), ('Barrios', 8, 'int')] + [(SHORT[c], 10, 'int') for c in MAIN] + \
         [('Válidos', 11, 'int')] + [(f'Barrio · {gr}', 11, 'pct') for gr in G3]
    for c in MAIN:
        for gr in G3: RC.append((f'{SHORT[c]} · {gr}', 11, 'pct'))
    RC.append(('Puestos', 8, 'int'))
    gran = add_totals([t for _, _, t in tots], MET); gran['bdisp'] = 'TOTAL 17 CIUDADES'
    def rr(a, name, nb):
        bc = barrio_edad3(a)
        row = [name, nb] + [a[c] for c in MAIN] + [base1v(a)] + [bc[gr] for gr in G3]
        for c in MAIN:
            cc = cand_edad3(a, BETA1.get(c, {})); row += [cc[gr] for gr in G3]
        return row + [a['puestos']]
    _resumen(ws_res, RC, [(rr(t, ciu, nb), False) for ciu, nb, t in tots] +
             [(rr(gran, 'TOTAL 17 CIUDADES', sum(x[1] for x in tots)), True)])
    save(wb, 'Edad_1V_2026_por_barrio_ciudades.xlsx', tots)

# ================================================================== FILE 4 · Edad 2V
def build_edad_2v():
    COLS = [('Barrio', 30, 'txt'), ('Comuna / Localidad', 24, 'txt'),
            ('Cepeda 2V', 11, 'int'), ('Abelardo 2V', 12, 'int'), ('Válidos 2V', 11, 'int'),
            ('Cepeda %', 9, 'pct'), ('Ganador', 11, 'txt')]
    for gr in G3: COLS.append((f'Barrio · {gr}', 11, 'pct'))
    for nm in ('Cepeda', 'Abelardo'):
        for gr in G3: COLS.append((f'{nm} · {gr}', 11, 'pct'))
    COLS += [('Δ Cep 18-30 vs 1V', 15, 'pp'), ('Δ Cep 61+ vs 1V', 14, 'pp'), ('Puestos', 8, 'int')]
    GCOL = 7

    def cep2_edad(a): return cand_edad3(a, B2)
    def abe2_edad(a): return cand_edad3(a, {gr: 1 - B2.get(gr, 0.5) for gr in EI5})
    def cep1_edad(a): return cand_edad3(a, B1H)   # Cepeda 1V cara a cara

    def rowfn(a):
        v = a['cep2'] + a['abe2']; cp = a['cep2'] / v if v else 0
        bc = barrio_edad3(a); ce = cep2_edad(a); ae = abe2_edad(a); c1 = cep1_edad(a)
        row = [a['bdisp'], a['cdisp'], a['cep2'], a['abe2'], v, cp,
               'Cepeda' if a['cep2'] >= a['abe2'] else 'Abelardo']
        row += [bc[gr] for gr in G3] + [ce[gr] for gr in G3] + [ae[gr] for gr in G3]
        row += [(ce['18-30'] - c1['18-30']) * 100, (ce['61+'] - c1['61+']) * 100, a['puestos']]
        return row

    wb = Workbook(); leeme_sheet(wb.active, LEEME_E2); wb.active.title = 'Leeme'
    ws_res = wb.create_sheet('Resumen'); tots = []
    for c5, ciudad in CITIES:
        rows = AGG.get(ciudad, {})
        items = sorted(rows.values(), key=lambda a: (a['ccode'], a['cdisp'], -(a['cep2'] + a['abe2'])))
        tot = add_totals(items, MET); tot['bdisp'] = f'TOTAL {ciudad.upper()}'
        write_sheet(wb.create_sheet(ciudad[:31]), COLS, [(a, False) for a in items] + [(tot, True)], rowfn,
                    gcol=GCOL, ganfn=lambda a: 'Cepeda' if a['cep2'] >= a['abe2'] else 'Abelardo')
        tots.append((ciudad, len(items), tot))
    RC = [('Ciudad', 16, 'txt'), ('Barrios', 8, 'int'), ('Cepeda 2V', 12, 'int'), ('Abelardo 2V', 12, 'int'),
          ('Cepeda %', 9, 'pct')] + [(f'Barrio · {gr}', 11, 'pct') for gr in G3]
    for nm in ('Cepeda', 'Abelardo'):
        for gr in G3: RC.append((f'{nm} · {gr}', 11, 'pct'))
    RC.append(('Puestos', 8, 'int'))
    gran = add_totals([t for _, _, t in tots], MET); gran['bdisp'] = 'TOTAL 17 CIUDADES'
    def rr(a, name, nb):
        v = a['cep2'] + a['abe2']; bc = barrio_edad3(a); ce = cep2_edad(a); ae = abe2_edad(a)
        return [name, nb, a['cep2'], a['abe2'], a['cep2'] / v if v else 0] + [bc[gr] for gr in G3] + \
               [ce[gr] for gr in G3] + [ae[gr] for gr in G3] + [a['puestos']]
    _resumen(ws_res, RC, [(rr(t, ciu, nb), False) for ciu, nb, t in tots] +
             [(rr(gran, 'TOTAL 17 CIUDADES', sum(x[1] for x in tots)), True)])
    save(wb, 'Edad_2V_2026_por_barrio_ciudades.xlsx', tots, twoc=True)

# ------------------------------------------------------------------ Resumen genérico + save
def _resumen(ws, RC, filas, gcol=None):
    for j, (t, w, _) in enumerate(RC, 1):
        c = ws.cell(1, j, t); c.font = HDRF; c.alignment = CEN; c.border = BORD
        c.fill = hdr_fill(t); ws.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for vals, bold in filas:
        for j, (v, (t, w, fmt)) in enumerate(zip(vals, RC), 1):
            c = ws.cell(r, j, v); cellfmt(c, fmt)
            if j == 1: c.alignment = LEFT
            if not bold and (t.startswith('% Muj') or t == '% Mujeres'): c.fill = MUJF
            elif not bold and (t.startswith('% Hom') or t == '% Hombres'): c.fill = HOMF
        if bold:
            for j in range(1, len(RC) + 1):
                cc = ws.cell(r, j); cc.font = Font(bold=True); cc.fill = PatternFill('solid', fgColor='ECE6D8')
        r += 1
    ws.freeze_panes = 'B2'; ws.auto_filter.ref = f"A1:{get_column_letter(len(RC))}{r-1}"; ws.sheet_view.showGridLines = False

def save(wb, fn, tots, twoc=False):
    path = os.path.join(OUT, fn); wb.save(path)
    nb = sum(x[1] for x in tots)
    print(f'OK -> {fn}  ({nb:,} filas-barrio, {len(tots)} ciudades)')

# ------------------------------------------------------------------ textos Leeme
_MET = ('Método (importante): el voto es SECRETO. El % de mujeres / la edad DEL BARRIO es la proyección', 'b')
LEEME_G1 = [
    ('Votación 1ª vuelta + % de mujeres por barrio y por candidato — 17 ciudades', 't'),
    ('Presidencial 2026 · 1ª vuelta 31-may · una hoja por ciudad · filas = barrio', 's'), ('', ''),
    ('Qué hay', 'h'),
    ('Los votos de los 13 candidatos en cada barrio + el % de mujeres del electorado del barrio + el % de', 'b'),
    ('mujeres ENTRE los votantes de cada candidato principal (Cepeda, Abelardo, Paloma, Fajardo).', 'b'), ('', ''),
    ('Método (el voto es secreto — esto es estimación)', 'h'),
    ('% Mujeres del barrio = proyección demográfica 2026 del electorado del puesto (raking a votantes', 'b'),
    ('reales), agregada por barrio. % Muj de cada candidato = se impone el sesgo de género NACIONAL del', 'b'),
    ('candidato (método robusto mesa-EF) sobre el resultado real del barrio y su composición de sexo,', 'b'),
    ('reconciliando con el voto. Barrio/comuna: master + reasignación por coordenadas (coordenadas > nombre).', 'b'), ('', ''),
    ('HALLAZGO IMPORTANTE — en 1ª vuelta NO hay brecha de género relevante por candidato', 'h'),
    ('El método robusto (efectos fijos de mesa dentro del puesto) NO encuentra diferencias de género', 'b'),
    ('apreciables entre candidatos en 1ª vuelta: los votantes de Cepeda, Abelardo, Paloma y Fajardo reflejan', 'b'),
    ('casi la misma proporción de mujeres que su barrio (brechas < 1 pp). OJO: la EI por puesto sugería', 'b'),
    ('brechas enormes (Cepeda ~+18 pp de mujeres), pero ese signo está inflado/invertido por falacia', 'b'),
    ('ecológica y NO se usa. La brecha de género real aparece en el DUELO de 2ª vuelta (archivo 2V): ahí las', 'b'),
    ('mujeres se inclinan ~6 pp hacia Abelardo. Los candidatos menores no se muestran (su voto sí está).', 'b'), ('', ''),
    ('Columnas', 'h'),
    ('13 columnas de candidato = votos. Válidos = candidatos + Blanco. % Mujeres = mujeres del electorado', 'b'),
    ('del barrio. % Muj Cepeda… = de los votantes de ese candidato, qué porción son mujeres (en 1V ≈ el', 'b'),
    ('% Mujeres del barrio, por lo dicho arriba).', 'b'), ('', ''),
    ('Generado por tools/segunda-vuelta-prec/build_barrios_genero_edad.py', 'n')]
LEEME_G2 = [
    ('Votación 2ª vuelta + % de mujeres por barrio y por candidato + variación vs 1ª — 17 ciudades', 't'),
    ('Presidencial 2026 · 2ª vuelta 21-jun · Cepeda vs Abelardo · filas = barrio', 's'), ('', ''),
    ('Qué hay', 'h'),
    ('Votos de Cepeda y Abelardo en 2ª vuelta + % de mujeres del barrio + % de mujeres entre los votantes', 'b'),
    ('de cada uno + cuánto varió respecto a la 1ª vuelta (el voto cara a cara y la composición de mujeres).', 'b'), ('', ''),
    ('Método (el voto es secreto — esto es estimación)', 'h'),
    ('% Mujeres del barrio = proyección demográfica 2026. % Muj Cepeda/Abelardo = se impone el gap de', 'b'),
    ('género del duelo (mujeres ~6 pp menos a la izquierda; gap por ciudad/comuna en las 4 grandes) sobre', 'b'),
    ('el resultado real del barrio y su composición de sexo, reconciliando con el voto. Δ Cepeda voto vs 1V =', 'b'),
    ('cuánto subió/bajó su voto cara a cara Cepeda-vs-Abelardo entre 1ª y 2ª vuelta (en pp). Δ % Muj Cepeda', 'b'),
    ('vs 1V = cuánto cambió la porción de mujeres entre sus votantes (con el mismo gap binario en ambas', 'b'),
    ('vueltas, la variación viene del cambio de su voto). Aproximación, no dato observado.', 'b'), ('', ''),
    ('Generado por tools/segunda-vuelta-prec/build_barrios_genero_edad.py', 'n')]
LEEME_E1 = [
    ('Votación 1ª vuelta + composición etaria por barrio y por candidato — 17 ciudades', 't'),
    ('Presidencial 2026 · 1ª vuelta 31-may · franjas 18-30 / 31-60 / 61+ · filas = barrio', 's'), ('', ''),
    ('Qué hay', 'h'),
    ('Votos de los 13 candidatos + composición por edad del electorado del barrio + composición por edad', 'b'),
    ('de los votantes de cada candidato principal (Cepeda, Abelardo, Paloma, Fajardo).', 'b'), ('', ''),
    ('Método (el voto es secreto — esto es estimación)', 'h'),
    ('Edad del barrio = proyección demográfica 2026 del electorado del puesto (perfil 2022 × envejecimiento', 'b'),
    ('DANE × raking a votantes reales), agregada por barrio, en 3 franjas: 18-30, 31-60, 61+. Edad de cada', 'b'),
    ('candidato = se aplica el perfil de edad NACIONAL del candidato (inferencia ecológica: p.ej. Cepeda', 'b'),
    ('joven, Abelardo mayor, Paloma concentrada en 46+) sobre la estructura de edad del barrio, reconciliando', 'b'),
    ('con su voto. NO mide el comportamiento propio del barrio. Los candidatos menores comparten un perfil', 'b'),
    ('residual y no se muestran por separado (su voto sí está en las columnas). Es una lectura, no un conteo.', 'b'), ('', ''),
    ('Columnas', 'h'),
    ('Barrio · 18-30 … = porción del electorado del barrio en esa franja. Cepeda · 18-30 … = de los votantes', 'b'),
    ('de Cepeda, qué porción está en esa franja (suman 100% por candidato).', 'b'), ('', ''),
    ('Generado por tools/segunda-vuelta-prec/build_barrios_genero_edad.py', 'n')]
LEEME_E2 = [
    ('Votación 2ª vuelta + composición etaria por barrio y por candidato + variación vs 1ª — 17 ciudades', 't'),
    ('Presidencial 2026 · 2ª vuelta 21-jun · franjas 18-30 / 31-60 / 61+ · filas = barrio', 's'), ('', ''),
    ('Qué hay', 'h'),
    ('Votos de Cepeda y Abelardo + composición por edad del barrio + composición por edad de los votantes', 'b'),
    ('de cada uno + cuánto varió el perfil de Cepeda respecto a la 1ª vuelta.', 'b'), ('', ''),
    ('Método (el voto es secreto — esto es estimación)', 'h'),
    ('Edad del barrio = proyección demográfica 2026 en 3 franjas. Edad de Cepeda/Abelardo = perfil de edad', 'b'),
    ('del duelo de 2ª vuelta (inferencia ecológica del balotaje) sobre la estructura de edad del barrio. Δ Cep', 'b'),
    ('18-30 / 61+ vs 1V = cuánto cambió la porción de jóvenes / mayores entre los votantes de Cepeda al pasar', 'b'),
    ('de 1ª a 2ª vuelta (absorbió centro en 2V). NO mide comportamiento propio del barrio. Aproximación.', 'b'), ('', ''),
    ('Generado por tools/segunda-vuelta-prec/build_barrios_genero_edad.py', 'n')]

# ------------------------------------------------------------------ run
build_genero_1v()
build_genero_2v()
build_edad_1v()
build_edad_2v()
print('\nlisto · 4 archivos en Bases de datos/output_2v/')
