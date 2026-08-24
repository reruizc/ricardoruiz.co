#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Mesas del preconteo 2V donde UN candidato obtuvo el 100% de la votacion valida
(el rival = 0 y el candidato > 0), con:
  - datos del preconteo de 2V por mesa,
  - % que saco ese candidato en 1V en esa misma mesa,
  - participacion estimada en 1V y en 2V.

Definicion de "100%": de la votacion valida del balotaje (Cepeda + Abelardo),
el 100% fue para el candidato objetivo, es decir el rival = 0 y el objetivo > 0.

Cambiar de candidato: editar el bloque SWEEP/RIVAL de abajo.

Fuentes:
  2V por mesa : Bases de datos/output_2v/detalle_nacional_presidencia_mesas.xlsx
  1V por mesa : Bases de datos/nuevos archivos 1v 2026/PRECONTEO_1V_2026_MESA_con_Claudia.csv
  Censo       : Bases de datos/COMUNAS_DATA.csv  (censo y # de mesas por puesto)
  Nombres     : Bases de datos/test-presidencial/divipola.json  (fallback dep/mun)
"""
import csv, re, json, unicodedata, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ===================== candidato que barre el 100% =====================
# SWEEP = candidato que obtiene el 100% ; RIVAL = el que debe quedar en 0.
SWEEP_LBL, SWEEP_1V, SWEEP_IS_CEP = 'Cepeda',   'Iván Cepeda',                True
RIVAL_LBL, RIVAL_1V                = 'Abelardo', 'Abelardo De La Espriella'
SWEEP_COL2V, RIVAL_COL2V = 10, 11   # indices en el xlsx 2V: 10=Cepeda, 11=Abelardo
OUT_NAME = f'Mesas_{SWEEP_LBL}_100pct_2V_vs_1V.xlsx'
# ======================================================================

BASE = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', '..', 'Bases de datos'))
XLSX_2V = os.path.join(BASE, 'output_2v', 'detalle_nacional_presidencia_mesas.xlsx')
CSV_1V  = os.path.join(BASE, 'nuevos archivos 1v 2026', 'PRECONTEO_1V_2026_MESA_con_Claudia.csv')
COMUNAS = os.path.join(BASE, 'COMUNAS_DATA.csv')
DIVIPOLA= os.path.join(BASE, 'test-presidencial', 'divipola.json')
OUT     = os.path.join(BASE, 'output_2v', OUT_NAME)

def norm(s):
    s = unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode().upper().strip()
    return re.sub(r'\s+', ' ', s)

def pi(v):
    try: return int(v)
    except (TypeError, ValueError): return None

def fmt(v, w): return f'{v:0{w}d}' if isinstance(v, int) else str(v)

# ---------- divipola: nombre -> codigo (fallback cuando falta el hipervinculo) ----------
dv = json.load(open(DIVIPOLA, encoding='utf-8'))
dep_by_name, mun_by = {}, {}
for d in dv['deptos']:
    dep_by_name[norm(d['nombre'])] = d['cod']
    for m in d['muns']:
        mun_by[(d['cod'], norm(m['nombre']))] = m['cod']

# ---------- censo por puesto: (dep,mun,zona,puesto) -> (censo, n_mesas, depn, munn) ----------
censo = {}
with open(COMUNAS, encoding='utf-8-sig') as f:
    for row in csv.DictReader(f, delimiter=';'):
        try:
            k = (int(row['dd']), int(row['mm']), int(row['zz']), int(row['pp']))
        except (ValueError, KeyError):
            continue
        censo[k] = (int(row['total']), int(row['mesas']), norm(row['departamento']), norm(row['municipio']))

# ---------- 1V por mesa: (dep,mun,zona,puesto,mesa) -> (cep1, abe1, validos1, urna1) ----------
CAND = ['Iván Cepeda','Santiago Botero','Abelardo De La Espriella','Mauricio Lizcano',
        'Miguel Uribe','Sondra Macollins','Roy Barreras','Carlos Caicedo','Gustavo Matamoros',
        'Paloma Valencia','Sergio Fajardo','Gilberto Murillo','Claudia López']
oneV = {}
pu_tot1 = {}     # (dep,mun,zona,puesto) -> votantes 1V acumulados en el puesto
with open(CSV_1V, encoding='utf-8-sig') as f:
    r = csv.reader(f); hdr = next(r); idx = {h: i for i, h in enumerate(hdr)}
    ic, ia, it = idx['Iván Cepeda'], idx['Abelardo De La Espriella'], idx['total_votos_urna']
    cand_idx = [idx[c] for c in CAND]
    for row in r:
        try:
            key = (int(row[0]), int(row[1]), int(row[2]), int(row[3]), int(row[4]))
        except ValueError:
            continue
        val = sum(int(row[j]) for j in cand_idx)
        urna = int(row[it])
        oneV[key] = (int(row[ic]), int(row[ia]), val, urna)
        pu_tot1[key[:4]] = pu_tot1.get(key[:4], 0) + urna

# ---------- 2V por mesa ----------
rx = re.compile(r'/pdf/(\d+)/(\d+)/(\d+)/(\d+)/(\d+)/')
wb = openpyxl.load_workbook(XLSX_2V, read_only=True)
ws = wb['Presidencia_Mesas']

def parse_codes(row):
    m = rx.search(str(row[24] or ''))
    if m:
        g = tuple(int(x) for x in m.groups())
        return g[0], g[1], g[2], g[3], g[4], 'E14'
    dc = dep_by_name.get(norm(row[1])); mc = mun_by.get((dc, norm(row[2]))) if dc else None
    if not (dc and mc):
        return None, None, None, None, None, 'FAIL'
    return int(dc), int(mc), pi(row[3]), pi(row[4]), pi(row[6]), 'nombre'

# Pass unico: acumular voto por PUESTO (todas las mesas) y recoger las mesas-100.
pu_rival2, pu_sweep2, pu_tot2, pu_n = {}, {}, {}, {}   # (dep,mun,zona,pto) -> rival/sweep/votantes 2V / nº mesas
hits = []
for row in ws.iter_rows(min_row=2, values_only=True):
    sweep2 = row[SWEEP_COL2V] or 0
    rival2 = row[RIVAL_COL2V] or 0
    tot2   = (row[7] or 0) + (row[8] or 0) + (row[9] or 0) + sweep2 + rival2
    dep, mun, zona, pto, mesa, src = parse_codes(row)
    if None not in (dep, mun, zona, pto):
        pk = (dep, mun, zona, pto)
        pu_rival2[pk] = pu_rival2.get(pk, 0) + rival2
        pu_sweep2[pk] = pu_sweep2.get(pk, 0) + sweep2
        pu_tot2[pk] = pu_tot2.get(pk, 0) + tot2
        pu_n[pk] = pu_n.get(pk, 0) + 1
    if rival2 == 0 and sweep2 > 0:
        hits.append((row, sweep2, rival2, dep, mun, zona, pto, mesa, src))

recs = []
warn_censo_mismatch = 0
for row, sweep2, rival2, dep, mun, zona, pto, mesa, src in hits:
    if dep is None:
        print('  [FAIL] sin codigo:', row[1], row[2], row[3], row[4], row[6]); continue
    dep_n, mun_n, pto_name = row[1], row[2], row[5]
    blanco, nomarc, nulos = row[7] or 0, row[8] or 0, row[9] or 0

    joinable = None not in (dep, mun, zona, pto, mesa)
    key = (dep, mun, zona, pto, mesa) if joinable else None
    one = oneV.get(key) if key else None          # (cep1, abe1, val1, urna1) o None
    pk = (dep, mun, zona, pto) if None not in (dep, mun, zona, pto) else None
    cz = censo.get(pk) if pk else None
    if cz and (cz[2] != norm(dep_n) or cz[3] != norm(mun_n)):
        warn_censo_mismatch += 1; cz = None

    zona_d = zona if zona is not None else row[3]
    pto_d  = pto  if pto  is not None else row[4]
    mesa_d = mesa if mesa is not None else row[6]

    tot2v = blanco + nomarc + nulos + sweep2 + rival2
    cep1v, abe1v, val1v, urna1v = (one if one else (None, None, None, None))
    sweep1v = (cep1v if SWEEP_IS_CEP else abe1v)
    rival1v = (abe1v if SWEEP_IS_CEP else cep1v)

    censo_pto = cz[0] if cz else None
    n_mesas   = cz[1] if cz else None
    # Participacion EXACTA a nivel PUESTO (votantes del puesto / censo del puesto).
    # Es la unidad mas fina con censo oficial; la turnout propia de la mesa va aparte
    # como conteo de votantes (1V/2V Votantes urna).
    vot_pto_2v = pu_tot2.get(pk, 0) if pk else None
    vot_pto_1v = pu_tot1.get(pk) if pk else None

    nota = ''
    if dep == 88: nota = 'Exterior / consulado'
    elif zona == 90: nota = 'Puesto censo (90)'
    elif zona == 98: nota = 'Carcel (98)'
    elif zona == 99: nota = 'Zona rural (99)'

    # Alerta = patron "Armenia": dado el peso del RIVAL en TODO el puesto en 2V, en esta mesa
    # se esperarian varios votos suyos, pero hay exactamente 0 -> probable error de
    # transcripcion del E14. Se usa el valor ESPERADO (no la suma bruta) para no marcar
    # puestos grandes con 1-2 votos del rival por mesa, ni el voto en bloque de resguardos
    # (Jambalo, Toribio): ahi el rival queda ~0 en TODO el puesto y el 100% es genuino.
    pu_val = (pu_sweep2.get(pk, 0) + pu_rival2.get(pk, 0)) if pk else 0
    rival_share = (pu_rival2.get(pk, 0) / pu_val) if pu_val else 0
    exp_rival = rival_share * sweep2          # votos del rival que cabria esperar en esta mesa
    if exp_rival >= 5:
        alerta = (f'Posible error: {RIVAL_LBL} tuvo {rival_share*100:.0f}% en este puesto en 2V; '
                  f'se esperarían ~{round(exp_rival)} votos suyos aquí pero hay 0')
        if rival1v and rival1v >= 5:
            alerta += f' ({rival1v} en 1V)'
    else:
        alerta = ''

    part1 = (vot_pto_1v / censo_pto if (censo_pto and vot_pto_1v is not None) else None)
    part2 = (vot_pto_2v / censo_pto if (censo_pto and vot_pto_2v is not None) else None)
    # puestos rurales diminutos: el censo oficial (a veces 1-5 inscritos) es menor que los
    # votos depositados -> participacion >100%. Es dato real de la fuente, no error de calculo.
    if (part2 is not None and part2 > 1.0) or (part1 is not None and part1 > 1.0):
        nota = (nota + ' · ' if nota else '') + 'censo oficial < votos'

    recs.append(dict(
        dep_n=dep_n, mun_n=mun_n, dep=dep, mun=mun, zona=zona_d, pto=pto_d,
        pto_name=pto_name, mesa=mesa_d,
        llave=f'{dep:02d}-{mun:03d}-{fmt(zona_d,2)}-{fmt(pto_d,2)}-{fmt(mesa_d,3)}',
        s2=sweep2, r2=rival2, bl2=blanco, nu2=nulos, nm2=nomarc, tot2=tot2v,
        s2_val=sweep2/(sweep2+rival2), s2_urna=(sweep2/tot2v if tot2v else None),
        s1=sweep1v, r1=rival1v, val1=val1v, urna1=urna1v,
        s1_val=(sweep1v/val1v if val1v else None),
        s1_urna=(sweep1v/urna1v if urna1v else None),
        censo_pto=censo_pto, n_mesas=n_mesas, part1=part1, part2=part2,
        nota=nota, alerta=alerta, src=src,
    ))

recs.sort(key=lambda d: (-d['s2'], d['dep_n'], d['mun_n']))
print(f'Mesas {SWEEP_LBL}-100: {len(recs)}  |  con 1V: {sum(1 for r in recs if r["s1"] is not None)}'
      f'  |  con censo: {sum(1 for r in recs if r["censo_pto"])}  |  colisiones censo descartadas: {warn_censo_mismatch}')

# ---------- escribir xlsx ----------
wbo = openpyxl.Workbook(); ws = wbo.active; ws.title = f'Mesas {SWEEP_LBL} 100%'
COLS = [
    ('Departamento',              'dep_n',     22, None),
    ('Municipio',                 'mun_n',     20, None),
    ('Cód. dep',                  'dep',        7, '00'),
    ('Cód. mun',                  'mun',        7, '000'),
    ('Zona',                      'zona',       6, '00'),
    ('Cód. puesto',               'pto',        8, '00'),
    ('Puesto',                    'pto_name',  26, None),
    ('Mesa',                      'mesa',       6, '0'),
    ('Llave',                     'llave',     16, None),
    (f'2V {RIVAL_LBL}',           'r2',         9, '#,##0'),
    (f'2V {SWEEP_LBL}',           's2',        10, '#,##0'),
    ('2V Blanco',                 'bl2',        8, '#,##0'),
    ('2V Nulos',                  'nu2',        8, '#,##0'),
    ('2V No marc.',               'nm2',        9, '#,##0'),
    ('2V Votantes (urna)',        'tot2',      13, '#,##0'),
    (f'2V {SWEEP_LBL} % válid.',  's2_val',    12, '0.0%'),
    (f'2V {SWEEP_LBL} % urna',    's2_urna',   12, '0.0%'),
    (f'1V {SWEEP_LBL} votos',     's1',        11, '#,##0'),
    (f'1V {RIVAL_LBL} votos',     'r1',        11, '#,##0'),
    ('1V Válidos',                'val1',       9, '#,##0'),
    ('1V Votantes (urna)',        'urna1',     13, '#,##0'),
    (f'1V {SWEEP_LBL} % válid.',  's1_val',    12, '0.0%'),
    (f'1V {SWEEP_LBL} % urna',    's1_urna',   12, '0.0%'),
    ('Censo puesto',              'censo_pto', 11, '#,##0'),
    ('Mesas en puesto',           'n_mesas',   10, '#,##0'),
    ('Participación 1V (puesto)', 'part1',     15, '0.0%'),
    ('Participación 2V (puesto)', 'part2',     15, '0.0%'),
    ('Nota',                      'nota',      20, None),
    ('Alerta',                    'alerta',    34, None),
]
hdr_fill = PatternFill('solid', fgColor='8A1E16'); hdr_font = Font(bold=True, color='FFFFFF', size=10)
thin = Side(style='thin', color='D8D2C4'); border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c, (title, _, _, _) in enumerate(COLS, 1):
    cell = ws.cell(1, c, title); cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cell.border = border
for ri, rec in enumerate(recs, 2):
    for c, (_, field, _, fmt) in enumerate(COLS, 1):
        cell = ws.cell(ri, c, rec[field])
        if fmt: cell.number_format = fmt
        cell.border = border
        if ri % 2 == 0: cell.fill = PatternFill('solid', fgColor='F4F0E7')
for c, (_, _, w, _) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = 'C2'; ws.row_dimensions[1].height = 30
last = f'{get_column_letter(len(COLS))}{len(recs)+1}'
tbl = Table(displayName='Mesas100', ref=f'A1:{last}')
tbl.tableStyleInfo = TableStyleInfo(name='TableStyleLight1', showRowStripes=False)
ws.add_table(tbl)

# ejemplo dinamico de alerta (la de mayor votacion marcada)
flag = next((r for r in recs if r['alerta']), None)
ej = ('' if not flag else
      f'Ejemplo: {flag["mun_n"].title()} ({flag["dep_n"].title()}) mesa {flag["mesa"]} — '
      f'{flag["tot2"]} votantes en 2V con {RIVAL_LBL} en 0, pero {flag["r1"]} votos a {RIVAL_LBL} en 1V.')

wm = wbo.create_sheet('Metodología')
meth = [
    (f'Mesas donde {SWEEP_LBL} obtuvo el 100% de la votación válida — Balotaje 2026', True),
    ('', False),
    ('Criterio de selección', True),
    (f'Mesa del preconteo de 2ª vuelta en la que {RIVAL_LBL} obtuvo 0 votos y {SWEEP_LBL} obtuvo', False),
    (f'más de 0. Es decir, de la votación válida (Cepeda + Abelardo) el 100% fue para {SWEEP_LBL}.', False),
    ('Los votos en blanco, nulos y no marcados NO cuentan como votación a un candidato; por eso', False),
    ('"% válid." puede ser 100% aunque "% urna" sea menor.', False),
    ('', False),
    (f'Total de mesas que cumplen el criterio: {len(recs)}.', False),
    ('', False),
    ('Fuentes', True),
    ('• 2V por mesa: detalle_nacional_presidencia_mesas.xlsx (preconteo Registraduría, columnas Pre_*).', False),
    ('  Solo Cepeda y Abelardo vienen por mesa. Los códigos dep/mun/zona/puesto/mesa se extraen', False),
    ('  del hipervínculo al E14; cuando falta, se resuelven por nombre (divipola).', False),
    ('• 1V por mesa: PRECONTEO_1V_2026_MESA_con_Claudia.csv (preconteo Registraduría 1ª vuelta).', False),
    ('• Censo y nº de mesas por puesto: COMUNAS_DATA.csv (Divipole). Misma codificación que el', False),
    ('  preconteo (Nariño=23, Risaralda=24, Consulados=88), verificada antes del cruce.', False),
    ('', False),
    ('Definición de las columnas de %', True),
    (f'• 2V/1V {SWEEP_LBL} % válid. = {SWEEP_LBL} / votos a candidatos (en 2V, Cepeda + Abelardo).', False),
    (f'• 2V/1V {SWEEP_LBL} % urna  = {SWEEP_LBL} / total depositado en la urna (incluye blanco/nulo/no marcado).', False),
    ('', False),
    ('Participación (importante)', True),
    ('El censo electoral oficial está disponible por PUESTO, no por mesa. Por eso la PARTICIPACIÓN se', False),
    ('calcula a nivel de puesto (exacta), y la "turnout" propia de cada mesa se da aparte como conteo', False),
    ('de votantes (columnas "1V/2V Votantes (urna)"):', False),
    ('   Participación 1V (puesto) = votantes 1V del puesto / censo del puesto', False),
    ('   Participación 2V (puesto) = votantes 2V del puesto / censo del puesto', False),
    ('Se eligió el puesto (y no una estimación por mesa = censo/nº mesas) porque las mesas no tienen', False),
    ('censo propio y esa estimación daba valores imposibles (>100%) en puestos con mesas desiguales.', False),
    ('La participación a nivel puesto sí es exacta y además es robusta a la renumeración de mesas', False),
    ('entre 1V y 2V. Mesas del exterior y puestos sin censo en COMUNAS_DATA quedan con participación vacía.', False),
    ('En unos pocos puestos rurales muy pequeños la participación supera 100% porque el censo oficial', False),
    ('(a veces de 1 a 5 inscritos) es menor que los votos depositados; es un dato real de la fuente, no', False),
    ('un error de cálculo. Esas filas quedan marcadas con "censo oficial < votos" en la columna Nota.', False),
    ('', False),
    ('Columna "Alerta" — posibles errores de preconteo', True),
    (f'Se marca una mesa cuando {RIVAL_LBL} sumó votos en las OTRAS mesas del mismo puesto en 2V', False),
    ('pero figura con 0 justo en esta mesa. Ese contraste dentro del puesto es la huella típica de un', False),
    ('error de transcripción del E14 (no se cargó la cifra del rival) y debería corregirse con el', False),
    ('escrutinio oficial.', False),
    (f'NO se marca el voto en bloque genuino: en resguardos indígenas y zonas donde TODO el puesto vota', False),
    (f'casi unánime por {SWEEP_LBL} (p. ej. Jambaló o Toribío en Cauca), {RIVAL_LBL} queda en ~0 en todas', False),
    ('las mesas del puesto y ese 100% es real, no un error.', False),
    (ej, False),
    ('', False),
    ('Nota: preconteo, no escrutinio oficial. La columna "Nota" marca exterior y zonas especiales (90/98/99).', False),
]
for ri, (txt, bold) in enumerate(meth, 1):
    cell = wm.cell(ri, 1, txt); cell.font = Font(bold=bold, size=12 if (bold and ri == 1) else 10)
wm.column_dimensions['A'].width = 105
wbo.save(OUT)
print('OK ->', OUT)

ext = sum(1 for r in recs if r['dep'] == 88)
print(f'  exterior/consulado: {ext}  |  territorio nacional: {len(recs)-ext}  |  con alerta: {sum(1 for r in recs if r["alerta"])}')
print(f'  Top 6 por votos de {SWEEP_LBL} en 2V:')
for r in recs[:6]:
    p1 = f'{r["s1_val"]*100:.0f}%' if r['s1_val'] is not None else 's/d'
    print(f'    {r["dep_n"][:18]:18} {r["mun_n"][:16]:16} mesa {r["mesa"]:>3}  {SWEEP_LBL}2V={r["s2"]:>3}  {SWEEP_LBL}1V%={p1}  {("["+r["alerta"]+"]") if r["alerta"] else ""}')
