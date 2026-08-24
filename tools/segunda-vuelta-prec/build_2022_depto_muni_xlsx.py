#!/usr/bin/env python3
"""
Resultados presidenciales 2022 (1ra y 2da vuelta) por DEPARTAMENTO y MUNICIPIO -> un Excel.
4 hojas: 1V Departamento · 1V Municipio · 2V Departamento · 2V Municipio (+ Leeme).

Fuente: GCS de la Registraduría por mesa (FINAL SUBIDA GCS/GCS_2022PRES1V.csv y 2V).
Nombres depto/municipio: test-presidencial/divipola.json (códigos electorales).
"""
import os, csv, json, unicodedata
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(__file__), '..', '..')
BD   = os.path.join(ROOT, 'Bases de datos')
GCS  = os.path.join(BD, 'FINAL SUBIDA GCS')
OUT  = os.path.join(BD, 'output_historicos')
OUTFILE = os.path.join(OUT, 'Resultados_Presidencial_2022_depto_municipio.xlsx')

def nrm(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s or '').upper().strip())
                   if unicodedata.category(c) != 'Mn')
def zf(v, n):
    v = str(v).strip()
    return v.zfill(n)

# ---- nombres ----
dv = json.load(open(os.path.join(BD, 'test-presidencial', 'divipola.json')))
# nombres canonicos de depto (divipola los trae sin tildes y con "Norte De San" truncado)
CANON_DEP = {'01':'Antioquia','03':'Atlántico','05':'Bolívar','07':'Boyacá','09':'Caldas','11':'Cauca',
 '12':'Cesar','13':'Córdoba','15':'Cundinamarca','16':'Bogotá D.C.','17':'Chocó','19':'Huila',
 '21':'Magdalena','23':'Nariño','24':'Risaralda','25':'Norte de Santander','26':'Quindío',
 '27':'Santander','28':'Sucre','29':'Tolima','31':'Valle del Cauca','40':'Arauca','44':'Caquetá',
 '46':'Casanare','48':'La Guajira','50':'Guainía','52':'Meta','54':'Guaviare',
 '56':'San Andrés y Providencia','60':'Amazonas','64':'Putumayo','68':'Vaupés','72':'Vichada',
 '88':'Consulados (exterior)'}
DEPNAME = {d['cod']: CANON_DEP.get(d['cod'], d['nombre']) for d in dv['deptos']}
MUNNAME = {}
for d in dv['deptos']:
    for m in d['muns']:
        MUNNAME[(d['cod'], m['cod'])] = m['nombre']

# ---- candidatos (orden por votación nacional) ----
ORDER1 = [('petro', 'Gustavo Petro'), ('rodolfo', 'Rodolfo Hernández'), ('fico', 'Federico Gutiérrez'),
          ('fajardo', 'Sergio Fajardo'), ('jmr', 'John Milton Rodríguez'), ('egomez', 'Enrique Gómez'),
          ('ingrid', 'Íngrid Betancourt'), ('lperez', 'Luis Pérez')]
ORDER2 = [('petro', 'Gustavo Petro'), ('rodolfo', 'Rodolfo Hernández')]
PCT1 = ['petro', 'rodolfo', 'fico', 'fajardo']   # % que se muestran en 1V
CANDCOLOR = {'petro': 'C0392B', 'rodolfo': 'E67E22', 'fico': '1F47CC', 'fajardo': '16A34A',
             'jmr': '7A6000', 'egomez': '7A6000', 'ingrid': '7A6000', 'lperez': '7A6000'}

def cand1v(des):
    d = nrm(des)
    if 'PETRO' in d: return 'petro'
    if 'RODOLFO' in d: return 'rodolfo'
    if 'FEDERICO' in d: return 'fico'
    if 'FAJARDO' in d: return 'fajardo'
    if 'JOHN' in d or 'MILTON' in d: return 'jmr'
    if 'ENRIQUE GOMEZ' in d: return 'egomez'
    if 'BETANCOURT' in d: return 'ingrid'
    if 'LUIS PEREZ' in d: return 'lperez'
    return None
def cand2v(des):
    d = nrm(des)
    if 'PETRO' in d: return 'petro'
    if 'RODOLFO' in d: return 'rodolfo'
    return None

def aggregate(fname, candfn):
    """-> (DEP, MUN). Cada uno: key -> dict(cand->votos, blanco, nulos, nomarc)."""
    DEP = defaultdict(lambda: defaultdict(int)); MUN = defaultdict(lambda: defaultdict(int))
    n = 0
    with open(os.path.join(GCS, fname), encoding='utf-8-sig') as f:
        for row in csv.DictReader(f, delimiter=';'):
            n += 1
            dep = zf(row['COD_DDE'], 2); mun = zf(row['COD_MME'], 3)
            v = int(row['NUM_VOT'] or 0); cc = str(row['COD_CAN']).strip()
            if cc == '996': key = 'blanco'
            elif cc == '997': key = 'nulos'
            elif cc in ('998', '999'): key = 'nomarc'
            else:
                key = candfn(row['DES_CAN'])
                if not key: continue
            DEP[dep][key] += v; MUN[(dep, mun)][key] += v
    return DEP, MUN, n

# ------------------------------ estilos ------------------------------
HDR = PatternFill('solid', fgColor='8A1E16'); HDRF = Font(bold=True, color='FFFFFF', size=10)
ZEBRA = PatternFill('solid', fgColor='F4F0E7')
thin = Side(style='thin', color='D9D2C5'); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment('center', vertical='center'); LEFT = Alignment('left', vertical='center')
WINFT = Font(color='FFFFFF', bold=True)

def cellfmt(c, fmt):
    c.border = BORD; c.alignment = CEN
    if fmt == 'int': c.number_format = '#,##0'
    elif fmt == 'pct': c.number_format = '0.0%'
    elif fmt == 'pp': c.number_format = '+0.0;-0.0'

def winner(agg, order):
    best = None; bv = -1
    for k, _ in order:
        if agg.get(k, 0) > bv: bv = agg.get(k, 0); best = k
    return best

def write_sheet(ws, rows_data, order, level, vuelta):
    """rows_data: lista de (label_cols[], agg). level: 'dep'|'mun'. vuelta: 1|2."""
    pre = [('Departamento', 22, 'txt')] + ([('Municipio', 24, 'txt')] if level == 'mun' else [])
    cols = pre[:]
    for k, name in order: cols.append((name, 13, 'int'))
    cols += [('Válidos', 12, 'int'), ('Blanco', 10, 'int'), ('Nulos', 9, 'int'),
             ('No marcados', 12, 'int'), ('Total votos', 13, 'int'), ('Ganador', 18, 'txt')]
    if vuelta == 2: cols.append(('Margen (pp)', 11, 'pp'))
    pctkeys = PCT1 if vuelta == 1 else ['petro', 'rodolfo']
    cmap = dict(order)
    for k in pctkeys: cols.append((f'{cmap[k].split()[-1]} %', 10, 'pct'))
    gcol = len(pre) + len(order) + 6  # columna Ganador (1-based)

    for j, (t, w, _) in enumerate(cols, 1):
        c = ws.cell(1, j, t); c.fill = HDR; c.font = HDRF; c.alignment = CEN; c.border = BORD
        ws.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for labels, agg in rows_data:
        validos = sum(agg.get(k, 0) for k, _ in order)
        total = validos + agg.get('blanco', 0) + agg.get('nulos', 0) + agg.get('nomarc', 0)
        win = winner(agg, order)
        is_total = labels[0].startswith('TOTAL')
        vals = list(labels)
        for k, _ in order: vals.append(agg.get(k, 0))
        vals += [validos, agg.get('blanco', 0), agg.get('nulos', 0), agg.get('nomarc', 0),
                 total, cmap.get(win, win)]
        if vuelta == 2:
            p = agg.get('petro', 0); ro = agg.get('rodolfo', 0)
            vals.append(abs(p - ro) / validos * 100 if validos else 0)
        for k in pctkeys: vals.append(agg.get(k, 0) / validos if validos else 0)
        for j, (v, (t, w, fmt)) in enumerate(zip(vals, cols), 1):
            c = ws.cell(r, j, v); cellfmt(c, fmt)
            if j <= len(pre): c.alignment = LEFT
            elif not is_total and r % 2 == 0: c.fill = ZEBRA
        gc = ws.cell(r, gcol); gc.fill = PatternFill('solid', fgColor=CANDCOLOR.get(win, '555555')); gc.font = WINFT
        if is_total:
            for j in range(1, len(cols) + 1):
                cc = ws.cell(r, j); cc.font = Font(bold=True)
                if j != gcol: cc.fill = PatternFill('solid', fgColor='ECE6D8')
        r += 1
    ws.freeze_panes = ('C2' if level == 'mun' else 'B2')
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{r-1}"

def leeme(ws, tot1, tot2):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3; ws.column_dimensions['B'].width = 110
    p1 = tot1['petro'] / sum(tot1.get(k, 0) for k, _ in ORDER1) * 100
    L = [('Elección Presidencial de Colombia 2022 — resultados por departamento y municipio', 't'),
         ('1ra vuelta (29-may-2022) y 2da vuelta (19-jun-2022) · una hoja por nivel y vuelta', 's'), ('', ''),
         ('Hojas', 'h'),
         ('1V Departamento · 1V Municipio · 2V Departamento · 2V Municipio.', 'b'),
         ('1ra vuelta: 8 candidatos. 2da vuelta: Gustavo Petro vs Rodolfo Hernández (ganó Petro).', 'b'), ('', ''),
         ('Fuente y columnas', 'h'),
         ('Votos por mesa de la Registraduría (archivos GCS), agregados por departamento y municipio.', 'b'),
         ('Nombres de depto/municipio desde el divipolar electoral. Válidos = suma de candidatos.', 'b'),
         ('Total votos = válidos + blanco + nulos + no marcados. % = sobre válidos. En 1V se muestran', 'b'),
         ('los % de los 4 primeros (Petro, Rodolfo, Fico, Fajardo); los votos están de todos.', 'b'),
         ('Ganador resaltado con el color del candidato. Filas ordenadas alfabéticamente; TOTAL al final.', 'b'), ('', ''),
         ('Control nacional', 'h'),
         (f'1V: Petro {tot1["petro"]:,} ({p1:.1f}%) · Rodolfo {tot1["rodolfo"]:,} · Fico {tot1["fico"]:,} · '
          f'Fajardo {tot1["fajardo"]:,}.', 'b'),
         (f'2V: Petro {tot2["petro"]:,} · Rodolfo {tot2["rodolfo"]:,} '
          f'(ganó Petro con {tot2["petro"]/(tot2["petro"]+tot2["rodolfo"])*100:.2f}% de los válidos).', 'b'),
         ('Incluye el departamento 88 (Consulados / voto en el exterior).', 'b'), ('', ''),
         ('Generado por tools/segunda-vuelta-prec/build_2022_depto_muni_xlsx.py', 'n')]
    r = 1
    for txt, k in L:
        c = ws.cell(r, 2, txt)
        c.font = (Font(bold=True, size=15, color='8A1E16') if k == 't' else
                  Font(size=11, italic=True, color='555555') if k == 's' else
                  Font(bold=True, size=11, color='1F47CC') if k == 'h' else
                  Font(size=9, italic=True, color='888888') if k == 'n' else Font(size=10))
        r += 1

def depname(d): return DEPNAME.get(d, f'Depto {d}')
def munname(d, m):
    if (d, m) in MUNNAME: return MUNNAME[(d, m)]
    return f'Consulado {m}' if d == '88' else f'Mun {d}-{m}'

def rows_dep(DEP):
    out = [([depname(d)], agg) for d, agg in DEP.items()]
    out.sort(key=lambda x: nrm(x[0][0]))
    tot = defaultdict(int)
    for _, agg in out:
        for k, v in agg.items(): tot[k] += v
    out.append((['TOTAL NACIONAL'], tot))
    return out
def rows_mun(MUN):
    out = [([depname(d), munname(d, m)], agg) for (d, m), agg in MUN.items()]
    out.sort(key=lambda x: (nrm(x[0][0]), nrm(x[0][1])))
    tot = defaultdict(int)
    for _, agg in out:
        for k, v in agg.items(): tot[k] += v
    out.append((['TOTAL NACIONAL', ''], tot))
    return out

# ------------------------------ build ------------------------------
print('Agregando 1ra vuelta 2022...', flush=True)
DEP1, MUN1, n1 = aggregate('GCS_2022PRES1V.csv', cand1v)
print(f'  {n1:,} filas · {len(DEP1)} deptos · {len(MUN1)} municipios')
print('Agregando 2da vuelta 2022...', flush=True)
DEP2, MUN2, n2 = aggregate('GCS_2022PRES2V.csv', cand2v)
print(f'  {n2:,} filas · {len(DEP2)} deptos · {len(MUN2)} municipios')

tot1 = defaultdict(int)
for agg in DEP1.values():
    for k, v in agg.items(): tot1[k] += v
tot2 = defaultdict(int)
for agg in DEP2.values():
    for k, v in agg.items(): tot2[k] += v

wb = Workbook()
leeme(wb.active, tot1, tot2); wb.active.title = 'Leeme'
write_sheet(wb.create_sheet('1V Departamento'), rows_dep(DEP1), ORDER1, 'dep', 1)
write_sheet(wb.create_sheet('1V Municipio'),    rows_mun(MUN1), ORDER1, 'mun', 1)
write_sheet(wb.create_sheet('2V Departamento'), rows_dep(DEP2), ORDER2, 'dep', 2)
write_sheet(wb.create_sheet('2V Municipio'),    rows_mun(MUN2), ORDER2, 'mun', 2)
os.makedirs(OUT, exist_ok=True)
wb.save(OUTFILE)
print(f'\nOK -> {OUTFILE}')
print(f'  1V nacional: Petro {tot1["petro"]:,} · Rodolfo {tot1["rodolfo"]:,} · Fico {tot1["fico"]:,}')
print(f'  2V nacional: Petro {tot2["petro"]:,} · Rodolfo {tot2["rodolfo"]:,} '
      f'({tot2["petro"]/(tot2["petro"]+tot2["rodolfo"])*100:.2f}% Petro)')
