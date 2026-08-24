#!/usr/bin/env python3
"""Paso 1 · extrae de Edadygenero.xlsx las mesas del CONGRESO en Cundinamarca
(dep 15) para 2018 y 2022, con el desglose JOINT sexo × edad por mesa.

Una sola pasada por el xlsx de 135 MB (openpyxl read_only).
Salida: output_cundi_edad_genero/edad-congreso-{2018,2022}-cundi.csv
"""
import csv
import os
import sys
import time

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import (BASE, OUT, DEP_CUNDI, H_COLS, M_COLS, H_INDEF, M_INDEF)

SRC = os.path.join(BASE, "Edadygenero.xlsx")

KEY_COLS = ["Cód. Depto", "Estadosnoborrar", "Cód. Municipio",
            "Cód. Comuna / Localidad", "Cód. Puesto de Votación", "Mesa"]
VAL_COLS = ["Cantidad de Sufragantes", "Cantidad de Indefinidos",
            "Edad Indefinida", H_INDEF, M_INDEF] + H_COLS + M_COLS
OUT_COLS = KEY_COLS + VAL_COLS


def main():
    os.makedirs(OUT, exist_ok=True)
    t0 = time.time()
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = None
    idx = {}
    writers, files = {}, {}
    counts = {2018: 0, 2022: 0}
    suf = {2018: 0, 2022: 0}
    tipos_seen = set()
    depnames = set()
    n = 0
    for row in ws.iter_rows(values_only=True):
        n += 1
        if header is None:
            header = list(row)
            idx = {c: header.index(c) for c in OUT_COLS}
            i_tipo = header.index("Datos de tipo de elección")
            i_anio = header.index("Año")
            for y in (2018, 2022):
                f = open(os.path.join(OUT, f"edad-congreso-{y}-cundi.csv"),
                         "w", newline="", encoding="utf-8")
                w = csv.writer(f)
                w.writerow(OUT_COLS)
                writers[y], files[y] = w, f
            continue

        dep = str(row[idx["Cód. Depto"]] or "").strip()
        if dep not in (DEP_CUNDI, str(int(DEP_CUNDI))):
            if n % 100000 == 0:
                print(f"  ... {n} filas ({time.time()-t0:.0f}s)", flush=True)
            continue
        tipo = str(row[i_tipo] or "")
        if "Congreso" not in tipo:
            continue
        anio_val = row[i_anio]
        anio = anio_val.year if hasattr(anio_val, "year") else int(str(anio_val)[:4])
        if anio not in writers:
            continue
        tipos_seen.add(tipo)
        depnames.add(str(row[idx["Estadosnoborrar"]]))
        writers[anio].writerow([row[idx[c]] for c in OUT_COLS])
        counts[anio] += 1
        try:
            suf[anio] += int(row[idx["Cantidad de Sufragantes"]] or 0)
        except (TypeError, ValueError):
            pass
        if n % 100000 == 0:
            print(f"  ... {n} filas ({time.time()-t0:.0f}s)", flush=True)

    for f in files.values():
        f.close()
    wb.close()
    print(f"\nTotal filas leídas: {n:,} en {time.time()-t0:.0f}s")
    print(f"Tipos de elección vistos (dep {DEP_CUNDI}): {sorted(tipos_seen)}")
    print(f"Nombre(s) de depto: {sorted(depnames)}")
    for y in (2018, 2022):
        print(f"  Congreso {y}: {counts[y]:,} mesas · {suf[y]:,} sufragantes "
              f"-> edad-congreso-{y}-cundi.csv")


if __name__ == "__main__":
    main()
