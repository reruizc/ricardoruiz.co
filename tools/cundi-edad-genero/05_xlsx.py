#!/usr/bin/env python3
"""Paso 5 · ensambla el Excel entregable.

Cundinamarca_Congreso_Edad_Genero_2022_2026.xlsx con hojas:
  Léeme · Resumen departamental · General 2022/2026 (sexo×edad observado/proy)
  Senado 2022/2026 · Cámara 2022/2026 (por partido, punto+cotas+IC) · Beta (motor)
"""
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import OUT, CELL_LABELS, GROUPS

XLSX = os.path.join(OUT, "Cundinamarca_Congreso_Edad_Genero_2022_2026.xlsx")

OX = "8a1e16"          # oxblood (header)
CREAM = "F4F0E7"       # zebra
INK = "1a1510"
HEAD_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=15, color=OX)
B_THIN = Side(style="thin", color="D8CFC0")
BORDER = Border(left=B_THIN, right=B_THIN, top=B_THIN, bottom=B_THIN)
SEXLAB = {"H": "Hombres", "M": "Mujeres"}
GROUP_ORDER = list(GROUPS)


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=OX)
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def finish(ws, header_row, ncol, first_data_row, n_data, freeze=None,
           num_cols=None, pct_cols=None):
    for i in range(n_data):
        r = first_data_row + i
        if i % 2 == 1:
            for c in range(1, ncol + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=CREAM)
        for c in range(1, ncol + 1):
            ws.cell(row=r, column=c).border = BORDER
    if num_cols:
        for c in num_cols:
            for i in range(n_data):
                ws.cell(row=first_data_row + i, column=c).number_format = "#,##0"
    if pct_cols:
        for c in pct_cols:
            for i in range(n_data):
                ws.cell(row=first_data_row + i, column=c).number_format = "0.0%"
    ws.freeze_panes = freeze or f"A{first_data_row}"
    last = f"{get_column_letter(ncol)}{header_row}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncol)}{first_data_row + n_data - 1}"


def widths(ws, w):
    for col, width in w.items():
        ws.column_dimensions[col].width = width


# ------------------------------------------------------------------ Léeme
def sheet_leeme(wb):
    ws = wb.active
    ws.title = "Léeme"
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 110})
    L = [
        ("T", "Votación por municipio, partido, sexo y rangos de edad"),
        ("T2", "Cámara y Senado · 2022 y 2026 · Departamento de Cundinamarca"),
        ("", ""),
        ("H", "Qué contiene"),
        ("p", "Para cada municipio de Cundinamarca (116), estima cuántos votos recibió cada lista "
              "—y el total «en general»— desglosados por las 10 celdas sexo × edad: Hombres y "
              "Mujeres en 18-25, 26-35, 36-45, 46-60 y 61+. Cubre Cámara y Senado, en 2022 y 2026."),
        ("", ""),
        ("H", "Lo que es DATO directo y lo que es ESTIMACIÓN"),
        ("p", "• «En general» (hojas General 2022/2026): el total de votantes por sexo × edad de cada "
              "municipio. En 2022 es DATO OBSERVADO (RNEC Edad y Género, mesa a mesa). En 2026 es "
              "PROYECCIÓN demográfica (no existe aún el Edad y Género 2026)."),
        ("p", "• «Por partido» (hojas Senado/Cámara): cuántos de los votos de cada lista vinieron de "
              "cada celda sexo × edad. Esto NO es un conteo: las urnas nunca registran por quién votó "
              "cada grupo. Es INFERENCIA ECOLÓGICA (modelo RxC tipo Goodman/King) calibrada con la "
              "variación entre las ~5.800 mesas (2022) / ~600 puestos (2026) del departamento."),
        ("", ""),
        ("H", "Cómo leer la incertidumbre (3 columnas por celda)"),
        ("p", "• punto = estimación central. Por construcción, SIEMPRE cuadra con los totales reales: "
              "la suma por partido da su votación real en el municipio, y la suma por celda da el total "
              "de votantes de esa celda."),
        ("p", "• cota_min / cota_max = cotas DURAS de Duncan-Davis (King): el rango que los datos "
              "permiten SIN ningún supuesto, solo por aritmética de márgenes. Es lo defendible al 100%."),
        ("p", "• ic_lo / ic_hi = intervalo 90% por bootstrap de municipios (incertidumbre del modelo "
              "de preferencias). Más estrecho que las cotas duras; depende del supuesto del modelo."),
        ("", ""),
        ("H", "La edad se identifica bien; el sexo casi no"),
        ("p", "La composición etaria varía mucho entre puestos y se correlaciona fuerte con el voto "
              "(p.ej. Pacto: corr con %18-25 = +0,40; con %61+ = −0,34), así que el PERFIL DE EDAD de "
              "cada lista es robusto. En cambio el % de mujeres por mesa NO predice el voto "
              "(correlación ≈ 0): el dato NO identifica una diferencia de sexo robusta. Por eso el "
              "reparto por sexo dentro de cada grupo de edad se mantiene cercano a la composición "
              "poblacional (gap H−M < 2 pp); publicar un sesgo de sexo fuerte sería un artefacto. "
              "La estimación de sexo más fina exigiría el método de efectos fijos de puesto a nivel "
              "mesa (pendiente)."),
        ("", ""),
        ("H", "Proyección 2026 (composición sexo × edad)"),
        ("p", "Tasa de participación por celda en 2022 (votantes/​población DANE 2022) aplicada a la "
              "población DANE 2026 por sexo × edad del departamento, y reajustada (raking IPF) a los "
              "votantes reales de cada puesto en 2026 (escrutinio declarado). Supuesto: la FORMA del "
              "perfil de participación por sexo y edad es estable entre 2022 y 2026; el NIVEL lo fija "
              "el dato real 2026. Validado en el modelo presidencial análogo (backtest 2018→2022: "
              "0,32 pp de error nacional por banda)."),
        ("", ""),
        ("H", "Grupos de listas"),
        ("p", "Se nombran las listas con ≥ 0,8% del voto válido+blanco del departamento. El resto "
              "—listas menores y las circunscripciones especiales (afro / indígena / CITREP)— se "
              "agrupan en «Otras listas». «Votos en blanco» va como columna propia. Nulos y no "
              "marcados se excluyen del universo repartido (no son atribuibles a una lista)."),
        ("", ""),
        ("H", "Fuentes"),
        ("p", "RNEC Edad y Género (sufragantes por mesa × sexo × edad, 2022) · Escrutinio Congreso "
              "2022 (GCS) · Declarados Congreso 2026 por mesa (RNEC) · Proyecciones de población DANE "
              "2018-2050 por área, sexo y edad. Departamento electoral RNEC 15 = Cundinamarca "
              "(no incluye Bogotá D.C., que es el dep 16)."),
        ("", ""),
        ("H", "Advertencia"),
        ("p", "Las cifras por partido × sexo × edad son estimaciones de comportamiento de GRUPO, no de "
              "individuos (Robinson 1950). Acompáñelas siempre de sus cotas. Borrador analítico."),
    ]
    r = 1
    for kind, txt in L:
        cell = ws.cell(row=r, column=2, value=txt)
        if kind == "T":
            cell.font = TITLE_FONT
        elif kind == "T2":
            cell.font = Font(bold=True, size=11, color=INK)
        elif kind == "H":
            cell.font = Font(bold=True, size=11, color=OX)
        else:
            cell.font = Font(size=10, color=INK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 14 * (1 + len(txt) // 95)
        r += 1


# ------------------------------------------------------------------ General
def sheet_general(wb, year, observed):
    comp = pd.read_csv(os.path.join(OUT, f"comp-{year}-mun.csv"), dtype={"mun": str})
    names = pd.read_csv(os.path.join(OUT, "mun-names-cundi.csv"), dtype=str) \
        .set_index("mun_code")["mun_name"]
    comp["municipio"] = comp["mun"].str[3:].map(names)
    comp["total"] = comp[CELL_LABELS].sum(axis=1)
    comp = comp.sort_values("municipio")
    title = (f"En general · votantes por sexo × edad por municipio · {year} · "
             + ("OBSERVADO (RNEC Edad y Género)" if observed
                else "PROYECTADO (DANE + participación 2022)"))
    ws = wb.create_sheet(f"General {year}")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    hdr = ["Cód", "Municipio"] + [f"{SEXLAB[l[0]]} {l[2:]}" for l in CELL_LABELS] + ["Total"]
    hr = 3
    for c, h in enumerate(hdr, 1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(hdr))
    r = hr + 1
    tot = [0] * 11
    for _, row in comp.iterrows():
        ws.cell(row=r, column=1, value=row["mun"][3:])
        ws.cell(row=r, column=2, value=row["municipio"])
        for i, l in enumerate(CELL_LABELS):
            ws.cell(row=r, column=3 + i, value=round(row[l]))
            tot[i] += row[l]
        ws.cell(row=r, column=13, value=round(row["total"]))
        tot[10] += row["total"]
        r += 1
    ws.cell(row=r, column=2, value="CUNDINAMARCA (total)").font = Font(bold=True)
    for i in range(11):
        cc = ws.cell(row=r, column=3 + i, value=round(tot[i]))
        cc.font = Font(bold=True)
    finish(ws, hr, len(hdr), hr + 1, len(comp) + 1,
           freeze="C4", num_cols=list(range(3, 14)))
    widths(ws, {"A": 6, "B": 26})
    for c in range(3, 14):
        ws.column_dimensions[get_column_letter(c)].width = 11


# ------------------------------------------------------------------ por partido
def sheet_party(wb, long, corp, year):
    names = pd.read_csv(os.path.join(OUT, "mun-names-cundi.csv"), dtype=str) \
        .set_index("mun_code")["mun_name"]
    d = long[(long["corp"] == corp) & (long["year"] == year)].copy()
    d["municipio"] = d["mun"].str[3:].map(names)
    d["sexo"] = d["sexo"].map(SEXLAB)
    d["grupo_cat"] = pd.Categorical(d["grupo"], GROUP_ORDER, ordered=True)
    # % que representa la celda dentro del voto del partido en el municipio
    ptot = d.groupby(["mun", "partido"])["punto"].transform("sum")
    d["pct_partido"] = d["punto"] / ptot.where(ptot > 0, 1)
    d = d.sort_values(["municipio", "partido", "sexo", "grupo_cat"])
    cols = [("Cód", "mun_s"), ("Municipio", "municipio"), ("Lista / partido", "partido"),
            ("Sexo", "sexo"), ("Edad", "grupo"),
            ("Votos (est.)", "punto"), ("% del partido", "pct_partido"),
            ("Cota mín", "cota_min"), ("Cota máx", "cota_max"),
            ("IC90 inf", "ic_lo"), ("IC90 sup", "ic_hi")]
    d["mun_s"] = d["mun"].str[3:]
    ws = wb.create_sheet(f"{corp.title()} {year}")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1,
            value=f"{corp.title()} {year} · votos por municipio × lista × sexo × edad "
                  f"(estimación ecológica con cotas)").font = TITLE_FONT
    hr = 3
    for c, (h, _) in enumerate(cols, 1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(cols))
    r = hr + 1
    for _, row in d.iterrows():
        for c, (_, key) in enumerate(cols, 1):
            v = row[key]
            if key in ("punto", "cota_min", "cota_max", "ic_lo", "ic_hi"):
                v = round(float(v))
            ws.cell(row=r, column=c, value=v)
        r += 1
    finish(ws, hr, len(cols), hr + 1, len(d), freeze="C4",
           num_cols=[6, 8, 9, 10, 11], pct_cols=[7])
    widths(ws, {"A": 6, "B": 22, "C": 40, "D": 9, "E": 8, "F": 11,
                "G": 11, "H": 10, "I": 10, "J": 10, "K": 10})


# ------------------------------------------------------------------ resumen depto
def sheet_resumen(wb, long):
    ws = wb.create_sheet("Resumen depto")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1,
            value="Resumen departamental · composición del electorado de cada lista "
                  "(% de sus votos por sexo y edad)").font = TITLE_FONT
    hdr = ["Corp.", "Año", "Lista / partido", "Votos", "% muj", "% 18-25",
           "% 26-35", "% 36-45", "% 46-60", "% 61+"]
    hr = 3
    for c, h in enumerate(hdr, 1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(hdr))
    r = hr + 1
    n = 0
    for corp in ("SENADO", "CAMARA"):
        for year in (2022, 2026):
            d = long[(long["corp"] == corp) & (long["year"] == year)]
            g = d.groupby("partido")["punto"].sum().sort_values(ascending=False)
            for party, tot in g.items():
                sub = d[d["partido"] == party]
                muj = sub[sub["sexo"] == "M"]["punto"].sum() / tot if tot else 0
                byg = sub.groupby("grupo")["punto"].sum()
                vals = [byg.get(grp, 0) / tot if tot else 0 for grp in GROUP_ORDER]
                ws.cell(row=r, column=1, value=corp.title())
                ws.cell(row=r, column=2, value=year)
                ws.cell(row=r, column=3, value=party)
                ws.cell(row=r, column=4, value=round(tot))
                ws.cell(row=r, column=5, value=muj)
                for i, v in enumerate(vals):
                    ws.cell(row=r, column=6 + i, value=v)
                r += 1
                n += 1
    finish(ws, hr, len(hdr), hr + 1, n, freeze="D4",
           num_cols=[4], pct_cols=[5, 6, 7, 8, 9, 10])
    widths(ws, {"A": 9, "B": 7, "C": 42, "D": 11})
    for c in range(5, 11):
        ws.column_dimensions[get_column_letter(c)].width = 9


# ------------------------------------------------------------------ beta motor
def sheet_beta(wb):
    b = pd.read_csv(os.path.join(OUT, "ei-cundi-beta-dept.csv"))
    ws = wb.create_sheet("Beta (motor)")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1,
            value="Motor · β = % del voto DENTRO de cada celda sexo×edad (nivel "
                  "departamental) con cotas duras Duncan-Davis").font = TITLE_FONT
    b["sexo"] = b["celda"].str[0].map(SEXLAB)
    b["grupo"] = b["celda"].str[2:]
    hdr = ["Corp.", "Año", "Lista / partido", "Sexo", "Edad",
           "β (%)", "cota dura mín (%)", "cota dura máx (%)"]
    hr = 3
    for c, h in enumerate(hdr, 1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(hdr))
    r = hr + 1
    for _, row in b.iterrows():
        ws.cell(row=r, column=1, value=row["corp"].title())
        ws.cell(row=r, column=2, value=int(row["year"]))
        ws.cell(row=r, column=3, value=row["partido"])
        ws.cell(row=r, column=4, value=row["sexo"])
        ws.cell(row=r, column=5, value=row["grupo"])
        ws.cell(row=r, column=6, value=row["beta"])
        ws.cell(row=r, column=7, value=row["dd_lo"])
        ws.cell(row=r, column=8, value=row["dd_hi"])
        r += 1
    finish(ws, hr, len(hdr), hr + 1, len(b), freeze="C4", pct_cols=[6, 7, 8])
    widths(ws, {"A": 9, "B": 7, "C": 42, "D": 9, "E": 8, "F": 9, "G": 14, "H": 14})


def main():
    long = pd.read_csv(os.path.join(OUT, "ei-cundi-long.csv"), dtype={"mun": str})
    wb = Workbook()
    sheet_leeme(wb)
    sheet_resumen(wb, long)
    sheet_general(wb, 2022, observed=True)
    sheet_general(wb, 2026, observed=False)
    for corp in ("SENADO", "CAMARA"):
        for year in (2022, 2026):
            sheet_party(wb, long, corp, year)
    sheet_beta(wb)
    wb.save(XLSX)
    print(f"-> {XLSX}")
    print(f"   hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
