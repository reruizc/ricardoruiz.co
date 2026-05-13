#!/usr/bin/env python3
"""
tools/build-bog-upl-censo.py

Construye el censo electoral por UPL de Bogotá para 2026, 2022 y 2018.
Reusa el mapeo puesto → UPL (point-in-polygon con PUESTOS_GEOREF.csv +
BOG-UPL.geojson) y suma el censo de cada puesto desde la fuente
correspondiente al año.

Fuentes de censo por puesto:
  2026 → COMUNAS_DATA.csv (Divipole oficial vigente, columnas dd,mm,zz,pp,total,mujeres,hombres)
  2022 → Divipol 23.09.2021.xlsx (mismo shape, leído con openpyxl)
  2018 → escalado proporcional desde 2022 con ratio Bogotá = censo_18_dep / censo_22_dep
         (los Divipole 2018 oficiales no están disponibles; el ratio sale de Wikipedia)

Outputs (en out_dir = output_ciudades/bogota):
  censo-upl-2026.json
  censo-upl-2022.json
  censo-upl-2018.json

Cada archivo:
  { year, ciudad_total, n_upl, por_upl: { uplCod: { uplCod, nombre, sector,
    censo, mujeres, hombres, n_puestos } } }

Uso:
  python3 tools/build-bog-upl-censo.py \
    "Bases de datos/PUESTOS_GEOREF.csv" \
    "CIUDADES/BOGOTA/BOG-UPL.geojson" \
    "Bases de datos/COMUNAS_DATA.csv" \
    "Bases de datos/Divipol 23.09.2021.xlsx" \
    "Bases de datos/output_ciudades/bogota"
"""

import csv, json, os, sys, re
from collections import defaultdict
import openpyxl

BOG_DEP, BOG_MUN = '16', '001'
BOG_2018_DEP_CENSO = 5_703_232
BOG_2022_DEP_CENSO = 5_630_748   # del puestos-censos-agg-2022.json


def pad2(s): return str(s).zfill(2)
def pad3(s): return str(s).zfill(3)


# ── point-in-polygon (ray casting) ───────────────────────────────────
def point_in_ring(x, y, ring):
    inside = False
    n = len(ring); j = n - 1
    for i in range(n):
        xi, yi = ring[i][0], ring[i][1]
        xj, yj = ring[j][0], ring[j][1]
        if ((yi > y) != (yj > y)) and \
           (x < (xj - xi) * (y - yi) / ((yj - yi) or 1e-12) + xi):
            inside = not inside
        j = i
    return inside


def point_in_feature(x, y, geom):
    polys = [geom['coordinates']] if geom['type'] == 'Polygon' else geom['coordinates']
    for poly in polys:
        if not poly or not point_in_ring(x, y, poly[0]): continue
        in_hole = any(point_in_ring(x, y, ring) for ring in poly[1:])
        if not in_hole: return True
    return False


def feature_bbox(geom):
    minx, miny, maxx, maxy = 1e9, 1e9, -1e9, -1e9
    polys = [geom['coordinates']] if geom['type'] == 'Polygon' else geom['coordinates']
    for poly in polys:
        for ring in poly:
            for x, y in ring:
                if x < minx: minx = x
                if x > maxx: maxx = x
                if y < miny: miny = y
                if y > maxy: maxy = y
    return minx, miny, maxx, maxy


# ── 1) Mapeo puesto → UPL ────────────────────────────────────────────
def load_puestos_bogota(csv_path):
    out = {}
    with open(csv_path, encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        header = next(reader); header[0] = header[0].replace('﻿', '')
        col = lambda n: next(i for i, h in enumerate(header) if h.strip().upper() == n.upper())
        iDep = col('DEPARTAMENTO'); iZon = col('ZONA'); iPue = col('PUESTO')
        iLat = col('LATITUD'); iLon = col('LONGITUD'); iBar = col('BARRIO')
        for row in reader:
            if not row or len(row) <= iLon: continue
            if 'BOGOTA' not in (row[iDep] or '').upper(): continue
            try:
                lat = float((row[iLat] or '').replace(',', '.'))
                lon = float((row[iLon] or '').replace(',', '.'))
            except ValueError:
                continue
            zz = pad2(row[iZon]); pp = pad2(row[iPue])
            out[f'{zz}-{pp}'] = {'lat': lat, 'lon': lon, 'barrio': row[iBar] or ''}
    return out


def build_puesto_to_upl(puestos, upl_geo):
    upls = []
    for f in upl_geo['features']:
        cod = f['properties'].get('CODIGO_UPL')
        if not cod: continue
        upls.append((cod, feature_bbox(f['geometry']), f['geometry']))
    out = {}; unmatched = []
    for zzpp, p in puestos.items():
        x, y = p['lon'], p['lat']; assigned = None
        for cod, bb, geom in upls:
            if x < bb[0] or x > bb[2] or y < bb[1] or y > bb[3]: continue
            if point_in_feature(x, y, geom):
                assigned = cod; break
        if assigned: out[zzpp] = assigned
        else: unmatched.append(zzpp)
    return out, unmatched


# ── 2) Censo por puesto ─────────────────────────────────────────────
def load_censo_2026(csv_path):
    out = {}
    with open(csv_path, encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        header = next(reader); header[0] = header[0].replace('﻿', '')
        col = lambda n: next(i for i, h in enumerate(header) if h.strip().lower() == n.lower())
        iDd = col('dd'); iMm = col('mm'); iZz = col('zz'); iPp = col('pp')
        iTot = col('total'); iMuj = col('mujeres'); iHom = col('hombres')
        for row in reader:
            if not row or len(row) <= iTot: continue
            dd = pad2(row[iDd]); mm = pad3(row[iMm])
            if dd != BOG_DEP or mm != BOG_MUN: continue
            zz = pad2(row[iZz]); pp = pad2(row[iPp])
            try: tot = int(row[iTot] or 0)
            except ValueError: tot = 0
            if tot <= 0: continue
            try: muj = int(row[iMuj] or 0)
            except ValueError: muj = 0
            try: hom = int(row[iHom] or 0)
            except ValueError: hom = 0
            out[f'{zz}-{pp}'] = {'total': tot, 'mujeres': muj, 'hombres': hom}
    return out


def load_censo_2022(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_idx = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if row and str(row[0]).strip().lower() == 'dd':
            for j, h in enumerate(row):
                if h is not None: header_idx[str(h).strip().lower()] = j
            header_row = i; break
    iDd = header_idx['dd']; iMm = header_idx['mm']
    iZz = header_idx['zz']; iPp = header_idx['pp']
    iTot = header_idx['total']
    iMuj = header_idx.get('mujeres'); iHom = header_idx.get('hombres')
    out = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or row[iDd] is None: continue
        dd = pad2(row[iDd]); mm = pad3(row[iMm])
        if dd != BOG_DEP or mm != BOG_MUN: continue
        zz = pad2(row[iZz]); pp = pad2(row[iPp])
        try: tot = int(row[iTot] or 0)
        except (TypeError, ValueError): tot = 0
        if tot <= 0: continue
        try: muj = int(row[iMuj] or 0) if iMuj is not None else 0
        except (TypeError, ValueError): muj = 0
        try: hom = int(row[iHom] or 0) if iHom is not None else 0
        except (TypeError, ValueError): hom = 0
        out[f'{zz}-{pp}'] = {'total': tot, 'mujeres': muj, 'hombres': hom}
    return out


# ── 3) Agregar por UPL ──────────────────────────────────────────────
def aggregate_upl(pue_to_upl, censo_by_pue, upl_info):
    by_upl = defaultdict(lambda: {'total': 0, 'mujeres': 0, 'hombres': 0, 'n_puestos': 0})
    n_matched = 0
    for zzpp, censo in censo_by_pue.items():
        upl_cod = pue_to_upl.get(zzpp)
        if not upl_cod: continue
        n_matched += 1
        b = by_upl[upl_cod]
        b['total'] += censo['total']
        b['mujeres'] += censo['mujeres']
        b['hombres'] += censo['hombres']
        b['n_puestos'] += 1
    por_upl = {}; ciudad_total = 0
    for cod, b in by_upl.items():
        info = upl_info.get(cod, {})
        por_upl[cod] = {
            'uplCod': cod, 'nombre': info.get('nombre', ''), 'sector': info.get('sector', ''),
            'censo': b['total'], 'mujeres': b['mujeres'], 'hombres': b['hombres'],
            'n_puestos': b['n_puestos'],
        }
        ciudad_total += b['total']
    return por_upl, ciudad_total, n_matched


def main():
    if len(sys.argv) < 6:
        print('Uso: build-bog-upl-censo.py <PUESTOS_GEOREF.csv> <BOG-UPL.geojson> <COMUNAS_DATA.csv> <Divipol2021.xlsx> <out-dir>', file=sys.stderr)
        sys.exit(1)
    csv_path, geo_path, censo26_path, divipol22_path, out_dir = sys.argv[1:6]
    os.makedirs(out_dir, exist_ok=True)

    print('[1] cargando puestos…')
    puestos = load_puestos_bogota(csv_path)
    print(f'    {len(puestos)} puestos con coords')

    upl_geo = json.load(open(geo_path))
    upl_info = {f['properties']['CODIGO_UPL']: {
                    'cod': f['properties']['CODIGO_UPL'],
                    'nombre': f['properties'].get('NOMBRE', ''),
                    'sector': f['properties'].get('SECTOR', ''),
                } for f in upl_geo['features'] if f['properties'].get('CODIGO_UPL')}

    print('[2] mapeo puesto → UPL…')
    pue_to_upl, unmatched = build_puesto_to_upl(puestos, upl_geo)
    print(f'    matched: {len(pue_to_upl)}/{len(puestos)} · unmatched: {len(unmatched)}')

    print('[3] censo 2026 (COMUNAS_DATA)…')
    censo26 = load_censo_2026(censo26_path)
    print(f'    {len(censo26)} puestos con censo 2026 en Bogotá')
    por_upl26, total26, m26 = aggregate_upl(pue_to_upl, censo26, upl_info)
    print(f'    {len(por_upl26)} UPL · ciudad total {total26:,} · {m26} puestos asignados')

    print('[4] censo 2022 (Divipol 23-09-2021)…')
    censo22 = load_censo_2022(divipol22_path)
    print(f'    {len(censo22)} puestos con censo 2022 en Bogotá')
    por_upl22, total22, m22 = aggregate_upl(pue_to_upl, censo22, upl_info)
    print(f'    {len(por_upl22)} UPL · ciudad total {total22:,} · {m22} puestos asignados')

    print('[5] censo 2018 (escalado proporcional desde 2022)…')
    ratio_18_22 = BOG_2018_DEP_CENSO / max(BOG_2022_DEP_CENSO, 1)
    por_upl18 = {}
    for cod, e in por_upl22.items():
        por_upl18[cod] = {**e,
            'censo':   round(e['censo']   * ratio_18_22),
            'mujeres': round(e['mujeres'] * ratio_18_22),
            'hombres': round(e['hombres'] * ratio_18_22),
        }
    total18 = sum(e['censo'] for e in por_upl18.values())
    print(f'    {len(por_upl18)} UPL · ciudad total {total18:,} · ratio {ratio_18_22:.4f}')

    for year, por, total in [(2026, por_upl26, total26), (2022, por_upl22, total22), (2018, por_upl18, total18)]:
        out_path = os.path.join(out_dir, f'censo-upl-{year}.json')
        json.dump({
            'year': year, 'depCod': BOG_DEP, 'munCod': BOG_MUN,
            'ciudad_total': total, 'n_upl': len(por),
            'por_upl': por,
            'fuente': 'aggregated por point-in-polygon puesto→UPL' if year != 2018 else 'escalado proporcional desde 2022',
        }, open(out_path, 'w'))
        print(f'✓ {out_path}')


if __name__ == '__main__':
    main()
