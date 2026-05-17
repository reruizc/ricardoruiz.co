"""Aggregator del archivo Edadygenero.xlsx (Registraduría, ~135 MB, 645k filas).

Stream-lee el xlsx con openpyxl read_only y agrega:
  - Por departamento (Cód. Depto)
  - Por (año, tipo de elección)
  - Cohortes de mujeres en buckets: 18-25, 26-40, 41-60, 60+

Salida: Bases de datos/output_observatorio_mujer/por-depto-edadgenero.json
Formato compacto:
  {
    "2018-pres1v": { "01": {nombre, total_muj, b18_25, b26_40, b41_60, b60p, total_hom}, ... },
    "2022-pres1v": { ... },
    "2022-cong":   { ... }
  }

Uso: python3 tools/build-observatorio-mujer/build-edadgenero.py
"""
import json
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SRC  = ROOT / "Bases de datos" / "Edadygenero.xlsx"
OUT  = ROOT / "Bases de datos" / "output_observatorio_mujer" / "por-depto-edadgenero.json"

# Combos (año, tipo) que queremos extraer
TARGETS = {
    (2018, "Presidencia 1V"): "2018-pres1v",
    (2018, "Congreso de la República"): "2018-cong",
    (2022, "Presidencia 1V"): "2022-pres1v",
    (2022, "Congreso de la República"): "2022-cong",
}

# Índices de columnas (0-based) según el header
C_DEPTO_NUM   = 1   # Cód. Depto
C_DEPTO_NAME  = 2   # Estadosnoborrar (nombre)
C_TIPO        = 3   # Datos de tipo de elección
C_ANO         = 4   # Año (datetime)
C_SUFRAG_TOT  = 10  # Cantidad de Sufragantes
C_HOMBRES     = 22  # Cantidad Hombres
C_MUJERES     = 23  # Cantidad de Mujeres
# Mujeres por bucket etario (índices)
C_MUJ_INDEF   = 36
C_MUJ_18_20   = 37
C_MUJ_21_25   = 38
C_MUJ_26_30   = 39
C_MUJ_31_35   = 40
C_MUJ_36_40   = 41
C_MUJ_41_45   = 42
C_MUJ_46_50   = 43
C_MUJ_51_55   = 44
C_MUJ_56_60   = 45
C_MUJ_60P     = 46

def empty_bucket():
    return {
        "nombre": "",
        "total_muj": 0,
        "b18_25":   0,
        "b26_40":   0,
        "b41_60":   0,
        "b60p":     0,
        "muj_indef":0,
        "total_hom":0,
    }

def num(v):
    if v is None: return 0
    try: return int(v)
    except (TypeError, ValueError):
        try: return int(float(v))
        except: return 0

def main():
    print(f"Leyendo {SRC} ...")
    wb = load_workbook(str(SRC), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    print(f"Hoja: {wb.sheetnames[0]} · {ws.max_row} filas")

    out = {key: defaultdict(empty_bucket) for key in TARGETS.values()}

    rows = ws.iter_rows(min_row=2, values_only=True)
    n = 0
    for row in rows:
        n += 1
        if n % 100_000 == 0:
            print(f"  fila {n:,}...")

        try:
            tipo = row[C_TIPO]
            ano  = row[C_ANO]
            if hasattr(ano, "year"): ano = ano.year
            else: ano = int(ano) if ano else None

            key = TARGETS.get((ano, tipo))
            if key is None: continue

            dep_num = num(row[C_DEPTO_NUM])
            if dep_num <= 0: continue
            dep_cod = str(dep_num).zfill(2)
            dep_nom = row[C_DEPTO_NAME] or ""

            b = out[key][dep_cod]
            if not b["nombre"]: b["nombre"] = dep_nom

            # Cohortes de mujeres
            m_18_20 = num(row[C_MUJ_18_20])
            m_21_25 = num(row[C_MUJ_21_25])
            m_26_30 = num(row[C_MUJ_26_30])
            m_31_35 = num(row[C_MUJ_31_35])
            m_36_40 = num(row[C_MUJ_36_40])
            m_41_45 = num(row[C_MUJ_41_45])
            m_46_50 = num(row[C_MUJ_46_50])
            m_51_55 = num(row[C_MUJ_51_55])
            m_56_60 = num(row[C_MUJ_56_60])
            m_60p   = num(row[C_MUJ_60P])
            m_indef = num(row[C_MUJ_INDEF])

            b["b18_25"] += m_18_20 + m_21_25
            b["b26_40"] += m_26_30 + m_31_35 + m_36_40
            b["b41_60"] += m_41_45 + m_46_50 + m_51_55 + m_56_60
            b["b60p"]   += m_60p
            b["muj_indef"] += m_indef
            b["total_muj"] += num(row[C_MUJERES])
            b["total_hom"] += num(row[C_HOMBRES])
        except Exception as e:
            if n < 10:
                print(f"  ERROR fila {n}: {e}")
            continue

    print(f"\n  total filas procesadas: {n:,}")

    # Convertir defaultdict a dict normal, agregar pct por bucket
    final = {}
    for key, dep_map in out.items():
        final[key] = {}
        for cod, b in sorted(dep_map.items()):
            tm = b["total_muj"] or 1
            final[key][cod] = {
                "nombre":     b["nombre"],
                "total_muj":  b["total_muj"],
                "total_hom":  b["total_hom"],
                "b18_25":     b["b18_25"],
                "b26_40":     b["b26_40"],
                "b41_60":     b["b41_60"],
                "b60p":       b["b60p"],
                "muj_indef":  b["muj_indef"],
                "pct_18_25":  round(b["b18_25"]/tm*100, 2),
                "pct_26_40":  round(b["b26_40"]/tm*100, 2),
                "pct_41_60":  round(b["b41_60"]/tm*100, 2),
                "pct_60p":    round(b["b60p"]/tm*100,   2),
            }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(final, f, ensure_ascii=False, indent=2)
    print(f"\n  escrito en {OUT}")
    print(f"  llaves: {list(final.keys())}")
    for key, deps in final.items():
        total_muj_nac = sum(d["total_muj"] for d in deps.values())
        print(f"  {key}: {len(deps)} deptos, {total_muj_nac:,} mujeres totales")

if __name__ == "__main__":
    main()
