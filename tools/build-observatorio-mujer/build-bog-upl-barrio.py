"""Cohortes etarias de mujeres por UPL y barrio (Bogotá) desde Edadygenero.

Procesa Edadygenero.xlsx (Pres 1V 2018 + 2022, filtrado a Bogotá dep=16
mun=001), agrega por puesto, y cruza:
  - puesto → barrio catastral usando bog-puesto-to-barrio.json (S3)
  - puesto → UPL usando point-in-polygon sobre BOG-UPL.json + coords
    de PUESTOS_GEOREF.csv

Salida:
  Bases de datos/output_observatorio_mujer/bog-upl-edadgenero.json
  Bases de datos/output_observatorio_mujer/bog-barrio-edadgenero.json

Estructura:
  { "UPL01": {nombre, sector, vot22:{muj, hom, b18, b26, b41, b60}, vot18:{...}} }
  { "008407": {nombre, loc_cod, vot22:{...}, vot18:{...}} }

Uso: python3 tools/build-observatorio-mujer/build-bog-upl-barrio.py
"""
import csv, json, urllib.request
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
EDAD_XLS = ROOT / "Bases de datos" / "Edadygenero.xlsx"
OUT_DIR  = ROOT / "Bases de datos" / "output_observatorio_mujer"
OUT_UPL  = OUT_DIR / "bog-upl-edadgenero.json"
OUT_BARR = OUT_DIR / "bog-barrio-edadgenero.json"

# S3 inputs
S3_BASE = "https://elecciones-2026.s3.us-east-1.amazonaws.com/ricardoruiz.co"
URL_UPL   = f"{S3_BASE}/congreso-2026/output/mapas-2026/Ciudades-COM-LOC/BOG-UPL.json"
URL_BARR  = f"{S3_BASE}/congreso-2026/output/mapas-2026/Ciudades-COM-LOC/BOG-BARRIOS-CATASTRALES.json"
URL_GEOREF = f"{S3_BASE}/congreso-2026/output/mapas-2026/PUESTOS_GEOREF.csv"
URL_PUESTO_BARRIO = f"{S3_BASE}/bases+de+datos/output_ciudades/bogota/bog-puesto-to-barrio.json"

# Edadygenero column indices
C_DEPTO_NUM=1; C_TIPO=3; C_ANO=4; C_MUN_NUM=5; C_COMUNA=6; C_PUESTO=7
C_HOMBRES=22; C_MUJERES=23
C_MUJ_18_20=37; C_MUJ_21_25=38; C_MUJ_26_30=39; C_MUJ_31_35=40; C_MUJ_36_40=41
C_MUJ_41_45=42; C_MUJ_46_50=43; C_MUJ_51_55=44; C_MUJ_56_60=45; C_MUJ_60P=46

TARGET_TIPO = "Presidencia 1V"
BOG_DEP_REG = 16  # Registraduría code

def num(v):
    if v is None: return 0
    try: return int(v)
    except (TypeError, ValueError):
        try: return int(float(v))
        except: return 0

def fetch_json(url):
    print(f"  fetch {url}")
    return json.loads(urllib.request.urlopen(url).read())

def fetch_csv_lines(url):
    print(f"  fetch {url}")
    return urllib.request.urlopen(url).read().decode('utf-8-sig').splitlines()

# ─── Point-in-polygon ────────────────────────────────────────────
def pip_ring(x, y, ring):
    """Ray casting sobre un anillo. ring: [[lng,lat], ...]"""
    inside = False
    j = len(ring) - 1
    for i in range(len(ring)):
        xi, yi = ring[i][0], ring[i][1]
        xj, yj = ring[j][0], ring[j][1]
        if ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / ((yj - yi) or 1e-12) + xi):
            inside = not inside
        j = i
    return inside

def pip_polygon(x, y, geom):
    """geom es GeoJSON Polygon o MultiPolygon."""
    t = geom.get('type')
    coords = geom.get('coordinates', [])
    if t == 'Polygon':
        polys = [coords]
    elif t == 'MultiPolygon':
        polys = coords
    else:
        return False
    for poly in polys:
        if not poly: continue
        outer = poly[0]
        if not pip_ring(x, y, outer): continue
        holes = poly[1:]
        in_hole = any(pip_ring(x, y, h) for h in holes)
        if not in_hole: return True
    return False

# ─── 1) Cargar geometrías + mapeos ──────────────────────────────
def norm_loc(s):
    """Normaliza nombre de localidad: mayus, sin tildes, sin puntos, trim."""
    return (s or "").upper().replace(".", "").replace(",", "").replace("Á","A").replace("É","E").replace("Í","I").replace("Ó","O").replace("Ú","U").replace("Ñ","N").strip()

def load_inputs():
    print("\n[1/4] Cargando inputs de S3")
    upl_geo  = fetch_json(URL_UPL)
    barr_geo = fetch_json(URL_BARR)
    barrio_data = {f['properties'].get('codigo'): f['properties'] for f in barr_geo.get('features',[]) if f['properties'].get('codigo')}
    upl_data    = {f['properties'].get('CODIGO_UPL'): f['properties'] for f in upl_geo.get('features',[]) if f['properties'].get('CODIGO_UPL')}
    puesto_to_barrio = fetch_json(URL_PUESTO_BARRIO)
    # PUESTOS_GEOREF csv → puesto_key=ZONA-PUESTO (zero-padded zz-pp) → (lat, lon)
    # Y también construir mapeo nombre_localidad → zona_cod para cruzar con Edadygenero (que trae nombre, no código).
    print("  parsing PUESTOS_GEOREF.csv ...")
    lines = fetch_csv_lines(URL_GEOREF)
    reader = csv.DictReader(lines, delimiter=';')
    coords_bog = {}
    loc_name_to_cod = {}
    for row in reader:
        codigo = (row.get('CÓDIGO COMPLETO') or '').strip()
        if not codigo or not codigo.startswith('16'): continue
        try:
            lat = float(row['LATITUD']); lng = float(row['LONGITUD'])
        except (KeyError, ValueError, TypeError):
            continue
        zz = (row.get('ZONA') or '').zfill(2)
        pp = (row.get('PUESTO') or '').zfill(2)
        coords_bog[f"{zz}-{pp}"] = (lat, lng)
        nombre_loc = row.get('NOMBRE COMUNA') or row.get('MUNICIPIO')
        if nombre_loc:
            loc_name_to_cod[norm_loc(nombre_loc)] = zz
    print(f"  puestos Bogotá georef: {len(coords_bog):,}")
    print(f"  localidades nombre→cod: {len(loc_name_to_cod)}")
    return upl_geo, upl_data, barrio_data, puesto_to_barrio, coords_bog, loc_name_to_cod

def puesto_to_upl(coords_bog, upl_geo):
    """Para cada puesto Bogotá, encontrar UPL via PIP."""
    print("\n[2/4] Asignando puestos a UPL (point-in-polygon)")
    mapping = {}
    features = upl_geo.get('features', [])
    n_miss = 0
    for i, (puesto_key, (lat, lng)) in enumerate(coords_bog.items()):
        found = None
        for feat in features:
            geom = feat.get('geometry') or {}
            if pip_polygon(lng, lat, geom):
                found = feat['properties'].get('CODIGO_UPL')
                break
        if found:
            mapping[puesto_key] = found
        else:
            n_miss += 1
        if (i+1) % 200 == 0: print(f"  {i+1}/{len(coords_bog)}...")
    print(f"  asignados {len(mapping):,}  sin UPL: {n_miss}")
    return mapping

# ─── 3) Stream Edadygenero ─────────────────────────────────────
def empty_year():
    return {"muj":0,"hom":0,"b18":0,"b26":0,"b41":0,"b60":0}

def stream_edadgenero(loc_name_to_cod):
    print(f"\n[3/4] Stream Edadygenero {EDAD_XLS.name}")
    wb = load_workbook(str(EDAD_XLS), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by_puesto = defaultdict(lambda: {2018: empty_year(), 2022: empty_year()})
    n = 0; n_kept = 0; n_no_loc = 0
    missing_loc_samples = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        n += 1
        if n % 100_000 == 0: print(f"  fila {n:,}...")
        try:
            if row[C_TIPO] != TARGET_TIPO: continue
            ano = row[C_ANO]
            if hasattr(ano, "year"): ano = ano.year
            else: ano = int(ano) if ano else None
            if ano not in (2018, 2022): continue
            d = num(row[C_DEPTO_NUM]); m = num(row[C_MUN_NUM])
            if d != BOG_DEP_REG or m != 1: continue
            # En Bogotá, col 6 trae nombre de localidad ("Antonio Nariño"), no código.
            zona_raw = row[C_COMUNA]
            zz = None
            if isinstance(zona_raw, str):
                zz = loc_name_to_cod.get(norm_loc(zona_raw))
                if not zz:
                    if len(missing_loc_samples) < 5:
                        missing_loc_samples.add(zona_raw)
                    n_no_loc += 1
                    continue
            else:
                zz = str(num(zona_raw)).zfill(2)
            pp = str(num(row[C_PUESTO])).zfill(2)
            key = f"{zz}-{pp}"
            b = by_puesto[key][ano]
            b["muj"] += num(row[C_MUJERES])
            b["hom"] += num(row[C_HOMBRES])
            b["b18"] += num(row[C_MUJ_18_20]) + num(row[C_MUJ_21_25])
            b["b26"] += num(row[C_MUJ_26_30]) + num(row[C_MUJ_31_35]) + num(row[C_MUJ_36_40])
            b["b41"] += num(row[C_MUJ_41_45]) + num(row[C_MUJ_46_50]) + num(row[C_MUJ_51_55]) + num(row[C_MUJ_56_60])
            b["b60"] += num(row[C_MUJ_60P])
            n_kept += 1
        except Exception:
            continue
    print(f"  filas total: {n:,}  bogotá pres 18/22: {n_kept:,}  puestos: {len(by_puesto):,}")
    if n_no_loc:
        print(f"  WARN: {n_no_loc} filas con localidad sin match. Ejemplos: {missing_loc_samples}")
    return by_puesto

# ─── 4) Agregar a UPL y Barrio ─────────────────────────────────
def aggregate(by_puesto, p_to_upl, p_to_barrio, upl_data, barrio_data):
    print("\n[4/4] Agregando a UPL y barrio")
    upl_acc  = defaultdict(lambda: {2018: empty_year(), 2022: empty_year(), "n_puestos":0})
    barr_acc = defaultdict(lambda: {2018: empty_year(), 2022: empty_year(), "n_puestos":0})
    for pk, years in by_puesto.items():
        upl_cod = p_to_upl.get(pk)
        barr_cod = p_to_barrio.get(pk)
        for y in (2018, 2022):
            src = years[y]
            for target_acc, cod in [(upl_acc, upl_cod), (barr_acc, barr_cod)]:
                if not cod: continue
                t = target_acc[cod][y]
                for k in ('muj','hom','b18','b26','b41','b60'):
                    t[k] += src[k]
        if upl_cod: upl_acc[upl_cod]["n_puestos"] += 1
        if barr_cod: barr_acc[barr_cod]["n_puestos"] += 1

    # Construir outputs con metadata
    out_upl = {}
    for cod, acc in upl_acc.items():
        info = upl_data.get(cod, {})
        out_upl[cod] = {
            "nombre": info.get("NOMBRE"),
            "sector": info.get("SECTOR"),
            "n_puestos": acc["n_puestos"],
            "vot18": acc[2018],
            "vot22": acc[2022]
        }
    out_barr = {}
    for cod, acc in barr_acc.items():
        info = barrio_data.get(cod, {})
        out_barr[cod] = {
            "nombre": info.get("nombre"),
            "loc_cod": info.get("loc_codigo"),
            "loc_nombre": info.get("loc_nombre"),
            "n_puestos": acc["n_puestos"],
            "vot18": acc[2018],
            "vot22": acc[2022]
        }
    return out_upl, out_barr

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    upl_geo, upl_data, barrio_data, p_to_barrio, coords_bog, loc_name_to_cod = load_inputs()
    p_to_upl = puesto_to_upl(coords_bog, upl_geo)
    by_puesto = stream_edadgenero(loc_name_to_cod)
    out_upl, out_barr = aggregate(by_puesto, p_to_upl, p_to_barrio, upl_data, barrio_data)

    with open(OUT_UPL, 'w', encoding='utf-8') as f:
        json.dump(out_upl, f, ensure_ascii=False, separators=(',', ':'))
    with open(OUT_BARR, 'w', encoding='utf-8') as f:
        json.dump(out_barr, f, ensure_ascii=False, separators=(',', ':'))
    print(f"\n  {OUT_UPL.name}: {len(out_upl)} UPLs, {OUT_UPL.stat().st_size:,} bytes")
    print(f"  {OUT_BARR.name}: {len(out_barr)} barrios, {OUT_BARR.stat().st_size:,} bytes")

    # Sanity Usaquen
    print("\nSanity (Usaquén UPL25):")
    s = out_upl.get('UPL25', {})
    print(f"  {s.get('nombre')}  n_puestos={s.get('n_puestos')}")
    v22 = s.get('vot22', {})
    print(f"  vot22 muj={v22.get('muj',0):,}  b18={v22.get('b18',0):,}  b60={v22.get('b60',0):,}")

if __name__ == "__main__":
    main()
