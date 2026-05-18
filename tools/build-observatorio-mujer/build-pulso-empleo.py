"""Parser del anexo DANE GEIH-MLS (mercado laboral según sexo) trimestral.

El archivo `anex-GEIHMLS-{mes1}-{mes3}{año}.xlsx` que el DANE publica
mensualmente (cada nuevo trimestre móvil agrega una columna) contiene
series 2007-presente, con tres bloques (Total / Hombres / Mujeres) en
cada hoja territorial. Los trimestres móviles son SOLAPADOS: una
columna por mes (Ene-Mar, Feb-Abr, ...). Cada columna corresponde al
mes final del TM (Ene-Mar → marzo).

Salida:
  Bases de datos/output_observatorio_mujer/dane-empleo-mensual.json

Forma:
  {
    "actualizado_a": "2026-03",
    "tms": ["2007-03","2007-04",...,"2026-03"],
    "nacional": {
      "total":   {"tgp":[...], "to":[...], "td":[...]},
      "hombres": {...},
      "mujeres": {...}
    },
    "ciudades": {
      "Bogotá D.C.": {"hombres":{...}, "mujeres":{...}},
      "Medellín A.M.": {...},
      ...
    }
  }

Uso: python3 tools/build-observatorio-mujer/build-pulso-empleo.py
"""
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "Bases de datos" / "output_observatorio_mujer" / "dane-empleo" / "anex-GEIHMLS-ene-mar2026.xlsx"
OUT = ROOT / "Bases de datos" / "output_observatorio_mujer" / "dane-empleo-mensual.json"

# Mapeo TM nombre → mes final (1..12).
MES_NUM = {
    'Ene': 1, 'Feb': 2, 'Mar': 3, 'Abr': 4, 'May': 5, 'Jun': 6,
    'Jul': 7, 'Ago': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dic': 12,
}

def parse_tm_end_month(tm_str):
    """Saca el último mes del rango TM. 'Ene - Mar' → 3. 'Nov 25 - Ene 26' → 1."""
    if not tm_str: return None
    s = str(tm_str).strip()
    # Tomar la parte después del último '-'
    parts = s.split('-')
    last = parts[-1].strip()
    # Quitar números de año pegados ('Ene 26' → 'Ene')
    last_clean = ''.join([c for c in last if c.isalpha() or c==' ' or c=='ñ'])
    last_clean = last_clean.strip().split()[0] if last_clean.strip() else ''
    return MES_NUM.get(last_clean)

def load_sheet_matrix(ws, max_row=None):
    """Lee la hoja completa a matriz [[val,...],...]. Read-only friendly."""
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
        rows.append(list(row))
    return rows

def build_tms_from_matrix(matrix, row_year, row_tm):
    """Igual que build_tms_from_header, pero sobre matriz pre-cargada."""
    out = []
    current_year = None
    yr = matrix[row_year - 1] if row_year - 1 < len(matrix) else []
    tm = matrix[row_tm - 1]   if row_tm - 1   < len(matrix) else []
    n_cols = max(len(yr), len(tm))
    for c in range(1, n_cols):  # skipping col 0 (label)
        y_cell = yr[c] if c < len(yr) else None
        if isinstance(y_cell, int) and 2000 <= y_cell <= 2050:
            current_year = y_cell
        m = parse_tm_end_month(tm[c] if c < len(tm) else None)
        if current_year is None or m is None: continue
        out.append((c, f"{current_year}-{m:02d}", tm[c] if c < len(tm) else None))
    return out

def read_series_from_matrix(matrix, row_data, tms_cols):
    """Lee la fila row_data (1-indexed → row_data-1) de la matriz."""
    r = matrix[row_data - 1] if row_data - 1 < len(matrix) else []
    out = []
    for (c, _, _) in tms_cols:
        v = r[c] if c < len(r) else None
        try: out.append(float(v) if v is not None else None)
        except (TypeError, ValueError): out.append(None)
    return out

# Para una hoja "P y T N" (Nacional): tres bloques de filas
NAC_BLOCKS = {
    "total":   { "header_year": 14, "header_tm": 15, "tgp": 17, "to": 18, "td": 19 },
    "hombres": { "header_year": 34, "header_tm": 35, "tgp": 37, "to": 38, "td": 39 },
    "mujeres": { "header_year": 54, "header_tm": 55, "tgp": 57, "to": 58, "td": 59 },
}

def process_nacional(wb):
    ws = wb['P y T N']
    print(f"  Cargando hoja 'P y T N' a matriz ...", flush=True)
    matrix = load_sheet_matrix(ws, max_row=82)
    tms_cols = build_tms_from_matrix(matrix, NAC_BLOCKS["total"]["header_year"], NAC_BLOCKS["total"]["header_tm"])
    tms_labels = [x[1] for x in tms_cols]
    out = {}
    for sexo, b in NAC_BLOCKS.items():
        out[sexo] = {
            "tgp": read_series_from_matrix(matrix, b["tgp"], tms_cols),
            "to":  read_series_from_matrix(matrix, b["to"],  tms_cols),
            "td":  read_series_from_matrix(matrix, b["td"],  tms_cols),
        }
    return tms_labels, out

def detect_city_blocks_from_matrix(matrix):
    """Lista (city_name, header_year_row(1idx), header_tm_row, data_row_start)."""
    blocks = []
    for i, row in enumerate(matrix):
        r = i + 1
        if not row: continue
        v = row[0]
        if isinstance(v, str) and v.strip() and v.strip() != 'Concepto':
            nxt = matrix[i+1][0] if i + 1 < len(matrix) and matrix[i+1] else None
            if isinstance(nxt, str) and nxt.strip() == 'Concepto':
                blocks.append((v.strip(), r + 1, r + 2, r + 3))
    return blocks

def process_ciudades(wb, sexo):
    sheet = 'Mujeres - 23 Ciud' if sexo == 'mujeres' else 'Hombres - 23 Ciud'
    ws = wb[sheet]
    print(f"  Cargando hoja '{sheet}' a matriz ...", flush=True)
    matrix = load_sheet_matrix(ws)
    print(f"    {len(matrix)} filas cargadas", flush=True)
    blocks = detect_city_blocks_from_matrix(matrix)
    out = {}
    for city, hy, htm, ds in blocks:
        tms_cols = build_tms_from_matrix(matrix, hy, htm)
        tgp_r = ds + 1
        to_r  = ds + 2
        td_r  = ds + 3
        out[city] = {
            "tgp": read_series_from_matrix(matrix, tgp_r, tms_cols),
            "to":  read_series_from_matrix(matrix, to_r,  tms_cols),
            "td":  read_series_from_matrix(matrix, td_r,  tms_cols),
        }
    return out

def main():
    print(f"Leyendo {SRC.name}")
    wb = load_workbook(str(SRC), read_only=True, data_only=True)
    tms_labels, nacional = process_nacional(wb)
    print(f"  Nacional: {len(tms_labels)} TMs, indicadores tgp/to/td x 3 sexos")
    # Para ciudades extraemos H y M y los anidamos por ciudad.
    cities_M = process_ciudades(wb, 'mujeres')
    cities_H = process_ciudades(wb, 'hombres')
    cities = {}
    for c in cities_M:
        cities[c] = { "mujeres": cities_M[c], "hombres": cities_H.get(c, {"tgp":[],"to":[],"td":[]}) }
    print(f"  Ciudades detectadas: {len(cities)}")

    actualizado = tms_labels[-1] if tms_labels else None
    out = {
        "actualizado_a": actualizado,
        "fuente": "DANE - GEIH Mercado Laboral según Sexo - Anexo trimestral mensual",
        "archivo": SRC.name,
        "tms": tms_labels,
        "nacional": nacional,
        "ciudades": cities,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',', ':'))
    print(f"\n  escrito en {OUT.name}  bytes: {OUT.stat().st_size:,}  ({OUT.stat().st_size/1024:.0f} KB)")

    # Sanity nacional último TM
    print(f"\nÚltimo TM: {actualizado}")
    print(f"  Total → TGP={nacional['total']['tgp'][-1]:.2f}  TO={nacional['total']['to'][-1]:.2f}  TD={nacional['total']['td'][-1]:.2f}")
    print(f"  Hombres → TGP={nacional['hombres']['tgp'][-1]:.2f}  TO={nacional['hombres']['to'][-1]:.2f}  TD={nacional['hombres']['td'][-1]:.2f}")
    print(f"  Mujeres → TGP={nacional['mujeres']['tgp'][-1]:.2f}  TO={nacional['mujeres']['to'][-1]:.2f}  TD={nacional['mujeres']['td'][-1]:.2f}")
    brecha = nacional['hombres']['tgp'][-1] - nacional['mujeres']['tgp'][-1]
    print(f"  Brecha TGP (H-M): {brecha:.2f}pp")

    # Sanity Bogotá
    if 'Bogotá D.C.' in cities:
        c = cities['Bogotá D.C.']
        print(f"\nBogotá D.C. último TM:")
        print(f"  Mujeres: TGP={c['mujeres']['tgp'][-1]:.2f}  TO={c['mujeres']['to'][-1]:.2f}  TD={c['mujeres']['td'][-1]:.2f}")
        print(f"  Hombres: TGP={c['hombres']['tgp'][-1]:.2f}  TO={c['hombres']['to'][-1]:.2f}  TD={c['hombres']['td'][-1]:.2f}")
    # Lista de ciudades
    print(f"\nCiudades:")
    for c in cities: print(f"  · {c}")

if __name__ == "__main__":
    main()
