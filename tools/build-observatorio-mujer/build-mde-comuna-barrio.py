"""Cohortes etarias de mujeres por comuna y barrio (Medellín) desde Edadygenero.

Procesa Edadygenero.xlsx (Pres 1V 2018+2022, dep=1 mun=1), agrega por
puesto, y cruza:
  - puesto → barrio oficial usando puestos_to_barrios.json (S3)
  - puesto → comuna política usando el mapeo zona electoral → comuna
    documentado en CLAUDE.md (estable 2015-2026):

  zona  → comuna
  01-02 → 01 Popular            17-18 → 09 Buenos Aires
  03-04 → 02 Santa Cruz         19-20 → 10 La Candelaria
  05-06 → 03 Manrique           21-22 → 11 Laureles Estadio
  07-08 → 04 Aranjuez           23-24 → 12 La América
  09-10 → 05 Castilla           25-26 → 13 San Javier
  11-12 → 06 Doce de Octubre    27-28 → 14 El Poblado
  13-14 → 07 Robledo            29    → 15 Guayabal
  15-16 → 08 Villa Hermosa      30-32 → 16 Belén
  99    → CORR (corregimientos agregados)
  90/98 → OTROS (zona consular / censo especial — se descartan)

Salida:
  Bases de datos/output_observatorio_mujer/mde-comuna-edadgenero.json
  Bases de datos/output_observatorio_mujer/mde-barrio-edadgenero.json

Uso: python3 tools/build-observatorio-mujer/build-mde-comuna-barrio.py
"""
import csv, json, urllib.request
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
EDAD_XLS = ROOT / "Bases de datos" / "Edadygenero.xlsx"
DIVI_CSV = ROOT / "Bases de datos" / "COMUNAS_DATA.csv"  # Divipole censo por puesto
OUT_DIR  = ROOT / "Bases de datos" / "output_observatorio_mujer"
OUT_COM  = OUT_DIR / "mde-comuna-edadgenero.json"
OUT_BARR = OUT_DIR / "mde-barrio-edadgenero.json"

S3_BASE  = "https://elecciones-2026.s3.us-east-1.amazonaws.com/ricardoruiz.co"
URL_P2B  = f"{S3_BASE}/bases+de+datos/puestos_to_barrios.json"

# Edadygenero columnas
C_DEPTO_NUM=1; C_TIPO=3; C_ANO=4; C_MUN_NUM=5; C_COMUNA=6; C_PUESTO=7
C_HOMBRES=22; C_MUJERES=23
C_MUJ_18_20=37; C_MUJ_21_25=38; C_MUJ_26_30=39; C_MUJ_31_35=40; C_MUJ_36_40=41
C_MUJ_41_45=42; C_MUJ_46_50=43; C_MUJ_51_55=44; C_MUJ_56_60=45; C_MUJ_60P=46

TARGET_TIPO = "Presidencia 1V"
MDE_DEP = 1
MDE_MUN = 1

# Zona electoral → comuna política (códigos 2 dígitos)
def zona_to_comuna(z):
    z = int(z)
    if z == 99: return "CORR"
    if z in (90, 98): return None  # OTROS / Consular — descartar
    if 1 <= z <= 2:   return "01"   # Popular
    if 3 <= z <= 4:   return "02"   # Santa Cruz
    if 5 <= z <= 6:   return "03"   # Manrique
    if 7 <= z <= 8:   return "04"   # Aranjuez
    if 9 <= z <= 10:  return "05"   # Castilla
    if 11 <= z <= 12: return "06"   # Doce de Octubre
    if 13 <= z <= 14: return "07"   # Robledo
    if 15 <= z <= 16: return "08"   # Villa Hermosa
    if 17 <= z <= 18: return "09"   # Buenos Aires
    if 19 <= z <= 20: return "10"   # La Candelaria
    if 21 <= z <= 22: return "11"   # Laureles Estadio
    if 23 <= z <= 24: return "12"   # La América
    if 25 <= z <= 26: return "13"   # San Javier
    if 27 <= z <= 28: return "14"   # El Poblado
    if z == 29:       return "15"   # Guayabal
    if 30 <= z <= 32: return "16"   # Belén
    return None

COMUNA_NAMES = {
    "01": "Popular", "02": "Santa Cruz", "03": "Manrique", "04": "Aranjuez",
    "05": "Castilla", "06": "Doce de Octubre", "07": "Robledo", "08": "Villa Hermosa",
    "09": "Buenos Aires", "10": "La Candelaria", "11": "Laureles Estadio",
    "12": "La América", "13": "San Javier", "14": "El Poblado", "15": "Guayabal",
    "16": "Belén", "CORR": "Corregimientos"
}

def num(v):
    if v is None: return 0
    try: return int(v)
    except (TypeError, ValueError):
        try: return int(float(v))
        except: return 0

def empty_year():
    return {"muj":0,"hom":0,"b18":0,"b26":0,"b41":0,"b60":0}

def build_censo_mde(p2b):
    """Censo Divipole 2024 agregado a comuna política + barrio Medellín."""
    print(f"\nCargando censo Divipole {DIVI_CSV.name}")
    com_censo = defaultdict(lambda: {"muj":0,"hom":0,"tot":0})
    bar_censo = defaultdict(lambda: {"muj":0,"hom":0,"tot":0})
    n_rows = 0; n_kept = 0
    with open(DIVI_CSV, encoding='utf-8-sig') as f:
        r = csv.DictReader(f, delimiter=';')
        for row in r:
            n_rows += 1
            dd = row['dd'].zfill(2); mm = row['mm'].zfill(3)
            if dd != "01" or mm != "001": continue
            zz_int = int(row['zz']); pp = row['pp'].zfill(2)
            zz = str(zz_int).zfill(2)
            ccod = zona_to_comuna(zz_int)
            if ccod is None: continue
            muj = int(row.get('mujeres') or 0)
            hom = int(row.get('hombres') or 0)
            tot = int(row.get('total') or 0)
            c = com_censo[ccod]
            c["muj"] += muj; c["hom"] += hom; c["tot"] += tot
            pkey = f"{zz}-{pp}"
            pb = p2b.get(pkey)
            if pb:
                bc = pb["codigo"]
                b = bar_censo[bc]
                b["muj"] += muj; b["hom"] += hom; b["tot"] += tot
            n_kept += 1
    print(f"  Medellín filas censo: {n_kept}  comunas: {len(com_censo)}  barrios: {len(bar_censo)}")
    return com_censo, bar_censo

def main():
    print(f"Cargando puestos_to_barrios.json")
    p2b = json.loads(urllib.request.urlopen(URL_P2B).read())
    print(f"  {len(p2b)} puestos mapeados a barrio Medellín")
    com_censo, bar_censo = build_censo_mde(p2b)

    print(f"\nStream {EDAD_XLS.name}")
    wb = load_workbook(str(EDAD_XLS), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    comuna_acc = defaultdict(lambda: {2018: empty_year(), 2022: empty_year(), "n_puestos": set()})
    barrio_acc = defaultdict(lambda: {2018: empty_year(), 2022: empty_year(), "n_puestos": set(), "nombre": "", "comuna": ""})
    n = 0; n_kept = 0; n_dropped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        n += 1
        if n % 100_000 == 0: print(f"  fila {n:,}...")
        try:
            if row[C_TIPO] != TARGET_TIPO: continue
            ano = row[C_ANO]
            if hasattr(ano, "year"): ano = ano.year
            else: ano = int(ano) if ano else None
            if ano not in (2018, 2022): continue
            d = num(row[C_DEPTO_NUM]); m = num(row[C_MUN_NUM])
            if d != MDE_DEP or m != MDE_MUN: continue
            zz_raw = num(row[C_COMUNA])
            comuna_cod = zona_to_comuna(zz_raw)
            if comuna_cod is None:
                n_dropped += 1
                continue
            zz = str(zz_raw).zfill(2)
            pp = str(num(row[C_PUESTO])).zfill(2)
            pkey = f"{zz}-{pp}"

            muj = num(row[C_MUJERES]); hom = num(row[C_HOMBRES])
            b18 = num(row[C_MUJ_18_20]) + num(row[C_MUJ_21_25])
            b26 = num(row[C_MUJ_26_30]) + num(row[C_MUJ_31_35]) + num(row[C_MUJ_36_40])
            b41 = num(row[C_MUJ_41_45]) + num(row[C_MUJ_46_50]) + num(row[C_MUJ_51_55]) + num(row[C_MUJ_56_60])
            b60 = num(row[C_MUJ_60P])

            # Acumular comuna
            ca = comuna_acc[comuna_cod][ano]
            ca["muj"] += muj; ca["hom"] += hom
            ca["b18"] += b18; ca["b26"] += b26; ca["b41"] += b41; ca["b60"] += b60
            comuna_acc[comuna_cod]["n_puestos"].add(pkey)

            # Acumular barrio (sólo si hay match en puestos_to_barrios)
            pb = p2b.get(pkey)
            if pb:
                bc = pb["codigo"]
                ba = barrio_acc[bc][ano]
                ba["muj"] += muj; ba["hom"] += hom
                ba["b18"] += b18; ba["b26"] += b26; ba["b41"] += b41; ba["b60"] += b60
                barrio_acc[bc]["n_puestos"].add(pkey)
                if not barrio_acc[bc]["nombre"]: barrio_acc[bc]["nombre"] = pb.get("nombre","")
                if not barrio_acc[bc]["comuna"]: barrio_acc[bc]["comuna"] = pb.get("comuna","")
            n_kept += 1
        except Exception as e:
            if n < 5: print(f"  ERROR fila {n}: {e}")
            continue
    print(f"  total filas: {n:,}  Medellín Pres 18/22: {n_kept:,}  dropped (90/98): {n_dropped}")
    print(f"  comunas: {len(comuna_acc)}  barrios: {len(barrio_acc)}")

    # Construir outputs (fusiona censo Divipole + cohortes Edadygenero)
    out_com = {}
    for cod, acc in sorted(comuna_acc.items()):
        cn = com_censo.get(cod, {"muj":0,"hom":0,"tot":0})
        out_com[cod] = {
            "nombre": COMUNA_NAMES.get(cod, cod),
            "n_puestos": len(acc["n_puestos"]),
            "censo_muj": cn["muj"], "censo_hom": cn["hom"], "censo_tot": cn["tot"],
            "vot18": acc[2018],
            "vot22": acc[2022]
        }
    out_barr = {}
    for cod, acc in sorted(barrio_acc.items()):
        bn = bar_censo.get(cod, {"muj":0,"hom":0,"tot":0})
        out_barr[cod] = {
            "nombre": acc["nombre"],
            "comuna": acc["comuna"],
            "n_puestos": len(acc["n_puestos"]),
            "censo_muj": bn["muj"], "censo_hom": bn["hom"], "censo_tot": bn["tot"],
            "vot18": acc[2018],
            "vot22": acc[2022]
        }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_COM, "w", encoding='utf-8') as f:
        json.dump(out_com, f, ensure_ascii=False, separators=(',', ':'))
    with open(OUT_BARR, "w", encoding='utf-8') as f:
        json.dump(out_barr, f, ensure_ascii=False, separators=(',', ':'))
    print(f"\n  {OUT_COM.name}: {len(out_com)} comunas, {OUT_COM.stat().st_size:,} bytes")
    print(f"  {OUT_BARR.name}: {len(out_barr)} barrios, {OUT_BARR.stat().st_size:,} bytes")

    # Sanity El Poblado (comuna 14)
    p = out_com.get("14", {})
    print(f"\nEl Poblado (com 14): n_puestos={p.get('n_puestos')}")
    v22 = p.get("vot22", {})
    print(f"  vot22 muj={v22.get('muj',0):,}  b18={v22.get('b18',0):,}  b60={v22.get('b60',0):,}")

if __name__ == "__main__":
    main()
