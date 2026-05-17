"""Aggregator del archivo DANE PPED-AreaSexoEdadMun por (depto, mun, año).

Proyecciones de población municipal 2018–2042, edad simple. Solo se
guarda el área 'Total' (Cabecera + Resto). Buckets de mujeres:
  12–17 (pre-electorales, entran al censo 2030–2036)
  18–28 (jóvenes)
  29–40 (adultas tempranas)
  41–59 (adultas)
  60+   (mayores)

Estructura del xlsx (verificada con peek):
  cols 0–5 : DP DPNOM MPIO DPMP AÑO ÁREA
  col  6   : TOTAL (suma H+M)
  col  7   : Hombres total
  col  8   : Mujeres total
  cols 9–109   : Hombres edad simple 0..100+
  cols 110–210 : Mujeres edad simple 0..100+
  cols 211–311 : Total edad simple 0..100+

Salida:
  Bases de datos/output_observatorio_mujer/dane-pob-mun.json
  { "05001": { nombre, depnom, dep_cod, "2018":{...}, "2026":{...}, ... } }

Uso: python3 tools/build-observatorio-mujer/build-dane-mun.py
"""
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SRC  = ROOT / "Bases de datos" / "output_observatorio_mujer" / "PPED-AreaSexoEdadMun-2018-2042_VP.xlsx"
OUT  = ROOT / "Bases de datos" / "output_observatorio_mujer" / "dane-pob-mun.json"

C_DP=0; C_DPNOM=1; C_MPIO=2; C_DPMP=3; C_ANO=4; C_AREA=5
C_TOT=6; C_HOM=7; C_MUJ=8
MUJ_BASE = 110   # col 110 = Mujeres 0 años; 110+X = Mujeres X años

# Buckets (inclusive, en años)
BUCKETS = [
    ("muj_12_17", 12, 17),
    ("muj_18_28", 18, 28),
    ("muj_29_40", 29, 40),
    ("muj_41_59", 41, 59),
    ("muj_60p",   60, 100),  # 100 incluye "100 años y más"
]

def num(v):
    if v is None: return 0
    try: return int(v)
    except (TypeError, ValueError):
        try: return int(float(v))
        except: return 0

def main():
    print(f"Leyendo {SRC} ({SRC.stat().st_size/1e6:.0f} MB) ...")
    wb = load_workbook(str(SRC), read_only=True, data_only=True)
    ws = wb['PobMunicipalxÁreaSexoEdad']

    out = {}
    n_rows = 0; n_kept = 0
    for row in ws.iter_rows(min_row=10, values_only=True):
        n_rows += 1
        if n_rows % 10_000 == 0: print(f"  fila {n_rows:,}...")

        try:
            area = row[C_AREA]
            if area != 'Total': continue
            mpio = str(row[C_MPIO] or '').strip()
            if len(mpio) != 5: continue
            ano  = row[C_ANO]
            if not isinstance(ano, int): ano = int(ano) if ano else None
            if ano is None: continue

            if mpio not in out:
                out[mpio] = {
                    "nombre": row[C_DPMP] or '',
                    "depnom": row[C_DPNOM] or '',
                    "dep_cod": str(row[C_DP] or '').zfill(2),
                    "years": {}
                }

            rec = {
                "tot": num(row[C_TOT]),
                "hom": num(row[C_HOM]),
                "muj": num(row[C_MUJ]),
            }
            for key, lo, hi in BUCKETS:
                rec[key] = sum(num(row[MUJ_BASE + x]) for x in range(lo, hi + 1))
            out[mpio]["years"][str(ano)] = rec
            n_kept += 1
        except Exception as e:
            if n_rows < 5: print(f"  ERROR fila {n_rows}: {e}")
            continue

    print(f"\n  total filas: {n_rows:,}  filas guardadas (área=Total): {n_kept:,}")
    print(f"  muns: {len(out):,}  años por mun: {len(next(iter(out.values()))['years'])}")

    # Mover years adentro como key directa para JSON más compacto
    flat = {}
    for mpio, d in sorted(out.items()):
        entry = {
            "nombre": d["nombre"],
            "depnom": d["depnom"],
            "dep_cod": d["dep_cod"],
        }
        entry.update(d["years"])  # años como keys top-level del mun
        flat[mpio] = entry

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding='utf-8') as f:
        json.dump(flat, f, ensure_ascii=False, separators=(',', ':'))
    print(f"\n  escrito en {OUT}")
    print(f"  bytes: {OUT.stat().st_size:,}  ({OUT.stat().st_size/1024/1024:.2f} MB)")

    # Sanity check
    sample_cod = "05001"  # Medellín
    if sample_cod in flat:
        d = flat[sample_cod]
        print(f"\n  Sample {sample_cod} ({d['nombre']}, {d['depnom']}):")
        for y in ['2018', '2026', '2030', '2042']:
            if y in d:
                r = d[y]
                print(f"    {y}: tot={r['tot']:>10,}  muj={r['muj']:>10,}  "
                      f"12-17={r['muj_12_17']:>7,}  18-28={r['muj_18_28']:>7,}  60+={r['muj_60p']:>7,}")

if __name__ == "__main__":
    main()
