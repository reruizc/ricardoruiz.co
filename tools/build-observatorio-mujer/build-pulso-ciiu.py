"""Parser de la hoja 'Ramas CIIU 4 N' del anexo DANE GEIH-MLS.

Población ocupada por rama de actividad económica (CIIU Rev. 4) y sexo,
serie trimestre móvil 2015-presente. Es el insumo del "abrebocas"
sectorial del módulo Pulso (estructura del empleo femenino + tendencia
simple — NO una proyección macro tipo DNP).

Estructura de la hoja (verificada):
  Bloque TOTAL    : header_year r13, header_tm r14, ocupada r16, ramas r18-31
  Bloque HOMBRES  : header_year r35, header_tm r36, ocupada r38, ramas r40-53
  Bloque MUJERES  : header_year r57, header_tm r58, ocupada r60, ramas r62-75
  (cada bloque: fila ocupada + 'No informa' + 14 ramas)

Salida:
  Bases de datos/output_observatorio_mujer/dane-ciiu-mensual.json

Uso: python3 tools/build-observatorio-mujer/build-pulso-ciiu.py
"""
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "Bases de datos" / "output_observatorio_mujer" / "dane-empleo" / "anex-GEIHMLS-mar-may2026.xlsx"
OUT = ROOT / "Bases de datos" / "output_observatorio_mujer" / "dane-ciiu-mensual.json"

MES_NUM = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}

# Nombres cortos para las 14 ramas (en orden de la hoja, fila ocupada + "No informa" excluidos)
RAMAS_CORTAS = [
    "Agricultura y pesca",
    "Minas y canteras",
    "Manufactura",
    "Electricidad, gas y agua",
    "Construcción",
    "Comercio",
    "Alojamiento y comida",
    "Transporte",
    "Información y comunic.",
    "Finanzas y seguros",
    "Inmobiliarias",
    "Profesionales y técnicas",
    "Admin. pública, educ. y salud",
    "Arte y entretenimiento",
]

# (header_year_row, header_tm_row, ocupada_row, primera_rama_row) 1-indexed
BLOCKS = {
    "total":   (13, 14, 16, 18),
    "hombres": (35, 36, 38, 40),
    "mujeres": (57, 58, 60, 62),
}

def parse_tm_end_month(tm):
    if not tm: return None
    last = str(tm).strip().split('-')[-1].strip()
    name = ''.join(c for c in last if c.isalpha() or c==' ').strip().split()
    return MES_NUM.get(name[0]) if name else None

def build_tms(matrix, ry, rtm):
    out = []
    cur = None
    yr = matrix[ry-1] if ry-1 < len(matrix) else []
    tm = matrix[rtm-1] if rtm-1 < len(matrix) else []
    n = max(len(yr), len(tm))
    for c in range(1, n):
        y = yr[c] if c < len(yr) else None
        if isinstance(y, int) and 2000 <= y <= 2050: cur = y
        m = parse_tm_end_month(tm[c] if c < len(tm) else None)
        if cur is None or m is None: continue
        out.append((c, f"{cur}-{m:02d}"))
    return out

def read_row(matrix, r, cols):
    row = matrix[r-1] if r-1 < len(matrix) else []
    out = []
    for (c,_) in cols:
        v = row[c] if c < len(row) else None
        try: out.append(round(float(v),2) if v is not None else None)
        except (TypeError,ValueError): out.append(None)
    return out

def main():
    print(f"Leyendo {SRC.name} hoja 'Ramas CIIU 4 N'")
    wb = load_workbook(str(SRC), read_only=True, data_only=True)
    ws = wb['Ramas CIIU 4 N']
    matrix = [list(r) for r in ws.iter_rows(min_row=1, max_row=87, values_only=True)]

    tms_cols = build_tms(matrix, *BLOCKS["mujeres"][:2])
    tms = [x[1] for x in tms_cols]
    print(f"  TMs: {len(tms)}  ({tms[0]} → {tms[-1]})")

    out = {
        "actualizado_a": tms[-1] if tms else None,
        "fuente": "DANE GEIH-MLS · Población ocupada por rama CIIU Rev.4 y sexo",
        "tms": tms,
        "ramas": RAMAS_CORTAS,
        "hombres": {}, "mujeres": {}, "total": {},
    }
    for sexo, (ry, rtm, occ_r, first_rama) in BLOCKS.items():
        cols = build_tms(matrix, ry, rtm)
        out[sexo]["_ocupada"] = read_row(matrix, occ_r, cols)
        # ocupada (occ_r), No informa (occ_r+1), luego 14 ramas
        for i, nombre in enumerate(RAMAS_CORTAS):
            out[sexo][nombre] = read_row(matrix, first_rama + i, cols)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',',':'))
    print(f"\n  escrito {OUT.name}  bytes: {OUT.stat().st_size:,}  ({OUT.stat().st_size/1024:.0f} KB)")

    # Sanity último TM: feminización por rama
    print(f"\nÚltimo TM {tms[-1]} · ocupadas por rama (miles) y % feminización:")
    def last(a):
        for v in reversed(a):
            if v is not None: return v
        return 0
    rows = []
    for nombre in RAMAS_CORTAS:
        mu = last(out["mujeres"][nombre]); ho = last(out["hombres"][nombre])
        tot = mu + ho
        pctf = (mu / tot * 100) if tot else 0
        rows.append((nombre, mu, ho, pctf))
    for n, mu, ho, p in sorted(rows, key=lambda x: -x[3]):
        print(f"  {n:32s} ♀{mu:8.0f}k  ♂{ho:8.0f}k  fem={p:5.1f}%")

if __name__ == "__main__":
    main()
