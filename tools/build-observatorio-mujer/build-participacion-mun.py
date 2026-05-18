"""Aggregator mun-level combinando Divipole 2024 + Edadygenero 2018/2022 Pres 1V.

Para el drill municipal del módulo de participación: emite un único JSON
keyado por código Registraduría 5 dígitos (dep 2 + mun 3 con padding)
con censo electoral femenino y participación histórica + cohortes
etarias de las votantes.

Salida:
  Bases de datos/output_observatorio_mujer/participacion-mun.json
  Estructura:
    "01-001": {
        "nombre": "Medellín", "depnom": "ANTIOQUIA", "dep_cod": "01",
        "censo_muj": ..., "censo_hom": ..., "censo_tot": ...,
        "vot22_muj": ..., "vot22_hom": ...,
        "v22_b18": ..., "v22_b26": ..., "v22_b41": ..., "v22_b60": ...,
        "vot18_muj": ..., "vot18_hom": ...,
        "v18_b18": ..., "v18_b26": ..., "v18_b41": ..., "v18_b60": ...
    }

Uso: python3 tools/build-observatorio-mujer/build-participacion-mun.py
"""
import csv, json
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DIVI_CSV = ROOT / "Bases de datos" / "COMUNAS_DATA.csv"
EDAD_XLS = ROOT / "Bases de datos" / "Edadygenero.xlsx"
OUT      = ROOT / "Bases de datos" / "output_observatorio_mujer" / "participacion-mun.json"

# Edadygenero columnas
C_DEPTO_NUM=1; C_DEPTO_NAME=2; C_TIPO=3; C_ANO=4; C_MUN_NUM=5
C_HOMBRES=22; C_MUJERES=23
C_MUJ_18_20=37; C_MUJ_21_25=38; C_MUJ_26_30=39; C_MUJ_31_35=40; C_MUJ_36_40=41
C_MUJ_41_45=42; C_MUJ_46_50=43; C_MUJ_51_55=44; C_MUJ_56_60=45; C_MUJ_60P=46

TARGET_TIPO = "Presidencia 1V"

def num(v):
    if v is None: return 0
    try: return int(v)
    except (TypeError, ValueError):
        try: return int(float(v))
        except: return 0

def title(s):
    if not s: return ""
    return " ".join(w.capitalize() for w in s.split())

def build_censo():
    """Censo Divipole 2024 agregado por (dep_reg, mun_reg)."""
    print(f"\n[1/2] Censo desde {DIVI_CSV}")
    agg = defaultdict(lambda: {"nombre":"", "depnom":"", "muj":0, "hom":0, "tot":0})
    with open(DIVI_CSV, encoding='utf-8-sig') as f:
        r = csv.DictReader(f, delimiter=';')
        for row in r:
            dd = row['dd'].zfill(2); mm = row['mm'].zfill(3)
            key = f"{dd}-{mm}"
            b = agg[key]
            b["muj"] += num(row.get('mujeres'))
            b["hom"] += num(row.get('hombres'))
            b["tot"] += num(row.get('total'))
            if not b["nombre"]: b["nombre"] = title(row.get('municipio',''))
            if not b["depnom"]: b["depnom"] = row.get('departamento','').upper()
    print(f"  {len(agg)} muns con censo")
    return agg

def build_edadgenero(censo):
    """Suma cohortes etarias de mujeres por (dep_reg, mun_reg) para 2018 y 2022 Pres 1V."""
    print(f"\n[2/2] Edadygenero desde {EDAD_XLS}")
    wb = load_workbook(str(EDAD_XLS), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    def empty(): return {"muj":0,"hom":0,"b18":0,"b26":0,"b41":0,"b60":0}
    vot22 = defaultdict(empty)
    vot18 = defaultdict(empty)

    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        n += 1
        if n % 100_000 == 0: print(f"  fila {n:,}...")
        try:
            if row[C_TIPO] != TARGET_TIPO: continue
            ano = row[C_ANO]
            if hasattr(ano, "year"): ano = ano.year
            else: ano = int(ano) if ano else None
            if ano not in (2018, 2022): continue

            d = num(row[C_DEPTO_NUM]); m = num(row[C_MUN_NUM])
            if d <= 0 or m <= 0: continue
            key = f"{str(d).zfill(2)}-{str(m).zfill(3)}"
            agg = vot22 if ano == 2022 else vot18
            b = agg[key]
            b["muj"] += num(row[C_MUJERES])
            b["hom"] += num(row[C_HOMBRES])
            b["b18"] += num(row[C_MUJ_18_20]) + num(row[C_MUJ_21_25])
            b["b26"] += num(row[C_MUJ_26_30]) + num(row[C_MUJ_31_35]) + num(row[C_MUJ_36_40])
            b["b41"] += num(row[C_MUJ_41_45]) + num(row[C_MUJ_46_50]) + num(row[C_MUJ_51_55]) + num(row[C_MUJ_56_60])
            b["b60"] += num(row[C_MUJ_60P])
        except Exception as e:
            if n < 5: print(f"  ERROR fila {n}: {e}")
    print(f"  filas: {n:,}  muns 2022: {len(vot22):,}  muns 2018: {len(vot18):,}")
    return vot22, vot18

def main():
    censo = build_censo()
    vot22, vot18 = build_edadgenero(censo)

    keys = set(censo.keys()) | set(vot22.keys()) | set(vot18.keys())
    out = {}
    for k in sorted(keys):
        c = censo.get(k, {"nombre":"","depnom":"","muj":0,"hom":0,"tot":0})
        v22 = vot22.get(k, {"muj":0,"hom":0,"b18":0,"b26":0,"b41":0,"b60":0})
        v18 = vot18.get(k, {"muj":0,"hom":0,"b18":0,"b26":0,"b41":0,"b60":0})
        dep_cod = k.split('-')[0]
        out[k] = {
            "nombre":   c["nombre"],
            "depnom":   c["depnom"],
            "dep_cod":  dep_cod,
            "censo_muj": c["muj"], "censo_hom": c["hom"], "censo_tot": c["tot"],
            "vot22_muj": v22["muj"], "vot22_hom": v22["hom"],
            "v22_b18": v22["b18"], "v22_b26": v22["b26"], "v22_b41": v22["b41"], "v22_b60": v22["b60"],
            "vot18_muj": v18["muj"], "vot18_hom": v18["hom"],
            "v18_b18": v18["b18"], "v18_b26": v18["b26"], "v18_b41": v18["b41"], "v18_b60": v18["b60"]
        }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',', ':'))
    print(f"\nEscrito en {OUT}")
    print(f"  muns: {len(out):,}  bytes: {OUT.stat().st_size:,}  ({OUT.stat().st_size/1024/1024:.2f} MB)")

    # Sanity
    print("\nSample (Medellín · 01-001):")
    sample = out.get("01-001")
    if sample:
        print(f"  nombre={sample['nombre']}  depnom={sample['depnom']}")
        print(f"  censo: {sample['censo_muj']:>9,} muj  +  {sample['censo_hom']:>9,} hom = {sample['censo_tot']:>10,}")
        print(f"  vot22: {sample['vot22_muj']:>9,} muj  cohortes b18={sample['v22_b18']:,}  b60={sample['v22_b60']:,}")
        print(f"  vot18: {sample['vot18_muj']:>9,} muj  cohortes b18={sample['v18_b18']:,}  b60={sample['v18_b60']:,}")

if __name__ == "__main__":
    main()
