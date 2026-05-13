#!/usr/bin/env python3
"""
tools/build-censo-divipole.py

Construye agregados del censo electoral a partir de un Divipole xlsx
oficial de la Registraduría (cualquier año). Genera:

  - {out}/puestos-censos-agg-{year}.json     ← nacional, por municipio
      { porMun: { 'dd-mm': int }, nacional: int, year: int }

  - {out}/output_ciudades/{city}/censo-comuna-{year}.json   ← 14 capitales
      Misma shape que el output de build-censo-comuna-ciudades.js
      (por_comuna, ciudad_total, n_comunas, n_corregimientos).

Uso:
  python3 tools/build-censo-divipole.py \
    "Bases de datos/Divipol 23.09.2021.xlsx" 2022 \
    "Bases de datos"

Pre-requisito: pip install openpyxl
"""

import json, sys, re, os
from collections import defaultdict
import openpyxl

CITIES = {
    'medellin':      {'depCod':'01', 'munCod':'001'},
    'bogota':        {'depCod':'16', 'munCod':'001'},
    'cali':          {'depCod':'31', 'munCod':'001'},
    'barranquilla':  {'depCod':'03', 'munCod':'001'},
    'ibague':        {'depCod':'29', 'munCod':'001'},
    'manizales':     {'depCod':'09', 'munCod':'001'},
    'pereira':       {'depCod':'24', 'munCod':'001'},
    'monteria':      {'depCod':'13', 'munCod':'001'},
    'bucaramanga':   {'depCod':'27', 'munCod':'001'},
    'cucuta':        {'depCod':'25', 'munCod':'001'},
    'neiva':         {'depCod':'19', 'munCod':'001'},
    'popayan':       {'depCod':'11', 'munCod':'001'},
    'sincelejo':     {'depCod':'28', 'munCod':'001'},
    'villavicencio': {'depCod':'52', 'munCod':'001'},
}

CORR_RE = re.compile(r'CORREGIMIENTO|CORR\.?\s+|\bVEREDA|\bRURAL\b', re.IGNORECASE)


def pad2(s):  return str(s).zfill(2)
def pad3(s):  return str(s).zfill(3)


def parse_comuna(raw):
    """`01COMUNA 1 POPULAR` → ('01', 'COMUNA 1 POPULAR')."""
    if not raw: return None
    m = re.match(r'^\s*(\d{2})\s*(.*)$', str(raw))
    if not m: return None
    cod = m.group(1)
    nombre = re.sub(r'\s+', ' ', m.group(2).strip())
    return cod, nombre


def main():
    if len(sys.argv) < 4:
        print('Uso: build-censo-divipole.py <xlsx> <year> <base-out-dir>', file=sys.stderr)
        sys.exit(1)
    xlsx_path, year, base_out = sys.argv[1], int(sys.argv[2]), sys.argv[3]
    if not os.path.exists(xlsx_path):
        print(f'No existe: {xlsx_path}', file=sys.stderr); sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row (col[0] == 'dd').
    header_row = None
    header_idx = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if row and str(row[0]).strip().lower() == 'dd':
            header_row = i
            for j, h in enumerate(row):
                if h is not None:
                    header_idx[str(h).strip().lower()] = j
            break
    if header_row is None:
        print('No encontró fila header (debe empezar con "dd").', file=sys.stderr); sys.exit(1)

    iDd = header_idx['dd']; iMm = header_idx['mm']
    iTot = header_idx['total']
    iComuna = header_idx.get('comuna')
    iMuj = header_idx.get('mujeres')
    iHom = header_idx.get('hombres')

    # Acumuladores
    por_mun = defaultdict(int)
    nacional = 0
    city_by_com = {k: {'byCom': {}, 'ciudad_total': 0, 'city': CITIES[k]} for k in CITIES}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or row[iDd] is None: continue
        dd = pad2(row[iDd]); mm = pad3(row[iMm])
        try: tot = int(row[iTot] or 0)
        except (TypeError, ValueError): tot = 0
        if tot <= 0: continue

        por_mun[f'{dd}-{mm}'] += tot
        nacional += tot

        # ¿es una de las 14 capitales?
        city_key = next((k for k, c in CITIES.items()
                         if c['depCod'] == dd and c['munCod'] == mm), None)
        if city_key is None or iComuna is None: continue
        parsed = parse_comuna(row[iComuna])
        if not parsed: continue
        cod, nombre = parsed
        muj = int(row[iMuj] or 0) if iMuj is not None else 0
        hom = int(row[iHom] or 0) if iHom is not None else 0

        bucket = city_by_com[city_key]['byCom']
        if cod not in bucket:
            bucket[cod] = {
                'comCod': cod, 'nombre': nombre,
                'tipo': 'corregimiento' if CORR_RE.search(nombre) else 'comuna',
                'censo': 0, 'mujeres': 0, 'hombres': 0, 'n_puestos': 0,
            }
        b = bucket[cod]
        b['censo']    += tot
        b['mujeres']  += muj
        b['hombres']  += hom
        b['n_puestos'] += 1
        city_by_com[city_key]['ciudad_total'] += tot

    os.makedirs(base_out, exist_ok=True)

    # 1) Nacional mun-level
    nac_path = os.path.join(base_out, f'puestos-censos-agg-{year}.json')
    with open(nac_path, 'w') as f:
        json.dump({'year': year, 'nacional': nacional, 'porMun': dict(por_mun)}, f)
    print(f'✓ nacional · {len(por_mun)} muns · censo total {nacional:,}')

    # 2) Ciudades comuna-level
    out_cities = os.path.join(base_out, 'output_ciudades')
    for city_key, r in city_by_com.items():
        cods = sorted(r['byCom'].keys())
        if not cods:
            print(f'[skip] {city_key} sin filas'); continue
        n_urb = sum(1 for c in cods if r['byCom'][c]['tipo'] != 'corregimiento')
        n_corr = len(cods) - n_urb
        city_dir = os.path.join(out_cities, city_key)
        os.makedirs(city_dir, exist_ok=True)
        out = {
            'city': city_key,
            'depCod': r['city']['depCod'],
            'munCod': r['city']['munCod'],
            'year': year,
            'ciudad_total': r['ciudad_total'],
            'n_comunas': n_urb,
            'n_corregimientos': n_corr,
            'por_comuna': r['byCom'],
        }
        with open(os.path.join(city_dir, f'censo-comuna-{year}.json'), 'w') as f:
            json.dump(out, f)
        suf = f' + {n_corr} corregimientos' if n_corr else ''
        print(f'✓ {city_key:14s} {n_urb:3d} comunas{suf} · censo total {r["ciudad_total"]:,}')


if __name__ == '__main__':
    main()
