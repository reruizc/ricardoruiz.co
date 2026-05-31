#!/usr/bin/env python3
"""
Genera XLSX 'amigables' para consultores: además de los códigos GCS originales,
agrega columnas con NOMBRES de departamento, municipio, puesto y comuna,
cruzando contra Bases de datos/PUESTOS_GEOREF.csv.

Columnas nuevas al final: DES_DDE, DES_MME, DES_PP, COD_COMUNA, DES_COMUNA.

Uso:
  python3 tools/build-xlsx-con-nombres/build.py [archivo1.csv [archivo2.csv ...]]

Si no se pasan archivos, procesa por default PRES1V 2022 y PRES2V 2022.

Sube los xlsx a S3 con sufijo `_CON_NOMBRES.xlsx` al lado del original.
"""
import csv
import re
import subprocess
import sys
import time
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook

SRC = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos")
GCS_DIR = SRC / "FINAL SUBIDA GCS"
GEOREF = SRC / "PUESTOS_GEOREF.csv"          # aporta el BARRIO
COMUNAS = SRC / "COMUNAS_DATA.csv"           # censo Divipole oficial: depto/mun/puesto/comuna (nacional + exterior)
OUT = SRC / "output_xlsx_con_nombres"
OUT.mkdir(parents=True, exist_ok=True)

_NUM_PREFIX = re.compile(r"^\d+")
def _strip_code_prefix(s: str) -> str:
    """'01COMUNA 1 POPULAR' → 'COMUNA 1 POPULAR'."""
    return _NUM_PREFIX.sub("", s or "").strip()

S3_BASE = "s3://elecciones-2026/ricardoruiz.co/DESCARGAS/raw"

# (gcs_name, categoria_s3, año)
DEFAULTS = [
    ("GCS_2022PRES1V.csv", "presidencial-1v", 2022),
    ("GCS_2022PRES2V.csv", "presidencial-2v", 2022),
]

SHEET_ROWS = 1_000_000


NO_DISP = "No disponible"

# Fallback hardcoded para deptos especiales de Registraduría que no aparecen
# en PUESTOS_GEOREF (consulados, atípicos, antiguos). Se complementa con el
# mapeo dinámico que se construye desde el propio CSV.
DEPTO_FALLBACK = {
    "88": "CONSULADOS",
    "99": "NO REGISTRADO",
}


def load_geo_lookup():
    """Construye los mapeos de nombres desde DOS fuentes:

      • COMUNAS_DATA.csv (censo Divipole oficial) → fuente PRIMARIA de
        depto/municipio/puesto/comuna. Cobertura nacional + exterior
        (consulados con país en la columna municipio).
      • PUESTOS_GEOREF.csv → aporta el BARRIO (que el censo no trae).

    Devuelve:
      1. lookup_full: (dde,mme,zz,pp) → {depto,municipio,puesto,comuna,barrio}
      2. depto_by_cod: COD_DDE → nombre depto (fallback estable por código)
      3. mun_by_cod:   (COD_DDE, COD_MME) → nombre municipio (fallback)
    """
    lookup = {}
    depto_by_cod = dict(DEPTO_FALLBACK)
    mun_by_cod = {}

    # ── Fuente primaria: COMUNAS_DATA (depto/municipio/puesto/comuna) ──
    print(f"Cargando {COMUNAS.name}…", flush=True)
    with COMUNAS.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            dde = (row.get("dd") or "").strip().zfill(2)
            mme = (row.get("mm") or "").strip().zfill(3)
            zz  = (row.get("zz") or "").strip().zfill(2)
            pp  = (row.get("pp") or "").strip().zfill(2)
            if not (dde and mme):
                continue
            depto_name = (row.get("departamento") or "").strip()
            mun_name   = (row.get("municipio") or "").strip()
            puesto     = (row.get("puesto") or "").strip()
            comuna     = _strip_code_prefix(row.get("comuna") or "")
            if depto_name and dde not in depto_by_cod:
                depto_by_cod[dde] = depto_name
            if mun_name and (dde, mme) not in mun_by_cod:
                mun_by_cod[(dde, mme)] = mun_name
            lookup[(dde, mme, zz, pp)] = {
                "depto":     depto_name,
                "municipio": mun_name,
                "puesto":    puesto,
                "comuna":    comuna,
                "barrio":    "",   # se completa con georef abajo
            }
    print(f"  → {len(lookup):,} puestos (censo) · {len(depto_by_cod)} deptos · {len(mun_by_cod):,} muns", flush=True)

    # ── Fuente de barrio: PUESTOS_GEOREF ──
    print(f"Cargando {GEOREF.name} (barrio)…", flush=True)
    geo_barrios = 0
    with GEOREF.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            codigo = (row.get("CÓDIGO COMPLETO") or "").strip()
            if len(codigo) != 9:
                continue
            try:
                dde, mme, zz, pp = codigo[0:2], codigo[2:5], codigo[5:7], codigo[7:9]
                int(dde); int(mme); int(zz); int(pp)
            except ValueError:
                continue
            barrio = (row.get("BARRIO") or "").strip() or (row.get("barrio.1") or "").strip()
            key = (dde, mme, zz, pp)
            if key in lookup:
                if barrio:
                    lookup[key]["barrio"] = barrio
                    geo_barrios += 1
            else:
                # Puesto en georef que no estaba en el censo: lo agregamos completo
                nom_comuna = _strip_code_prefix(row.get("NOMBRE COMUNA") or "")
                lookup[key] = {
                    "depto":     (row.get("DEPARTAMENTO") or "").strip(),
                    "municipio": (row.get("MUNICIPIO") or "").strip(),
                    "puesto":    (row.get("NOMBRE PUESTO") or "").strip(),
                    "comuna":    nom_comuna,
                    "barrio":    barrio,
                }
    print(f"  → {geo_barrios:,} barrios cruzados", flush=True)
    return lookup, depto_by_cod, mun_by_cod


def resolve_names(dde, mme, zz, pp, lookup_full, depto_by_cod, mun_by_cod):
    """Resuelve los 5 campos de nombre con cascada:
       1. match exacto del puesto → todos los valores del lookup_full
       2. para depto/municipio sin match exacto: fallback a los mapeos por
          código (siempre disponible para los 33 deptos y municipios
          que aparecen al menos una vez en georef).
       3. cualquier campo que quede vacío → NO_DISP.
    """
    info = lookup_full.get((dde, mme, zz, pp))
    if info:
        depto     = info["depto"]     or depto_by_cod.get(dde, "") or NO_DISP
        municipio = info["municipio"] or mun_by_cod.get((dde, mme), "") or NO_DISP
        puesto    = info["puesto"]    or NO_DISP
        comuna    = info["comuna"]    or NO_DISP
        barrio    = info["barrio"]    or NO_DISP
        return depto, municipio, puesto, comuna, barrio, True

    # Sin match exacto: rellenar lo que se pueda por código
    depto     = depto_by_cod.get(dde, "") or NO_DISP
    municipio = mun_by_cod.get((dde, mme), "") or NO_DISP
    return depto, municipio, NO_DISP, NO_DISP, NO_DISP, False


def enrich_por_puesto(gcs_in: Path, xlsx_out: Path, lookup_data):
    """Agregado por puesto + cruce con nombres. Output SIN códigos (sólo
    nombres amigables) y con encabezados legibles para consultores."""
    lookup_full, depto_by_cod, mun_by_cod = lookup_data
    misses = 0
    hits = 0

    with gcs_in.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";")
        headers = next(reader)
        col = {h: i for i, h in enumerate(headers)}
        # Columnas que necesitamos del GCS original
        for needed in ("COD_DDE","COD_MME","COD_ZZ","COD_PP","NUM_VOT",
                       "DES_COR","DES_CIR","DES_PAR","DES_CAN","FUENTE","FEC_ELEC"):
            if needed not in col:
                raise SystemExit(f"Falta columna en {gcs_in.name}: {needed}")

        # Groupby por (dde, mme, zz, pp, partido, candidato, corporación,
        # circunscripción, fuente, fecha). DES_MS se descarta porque al
        # agregar por puesto no aplica.
        agg = defaultdict(int)
        for row in reader:
            try:
                vot = int(row[col["NUM_VOT"]].strip() or "0")
            except (ValueError, IndexError):
                vot = 0
            try:
                key = (
                    row[col["FUENTE"]],
                    row[col["FEC_ELEC"]],
                    row[col["DES_COR"]],
                    row[col["DES_CIR"]],
                    row[col["COD_DDE"]].strip().zfill(2),
                    row[col["COD_MME"]].strip().zfill(3),
                    row[col["COD_ZZ"]].strip().zfill(2),
                    row[col["COD_PP"]].strip().zfill(2),
                    row[col["DES_PAR"]],
                    row[col["DES_CAN"]],
                )
            except IndexError:
                continue
            agg[key] += vot

    out_header = [
        "Fuente", "Fecha elección", "Corporación", "Circunscripción",
        "Departamento", "Municipio", "Comuna", "Puesto", "Barrio",
        "Partido", "Candidato", "Votos",
    ]

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Datos")
    ws.append(out_header)
    sheet_idx = 1
    rows_in_sheet = 1

    for key, votos in agg.items():
        (fuente, fecha, corp, circ, dde, mme, zz, pp, partido, candidato) = key
        depto, municipio, puesto, comuna, barrio, hit = resolve_names(
            dde, mme, zz, pp, lookup_full, depto_by_cod, mun_by_cod
        )
        if hit:
            hits += 1
        else:
            misses += 1
        enriched = [
            fuente, fecha, corp, circ,
            depto, municipio, comuna, puesto, barrio,
            partido, candidato, votos,
        ]
        if rows_in_sheet >= SHEET_ROWS:
            sheet_idx += 1
            ws = wb.create_sheet(f"Datos {sheet_idx}")
            ws.append(out_header)
            rows_in_sheet = 1
        ws.append(enriched)
        rows_in_sheet += 1

    wb.save(xlsx_out)
    return {
        "puestos_filas": len(agg),
        "hits": hits, "misses": misses,
        "match_pct": round(hits / (hits + misses) * 100, 2) if (hits + misses) else 0,
        "mb": round(xlsx_out.stat().st_size / (1024 * 1024), 2),
        "sheets": sheet_idx,
    }


def enrich(gcs_in: Path, xlsx_out: Path, lookup_data):
    """Por mesa + cruce con nombres. Sólo columnas amigables (sin códigos)."""
    lookup_full, depto_by_cod, mun_by_cod = lookup_data
    misses = 0
    hits = 0

    out_header = [
        "Fuente", "Fecha elección", "Corporación", "Circunscripción",
        "Departamento", "Municipio", "Comuna", "Puesto", "Barrio", "Mesa",
        "Partido", "Candidato", "Votos",
    ]

    with gcs_in.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";")
        headers = next(reader)
        col = {h: i for i, h in enumerate(headers)}
        for needed in ("COD_DDE","COD_MME","COD_ZZ","COD_PP","NUM_VOT",
                       "DES_COR","DES_CIR","DES_PAR","DES_CAN","DES_MS",
                       "FUENTE","FEC_ELEC"):
            if needed not in col:
                raise SystemExit(f"Falta columna en {gcs_in.name}: {needed}")

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Datos")
        ws.append(out_header)
        rows_in_sheet = 1
        sheet_idx = 1

        for row in reader:
            try:
                dde = row[col["COD_DDE"]].strip().zfill(2)
                mme = row[col["COD_MME"]].strip().zfill(3)
                zz  = row[col["COD_ZZ"]].strip().zfill(2)
                pp  = row[col["COD_PP"]].strip().zfill(2)
            except IndexError:
                continue
            depto, municipio, puesto, comuna, barrio, hit = resolve_names(
                dde, mme, zz, pp, lookup_full, depto_by_cod, mun_by_cod
            )
            if hit: hits += 1
            else:   misses += 1
            try:
                votos_raw = row[col["NUM_VOT"]].strip()
                votos = int(votos_raw) if votos_raw else 0
            except (ValueError, IndexError):
                votos = 0
            enriched = [
                row[col["FUENTE"]], row[col["FEC_ELEC"]],
                row[col["DES_COR"]], row[col["DES_CIR"]],
                depto, municipio, comuna, puesto, barrio,
                row[col["DES_MS"]],
                row[col["DES_PAR"]], row[col["DES_CAN"]], votos,
            ]
            if rows_in_sheet >= SHEET_ROWS:
                sheet_idx += 1
                ws = wb.create_sheet(f"Datos {sheet_idx}")
                ws.append(out_header)
                rows_in_sheet = 1
            ws.append(enriched)
            rows_in_sheet += 1

        wb.save(xlsx_out)
    return {
        "hits": hits, "misses": misses,
        "match_pct": round(hits / (hits + misses) * 100, 2) if (hits + misses) else 0,
        "mb": round(xlsx_out.stat().st_size / (1024 * 1024), 2),
        "sheets": sheet_idx,
    }


def s3_upload(local_path: Path, cat: str, año: int):
    s3_key = f"{S3_BASE}/{cat}/{año}/{local_path.name}"
    cmd = ["aws", "s3", "cp", str(local_path), s3_key,
           "--content-type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
           "--cache-control", "public, max-age=86400",
           "--no-progress"]
    try:
        subprocess.run(cmd, check=True, capture_output=True, text=True)
        return s3_key.replace("s3://elecciones-2026/", "https://elecciones-2026.s3.us-east-1.amazonaws.com/")
    except subprocess.CalledProcessError as e:
        print(f"   S3 upload FAIL ({local_path.name}): {e.stderr.strip()[:200]}", flush=True)
        return None


def main():
    only = [a for a in sys.argv[1:] if not a.startswith("--")]
    upload   = "--no-upload" not in sys.argv
    por_puesto = "--puesto" in sys.argv

    archivos = DEFAULTS
    if only:
        archivos = [(n, c, a) for (n, c, a) in DEFAULTS if n in only]
        for n in only:
            if not any(x[0] == n for x in archivos):
                cat, año = ("custom", 0)
                archivos.append((n, cat, año))

    lookup_data = load_geo_lookup()  # tuple (lookup_full, depto_by_cod, mun_by_cod)
    suffix = "_PUESTO_CON_NOMBRES.xlsx" if por_puesto else "_CON_NOMBRES.xlsx"
    mode = "POR PUESTO + NOMBRES" if por_puesto else "POR MESA + NOMBRES"
    print(f"Modo: {mode}", flush=True)

    for idx, (csv_name, cat, año) in enumerate(archivos, 1):
        csv_in = GCS_DIR / csv_name
        if not csv_in.exists():
            print(f"[{idx}/{len(archivos)}] SKIP (no existe): {csv_name}")
            continue
        out_dir = OUT / cat / str(año)
        out_dir.mkdir(parents=True, exist_ok=True)
        out_name = csv_name.replace(".csv", suffix)
        xlsx_out = out_dir / out_name
        print(f"[{idx}/{len(archivos)}] {csv_name} → {out_name}", flush=True)
        t0 = time.time()
        if por_puesto:
            stats = enrich_por_puesto(csv_in, xlsx_out, lookup_data)
            stats["sec"] = round(time.time() - t0, 1)
            print(f"   OK · {stats['puestos_filas']:,} filas (puestos×candidatos) · {stats['hits']:,} hits / {stats['misses']:,} misses ({stats['match_pct']}% match) · {stats['mb']} MB · {stats['sheets']} hojas · {stats['sec']}s", flush=True)
        else:
            stats = enrich(csv_in, xlsx_out, lookup_data)
            stats["sec"] = round(time.time() - t0, 1)
            print(f"   OK · {stats['hits']:,} hits / {stats['misses']:,} misses ({stats['match_pct']}% match) · {stats['mb']} MB · {stats['sheets']} hojas · {stats['sec']}s", flush=True)
        if upload:
            url = s3_upload(xlsx_out, cat, año)
            if url:
                print(f"   ↑ {url}", flush=True)


if __name__ == "__main__":
    main()
