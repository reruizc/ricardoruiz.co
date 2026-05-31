#!/usr/bin/env python3
"""
Agrega cada CSV crudo GCS de la Registraduría a nivel PUESTO (no mesa).

Salida: GCS_<año><tipo>_PUESTO.csv + .xlsx, paralelos al original,
con groupby (FUENTE, FEC_ELEC, COD_COR, DES_COR, COD_CIR, DES_CIR,
COD_DDE, COD_MME, COD_ZZ, COD_PP, COD_PAR, DES_PAR, COD_CAN, DES_CAN)
y SUM(NUM_VOT). Se elimina DES_MS (mesa). 15 columnas.

Para los archivos grandes (Congreso, Territoriales), el agregado por
puesto reduce ~85% las filas y CABE en una hoja Excel (1.048.576 max),
así que entregamos XLSX de un golpe sin partir por depto.

Stream-friendly: usa defaultdict en memoria. Para 50M filas de entrada
necesita ~2-4 GB de RAM (los agregados únicos por puesto son ~5-10M).
"""
import csv
import subprocess
import sys
import time
from collections import defaultdict
from pathlib import Path

S3_BASE = "s3://elecciones-2026/ricardoruiz.co/DESCARGAS/raw"

SRC = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos/FINAL SUBIDA GCS")
OUT = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos/output_crudos_puesto")
OUT.mkdir(parents=True, exist_ok=True)

# Todos los 25 archivos (chicos + grandes). Para los grandes el groupby
# cabe en una hoja XLSX porque agregamos por puesto, no por mesa.
ARCHIVOS = [
    ("GCS_2010PRES1V.csv",   "presidencial-1v",     2010),
    ("GCS_2014PRES1V.csv",   "presidencial-1v",     2014),
    ("GCS_2018PRES1V.csv",   "presidencial-1v",     2018),
    ("GCS_2022PRES1V.csv",   "presidencial-1v",     2022),
    ("GCS_2010PRES2V.csv",   "presidencial-2v",     2010),
    ("GCS_2014PRES2V.csv",   "presidencial-2v",     2014),
    ("GCS_2018PRES2V.csv",   "presidencial-2v",     2018),
    ("GCS_2022PRES2V.csv",   "presidencial-2v",     2022),
    ("GCS_2016PLEB.csv",     "plebiscito",          2016),
    ("GCS_2019JAL.csv",      "jal",                 2019),
    ("GCS_2023JAL.csv",      "jal",                 2023),
    ("GCS_2021CLMJ.csv",     "consejos-juventud",   2021),
    ("GCS_2025CLMJ.csv",     "consejos-juventud",   2025),
    ("GCS_2022CONSU.csv",    "consultas",           2022),
    ("GCS_2025CONSU.csv",    "consultas",           2025),
    ("GCS_2025CONSU_CAM.csv","consultas",           2025),
    ("GCS_2025CONSU_SEN.csv","consultas",           2025),
    ("GCS_2025ATIP_MAG.csv", "atipicas",            2025),
    # Grandes
    ("GCS_2014CON.csv",      "congreso",            2014),
    ("GCS_2018CON.csv",      "congreso",            2018),
    ("GCS_2022CON.csv",      "congreso",            2022),
    ("GCS_2011TER.csv",      "territoriales",       2011),
    ("GCS_2015TER.csv",      "territoriales",       2015),
    ("GCS_2019TER.csv",      "territoriales",       2019),
    ("GCS_2023TER.csv",      "territoriales",       2023),
]

# Columnas estándar GCS. DES_MS se elimina al agregar por puesto.
HEADER_MESA = [
    "FUENTE", "FEC_ELEC", "COD_COR", "DES_COR", "COD_CIR", "DES_CIR",
    "COD_DDE", "COD_MME", "COD_ZZ", "COD_PP", "DES_MS",
    "COD_PAR", "DES_PAR", "COD_CAN", "DES_CAN", "NUM_VOT",
]
HEADER_PUESTO = [c for c in HEADER_MESA if c != "DES_MS"]
KEY_COLS = HEADER_PUESTO[:-1]  # todas menos NUM_VOT
KEY_IDX_IN_MESA = [HEADER_MESA.index(c) for c in KEY_COLS]
VOT_IDX_IN_MESA = HEADER_MESA.index("NUM_VOT")

SHEET_ROWS = 1_000_000


def aggregate_to_puesto(csv_in: Path):
    """Devuelve un iterador (lazy-ish) de filas agregadas por puesto.

    Internamente arma un defaultdict en memoria — necesario para hacer
    el groupby sin pasar dos veces por el CSV. Para los archivos más
    grandes (Territoriales ~30M filas) el dict resultante tiene ~5-8M
    entradas, manejable en RAM moderna.
    """
    agg = defaultdict(int)
    with csv_in.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        header_seen = False
        for row in reader:
            if not header_seen:
                # Verifica que el header sea el esperado o se aproxime.
                # Algunos CSV antiguos (2010/2011) pueden traer las columnas en
                # otro orden — manejamos eso buscando por nombre.
                map_idx = {col: row.index(col) if col in row else None for col in HEADER_MESA}
                # Recomputar índices reales (puede que difieran del HEADER_MESA fijo).
                key_idx = [map_idx[c] for c in KEY_COLS]
                vot_idx = map_idx["NUM_VOT"]
                header_seen = True
                continue
            try:
                vot_raw = row[vot_idx] if vot_idx is not None else "0"
                vot = int(vot_raw) if vot_raw and vot_raw.strip() else 0
            except (ValueError, IndexError):
                vot = 0
            try:
                key = tuple(row[i] if i is not None else "" for i in key_idx)
            except IndexError:
                continue
            agg[key] += vot
    return agg


def write_csv(agg: dict, out_path: Path):
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(HEADER_PUESTO)
        for key, votos in agg.items():
            writer.writerow(list(key) + [votos])


def write_xlsx(agg: dict, out_path: Path):
    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Datos")
    ws.append(HEADER_PUESTO)
    rows_in_sheet = 1
    sheet_idx = 1
    for key, votos in agg.items():
        if rows_in_sheet >= SHEET_ROWS:
            sheet_idx += 1
            ws = wb.create_sheet(f"Datos {sheet_idx}")
            ws.append(HEADER_PUESTO)
            rows_in_sheet = 1
        ws.append(list(key) + [votos])
        rows_in_sheet += 1
    wb.save(out_path)


def s3_upload_and_delete(local_path: Path, cat: str, año: int):
    """Sube un archivo a S3 manteniendo la jerarquía cat/año/, y borra local."""
    fname = local_path.name
    s3_key = f"{S3_BASE}/{cat}/{año}/{fname}"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" \
        if local_path.suffix == ".xlsx" else "text/csv"
    cmd = ["aws", "s3", "cp", str(local_path), s3_key,
           "--content-type", content_type,
           "--cache-control", "public, max-age=86400",
           "--no-progress"]
    try:
        subprocess.run(cmd, check=True, capture_output=True, text=True)
        local_path.unlink()
        return True
    except subprocess.CalledProcessError as e:
        print(f"   S3 upload FAIL ({fname}): {e.stderr.strip()[:200]}", flush=True)
        return False


def main():
    skip_existing  = "--skip-existing" in sys.argv
    upload_each    = "--upload-and-delete" in sys.argv
    no_xlsx_large  = "--no-xlsx-on-large" in sys.argv  # Congreso/Territoriales sin XLSX
    LARGE_CATS = {"congreso", "territoriales"}
    # --only=NAME1.csv,NAME2.csv → procesar SOLO esos archivos
    only_set = set()
    for arg in sys.argv:
        if arg.startswith("--only="):
            only_set = {x.strip() for x in arg.split("=", 1)[1].split(",") if x.strip()}
    archivos = [a for a in ARCHIVOS if not only_set or a[0] in only_set]
    print(f"Output dir: {OUT}{' · skip-existing ON' if skip_existing else ''}{' · upload+delete ON' if upload_each else ''}{f' · only {len(archivos)} archivos' if only_set else ''}")
    total = len(archivos)
    results = []
    for idx, (csv_name, cat, año) in enumerate(archivos, 1):
        csv_in = SRC / csv_name
        if not csv_in.exists():
            print(f"[{idx}/{total}] SKIP (no existe): {csv_name}")
            continue
        out_dir = OUT / cat / str(año)
        out_dir.mkdir(parents=True, exist_ok=True)
        base = csv_name.replace(".csv", "_PUESTO")
        csv_out  = out_dir / (base + ".csv")
        xlsx_out = out_dir / (base + ".xlsx")
        if skip_existing and csv_out.exists() and xlsx_out.exists():
            print(f"[{idx}/{total}] SKIP (ya existe): {csv_name}", flush=True)
            results.append((csv_name, {"skipped": True}))
            continue
        print(f"[{idx}/{total}] {csv_name} ({csv_in.stat().st_size / (1024*1024):.0f} MB)", flush=True)
        t0 = time.time()
        try:
            skip_xlsx = no_xlsx_large and (cat in LARGE_CATS)
            agg = aggregate_to_puesto(csv_in)
            write_csv(agg, csv_out)
            if not skip_xlsx:
                write_xlsx(agg, xlsx_out)
            t = time.time() - t0
            stats = {
                "puestos": len(agg),
                "csv_mb": round(csv_out.stat().st_size / (1024 * 1024), 2),
                "xlsx_mb": round(xlsx_out.stat().st_size / (1024 * 1024), 2) if not skip_xlsx else 0,
                "sec": round(t, 1),
                "skip_xlsx": skip_xlsx,
            }
            results.append((csv_name, stats))
            xlsx_label = f"XLSX {stats['xlsx_mb']} MB" if not skip_xlsx else "XLSX skipped (cat grande)"
            print(f"   OK · {stats['puestos']:,} filas · CSV {stats['csv_mb']} MB · {xlsx_label} · {stats['sec']}s", flush=True)
            del agg
            if upload_each:
                ok_csv  = s3_upload_and_delete(csv_out, cat, año)
                if not skip_xlsx:
                    ok_xlsx = s3_upload_and_delete(xlsx_out, cat, año)
                if ok_csv:
                    print(f"   ↑ subido a S3 y liberado local", flush=True)
        except Exception as e:
            print(f"   FAIL: {e}", flush=True)
            results.append((csv_name, {"error": str(e)}))

    print("\n=== RESUMEN ===")
    for name, stats in results:
        if "error" in stats:
            print(f"  {name}: ERROR {stats['error']}")
        else:
            print(f"  {name}: {stats['puestos']:,} filas · CSV {stats['csv_mb']} MB · XLSX {stats['xlsx_mb']} MB · {stats['sec']}s")


if __name__ == "__main__":
    main()
