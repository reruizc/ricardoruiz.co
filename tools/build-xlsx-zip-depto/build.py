#!/usr/bin/env python3
"""
Genera el XLSX mesa de los archivos crudos GRANDES (Congreso, Territoriales)
particionado por departamento (COD_DDE) y empaquetado en un único ZIP.

Esos archivos no caben en una hoja Excel completa (1.048.576 filas max),
pero al partir por depto cada archivo individual sí cabe cómodamente,
incluso con margen para 2 hojas si algún depto extremo lo necesita.

Salida: GCS_<año><tipo>_BY_DEPTO.zip con 33 archivos depto_<COD_DDE>.xlsx
        adentro, paralelos al .csv en S3 con sufijo _BY_DEPTO.zip.

Soporta:
  --only=GCS_xxx.csv,GCS_yyy.csv   procesar sólo esos archivos
  --upload-and-delete              tras generar el ZIP, subir a S3 y borrar local
"""
import csv
import json
import shutil
import subprocess
import sys
import time
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

SRC = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos/FINAL SUBIDA GCS")
OUT = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos/output_xlsx_zip_depto")
OUT.mkdir(parents=True, exist_ok=True)

# Nombres oficiales dep/mun/puesto (códigos Registraduría) para las columnas
# DES_DDE / DES_MME / DES_PP que se insertan junto a sus códigos en el XLSX.
# Mismas fuentes que tools/build-csv-names/build.py.
DIVIPOLA = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos/test-presidencial/divipola.json")
GEOREF = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos/PUESTOS_GEOREF.csv")
DIVIPOL21 = Path("/Users/ricardoruiz/ricardoruiz.co/Bases de datos/Divipol 23.09.2021.xlsx")

_div = json.loads(DIVIPOLA.read_text(encoding="utf-8"))
DEP_NOMBRE = {}
MUN_NOMBRE = {}
for _d in _div["deptos"]:
    DEP_NOMBRE[_d["cod"].zfill(2)] = _d["nombre"]
    for _m in _d.get("muns", []):
        MUN_NOMBRE[f'{_d["cod"].zfill(2)}-{_m["cod"].zfill(3)}'] = _m["nombre"]

PUESTO_NOMBRE = {}
def _load_puesto_nombres():
    """Divipol 2021 como fallback; el georef 2026 (primario) lo pisa después."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(DIVIPOL21, read_only=True)
        ws = wb.active
        for i, r in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= 5:
                continue
            try:
                dd, mm, zz, pp, _dep, _mun, puesto = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
                if dd is None or puesto is None:
                    continue
                k = str(dd).zfill(2) + str(mm).zfill(3) + str(zz).zfill(2) + str(pp).zfill(2)
                PUESTO_NOMBRE[k] = str(puesto).strip()
            except Exception:
                pass
        wb.close()
    except Exception as e:
        print(f"aviso: Divipol 2021 no cargó ({e}); sigo solo con georef 2026", flush=True)
    with GEOREF.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f, delimiter=";"):
            code = (row.get("CÓDIGO COMPLETO") or "").strip()
            name = (row.get("NOMBRE PUESTO") or "").strip()
            if code and name:
                PUESTO_NOMBRE[code.zfill(9)] = name
_load_puesto_nombres()

S3_BASE = "s3://elecciones-2026/ricardoruiz.co/DESCARGAS/raw"

# Los 7 archivos grandes que requieren ZIP por depto.
ARCHIVOS = [
    ("GCS_2014CON.csv",  "congreso",      2014),
    ("GCS_2018CON.csv",  "congreso",      2018),
    ("GCS_2022CON.csv",  "congreso",      2022),
    ("GCS_2011TER.csv",  "territoriales", 2011),
    ("GCS_2015TER.csv",  "territoriales", 2015),
    ("GCS_2019TER.csv",  "territoriales", 2019),
    ("GCS_2023TER.csv",  "territoriales", 2023),
]

SHEET_ROWS = 1_000_000

# Mapeo COD_DDE → nombre de depto (códigos Registraduría). Sólo para etiquetar
# el archivo dentro del ZIP. Cualquier código fuera de este mapeo queda con
# un fallback "depto_XX".
DEPTOS = {
    "01": "ANTIOQUIA",    "03": "ATLANTICO",    "05": "BOLIVAR",
    "07": "BOYACA",       "09": "CALDAS",       "11": "CAUCA",
    "12": "CESAR",        "13": "CORDOBA",      "15": "CUNDINAMARCA",
    "17": "CHOCO",        "18": "HUILA",        "19": "LA_GUAJIRA",
    "20": "MAGDALENA",    "21": "META",         "22": "NARINO",
    "23": "NORTE_DE_SANTANDER", "25": "QUINDIO", "26": "RISARALDA",
    "27": "SANTANDER",    "28": "SUCRE",        "29": "TOLIMA",
    "30": "VALLE_DEL_CAUCA", "32": "ARAUCA",    "33": "CASANARE",
    "34": "PUTUMAYO",     "35": "SAN_ANDRES",   "36": "AMAZONAS",
    "37": "GUAINIA",      "38": "GUAVIARE",     "39": "VAUPES",
    "40": "VICHADA",      "16": "BOGOTA_DC",    "88": "CONSULADOS",
}


def s3_upload_and_delete(local_path: Path, cat: str, año: int) -> bool:
    fname = local_path.name
    s3_key = f"{S3_BASE}/{cat}/{año}/{fname}"
    content_type = "application/zip"
    cmd = ["aws", "s3", "cp", str(local_path), s3_key,
           "--content-type", content_type,
           "--cache-control", "public, max-age=86400",
           "--no-progress"]
    try:
        subprocess.run(cmd, check=True, capture_output=True, text=True)
        local_path.unlink()
        return True
    except subprocess.CalledProcessError as e:
        print(f"   S3 upload FAIL ({fname}): {e.stderr.strip()[:200]}", flush=True)
        return False


def process_file(csv_in: Path, zip_out: Path):
    """Stream CSV → defaultdict por COD_DDE → 1 XLSX por depto → ZIP."""
    # PASO 1: leer todo en memoria agrupado por COD_DDE
    print(f"   [1/3] Streaming CSV...", flush=True)
    by_depto = defaultdict(list)
    headers = None
    cod_dde_idx = None

    cod_mme_idx = None
    with csv_in.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        for i, row in enumerate(reader):
            if i == 0:
                headers = row
                try:
                    cod_dde_idx = headers.index("COD_DDE")
                except ValueError:
                    raise SystemExit(f"COD_DDE column missing in {csv_in.name}")
                try:
                    cod_mme_idx = headers.index("COD_MME")
                except ValueError:
                    cod_mme_idx = None
                try:
                    cod_zz_idx = headers.index("COD_ZZ")
                    cod_pp_idx = headers.index("COD_PP")
                except ValueError:
                    cod_zz_idx = cod_pp_idx = None
                continue
            try:
                cod_dde = row[cod_dde_idx].strip()
            except IndexError:
                cod_dde = ""
            if not cod_dde:
                cod_dde = "00"
            cod_dde = cod_dde.zfill(2)
            by_depto[cod_dde].append(tuple(row))

    n_deptos = len(by_depto)
    total_rows = sum(len(v) for v in by_depto.values())
    print(f"        {total_rows:,} filas en {n_deptos} deptos", flush=True)

    # PASO 2: generar XLSX por depto
    # Se insertan DES_DDE (depto) tras COD_DDE, DES_MME (municipio) tras
    # COD_MME y DES_PP (puesto) tras COD_PP — misma convención DES_* del
    # formato oficial (DES_COR, DES_PAR, DES_CAN...).
    ins_names = cod_mme_idx is not None and cod_mme_idx > cod_dde_idx
    ins_pp = (ins_names and cod_pp_idx is not None and cod_zz_idx is not None
              and cod_pp_idx > cod_mme_idx)
    if ins_names:
        headers_out = (headers[:cod_dde_idx+1] + ["DES_DDE"]
                       + headers[cod_dde_idx+1:cod_mme_idx+1] + ["DES_MME"])
        if ins_pp:
            headers_out += headers[cod_mme_idx+1:cod_pp_idx+1] + ["DES_PP"] + headers[cod_pp_idx+1:]
        else:
            headers_out += headers[cod_mme_idx+1:]
    else:
        headers_out = list(headers)

    tmpdir = zip_out.parent / f"_tmp_{zip_out.stem}"
    tmpdir.mkdir(parents=True, exist_ok=True)
    xlsx_files = []
    print(f"   [2/3] Generando XLSX por depto...", flush=True)
    for cod_dde in sorted(by_depto.keys()):
        rows = by_depto[cod_dde]
        depto_name = DEPTOS.get(cod_dde, f"depto_{cod_dde}")
        dep_nombre = DEP_NOMBRE.get(cod_dde, depto_name.replace("_", " ").title())
        xlsx_path = tmpdir / f"{cod_dde}_{depto_name}.xlsx"
        wb = Workbook(write_only=True)
        sheet_idx = 0
        rows_in_sheet = 0
        ws = None
        for row in rows:
            if ws is None or rows_in_sheet >= SHEET_ROWS:
                sheet_idx += 1
                ws = wb.create_sheet(f"Datos {sheet_idx}" if sheet_idx > 1 else "Datos")
                ws.append(headers_out)
                rows_in_sheet = 1
            if ins_names:
                mun_cod = (row[cod_mme_idx].strip() if cod_mme_idx < len(row) else "").zfill(3)
                mun_nombre = MUN_NOMBRE.get(f"{cod_dde}-{mun_cod}", "")
                out = (list(row[:cod_dde_idx+1]) + [dep_nombre]
                       + list(row[cod_dde_idx+1:cod_mme_idx+1]) + [mun_nombre])
                if ins_pp:
                    zz2 = (row[cod_zz_idx].strip() if cod_zz_idx < len(row) else "").zfill(2)
                    pp2 = (row[cod_pp_idx].strip() if cod_pp_idx < len(row) else "").zfill(2)
                    pp_nombre = PUESTO_NOMBRE.get(cod_dde + mun_cod + zz2 + pp2, "")
                    out += list(row[cod_mme_idx+1:cod_pp_idx+1]) + [pp_nombre] + list(row[cod_pp_idx+1:])
                else:
                    out += list(row[cod_mme_idx+1:])
            else:
                out = list(row)
            ws.append(out)
            rows_in_sheet += 1
        wb.save(xlsx_path)
        xlsx_files.append((cod_dde, xlsx_path))
        # Liberar memoria del depto procesado
        del by_depto[cod_dde]

    # PASO 3: empaquetar a ZIP
    print(f"   [3/3] Empaquetando ZIP...", flush=True)
    with zipfile.ZipFile(zip_out, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for cod_dde, xlsx in xlsx_files:
            zf.write(xlsx, arcname=xlsx.name)
            xlsx.unlink()  # liberar después de meter al zip

    # Borrar tmpdir
    try:
        tmpdir.rmdir()
    except OSError:
        shutil.rmtree(tmpdir, ignore_errors=True)

    return {
        "deptos": n_deptos,
        "filas": total_rows,
        "zip_mb": round(zip_out.stat().st_size / (1024 * 1024), 2),
    }


def main():
    upload_each = "--upload-and-delete" in sys.argv
    only_set = set()
    for arg in sys.argv:
        if arg.startswith("--only="):
            only_set = {x.strip() for x in arg.split("=", 1)[1].split(",") if x.strip()}
    archivos = [a for a in ARCHIVOS if not only_set or a[0] in only_set]
    print(f"Output dir: {OUT}{' · upload+delete ON' if upload_each else ''}{f' · only {len(archivos)} archivos' if only_set else ''}")
    total = len(archivos)
    results = []
    for idx, (csv_name, cat, año) in enumerate(archivos, 1):
        csv_in = SRC / csv_name
        if not csv_in.exists():
            print(f"[{idx}/{total}] SKIP (no existe): {csv_name}")
            continue
        out_dir = OUT / cat / str(año)
        out_dir.mkdir(parents=True, exist_ok=True)
        base = csv_name.replace(".csv", "_BY_DEPTO")
        zip_out = out_dir / (base + ".zip")
        print(f"[{idx}/{total}] {csv_name} ({csv_in.stat().st_size / (1024*1024):.0f} MB)", flush=True)
        t0 = time.time()
        try:
            stats = process_file(csv_in, zip_out)
            stats["sec"] = round(time.time() - t0, 1)
            results.append((csv_name, stats))
            print(f"   OK · {stats['deptos']} deptos · {stats['filas']:,} filas · ZIP {stats['zip_mb']} MB · {stats['sec']}s", flush=True)
            if upload_each:
                ok = s3_upload_and_delete(zip_out, cat, año)
                if ok:
                    print(f"   ↑ subido a S3 y liberado local", flush=True)
        except Exception as e:
            print(f"   FAIL: {e}", flush=True)
            import traceback; traceback.print_exc()
            results.append((csv_name, {"error": str(e)}))

    print("\n=== RESUMEN ===")
    for name, stats in results:
        if "error" in stats:
            print(f"  {name}: ERROR {stats['error']}")
        else:
            print(f"  {name}: {stats['deptos']} deptos · {stats['filas']:,} filas · {stats['zip_mb']} MB · {stats['sec']}s")


if __name__ == "__main__":
    main()
