#!/usr/bin/env python3
"""
dedup_revision.py — arma la LISTA DE REVISIÓN de personas que quedaron partidas
en varias fichas por variantes del nombre. No modifica nada: solo produce el
Excel/CSV para que Ricardo marque cuáles son la misma persona.

Dos reglas del dominio (dadas por Ricardo, jul-2026), que son las que hacen
utilizable el resultado:

1. **El nombre corto es "primer nombre + primer apellido".** Así se abrevia en
   presidenciales y en prensa: GUSTAVO PETRO = GUSTAVO (1er nombre) + PETRO
   (1er apellido) de GUSTAVO FRANCISCO PETRO URREGO. Por eso el candidato largo
   debe tener el apellido del corto en posición de PRIMER apellido (el penúltimo
   token), no como segundo apellido — ahí está la diferencia entre Petro y
   "GUSTAVO MIGUEL OSORIO PETRO", que es otra persona.

2. **Las carreras políticas son departamentales.** Nadie es diputado del
   Magdalena y luego concejal en Nariño. Si las dos fichas son de cargos
   DEPARTAMENTALES (asamblea · concejo · JAL · alcaldía · gobernación · cámara
   territorial) y los departamentos no coinciden, NO son la misma persona.
   La excepción son los cargos NACIONALES (Senado, Presidencia, consultas,
   circunscripciones especiales): ahí el departamento no aplica y la pareja
   queda para revisión humana en vez de descartarse.

El departamento sale del SLUG, que lo trae codificado en todas las fuentes que
generamos (ASAM2011-{dde}-… · CONC2019-{dde}-{mme}-… · CON2018-C-{dde}-…);
para `endoso`, que usa slugs por nombre, se cae a la circunscripción.

Uso:
  python3 tools/analisis-candidato/dedup_revision.py            # resumen
  python3 tools/analisis-candidato/dedup_revision.py --xlsx     # + Excel de revisión
"""
import csv, json, os, re, sys, unicodedata
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
BD = os.path.join(ROOT, 'Bases de datos')
OUT_DIR = os.path.join(BD, 'dedup-candidatos')
SC = os.environ.get('IDX_DIR', '/private/tmp/claude-501/'
    '-Users-ricardoruiz-ricardoruiz-co/9350b593-0e4b-4382-9e4e-6881b758fa4d/scratchpad/idx')

FUENTES_LOCALES = [
    ('asam2011', 'output_asamblea_2011/index-asamblea-2011.json'),
    ('asam2015', 'output_asamblea_2015/index-asamblea-2015.json'),
    ('asam2019', 'output_asamblea_2019/index-asamblea-2019.json'),
    ('conc2011', 'output_concejo_2011/index-concejo-2011.json'),
    ('conc2015', 'output_concejo_2015/index-concejo-2015.json'),
    ('conc2019', 'output_concejo_2019/index-concejo-2019.json'),
    ('jal2011',  'output_jal_2011/index-jal-2011.json'),
    ('jal2015',  'output_jal_2015/index-jal-2015.json'),
    ('jal2019',  'output_jal_2019/index-jal-2019.json'),
    ('pres2010', 'output_pres_2010/index-pres-2010.json'),
    ('pres2014', 'output_pres_2014/index-pres-2014.json'),
    ('pres2018', 'output_pres_2018/index-pres-2018.json'),
    ('pres2022', 'output_pres_2022/index-pres-2022.json'),
    ('consu2022','output_consu_2022/index-consu-2022.json'),
]
# Alcaldías y gobernaciones se suman en cuanto existan sus índices.
for _a in ('2011', '2015', '2019', '2023'):
    FUENTES_LOCALES.append((f'gob{_a}', f'output_gobernacion_{_a}/index-gobernacion-{_a}.json'))
    FUENTES_LOCALES.append((f'alc{_a}', f'output_alcaldia_{_a}/index-alcaldia-{_a}.json'))

FUENTES_S3 = [
    ('endoso',   'endoso_index.json'),
    ('con2014',  'congreso-2014_index-congreso-2014.json'),
    ('con2018',  'congreso-2018_index-congreso-2018.json'),
    ('con2022',  'congreso-2022_index-congreso-2022.json'),
    ('asamblea', 'asamblea-2023_index-asamblea-2023.json'),
    ('concejo',  'concejo-2023_index-concejo-2023.json'),
    ('jal',      'jal-2023_index-jal-2023.json'),
]

DEP_NAMES = {
    '01': 'ANTIOQUIA', '03': 'ATLANTICO', '05': 'BOLIVAR', '07': 'BOYACA',
    '09': 'CALDAS', '11': 'CAUCA', '12': 'CESAR', '13': 'CORDOBA',
    '15': 'CUNDINAMARCA', '17': 'CHOCO', '19': 'HUILA', '21': 'MAGDALENA',
    '23': 'NARINO', '24': 'RISARALDA', '25': 'NORTE DE SANTANDER',
    '26': 'QUINDIO', '27': 'SANTANDER', '28': 'SUCRE', '29': 'TOLIMA',
    '31': 'VALLE DEL CAUCA', '40': 'ARAUCA', '44': 'CAQUETA', '46': 'CASANARE',
    '48': 'LA GUAJIRA', '50': 'GUAINIA', '52': 'META', '54': 'GUAVIARE',
    '56': 'SAN ANDRES', '60': 'AMAZONAS', '64': 'PUTUMAYO', '68': 'VAUPES',
    '72': 'VICHADA', '88': 'EXTERIOR', '16': 'BOGOTA D.C.',
}
NACIONAL = '·NACIONAL·'
# Año de cada fuente. Las que llevan el año en la etiqueta se derivan solas;
# estas cuatro no lo traen en el nombre.
ANIO_FUENTE = {'endoso': 2026, 'asamblea': 2023, 'concejo': 2023, 'jal': 2023}
PARTIDO_RE = ('PARTIDO', 'MOVIMIENTO', 'COALICION', 'ALIANZA', 'LISTA', 'PACTO',
              'GSC', 'GRUPO', 'CONSULTA', 'VOTO', 'TARJETON', 'CANDIDATO')
PARTICULAS = {'DE', 'DEL', 'LA', 'LAS', 'LOS', 'Y', 'DA', 'DI', 'VAN', 'SAN'}

# slug → departamento
RE_TER = re.compile(r'^(?:ASAM|CONC|JAL|ALC|GOB)\d{4}-(\d+)-')
RE_CAM = re.compile(r'^CON\d{4}-C-(\d+)-')
RE_NAC = re.compile(r'^(?:PRES|CONSU)\d{4}-|^CON\d{4}-(?:S|SI|CA|CI|CE|CT)-')


def norm(s):
    s = unicodedata.normalize('NFD', s or '').encode('ascii', 'ignore').decode()
    return ' '.join(s.upper().split())


def es_partido(nombre, partido):
    n = norm(nombre)
    if not n:
        return True
    if n == norm(partido):
        return True
    return n.split()[0] in PARTIDO_RE


def sin_particulas(toks):
    return [t for t in toks if t not in PARTICULAS]


def es_subsecuencia(cortos, largos):
    it = iter(largos)
    return all(t in it for t in cortos)


def depto_de(c):
    """Departamento de una candidatura, o NACIONAL si el cargo no es territorial."""
    slug = c.get('slug') or ''
    if RE_NAC.match(slug):
        return NACIONAL
    m = RE_TER.match(slug) or RE_CAM.match(slug)
    if m:
        return DEP_NAMES.get(m.group(1).zfill(2), 'DEP ' + m.group(1))
    circ = norm(c.get('circunscripcion', ''))
    if not circ or circ == 'NACIONAL':
        return NACIONAL
    if '(' in circ:                       # "MUNICIPIO (DEPARTAMENTO)"
        return circ.split('(')[-1].rstrip(')').strip()
    return circ


def cargar():
    reg = []
    faltan = []
    for etq, rel in FUENTES_LOCALES:
        p = os.path.join(BD, rel)
        if not os.path.exists(p):
            faltan.append(etq); continue
        for c in json.load(open(p, encoding='utf-8')).get('candidatos', []):
            reg.append((etq, c))
    for etq, fn in FUENTES_S3:
        p = os.path.join(SC, fn)
        if not os.path.exists(p):
            faltan.append(etq); continue
        d = json.load(open(p, encoding='utf-8'))
        for c in (d if isinstance(d, list) else d.get('candidatos', [])):
            reg.append((etq, c))
    if faltan:
        print(f'· fuentes ausentes (se omiten): {", ".join(faltan)}', file=sys.stderr)
    return reg


LOCAL = ('CONCEJO', 'JAL', 'ASAMBLEA', 'ALCALDIA', 'ALCALDÍA')
NAC_CARGOS = ('SENADO', 'PRESIDENCIA', 'CONSULTA', 'CONSULTAS')


def confianza(corto, largo, grupo, depto_comun, por_nombre, votos, cargos):
    """ALTA/MEDIA/BAJA + motivo, para que la revisión humana sea rápida.

    El patrón de falso positivo que queda tras la regla departamental es
    siempre el mismo: un concejal minúsculo (decenas de votos) en un
    departamento cualquiera contra una figura NACIONAL (senador, presidencial).
    La RNEC no abrevia el nombre de un concejal de 40 votos — las formas cortas
    aparecen porque el personaje se INSCRIBIÓ así, y eso solo pasa con gente
    conocida. Así que "corto local y diminuto vs largo nacional" es sospechoso.
    """
    if depto_comun:
        return ('ALTA', f'mismo departamento ({depto_comun[0]})'
                + (' y los dos apellidos coinciden' if grupo == 'A' else ''))
    cc, cl = cargos(corto), cargos(largo)
    corto_local = bool(cc) and all(any(x in c for x in LOCAL) for c in cc)
    largo_nac = bool(cl) and any(any(x in c for x in NAC_CARGOS) for c in cl)
    vc, vl = votos(corto), votos(largo)
    # Solo es sospechoso si el corto es PEQUEÑO: la RNEC no abrevia el nombre de
    # un concejal de 40 votos. Si el corto es una figura regional grande
    # (gobernación, alcaldía) que luego salta a lo nacional, es lo normal de una
    # carrera política — Alan Jara gobernador del Meta y después Senado.
    if corto_local and largo_nac and vc < 5000:
        return ('BAJA', f'el corto es local y diminuto ({vc:,} votos) frente a un largo '
                        f'nacional — la RNEC no abrevia a un candidato pequeño')
    corto_nac = bool(cc) and any(any(x in c for x in NAC_CARGOS) for c in cc)
    if corto_nac:
        return ('MEDIA', 'el corto es una candidatura nacional: así se inscriben '
                         'las figuras conocidas')
    if vc >= 20000 or vl >= 20000:
        return ('MEDIA', f'sin departamento en común, pero el peso electoral '
                         f'({max(vc, vl):,} votos) es de figura regional que salta a lo nacional')
    if grupo == 'A':
        return ('MEDIA', 'coinciden los dos apellidos, pero no hay departamento en común')
    return ('BAJA', 'sin departamento en común y sin señal nacional')


def main():
    reg = cargar()
    por_nombre = defaultdict(list)
    for etq, c in reg:
        nom = c.get('nombre') or ''
        if es_partido(nom, c.get('partido', '')):
            continue
        n = norm(nom)
        if len(n.split()) < 2:
            continue
        por_nombre[n].append((etq, c))

    nombres = list(por_nombre)
    print(f'{len(reg):,} candidaturas · {len(nombres):,} nombres distintos\n')

    def deptos(nom):
        return {depto_de(c) for _, c in por_nombre[nom]}

    def votos(nom):
        return sum(c.get('votos', 0) for _, c in por_nombre[nom])

    def cargos(nom):
        return sorted({(c.get('corp') or '').split(' · ')[0] for _, c in por_nombre[nom]})

    def anios(nom):
        out = set()
        for etq, _ in por_nombre[nom]:
            if etq in ANIO_FUENTE:
                out.add(ANIO_FUENTE[etq])
            else:
                m = re.search(r'(20\d{2})$', etq)
                if m:
                    out.add(int(m.group(1)))
        return sorted(out)

    def cargo_anio(nom):
        """{(año, cargo)} — para detectar que compiten por lo mismo el mismo año."""
        out = set()
        for etq, c in por_nombre[nom]:
            a = ANIO_FUENTE.get(etq)
            if a is None:
                m = re.search(r'(20\d{2})$', etq)
                a = int(m.group(1)) if m else None
            out.add((a, (c.get('corp') or '').split(' · ')[0]))
        return out

    def choca_depto(a, b):
        """True si ambos son territoriales y NO comparten departamento."""
        da = {d for d in deptos(a) if d != NACIONAL}
        db = {d for d in deptos(b) if d != NACIONAL}
        return bool(da) and bool(db) and not (da & db)

    # Índices
    por_ape2 = defaultdict(list)    # (1er nombre, 2 apellidos) → nombres  [grupo A]
    por_penul = defaultdict(list)   # (1er nombre, 1er apellido) → largos  [grupo B]
    for n in nombres:
        t = sin_particulas(n.split())
        if len(t) >= 3:
            por_ape2[(t[0], t[-2], t[-1])].append(n)
            por_penul[(t[0], t[-2])].append(n)

    filas = []
    descartados_depto = 0
    for n in nombres:
        t = sin_particulas(n.split())
        if len(t) < 2:
            continue
        cands = []
        grupo = None
        if len(t) >= 3:   # A · coinciden los DOS apellidos
            for m in por_ape2[(t[0], t[-2], t[-1])]:
                tm = sin_particulas(m.split())
                if len(tm) > len(t) and es_subsecuencia(t, tm):
                    cands.append(('A', m))
        for m in por_penul[(t[0], t[-1])]:   # B · corto = 1er nombre + 1er apellido
            tm = sin_particulas(m.split())
            if len(tm) > len(t) and es_subsecuencia(t, tm):
                cands.append(('B', m))
        # dedup manteniendo A sobre B
        vistos, únicos = set(), []
        for g, m in cands:
            if m not in vistos:
                vistos.add(m); únicos.append((g, m))
        if not únicos:
            continue
        # Regla 2: descartar los que chocan de departamento
        vivos = []
        for g, m in únicos:
            if choca_depto(n, m):
                descartados_depto += 1
            else:
                vivos.append((g, m))
        if not vivos:
            continue
        for g, m in vivos:
            da, db = deptos(n), deptos(m)
            comp = sorted((da & db) - {NACIONAL})
            conf, motivo = confianza(n, m, g, comp, por_nombre, votos, cargos)
            ac, al = anios(n), anios(m)
            # Misma corporación el mismo año ⇒ no puede ser la misma persona.
            choque = sorted(cargo_anio(n) & cargo_anio(m))
            if choque:
                conf = 'BAJA'
                motivo = (f'compiten por lo MISMO el mismo año '
                          f'({choque[0][0]} · {choque[0][1]}) — no pueden ser la misma persona')
            filas.append({
                'confianza': conf, 'motivo': motivo,
                'anios_corto': ' '.join(str(a) for a in ac),
                'anios_largo': ' '.join(str(a) for a in al),
                'rango': f'{min(ac + al)}–{max(ac + al)}' if (ac or al) else '',
                'alerta': 'MISMO AÑO Y CARGO' if choque else '',
                'grupo': g,
                'estado': 'UNICO' if len(vivos) == 1 else 'AMBIGUO',
                'n_opciones': len(vivos),
                'nombre_corto': n, 'votos_corto': votos(n),
                'cargos_corto': ' / '.join(cargos(n)),
                'deptos_corto': ' / '.join(sorted(d for d in da if d != NACIONAL)) or 'NACIONAL',
                'nombre_largo': m, 'votos_largo': votos(m),
                'cargos_largo': ' / '.join(cargos(m)),
                'deptos_largo': ' / '.join(sorted(d for d in db if d != NACIONAL)) or 'NACIONAL',
                'depto_comun': ' / '.join(comp) or ('—' if NACIONAL in da | db else ''),
                'votos_total': votos(n) + votos(m),
            })

    uni = [f for f in filas if f['estado'] == 'UNICO']
    amb = [f for f in filas if f['estado'] == 'AMBIGUO']
    A = [f for f in uni if f['grupo'] == 'A']
    B = [f for f in uni if f['grupo'] == 'B']
    Bdep = [f for f in B if f['depto_comun'] not in ('', '—')]

    print(f'Descartados por REGLA DEPARTAMENTAL: {descartados_depto:,} pares\n')
    print(f'A · dos apellidos iguales, solo cambia el 2º nombre : {len(A):,} únicos')
    print(f'B · corto = 1er nombre + 1er apellido               : {len(B):,} únicos')
    print(f'      de esos, con DEPARTAMENTO EN COMÚN            : {len(Bdep):,}  ← los más confiables')
    print(f'ambiguos (varios candidatos largos)                 : {len(amb):,}\n')

    ORDEN = {'ALTA': 0, 'MEDIA': 1, 'BAJA': 2}
    filas.sort(key=lambda f: (ORDEN[f['confianza']], -f['votos_total']))
    for cf in ('ALTA', 'MEDIA', 'BAJA'):
        n_ = sum(1 for f in filas if f['confianza'] == cf and f['estado'] == 'UNICO')
        print(f'   confianza {cf:<6}: {n_:,} únicos')
    os.makedirs(OUT_DIR, exist_ok=True)
    cols = ['confianza', 'alerta', 'motivo', 'rango', 'grupo', 'estado', 'n_opciones',
            'nombre_corto', 'anios_corto', 'votos_corto', 'cargos_corto', 'deptos_corto',
            'nombre_largo', 'anios_largo', 'votos_largo', 'cargos_largo', 'deptos_largo',
            'depto_comun', 'votos_total']
    csv_p = os.path.join(OUT_DIR, 'revision-duplicados.csv')
    with open(csv_p, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['¿MISMA PERSONA? (SI/NO)'] + cols)
        w.writeheader()
        for r in filas:
            w.writerow({'¿MISMA PERSONA? (SI/NO)': '', **r})
    print(f'→ {csv_p}  ({len(filas):,} filas)')

    if '--xlsx' in sys.argv:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            print('  (openpyxl no disponible; queda solo el CSV)'); return
        wb = Workbook(); ws = wb.active; ws.title = 'Revisión'
        hdr = ['¿MISMA PERSONA?'] + cols
        ws.append(hdr)
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='8A1E16')
            c.alignment = Alignment(vertical='center', wrap_text=True)
        for r in filas:
            ws.append([''] + [r[k] for k in cols])
        anchos = {'A': 16, 'B': 11, 'C': 19, 'D': 48, 'E': 12,
                  'I': 30, 'J': 17, 'L': 22, 'M': 20,
                  'N': 34, 'O': 17, 'Q': 22, 'R': 20, 'S': 18}
        for col, w_ in anchos.items():
            ws.column_dimensions[col].width = w_
        ws.freeze_panes = 'B2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(hdr))}{len(filas)+1}'
        colores = {'ALTA': 'E8F5E9', 'MEDIA': 'FFF8E1', 'BAJA': 'FFEBEE'}
        for i, r in enumerate(filas, start=2):
            fill = PatternFill('solid', fgColor=colores[r['confianza']])
            for c in ws[i]:
                c.fill = fill
        xp = os.path.join(OUT_DIR, 'revision-duplicados.xlsx')
        wb.save(xp)
        print(f'→ {xp}   (verde = alta confianza: grupo A o departamento en común)')


if __name__ == '__main__':
    main()
