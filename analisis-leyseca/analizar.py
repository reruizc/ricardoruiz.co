#!/usr/bin/env python3
"""
Analisis ley seca · presidenciales 1V Colombia.
Compara delitos (homicidios + lesiones personales) en el fin de semana
electoral vs el promedio del mismo día de semana del año.

Fechas presidenciales 1V:
  2010: domingo 30 de mayo
  2014: domingo 25 de mayo
  2018: domingo 27 de mayo
  2022: domingo 29 de mayo

Ley seca histórica nacional: desde sábado 6pm hasta lunes 6am.
Bogotá 2026: ampliada al viernes 6pm (decisión Alcaldía Galán).
"""

import openpyxl
from datetime import datetime, date, timedelta
from collections import defaultdict
import unicodedata, json, os, statistics

YEARS = [2010, 2014, 2018, 2022]
ELECCION = {
    2010: date(2010, 5, 30),
    2014: date(2014, 5, 25),
    2018: date(2018, 5, 27),
    2022: date(2022, 5, 29),
}

FILES = {
    'homicidios': {
        2010: 'Homicidio Intencional 2010.xlsx',
        2014: 'Homicidio Intencional 2014.xlsx',
        2018: 'Homicidio Intencional 2018.xlsx',
        2022: 'Homicidio Intencional 2022.xlsx',
    },
    'lesiones': {
        2010: 'lesiones_personales_2010_0.xlsx',
        2014: 'lesiones_personales_2014_0.xlsx',
        2018: 'lesiones_personales_2018_0.xlsx',
        2022: 'lesiones_personales_2022.xlsx',
    },
}

# Cada archivo trae el header en una fila distinta. Detectamos por keyword.
HEADER_KEYS = {'FECHA', 'DEPARTAMENTO', 'MUNICIPIO', 'CANTIDAD'}


def norm(s):
    if s is None:
        return ''
    s = str(s).strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s


def es_bogota(depto, mun):
    d = norm(depto); m = norm(mun)
    # Bogotá aparece como mun "BOGOTA D.C. (CT)" en depto "CUNDINAMARCA" o "BOGOTA D.C."
    if 'BOGOTA' in d:
        return True
    if 'BOGOTA' in m:
        return True
    return False


def parse_fecha(v, year):
    """v puede ser datetime, str 'YYYYMMDD', o float serial Excel."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial date
        try:
            base = datetime(1899, 12, 30)
            d = base + timedelta(days=int(v))
            return d.date()
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    # try YYYYMMDD
    if len(s) == 8 and s.isdigit():
        try:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except Exception:
            return None
    # try ISO
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        pass
    return None


def cargar(path, year):
    """Devuelve dict[(date, scope)] -> count, donde scope in {'nacional','bogota'}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # detectar fila de header
    header_row = None
    header = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        vals = [norm(v) for v in row if v is not None]
        if not vals:
            continue
        joined = ' | '.join(vals)
        if 'FECHA' in joined and ('DEPARTAMENTO' in joined or 'DEPTO' in joined) and 'CANTIDAD' in joined:
            header_row = i
            header = [norm(v) for v in row]
            break
    if header_row is None:
        raise RuntimeError(f'No header detectado en {path}')

    # localizar índices
    def idx_of(*keys):
        for k in keys:
            for j, h in enumerate(header):
                if k in h:
                    return j
        return None

    i_dep = idx_of('DEPARTAMENTO', 'DEPTO')
    i_mun = idx_of('MUNICIPIO', 'MUNICIPI', 'MUNICICPIO', 'MUNIC')
    i_fec = idx_of('FECHA')
    i_cant = idx_of('CANTIDAD')

    counts = defaultdict(int)
    bad = 0
    total = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or row[i_fec] is None:
            continue
        f = parse_fecha(row[i_fec], year)
        if f is None or f.year != year:
            bad += 1
            continue
        try:
            c = int(row[i_cant]) if row[i_cant] is not None else 1
        except Exception:
            c = 1
        counts[(f, 'nacional')] += c
        if es_bogota(row[i_dep], row[i_mun]):
            counts[(f, 'bogota')] += c
        total += c
    wb.close()
    return counts, total, bad


def analizar():
    # series[year][delito][scope][date] = count
    series = {y: {} for y in YEARS}
    for delito, files in FILES.items():
        for y, path in files.items():
            print(f'· cargando {delito} {y}...', flush=True)
            counts, total, bad = cargar(path, y)
            print(f'  total nacional {y}: {total:,} (descartados fecha mala: {bad:,})', flush=True)
            series[y].setdefault(delito, defaultdict(int))
            for (f, sc), c in counts.items():
                series[y][delito][(sc, f)] = c

    return series


def baseline_dow(serie_dict, year, dow, excluir):
    """Promedio diario para un dow específico en el año, excluyendo fechas dadas."""
    vals = []
    d = date(year, 1, 1)
    end = date(year, 12, 31)
    while d <= end:
        if d.weekday() == dow and d not in excluir:
            vals.append(serie_dict.get(d, 0))
        d += timedelta(days=1)
    return statistics.mean(vals), statistics.stdev(vals), len(vals)


def reportar(series):
    DAYS_LABEL = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    out = []
    for y in YEARS:
        eleccion = ELECCION[y]
        vie = eleccion - timedelta(days=2)
        sab = eleccion - timedelta(days=1)
        dom = eleccion
        lun = eleccion + timedelta(days=1)
        excl = {vie, sab, dom, lun}
        block = {'year': y, 'eleccion': eleccion.isoformat(), 'detalle': {}}
        for delito in ['homicidios', 'lesiones']:
            # construir dict[date] -> count para cada scope
            data_nac = defaultdict(int)
            data_bog = defaultdict(int)
            for (sc, f), c in series[y][delito].items():
                if sc == 'nacional':
                    data_nac[f] += c
                else:
                    data_bog[f] += c

            block_delito = {}
            for scope_name, data in [('nacional', data_nac), ('bogota', data_bog)]:
                day_rows = []
                for label, d in [('viernes', vie), ('sabado', sab), ('domingo', dom), ('lunes', lun)]:
                    cant = data.get(d, 0)
                    mean, std, n = baseline_dow(data, y, d.weekday(), excl)
                    delta = cant - mean
                    pct = (delta / mean * 100) if mean > 0 else 0
                    day_rows.append({
                        'dia': label,
                        'fecha': d.isoformat(),
                        'cant': cant,
                        'baseline_mean': round(mean, 2),
                        'baseline_std': round(std, 2),
                        'baseline_n': n,
                        'delta': round(delta, 2),
                        'pct': round(pct, 1),
                    })
                # totales fin de semana
                total_fds = sum(r['cant'] for r in day_rows)
                base_fds = sum(r['baseline_mean'] for r in day_rows)
                block_delito[scope_name] = {
                    'dias': day_rows,
                    'total_fds': total_fds,
                    'baseline_fds': round(base_fds, 2),
                    'pct_fds': round((total_fds - base_fds) / base_fds * 100, 1) if base_fds > 0 else 0,
                }
            block['detalle'][delito] = block_delito
        out.append(block)
    return out


if __name__ == '__main__':
    series = analizar()
    rep = reportar(series)
    with open('reporte.json', 'w', encoding='utf-8') as f:
        json.dump(rep, f, indent=2, ensure_ascii=False)
    print('\n--- RESUMEN ---\n')
    for b in rep:
        y = b['year']
        print(f'== {y} (elección {b["eleccion"]}) ==')
        for delito in ['homicidios', 'lesiones']:
            for scope in ['nacional', 'bogota']:
                d = b['detalle'][delito][scope]
                print(f'  {delito:10} · {scope:8} · fds total {d["total_fds"]:>5} vs baseline {d["baseline_fds"]:>7}  ({d["pct_fds"]:+.1f}%)')
                for r in d['dias']:
                    print(f'    {r["dia"]:8} {r["fecha"]}: {r["cant"]:>4}  (baseline {r["baseline_mean"]:>5.1f} ± {r["baseline_std"]:>4.1f}, n={r["baseline_n"]:>2})  {r["pct"]:+5.1f}%')
        print()
